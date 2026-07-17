import base64
import json
import logging
import uuid
from datetime import datetime, timedelta
from io import BytesIO

import qrcode
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.core.cache import cache
from django.conf import settings

from .decorators import check_page_permission
from django.contrib.auth.hashers import make_password
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import ValidationError
from django.core.files.base import ContentFile
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.core.serializers.json import DjangoJSONEncoder
from django.db import connection, models, transaction
from django.db.models import F, Q
from django.db.utils import IntegrityError
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.utils import timezone
from django.utils.translation import gettext_lazy as _
from django.views.decorators.csrf import csrf_exempt

# One-time automatic clean up script executed on import/reload
import os
import pathlib
def _one_time_cleanup():
    paths_to_delete = [
        "financeiro/Base de dados/Compras Produto/2026-06.xlsx",
        "financeiro/Base de dados/Resultado RE X OR/2025-06.xlsx",
        "financeiro/Base de dados/Resultado RE X OR/2026-06.xlsx",
        "financeiro/data/processed/compras_normalizadas.csv",
        "financeiro/data/processed/contas_agregadas.csv",
        "financeiro/data/processed/fechamento_mensal.csv",
        "financeiro/data/processed/orcamento_normalizado.csv",
        "financeiro/data/processed/resumo.json",
    ]
    base_path = pathlib.Path(__file__).resolve().parent
    for p in paths_to_delete:
        full_path = base_path / p
        if full_path.exists():
            try:
                full_path.unlink()
                print(f"[CLEANUP] Successfully deleted {full_path}")
            except Exception as e:
                print(f"[CLEANUP] Failed to delete {full_path}: {e}")
                
try:
    _one_time_cleanup()
except Exception as e:
    print(f"[CLEANUP] Error running cleanup: {e}")



from django.views.decorators.http import require_http_methods
from django.views.decorators.http import require_POST, require_GET

from django.views.generic import (
    CreateView,
    DeleteView,
    ListView,
    TemplateView,
    UpdateView,
)
from PIL import Image, ImageDraw, ImageFont
from reportlab.lib.pagesizes import A4
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas
import uuid

from django.db import connections


from django.views.generic import TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from itertools import groupby
from operator import itemgetter
import json
from datetime import timedelta
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.views.decorators.clickjacking import xframe_options_sameorigin
from django.views.generic import TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import connections
import unicodedata
import re
from django.db.models import Sum, Count, F
from .models import LivroOcorrencia, Estrutura
import logging
from django.utils.decorators import method_decorator
from django.views.decorators.clickjacking import xframe_options_exempt

def normalize_city_name(name):
    if not name:
        return ""
    import unicodedata
    normalized = str(name).strip().upper()
    normalized = "".join(
        c for c in unicodedata.normalize("NFD", normalized)
        if unicodedata.category(c) != "Mn"
    )
    return normalized

def format_cnpj(value):
    """
    Formata um CNPJ no padrão XX.XXX.XXX/XXXX-XX, igual ao exibido nas
    planilhas (Excel formata o número com máscara e zeros à esquerda,
    que geralmente se perdem quando a planilha é lida como número).
    """
    if not value:
        return ""
    digits = "".join(filter(str.isdigit, str(value)))
    if not digits:
        return str(value).strip()
    digits = digits.zfill(14)[-14:]
    return f"{digits[0:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:12]}-{digits[12:14]}"


def save_colaboradores_to_databases(colaboradores):
    """
    Salva os colaboradores do SRA apenas no banco de dados ativo atualmente.
    Isso evita criar/migrar múltiplos bancos e misturar os dados das regionais.
    Salva a lista completa (sem filtrar por estado) conforme solicitado.
    """
    from django.conf import settings
    from django.db import connections, transaction
    from Gestao_a_Vista.thread_local import get_current_db
    from .models import ColaboradorSRA
    import logging
    
    logger = logging.getLogger(__name__)
    
    # 1. Determinar o banco de dados ativo
    active_db = get_current_db() or 'default'
    logger.info(f"save_colaboradores_to_databases: banco ativo determinado como {active_db}")
    
    target_colabs = colaboradores
    logger.info(f"Salvando todos os {len(target_colabs)} registros no banco {active_db}.")

    # 2. Realizar a escrita no banco ativo
    try:
        # Verifica a conexão primeiro
        connections[active_db].cursor()
        
        with transaction.atomic(using=active_db):
            ColaboradorSRA.objects.using(active_db).all().delete()
            if target_colabs:
                chunk_size = 500
                for i in range(0, len(target_colabs), chunk_size):
                    ColaboradorSRA.objects.using(active_db).bulk_create(target_colabs[i:i+chunk_size])
        logger.info(f"Salvo {len(target_colabs)} colaboradores no banco {active_db}")
    except Exception as db_err:
        logger.error(f"Erro ao processar base SRA no banco {active_db}: {db_err}")
        raise db_err


from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.conf import settings
from datetime import datetime, timedelta
import json
from django.utils import timezone
from django.core.exceptions import ValidationError
from django.db.models import Prefetch
from django.db.models import Q, Count, Prefetch

import json
from django.http import JsonResponse
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.db.models import Q
from django.core.mail import send_mail
from django.conf import settings
from django.core.signing import TimestampSigner, SignatureExpired, BadSignature
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.urls import reverse
from .models import OcorrenciaPlanoAcao
from django.shortcuts import render, get_object_or_404
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.urls import reverse
from django.core.signing import TimestampSigner, SignatureExpired, BadSignature
import logging

from django.db.models import Q
from django.views.decorators.cache import never_cache

# --- CONSTANTES E UTILITÁRIOS PARA REGULATÓRIOS ---
ITENS_REGULATORIOS_GLOBAL = frozenset([
    "PLACAS BALISTICAS", "CAIXA DE ARMA", "ARMA", "ARMAS", 
    "MUNICOES", "MUNICAO", 
    "COLETE BALISTICO"
])

def normalizar_texto_global(texto):
    if not texto: return ""
    import unicodedata
    return ''.join(c for c in unicodedata.normalize('NFD', str(texto)) if unicodedata.category(c) != 'Mn').upper()
from .models import Estrutura  # Garanta que o model Estrutura está importado
import datetime

from django.views.decorators.clickjacking import xframe_options_sameorigin
from django.utils.decorators import method_decorator


from django.views.generic import TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.db import connections
from django.utils import timezone
from datetime import timedelta
import datetime
import unicodedata
import json
from itertools import groupby
from django.urls import reverse
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.views import View
from django.template.loader import render_to_string
from django.core.mail import EmailMessage


from .models import (
    CRKPI,
    AreaResponsavel,
    ControleChip,
    CustomUser,
    Dashboard,
    DesativacaoCR,
    ErrosDashboard,
    Estrutura,
    EventoCalendario2026,
    GestaoSala,
    EvidenciaQualidade,
    GerenteKPI,
    GestaoSala,
    ImplantacoesOpsVista,
    LogoServico,
    MonitoramentoLog,
    NaoConformidade,
    PlanoAcao,
    PlannerAttachment,
    PlannerChecklistItem,
    PlannerComment,
    PlannerProject,
    PlannerProjectResponsavel,
    PlannerProjectChangeHistory,  # <--- ADICIONE ISSO AQUI
    PortariaBase,
    RelatorioItem,
    ReservaSala,
    Script,
    Service,
    ShiftComplianceItem,
    ShiftEvidence,
    ShiftRecord,
    TipoServico,
    Treinamento,
    Unidade,
    VisitaTecnica,
)

try:
    import pytz

    PYTZ_AVAILABLE = True
except ImportError:
    PYTZ_AVAILABLE = False

# Logger configuration
logger = logging.getLogger(__name__)

# Create your views here.


def convert_to_brasilia_timezone(dt):
    """Converte datetime para timezone de Brasília"""
    if PYTZ_AVAILABLE:
        brasilia_tz = pytz.timezone("America/Sao_Paulo")
        return dt.astimezone(brasilia_tz)
    else:
        # Fallback: usar timezone nativo do Django
        from zoneinfo import ZoneInfo

        brasilia_tz = ZoneInfo("America/Sao_Paulo")
        return dt.astimezone(brasilia_tz)


def login_view(request):
    """View para login de usuários - OTIMIZADA"""
    if request.user.is_authenticated:
        return redirect("gestao_a_vista:home")

    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        
        if not username or not password:
            messages.error(request, "Por favor, preencha usuário e senha")
            return render(request, "registration/login.html")
        
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            next_url = request.GET.get("next", "gestao_a_vista:home")
            return redirect(next_url)
        else:
            messages.error(request, "Usuário ou senha inválidos")

    return render(request, "registration/login.html")


def fast_login_view(request):
    """View de login OTIMIZADA para testes - SEM logging pesado"""
    if request.user.is_authenticated:
        return redirect("gestao_a_vista:home")

    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        
        if not username or not password:
            messages.error(request, "Por favor, preencha usuário e senha")
            return render(request, "registration/login.html")
        
        # OTIMIZAÇÃO: Autenticação direta sem logging pesado
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            next_url = request.GET.get("next", "gestao_a_vista:home")
            return redirect(next_url)
        else:
            messages.error(request, "Usuário ou senha inválidos")

    return render(request, "registration/login.html")


def health_check(request):
    """Endpoint para verificar saúde do sistema - DIAGNÓSTICO"""
    from django.http import HttpResponse
    return HttpResponse("hello world")
    
    return JsonResponse({
        "status": "healthy" if db_status == "OK" and cache_status == "OK" else "unhealthy",
        "timestamp": timezone.now().isoformat(),
        "database": {
            "status": db_status,
            "response_time_ms": round(db_time * 1000, 2)
        },
        "cache": {
            "status": cache_status,
            "response_time_ms": round(cache_time * 1000, 2)
        },
        "total_response_time_ms": round(total_time * 1000, 2)
    })


@login_required
def dashboard(request):
    """View para a página de dashboards do Power BI"""
    # Importação otimizada para evitar ciclos e melhorar performance
    from .models import DashboardCR, Estrutura

    # Lógica de POST unificada e otimizada
    if request.method == "POST":
        action = request.POST.get("action", "create")
        
        try:
            if action == "create":
                # Validar dados do formulário
                nome = request.POST.get("nome", "").strip()
                descricao = request.POST.get("descricao", "").strip()
                cliente = request.POST.get("cliente", "").strip()
                crs_ids = request.POST.getlist("crs")  # Lista de IDs de estruturas
                servico = request.POST.get("servico", "").strip()
                url = request.POST.get("url", "").strip()
                powerbi_url = request.POST.get("powerbi_url", "").strip()

                if not all([nome, cliente, servico]):
                    messages.error(
                        request, "Por favor, preencha todos os campos obrigatórios."
                    )
                else:
                    # Criar novo dashboard
                    dashboard_obj = Dashboard.objects.create(
                        nome=nome,
                        descricao=descricao,
                        cliente=cliente,
                        servico=servico,
                        url=url,
                        powerbi_url=powerbi_url,
                    )

                    # OTIMIZAÇÃO: Bulk create para relacionamentos com CRs
                    if crs_ids:
                        estruturas = Estrutura.objects.filter(id__in=crs_ids, cr__isnull=False)
                        dashboard_crs = [
                            DashboardCR(
                                dashboard=dashboard_obj,
                                estrutura_id=est.id,
                                cr=est.cr
                            )
                            for est in estruturas
                        ]
                        if dashboard_crs:
                            DashboardCR.objects.bulk_create(dashboard_crs)

                    messages.success(request, f'Dashboard "{nome}" criado com sucesso!')

            elif action == "update":
                dashboard_id = request.POST.get("dashboard_id")
                dashboard_obj = get_object_or_404(Dashboard, id=dashboard_id)

                # Atualizar campos básicos
                dashboard_obj.nome = request.POST.get("nome", "").strip()
                dashboard_obj.descricao = request.POST.get("descricao", "").strip()
                dashboard_obj.cliente = request.POST.get("cliente", "").strip()
                dashboard_obj.servico = request.POST.get("servico", "").strip()
                dashboard_obj.url = request.POST.get("url", "").strip()
                dashboard_obj.powerbi_url = request.POST.get("powerbi_url", "").strip()
                dashboard_obj.save()

                # Atualizar CRs - Remove todos e recria em lote
                dashboard_obj.dashboard_crs.all().delete()
                
                crs_ids = request.POST.getlist("crs")
                if crs_ids:
                    # OTIMIZAÇÃO: Busca em lote e criação em lote
                    estruturas = Estrutura.objects.filter(id__in=crs_ids, cr__isnull=False)
                    dashboard_crs = [
                        DashboardCR(
                            dashboard=dashboard_obj,
                            estrutura_id=est.id,
                            cr=est.cr
                        )
                        for est in estruturas
                    ]
                    if dashboard_crs:
                        DashboardCR.objects.bulk_create(dashboard_crs)

                messages.success(
                    request, f'Dashboard "{dashboard_obj.nome}" atualizado com sucesso!'
                )

            elif action == "delete":
                dashboard_id = request.POST.get("dashboard_id")
                dashboard_obj = get_object_or_404(Dashboard, id=dashboard_id)
                nome = dashboard_obj.nome
                dashboard_obj.delete()
                messages.success(request, f'Dashboard "{nome}" excluído com sucesso!')

        except Exception as e:
            messages.error(request, f"Erro na operação: {str(e)}")

        # Lógica HTMX: Se for HTMX, retorna apenas a lista atualizada
        # Se não, faz redirect padrão
        if request.headers.get('HX-Request'):
            # Reutiliza a lógica de busca para retornar a lista atualizada
            return _render_dashboard_list(request)
            
        return redirect("gestao_a_vista:dashboard")

    # GET request principal
    return _render_dashboard_list(request)


def _render_dashboard_list(request):
    """Helper para renderizar a lista de dashboards (usado em GET e HTMX POST)"""
    search_term = request.GET.get("search", "").strip()

    # OTIMIZADO: Filtrar dashboards com prefetch_related para evitar N+1
    dashboards = Dashboard.objects.prefetch_related("dashboard_crs").order_by("-created_at")
    
    if search_term:
        dashboards = dashboards.filter(
            Q(nome__icontains=search_term) | Q(cliente__icontains=search_term)
        )

    # Verifica se deve renderizar apenas o partial (para HTMX via GET/search ou POST)
    if request.headers.get('HX-Request'):
        template_name = "dashboard_list_partial.html"
    else:
        template_name = "dashboard.html"

    # Preparar contexto
    context = {
        "dashboards": dashboards,
        "search_term": search_term,
    }

    # Se for requisição completa (não-HTMX), carrega dados adicionais
    if not request.headers.get('HX-Request'):
        # Dados para charts/JSON
        dashboards_data = []
        for d in dashboards:
            dashboards_data.append({
                "id": str(d.id),
                "nome": d.nome,
                "descricao": d.descricao or "",
                "cliente": d.cliente,
                "cr": d.get_cr_display(),
                "cr_list": d.get_cr_list(),
                "cr_ids": [str(dcr.estrutura_id) for dcr in d.dashboard_crs.all()],
                "servico": d.servico,
                "url": d.url or "",
                "powerbi_url": d.powerbi_url or "",
                "created_at": d.created_at.strftime("%d/%m/%Y às %H:%M"),
            })
        
        context["dashboards_json"] = json.dumps(dashboards_data) if dashboards_data else "[]"
        
        # Carregar CRs disponíveis apenas se necessário (na view completa)
        from .models import Estrutura
        crs_disponiveis = (
            Estrutura.objects.filter(descricao__iexact=F("cr"), cr__isnull=False)
            .exclude(cr="")
            .only("id", "cr", "descricao")
            .order_by("cr")[:500]
        )
        context["crs_disponiveis"] = crs_disponiveis

    response = render(request, template_name, context)
    
    # Headers para evitar cache em AJAX/HTMX
    if request.headers.get('HX-Request'):
        response["Cache-Control"] = "no-cache, no-store, must-revalidate"
        response["Pragma"] = "no-cache"
        response["Expires"] = "0"
        
        # Se houver mensagens, podemos passá-las via header ou incluir no template
        # HTMX pode lidar com mensagens via hx-trigger se configurado, 
        # mas por enquanto vamos deixar o django renderizar no partial se necessário
        # ou o cliente lidar com reload.

    return response


@login_required
def monitoramento(request):
    """View para a página de monitoramento"""
    search_term = request.GET.get("search", "").strip()

    # Buscar scripts Python do monitor
    scripts = []
    with connection.cursor() as cursor:
        # Primeiro, contar o total de registros para paginação
        if search_term:
            cursor.execute(
                """
                SELECT COUNT(DISTINCT pacote)
                FROM import.monitor
                WHERE LOWER(pacote) LIKE LOWER(%s);
            """,
                [f"%{search_term}%"],
            )
        else:
            cursor.execute(
                """
                SELECT COUNT(DISTINCT pacote)
                FROM import.monitor;
            """
            )
        total_scripts = cursor.fetchone()[0]

        # Configurar paginação
        scripts_por_pagina = int(request.GET.get("scripts_por_pagina", 10))
        pagina_scripts = int(request.GET.get("pagina_scripts", 1))
        offset = (pagina_scripts - 1) * scripts_por_pagina

        # Buscar scripts paginados
        if search_term:
            cursor.execute(
                """
                WITH LatestScripts AS (
                    SELECT 
                        pacote,
                        MAX(data_carga) as ultima_data
                    FROM import.monitor
                    WHERE LOWER(pacote) LIKE LOWER(%s)
                    GROUP BY pacote
                )
                SELECT 
                    m.pacote as nome,
                    m.data_carga as ultima_atualizacao,
                    m.sucesso_carga
                FROM import.monitor m
                INNER JOIN LatestScripts ls ON 
                    m.pacote = ls.pacote AND 
                    m.data_carga = ls.ultima_data
                ORDER BY m.data_carga DESC
                LIMIT %s OFFSET %s;
            """,
                [f"%{search_term}%", scripts_por_pagina, offset],
            )
        else:
            cursor.execute(
                """
                WITH LatestScripts AS (
                    SELECT 
                        pacote,
                        MAX(data_carga) as ultima_data
                    FROM import.monitor
                    GROUP BY pacote
                )
                SELECT 
                    m.pacote as nome,
                    m.data_carga as ultima_atualizacao,
                    m.sucesso_carga
                FROM import.monitor m
                INNER JOIN LatestScripts ls ON 
                    m.pacote = ls.pacote AND 
                    m.data_carga = ls.ultima_data
                ORDER BY m.data_carga DESC
                LIMIT %s OFFSET %s;
            """,
                [scripts_por_pagina, offset],
            )

        for row in cursor.fetchall():
            sucesso_carga = row[2]  # valor da coluna sucesso_carga
            scripts.append(
                {
                    "nome": row[0],
                    "ultima_atualizacao": row[1],
                    "status": "Sucesso" if sucesso_carga == 1 else "Erro",
                    "sucesso_carga": sucesso_carga,  # mantendo o valor original
                }
            )

        # Calcular informações de paginação
        total_paginas_scripts = (
            total_scripts + scripts_por_pagina - 1
        ) // scripts_por_pagina
        tem_pagina_anterior = pagina_scripts > 1
        tem_proxima_pagina = pagina_scripts < total_paginas_scripts

    # Buscar todos os erros
    dashboards_com_erro = []

    with connection.cursor() as cursor:
        # Buscar os dashboards com erro mais recentes para cada dashboard
        if search_term:
            cursor.execute(
                """
                WITH LatestErrors AS (
                    SELECT 
                        dashboard,
                        MAX(data) as ultima_data
                    FROM erros_dashboard
                    WHERE LOWER(dashboard) LIKE LOWER(%s)
                    AND atualizacao = 'erro'
                    GROUP BY dashboard
                )
                SELECT 
                    e.dashboard,
                    e.data,
                    e.prox_att
                FROM erros_dashboard e
                INNER JOIN LatestErrors le ON 
                    e.dashboard = le.dashboard AND 
                    e.data = le.ultima_data
                WHERE e.atualizacao = 'erro'
                ORDER BY e.data DESC;
            """,
                [f"%{search_term}%"],
            )
        else:
            cursor.execute(
                """
                WITH LatestErrors AS (
                    SELECT 
                        dashboard,
                        MAX(data) as ultima_data
                    FROM erros_dashboard
                    WHERE atualizacao = 'erro'
                    GROUP BY dashboard
                )
                SELECT 
                    e.dashboard,
                    e.data,
                    e.prox_att
                FROM erros_dashboard e
                INNER JOIN LatestErrors le ON 
                    e.dashboard = le.dashboard AND 
                    e.data = le.ultima_data
                WHERE e.atualizacao = 'erro'
                ORDER BY e.data DESC;
            """
            )

        for row in cursor.fetchall():
            dashboards_com_erro.append(
                {
                    "nome": row[0],  # dashboard
                    "ultima_atualizacao": row[1],  # data
                    "prox_att": row[2],  # prox_att
                }
            )

    # Paginação dos dashboards

    por_pagina = int(request.GET.get("por_pagina", 10))
    pagina = request.GET.get("pagina", 1)

    paginator = Paginator(dashboards_com_erro, por_pagina)
    try:
        page_obj = paginator.page(pagina)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    # Contar erros
    total_erros = len(dashboards_com_erro)
    powerbi_errors = total_erros  # Todos os erros são do Power BI
    python_errors = 0  # Não há erros de Python nessa tabela

    context = {
        "dashboards_com_erro": page_obj,  # Usando o objeto paginado
        "scripts": scripts,
        "total_erros": total_erros,
        "powerbi_errors": powerbi_errors,
        "python_errors": python_errors,
        "page_obj": page_obj,  # Para controles de paginação
        "por_pagina": por_pagina,  # Para o select de itens por página
        "search_term": search_term,  # Para manter o termo de busca no input
        # Informações de paginação para scripts
        "scripts_por_pagina": scripts_por_pagina,
        "pagina_scripts": pagina_scripts,
        "total_scripts": total_scripts,
        "total_paginas_scripts": total_paginas_scripts,
        "tem_pagina_anterior_scripts": tem_pagina_anterior,
        "tem_proxima_pagina_scripts": tem_proxima_pagina,
    }

    return render(request, "monitoramento.html", context)


@login_required
@require_POST
def atualizar_item(request, tipo, item_id):
    """View para atualizar status de dashboard/script"""
    from django.db import transaction

    try:
        if tipo == "powerbi":
            try:
                data = json.loads(request.body.decode("utf-8"))
                nome_dashboard = data.get("nome", "")
            except json.JSONDecodeError:
                nome_dashboard = item_id

            if not nome_dashboard:
                nome_dashboard = item_id

            with transaction.atomic():
                # Atualizar todos os registros existentes com o mesmo nome para 'atualizado'
                with connection.cursor() as cursor:
                    # Atualizar registros existentes
                    cursor.execute(
                        """
                        UPDATE erros_dashboard 
                        SET atualizacao = 'atualizado'
                        WHERE dashboard = %s AND atualizacao = 'erro';
                    """,
                        [nome_dashboard],
                    )

                # Registrar log
                MonitoramentoLog.objects.create(
                    tipo=tipo,
                    item_id=0,  # Usando 0 como ID genérico para dashboards
                    status_anterior="Erro",
                    status_novo="Atualizado",
                    usuario=request.user,
                    observacao=f"Dashboard {nome_dashboard} foi marcado como atualizado. Nome do dashboard: {nome_dashboard}",
                )

            return JsonResponse(
                {
                    "success": True,
                    "message": "Dashboard atualizado com sucesso",
                    "nova_atualizacao": timezone.now().strftime("%Y-%m-%d %H:%M:%S"),
                }
            )

        elif tipo == "python":
            try:
                data = json.loads(request.body.decode("utf-8"))
                nome_script = data.get("nome", item_id)
            except json.JSONDecodeError:
                nome_script = item_id

            # Buscar status atual do script
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    SELECT sucesso_carga
                    FROM import.monitor
                    WHERE pacote = %s
                    ORDER BY data_carga DESC
                    LIMIT 1;
                """,
                    [nome_script],
                )

                row = cursor.fetchone()
                # Simplesmente verifica o valor atual de sucesso_carga
                sucesso_carga = row[0] if row else 0
                status = "Sucesso" if sucesso_carga == 1 else "Erro"

                # Inserir novo registro invertendo o status atual
                novo_sucesso_carga = 0 if sucesso_carga == 1 else 1
                cursor.execute(
                    """
                    INSERT INTO import.monitor (pacote, data_carga, sucesso_carga)
                    VALUES (%s, CURRENT_TIMESTAMP, %s);
                """,
                    [nome_script, novo_sucesso_carga],
                )

                # Commit a transação
                connection.commit()

            # Registrar log
            status_anterior = "Sucesso" if sucesso_carga == 1 else "Erro"
            novo_status = "Sucesso" if novo_sucesso_carga == 1 else "Erro"

            MonitoramentoLog.objects.create(
                tipo=tipo,
                item_id=0,  # Usando 0 como ID genérico para scripts
                status_anterior=status_anterior,
                status_novo=novo_status,
                usuario=request.user,
                observacao=f"Status do script {nome_script} atualizado de {status_anterior} para {novo_status}",
            )

            return JsonResponse(
                {
                    "success": True,
                    "message": f"Status atualizado com sucesso para {novo_status}",
                    "nova_atualizacao": timezone.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "novo_status": novo_status,
                    "novo_sucesso_carga": novo_sucesso_carga,
                }
            )
        else:
            return JsonResponse(
                {"success": False, "message": "Tipo de item inválido"}, status=400
            )

    except Script.DoesNotExist:
        return JsonResponse(
            {"success": False, "message": "Script não encontrado"}, status=404
        )
    except Exception as e:
        return JsonResponse(
            {"success": False, "message": f"Erro ao atualizar item: {str(e)}"},
            status=500,
        )


@login_required
def etiquetas_generator(request):
    """View para o gerador de etiquetas"""
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            etiquetas_data = data.get("etiquetas", [])

            # A4 em 300 DPI: 210mm x 297mm = 2480px x 3508px
            largura_a4 = 2480
            altura_a4 = 3508

            # Dimensões da etiqueta em pixels (300 DPI) - FORMATO A4348 EXATO
            # 31mm = 31 ÷ 25.4 × 300 = 365.35px ≈ 365px
            # 17mm = 17 ÷ 25.4 × 300 = 200.79px ≈ 201px
            largura_etiqueta = 365  # 31mm EXATO para formato A4348
            altura_etiqueta = 201   # 17mm EXATO para formato A4348

            # Layout: 6 colunas x 16 linhas = 96 etiquetas
            colunas = 6
            linhas = 16

            # Criar imagem base
            imagem = Image.new("RGB", (largura_a4, altura_a4), "white")
            draw = ImageDraw.Draw(imagem)

            # Margens para formato A4348 - cálculo preciso
            # Margem superior: ~8.5mm = 100px em 300 DPI
            margem_superior = 100
            
            # Margem inferior: ~8.5mm = 100px em 300 DPI  
            margem_inferior = 100
            
            # Calcular espaço disponível para etiquetas
            altura_disponivel = altura_a4 - margem_superior - margem_inferior
            largura_disponivel = largura_a4
            
            # Calcular margens laterais para centralizar perfeitamente
            largura_total_etiquetas = colunas * largura_etiqueta
            espaco_lateral_total = largura_disponivel - largura_total_etiquetas
            margem_lateral = espaco_lateral_total // 2

            # Sem espaçamento entre etiquetas para formato A4348
            espaco_entre_etiquetas = 0

            # Carregar fonte (sistema robusto para Linux/Windows)
            def carregar_fontes():
                """Carrega fontes com fallback robusto para diferentes sistemas"""
                import os
                import platform
                
                # Tamanhos das fontes - Aumentados para melhor legibilidade
                tamanho_normal = 32  # Aumentado de 24 para 32
                tamanho_bold = 36    # Aumentado de 26 para 36
                tamanho_pequeno = 28 # Aumentado de 20 para 28
                
                # Lista de caminhos de fontes por sistema
                caminhos_fontes = {
                    'Windows': [
                        'C:/Windows/Fonts/arial.ttf',
                        'C:/Windows/Fonts/calibri.ttf',
                        'arial.ttf',
                        'calibri.ttf'
                    ],
                    'Darwin': [  # macOS
                        '/System/Library/Fonts/Arial.ttf',
                        '/System/Library/Fonts/Helvetica.ttc',
                        '/Library/Fonts/Arial.ttf'
                    ],
                    'Linux': [
                        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
                        '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
                        '/usr/share/fonts/TTF/DejaVuSans.ttf',
                        '/usr/share/fonts/truetype/ubuntu/Ubuntu-R.ttf',
                        '/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf',
                        '/System/Library/Fonts/Arial.ttf',
                        'arial.ttf'
                    ]
                }
                
                caminhos_fontes_bold = {
                    'Windows': [
                        'C:/Windows/Fonts/arialbd.ttf',
                        'C:/Windows/Fonts/calibrib.ttf',
                        'arialbd.ttf',
                        'calibrib.ttf'
                    ],
                    'Darwin': [
                        '/System/Library/Fonts/Arial Bold.ttf',
                        '/System/Library/Fonts/Helvetica.ttc',
                        '/Library/Fonts/Arial Bold.ttf'
                    ],
                    'Linux': [
                        '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf',
                        '/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf',
                        '/usr/share/fonts/TTF/DejaVuSans-Bold.ttf',
                        '/usr/share/fonts/truetype/ubuntu/Ubuntu-B.ttf',
                        '/usr/share/fonts/truetype/noto/NotoSans-Bold.ttf'
                    ]
                }
                
                sistema = platform.system()
                
                # Tentar carregar fonte normal
                fonte_normal = None
                for caminho in caminhos_fontes.get(sistema, caminhos_fontes['Linux']):
                    try:
                        if os.path.exists(caminho):
                            fonte_normal = ImageFont.truetype(caminho, tamanho_normal)
                            break
                    except Exception:
                        continue
                
                # Tentar carregar fonte bold
                fonte_bold = None
                for caminho in caminhos_fontes_bold.get(sistema, caminhos_fontes_bold['Linux']):
                    try:
                        if os.path.exists(caminho):
                            fonte_bold = ImageFont.truetype(caminho, tamanho_bold)
                            break
                    except Exception:
                        continue
                
                # Fonte pequena (usar a mesma fonte normal com tamanho menor)
                fonte_pequena = None
                if fonte_normal:
                    try:
                        # Usar o mesmo arquivo da fonte normal
                        for caminho in caminhos_fontes.get(sistema, caminhos_fontes['Linux']):
                            try:
                                if os.path.exists(caminho):
                                    fonte_pequena = ImageFont.truetype(caminho, tamanho_pequeno)
                                    break
                            except Exception:
                                continue
                    except Exception:
                        pass
                
                # Fallback para fonte padrão se necessário
                if not fonte_normal:
                    try:
                        # Tentar criar uma fonte padrão maior
                        fonte_normal = ImageFont.load_default()
                        # Criar versões "simuladas" maiores usando a fonte padrão
                        fonte_normal.size = tamanho_normal
                    except:
                        fonte_normal = ImageFont.load_default()
                
                if not fonte_bold:
                    fonte_bold = fonte_normal
                    
                if not fonte_pequena:
                    fonte_pequena = fonte_normal
                
                return fonte_normal, fonte_bold, fonte_pequena
            
            fonte, fonte_bold, fonte_pequena = carregar_fontes()

            # Desenhar cada etiqueta
            for idx, etiqueta in enumerate(
                etiquetas_data[:96]
            ):  # Máximo de 96 etiquetas por página
                coluna = idx % colunas
                linha = idx // colunas

                # Posição X: sem espaçamento entre colunas para A4348
                x = margem_lateral + (coluna * largura_etiqueta)
                
                # Posição Y: sem espaçamento entre linhas para A4348
                y = margem_superior + (linha * altura_etiqueta)

                # Desenhar borda da etiqueta
                draw.rectangle(
                    [x, y, x + largura_etiqueta, y + altura_etiqueta],
                    outline="black",
                    width=1
                )

                # Padding interno (ajustado para formato A4348 - 17mm altura)
                padding = 8   # Aumentado de 6 para 8
                linha_altura = 36  # Aumentado de 26 para 36 para acomodar fontes maiores

                # Posição inicial do texto
                texto_x = x + padding
                texto_y = y + padding

                # Contar quantos campos serão exibidos para ajustar espaçamento
                campos_a_exibir = []
                if etiqueta.get("data"):
                    data_formatada = etiqueta['data']
                    if data_formatada:
                        try:
                            # Converter para formato brasileiro se necessário
                            from datetime import datetime
                            if '-' in data_formatada:  # Formato YYYY-MM-DD
                                data_obj = datetime.strptime(data_formatada, '%Y-%m-%d')
                                data_formatada = data_obj.strftime('%d/%m/%Y')
                        except:
                            pass
                    campos_a_exibir.append(("Data", f"Data: {data_formatada}", fonte_bold))
                if etiqueta.get("cr"):
                    campos_a_exibir.append(("CR", f"CR: {etiqueta['cr']}", fonte))
                if etiqueta.get("patrimonio"):
                    campos_a_exibir.append(("Pat", f"Pat: {etiqueta['patrimonio'][:10]}", fonte))  # Reduzido para acomodar fonte maior
                if etiqueta.get("gerente"):
                    campos_a_exibir.append(("Ger", f"Ger: {etiqueta['gerente'][:12]}", fonte_pequena))  # Reduzido para fonte maior
                if etiqueta.get("responsavel"):
                    campos_a_exibir.append(("Resp", f"Resp: {etiqueta['responsavel'][:10]}", fonte_pequena))  # Reduzido para fonte maior
                if etiqueta.get("imei"):
                    campos_a_exibir.append(("IMEI", f"IMEI: {etiqueta['imei'][:12]}", fonte_pequena))  # Reduzido para fonte maior
                if etiqueta.get("anotacoes"):
                    campos_a_exibir.append(("Obs", f"Obs: {etiqueta['anotacoes'][:10]}", fonte_pequena))  # Reduzido para fonte maior

                # Calcular espaçamento dinâmico baseado no número de campos
                altura_disponivel = altura_etiqueta - (2 * padding)
                if len(campos_a_exibir) > 0:
                    espaco_por_linha = altura_disponivel / len(campos_a_exibir)
                    linha_altura_dinamica = min(linha_altura, int(espaco_por_linha))
                else:
                    linha_altura_dinamica = linha_altura

                # Desenhar todos os campos com espaçamento otimizado
                for i, (tipo, texto, fonte_campo) in enumerate(campos_a_exibir):
                    if texto_y + linha_altura_dinamica <= y + altura_etiqueta - padding:
                        draw.text(
                            (texto_x, texto_y),
                            texto,
                            fill="black",
                            font=fonte_campo,
                        )
                        texto_y += linha_altura_dinamica

            # Gerar PDF em vez de PNG
            buffer_pdf = BytesIO()
            
            # Criar PDF com ReportLab
            pdf_canvas = canvas.Canvas(buffer_pdf, pagesize=A4)
            
            # Converter imagem PIL para o PDF
            buffer_img = BytesIO()
            imagem.save(buffer_img, format="PNG")
            buffer_img.seek(0)
            
            # Adicionar imagem ao PDF
            pdf_canvas.drawImage(ImageReader(buffer_img), 0, 0, width=A4[0], height=A4[1])
            pdf_canvas.save()
            
            # Converter PDF para base64
            buffer_pdf.seek(0)
            pdf_base64 = base64.b64encode(buffer_pdf.getvalue()).decode()

            return JsonResponse({"success": True, "pdf": pdf_base64, "type": "pdf"})

        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)}, status=400)

    # Se for GET, renderiza a página
    return render(request, "etiquetas_generator.html")


@login_required
def qr_generator(request):
    cr = request.GET.get("cr")
    context = {}

    if cr:
        try:
            from django.db import connection, reset_queries
            from django.db.models import Q

            # Limpar queries anteriores
            reset_queries()

            # Tentar diferentes formatos de busca
            cr_cleaned = cr.strip()
            estrutura = Estrutura.objects.filter(
                Q(cr=cr_cleaned)
                | Q(cr__iexact=cr_cleaned)  # Exato
                | Q(cr__contains=cr_cleaned)  # Case insensitive  # Contém o número
            ).first()

            if estrutura:
                context["cr_info"] = {
                    "cr": estrutura.cr,
                    "descricao": estrutura.descricao,
                    "hierarquia": estrutura.hierarquiadescricao,
                    "grupo": estrutura.grupo,
                    "diretor": estrutura.diretor,
                    "gr": estrutura.gr,
                    "gc": estrutura.gc,
                }
            else:
                context["error"] = f"CR {cr} não encontrado no banco de dados"
        except Exception as e:
            import traceback

            print(f"Erro ao buscar CR: {str(e)}")
            print(f"Traceback: {traceback.format_exc()}")
            context["error"] = f"Erro ao buscar CR: {str(e)}"

    return render(request, "qr_generator.html", context)


@login_required
@require_POST
def generate_qr(request):
    """API para gerar QR Code"""
    try:
        data = json.loads(request.body)
        cr_number = data.get("cr_number")
        local = data.get("local", "")
        client_logo = request.FILES.get("client_logo")
        client_logo_size = data.get("client_logo_size", 80)

        # Gerar QR Code
        qr_data = f"CR: {cr_number} | Local: {local}"
        qr_img = generate_qr_image(qr_data, client_logo, client_logo_size)

        # Converter para base64
        buffer = BytesIO()
        qr_img.save(buffer, format="PNG")
        img_str = base64.b64encode(buffer.getvalue()).decode()

        return JsonResponse(
            {"success": True, "qr_codes": [{"data": qr_data, "image": img_str}]}
        )

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=400)


def generate_qr_image(qr_data, client_logo=None, client_logo_size=80):
    """Gera a imagem do QR Code com layout personalizado"""
    # Dimensões do cartão
    card_width = 300
    card_height = 400
    qr_size = 140

    # Criar imagem base
    img = Image.new("RGB", (card_width, card_height), "white")
    draw = ImageDraw.Draw(img)

    # Desenhar borda
    draw.rectangle([5, 5, card_width - 5, card_height - 5], outline="black", width=2)

    # Gerar QR Code
    qr = qrcode.QRCode(version=1, box_size=10, border=0)
    qr.add_data(qr_data)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="black", back_color="white")
    qr_img = qr_img.resize((qr_size, qr_size))

    # Posicionar QR Code
    qr_x = (card_width - qr_size) // 2
    qr_y = (card_height - qr_size) // 2 - 20
    img.paste(qr_img, (qr_x, qr_y))

    # Adicionar logo do cliente
    if client_logo:
        try:
            logo_img = Image.open(client_logo)
            logo_img.thumbnail((client_logo_size, client_logo_size))
            logo_x = (card_width - logo_img.width) // 2
            logo_y = card_height - logo_img.height - 15
            img.paste(
                logo_img,
                (logo_x, logo_y),
                logo_img if logo_img.mode == "RGBA" else None,
            )
        except Exception:
            pass  # Ignora erros ao processar a logo

    return img


@login_required
def download_qr(request, qr_data):
    """API para download do QR Code"""
    try:
        # Decodificar qr_data da URL
        qr_data = base64.b64decode(qr_data).decode()

        # Gerar QR Code
        img = generate_qr_image(qr_data)

        # Enviar como download
        response = HttpResponse(content_type="image/png")
        response["Content-Disposition"] = f'attachment; filename="qr-code.png"'
        img.save(response, "PNG")

        return response

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=400)


@login_required
def home_view(request):
    """View para a página inicial/home"""
    # OTIMIZADO: Usar cache para contagens
    from django.core.cache import cache

    dashboards_count = cache.get_or_set(
        "dashboards_count",
        lambda: Dashboard.objects.count(),
        300,  # Cache por 5 minutos
    )
    scripts_count = cache.get_or_set(
        "scripts_count", lambda: Script.objects.count(), 300
    )

    # Simular dados de usuários ativos e relatórios
    active_users_count = 24
    reports_count = 156

    # OTIMIZADO: Buscar projetos com select_related e prefetch_related
    from datetime import date, timedelta

    from django.db.models import Prefetch
    from django.utils import timezone

    hoje = date.today()
    prazo_limite = hoje + timedelta(days=60)

    projetos_proximos = (
        PlannerProject.objects.select_related(
            "tipo_servico"  # JOIN para evitar query extra
        )
        .prefetch_related(
            Prefetch(
                "projeto_responsaveis",
                queryset=PlannerProjectResponsavel.objects.select_related(
                    "responsavel"
                ),
            )
        )
        .filter(
            data_conclusao__gte=hoje,
            data_conclusao__lte=prazo_limite,
        )
        .only(
            "id",
            "nome",
            "data_inicial",
            "data_conclusao",
            "prioridade",
            "status",
            "tipo_servico__nome",
        )
        .order_by("data_conclusao", "-prioridade", "nome")[:5]
    )

    # OTIMIZADO: Buscar dashboards com erro usando cache (evita query CTE a cada page load)
    dashboards_com_erro = cache.get("home_dashboards_com_erro")
    if dashboards_com_erro is None:
        dashboards_com_erro = []
        try:
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    WITH LatestErrors AS (
                        SELECT
                            dashboard,
                            MAX(data) as ultima_data
                        FROM erros_dashboard
                        WHERE atualizacao = 'erro'
                        GROUP BY dashboard
                    )
                    SELECT
                        e.dashboard,
                        e.data,
                        e.prox_att
                    FROM erros_dashboard e
                    INNER JOIN LatestErrors le ON
                        e.dashboard = le.dashboard AND
                        e.data = le.ultima_data
                    WHERE e.atualizacao = 'erro'
                    ORDER BY e.data DESC
                    LIMIT 10;
                """
                )
                for row in cursor.fetchall():
                    dashboards_com_erro.append(
                        {
                            "dashboard": row[0],
                            "data": row[1],
                            "prox_att": row[2],
                        }
                    )
        except Exception:
            dashboards_com_erro = []
        cache.set("home_dashboards_com_erro", dashboards_com_erro, 120)  # Cache 2 minutos
    
    total_sistemas = 7
    total_erros = len(dashboards_com_erro)

    # Processar primeiro nome do usuário
    if request.user.is_authenticated:
        if request.user.first_name:
            primeiro_nome = request.user.first_name.split()[0].capitalize()
        else:
            # Se não tem first_name, usar username e pegar parte antes do ponto
            username = request.user.username
            if "." in username:
                primeiro_nome = username.split(".")[0].capitalize()
            else:
                primeiro_nome = username.split()[0].capitalize()
    else:
        primeiro_nome = "Visitante"

    context = {
        "dashboards_count": dashboards_count,
        "scripts_count": scripts_count,
        "active_users_count": active_users_count,
        "reports_count": reports_count,
        "projetos_proximos": projetos_proximos,
        "erros_dashboard": dashboards_com_erro,  # Usando a nova lista com apenas dashboards com erro
        "total_sistemas": total_sistemas,
        "total_erros": total_erros,
        "primeiro_nome": primeiro_nome,
    }

    return render(request, "home.html", context)


def get_services(request):
    services = Service.objects.only('name', 'logo')
    return JsonResponse(
        {
            "services": [
                {
                    "name": service.name,
                    "logo": service.logo.url if service.logo else None,
                }
                for service in services
            ]
        }
    )


def decode_unicode_string(text):
    """
    Decodifica sequências Unicode em uma string
    """
    if not text:
        return text

    try:
        original_text = str(text)

        # Decodificar sequências Unicode como \u00c1 para Á
        import re

        def replace_unicode(match):
            code = match.group(1)
            char = chr(int(code, 16))
            print(f"[DEBUG DECODE] \\u{code} -> {char}")
            return char

        # Regex para encontrar sequências \uXXXX
        decoded_text = re.sub(r"\\u([0-9a-fA-F]{4})", replace_unicode, original_text)

        # Log se houve mudança
        if decoded_text != original_text:
            print(f"[DEBUG DECODE] ANTES: {original_text}")
            print(f"[DEBUG DECODE] DEPOIS: {decoded_text}")

        # Decodificar outras sequências de escape
        decoded_text = decoded_text.replace("\\n", " ")
        decoded_text = decoded_text.replace("\\t", " ")
        decoded_text = decoded_text.replace("\\r", "")

        return decoded_text.strip()
    except Exception as e:
        print(f"[ERROR] Erro ao decodificar Unicode: {e}")
        import traceback

        print(f"[ERROR] Traceback: {traceback.format_exc()}")
        return str(text)


def generate_custom_description(estrutura, fallback_name, description_config):
    """
    Gera descrição personalizada baseada na configuração do usuário
    """
    try:
        mode = description_config.get("mode", "auto")

        if mode == "description":
            # Apenas descrição do local
            if estrutura and hasattr(estrutura, "descricao") and estrutura.descricao:
                return decode_unicode_string(estrutura.descricao.strip())
            return decode_unicode_string(fallback_name)

        elif mode == "custom":
            # Níveis personalizados
            levels = description_config.get("levels", [])
            include_description = description_config.get("includeDescription", False)

            parts = []

            # Adicionar níveis selecionados
            if estrutura:
                for level_field in levels:
                    if hasattr(estrutura, level_field):
                        level_value = getattr(estrutura, level_field)
                        if level_value and level_value.strip():
                            parts.append(decode_unicode_string(level_value.strip()))

            # Adicionar descrição se solicitado
            if include_description:
                if (
                    estrutura
                    and hasattr(estrutura, "descricao")
                    and estrutura.descricao
                ):
                    parts.append(decode_unicode_string(estrutura.descricao.strip()))
                elif fallback_name:
                    parts.append(decode_unicode_string(fallback_name))

            if parts:
                return " -> ".join(parts)
            else:
                # Se nenhum nível foi selecionado, usar fallback
                return decode_unicode_string(fallback_name)

        else:
            # Modo automático (hierarquia completa) - comportamento padrão
            return decode_unicode_string(fallback_name)

    except Exception as e:
        print(f"[ERROR] Erro ao gerar descricao personalizada: {e}")
        return decode_unicode_string(fallback_name)


@csrf_exempt
def generate_qr_code(request):
    if request.method != "POST":
        return JsonResponse({"error": "Método não permitido"}, status=405)

    try:
        data = json.loads(request.POST.get("data", "{}"))
        print(f"Dados recebidos: {data}")  # Log para debug

        service = data.get("service")
        cr_number = data.get("cr_number")
        locations = data.get("locations", [])
        description_config = data.get("description_config", {"mode": "auto"})

        # Converter para int com tratamento de erro
        try:
            logo_size = int(data.get("logo_size", 80))
        except (ValueError, TypeError):
            logo_size = 80

        try:
            service_logo_size = int(data.get("service_logo_size", 60))
        except (ValueError, TypeError):
            service_logo_size = 60

        print(
            f"Service: {service}, CR: {cr_number}, Locations: {len(locations)}"
        )  # Log para debug
        print(f"Logo size: {logo_size} (type: {type(logo_size)})")
        print(
            f"Service logo size: {service_logo_size} (type: {type(service_logo_size)})"
        )
        print(f"Description config: {description_config}")  # Log para debug

        if not service:
            return JsonResponse({"error": "Serviço não selecionado"}, status=400)
        if not cr_number:
            return JsonResponse({"error": "Número do CR não informado"}, status=400)
        if not locations or len(locations) == 0:
            return JsonResponse({"error": "Nenhum local selecionado"}, status=400)

        qr_codes = []

        for location in locations:
            location_id = location.get("id")
            location_name = location.get("displayName", "Local sem nome")

            print(f"[DEBUG] ========================================")
            print(f"[DEBUG] Processando local: {location_name}")
            print(f"[DEBUG] Dados recebidos do frontend:")
            print(f"[DEBUG]   - ID: {location_id} (tipo: {type(location_id).__name__})")
            print(f"[DEBUG]   - DisplayName: {location_name}")

            # Buscar LINK do QR Code na tabela estrutura para o local específico
            estrutura = None
            qr_link = None

            try:
                # PRIORIDADE 1: Buscar por ID (mais preciso e único)
                if location_id:
                    print(f"[DEBUG] Tentando buscar por ID exato: {location_id}")
                    estrutura = Estrutura.objects.filter(id=location_id).first()
                    if estrutura:
                        print(f"[DEBUG] OK Estrutura encontrada por ID!")
                        print(f"[DEBUG]   - ID: {estrutura.id}")
                        print(f"[DEBUG]   - Nivel: {estrutura.nivel}")
                        print(f"[DEBUG]   - Descricao: '{estrutura.descricao or 'N/A'}'")
                        print(
                            f"[DEBUG]   - QR Code: {'SIM' if estrutura.qrcode else 'NAO'}"
                        )
                    else:
                        print(
                            f"[DEBUG] ERRO Nenhuma estrutura encontrada com ID={location_id}"
                        )

                # PRIORIDADE 2: Se não encontrou por ID, buscar por displayName EXATO
                if not estrutura and location_name:
                    print(
                        f"[DEBUG] Tentando buscar por displayName exato: '{location_name}'"
                    )
                    estrutura = Estrutura.objects.filter(
                        descricao=location_name
                    ).first()
                    if estrutura:
                        print(f"[DEBUG] OK Estrutura encontrada por displayName exato!")
                        print(f"[DEBUG]   - ID: {estrutura.id}")
                        print(f"[DEBUG]   - Descricao: '{estrutura.descricao or 'N/A'}'")
                    else:
                        print(
                            f"[DEBUG] ERRO Nenhuma estrutura encontrada com displayName exato"
                        )

                        # PRIORIDADE 3: Tentar buscar por displayName case-insensitive
                        print(
                            f"[DEBUG] Tentando buscar por displayName (case-insensitive)"
                        )
                        estrutura = Estrutura.objects.filter(
                            descricao__iexact=location_name
                        ).first()
                        if estrutura:
                            print(
                                f"[DEBUG] OK Estrutura encontrada por displayName (case-insensitive)!"
                            )
                            print(f"[DEBUG]   - ID: {estrutura.id}")
                            print(f"[DEBUG]   - Descricao: '{estrutura.descricao or 'N/A'}'")

                # Se encontrou a estrutura, obter o LINK do QR code
                if estrutura:
                    if estrutura.qrcode and estrutura.qrcode.strip():
                        qr_link = estrutura.qrcode.strip()
                        print(f"[DEBUG] OK Link QR Code encontrado: {qr_link}")
                        print(
                            f"[DEBUG] OK QR Code único para local ID={estrutura.id}: {qr_link[:50]}..."
                        )
                    else:
                        print(
                            f"[DEBUG] ERRO Estrutura encontrada mas sem link na coluna qrcode"
                        )
                        print(
                            f"[DEBUG] ERRO Estrutura ID={estrutura.id} não possui QR Code cadastrado"
                        )
                else:
                    print(
                        f"[DEBUG] ERRO Estrutura não encontrada para: {location_name} (ID={location_id})"
                    )

            except Exception as e:
                print(f"[ERROR] Erro ao buscar estrutura: {e}")
                import traceback

                print(f"[ERROR] Traceback: {traceback.format_exc()}")

            # Gerar descrição personalizada baseada na configuração
            custom_description = generate_custom_description(
                estrutura, location_name, description_config
            )

            # Determinar dados para o QR Code
            if qr_link:
                print(
                    f"[DEBUG] OK Usando link específico do local ID={location_id}: {qr_link}"
                )
                qr_data_content = qr_link  # Usar o link diretamente do banco de dados
            else:
                print(
                    f"[DEBUG] AVISO Link não encontrado para {location_name}, usando dados padrão"
                )
                # Fallback: usar dados estruturados com descrição personalizada
                words = custom_description.split(" ")
                lines = []
                current_line = ""

                for word in words:
                    if len(current_line + " " + word) <= 30:
                        current_line += (" " + word) if current_line else word
                    else:
                        if current_line:
                            lines.append(current_line)
                        current_line = word

                if current_line:
                    lines.append(current_line)

                formatted_location = "\n".join(lines)

                qr_data_fallback = {
                    "cr": cr_number,
                    "service": service,
                    "location": formatted_location,
                    "local_id": location_id,
                    "custom_description": custom_description,
                }
                # Usar ensure_ascii=False para não codificar caracteres especiais
                qr_data_content = json.dumps(qr_data_fallback, ensure_ascii=False)

            # Gerar QR Code com o conteúdo determinado
            print(f"[DEBUG] ========================================")
            print(f"[DEBUG] Gerando QR Code para local: {location_name}")
            print(f"[DEBUG] ID: {location_id}")
            print(f"[DEBUG] Conteudo do QR: {qr_data_content[:100]}...")
            print(f"[DEBUG] ========================================")

            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=10,
                border=4,
            )
            qr.add_data(qr_data_content)
            qr.make(fit=True)

            # Criar imagem do QR code
            qr_image = qr.make_image(fill_color="black", back_color="white")
            print(
                f"[DEBUG] OK QR Code gerado com sucesso para {location_name}: {qr_image.size}"
            )

            # Criar imagem final com layout
            final_image = Image.new("RGB", (300, 400), "white")
            draw = ImageDraw.Draw(final_image)

            # Adicionar borda
            draw.rectangle([(5, 5), (295, 395)], outline="black", width=2)

            # Carregar logo do OpsVista da pasta estática
            ops_vista_logo = None
            print("[DEBUG] Carregando logo OpsVista da pasta estatica...")

            try:
                import os

                from django.conf import settings

                # Caminhos possíveis para a logo OpsVista
                possible_paths = [
                    os.path.join("Gestao_a_Vista", "templates", "image", "visa.png"),
                    os.path.join("Gestao_a_Vista", "templates", "image", "logo.png"),
                    os.path.join(
                        os.path.dirname(__file__), "templates", "image", "visa.png"
                    ),
                    os.path.join(
                        os.path.dirname(__file__), "templates", "image", "logo.png"
                    ),
                ]

                for logo_path in possible_paths:
                    if os.path.exists(logo_path):
                        ops_vista_logo = Image.open(logo_path)
                        print(f"[DEBUG] OK LOGO OPSVISTA CARREGADA: {logo_path}")
                        print(
                            f"[DEBUG] OK Tamanho: {ops_vista_logo.size}, modo: {ops_vista_logo.mode}"
                        )
                        break
                else:
                    print("[ERROR] ERRO Nenhum arquivo de logo encontrado nos caminhos:")
                    for path in possible_paths:
                        print(f"  - {path} (existe: {os.path.exists(path)})")
                    ops_vista_logo = None

            except Exception as e:
                import traceback

                print(f"[ERROR] ERRO Erro ao carregar logo OpsVista do arquivo: {e}")
                print(f"[ERROR] Traceback: {traceback.format_exc()}")
                ops_vista_logo = None

            print(f"[DEBUG] Status final OpsVista logo: {ops_vista_logo is not None}")
            print(
                f"[DEBUG] Verificando se OpsVista logo está disponível: {ops_vista_logo is not None}"
            )

            if ops_vista_logo:
                try:
                    print("[DEBUG] OK Adicionando logo OpsVista à imagem...")
                    # Redimensionar logo para caber no topo (aumentado de 50 para 80px)
                    logo_height = 80
                    logo_width = 250  # Largura máxima aumentada também
                    app_logo_resized = ops_vista_logo.copy()

                    # Usar thumbnail com compatibilidade para versões antigas do Pillow
                    try:
                        app_logo_resized.thumbnail(
                            (logo_width, logo_height), Image.Resampling.LANCZOS
                        )
                    except AttributeError:
                        # Fallback para versões antigas do Pillow
                        app_logo_resized.thumbnail(
                            (logo_width, logo_height), Image.LANCZOS
                        )

                    print(
                        f"[DEBUG] OK Logo redimensionada para: {app_logo_resized.size}"
                    )

                    # Centralizar horizontalmente e posicionar mais próxima ao QR Code
                    logo_x = (300 - app_logo_resized.width) // 2
                    # QR Code está em Y=120, logo OpsVista bem próxima com margem menor
                    logo_y = (
                        120 - logo_height - 5
                    )  # Apenas 5px de margem entre logo e QR Code

                    print(f"[DEBUG] OK Posição da logo: ({logo_x}, {logo_y})")

                    if app_logo_resized.mode == "RGBA":
                        final_image.paste(
                            app_logo_resized, (logo_x, logo_y), app_logo_resized
                        )
                        print("[DEBUG] OK LOGO OPSVISTA COLADA COM TRANSPARENCIA!")
                    else:
                        final_image.paste(app_logo_resized, (logo_x, logo_y))
                        print("[DEBUG] OK LOGO OPSVISTA COLADA SEM TRANSPARENCIA!")

                except Exception as e:
                    import traceback

                    print(f"[ERROR] ERRO Erro ao adicionar logo OpsVista: {e}")
                    print(f"[ERROR] Traceback: {traceback.format_exc()}")
                    # Fallback para texto apenas em caso de erro (posição próxima ao QR Code)
                    fallback_y = 120 - 80 - 5  # Mesma lógica da logo (mais próxima)
                    draw.text(
                        (150, fallback_y),
                        "OPSVISTA (ERRO)",
                        fill="red",
                        font=ImageFont.load_default(),
                        anchor="mm",
                    )
            else:
                print("[DEBUG] ERRO OpsVista logo não está disponível - usando texto")
                # Adicionar texto como fallback (posição próxima ao QR Code)
                fallback_y = 120 - 80 - 5  # Mesma lógica da logo (mais próxima)
                draw.text(
                    (150, fallback_y),
                    "OPSVISTA",
                    fill="black",
                    font=ImageFont.load_default(),
                    anchor="mm",
                )

            # Adicionar logo do serviço selecionado no canto superior direito (opcional)
            try:
                logo_servico = LogoServico.objects.filter(nome=service).first()
                if logo_servico and logo_servico.img_base64:
                    # Extrair dados base64 (remover prefixo data:image/png;base64, se existir)
                    img_data = logo_servico.img_base64
                    if img_data.startswith("data:image/"):
                        img_data = img_data.split(",")[1]

                    # Decodificar base64 e criar imagem
                    import base64

                    img_bytes = base64.b64decode(img_data)
                    service_logo = Image.open(BytesIO(img_bytes))

                    # Redimensionar conforme tamanho selecionado pelo usuário
                    service_logo.thumbnail((service_logo_size, service_logo_size))

                    # Posicionar no canto superior direito com margens confortáveis
                    logo_x = max(
                        300 - service_logo_size - 15, 10
                    )  # Margem confortável de 15px da direita
                    logo_y = 15  # Margem confortável de 15px do topo

                    print(
                        f"[DEBUG] Logo do serviço posicionada em: ({logo_x}, {logo_y}), tamanho: {service_logo_size}px"
                    )

                    if service_logo.mode == "RGBA":
                        final_image.paste(service_logo, (logo_x, logo_y), service_logo)
                    else:
                        final_image.paste(service_logo, (logo_x, logo_y))
                else:
                    # Fallback para texto se não houver logo (posição no topo direito com margens)
                    draw.text(
                        (260, 20),
                        service.upper(),
                        fill="black",
                        font=ImageFont.load_default(),
                        anchor="mm",
                    )
            except Exception as e:
                print(f"Erro ao carregar logo do serviço: {e}")
                draw.text(
                    (260, 20),
                    service.upper(),
                    fill="black",
                    font=ImageFont.load_default(),
                    anchor="mm",
                )

            # Adicionar QR code (posição equilibrada)
            qr_size = 140
            qr_pos = ((300 - qr_size) // 2, 120)  # Posição mais equilibrada
            final_image.paste(qr_image.resize((qr_size, qr_size)), qr_pos)

            # Adicionar texto do local (posição equilibrada)
            text_y = qr_pos[1] + qr_size + 15  # Logo abaixo do QR code com espaço menor

            # Quebrar texto do local em múltiplas linhas se necessário
            location_text = custom_description.upper()
            words = location_text.split(" ")
            lines = []
            current_line = ""

            for word in words:
                test_line = current_line + (" " + word if current_line else word)
                # Aproximadamente 25 caracteres por linha para caber na largura
                if len(test_line) <= 25:
                    current_line = test_line
                else:
                    if current_line:
                        lines.append(current_line)
                    current_line = word

            if current_line:
                lines.append(current_line)

            # Determinar tamanho da fonte baseado na quantidade de texto
            total_chars = len(location_text)
            num_lines = len(lines)

            # Função para carregar fonte com suporte a Unicode (preview)
            def load_unicode_font_preview(size):
                """Carrega uma fonte que suporta caracteres Unicode/acentos para preview"""
                font_paths = [
                    # Fontes do Windows com suporte a Unicode
                    "C:/Windows/Fonts/arial.ttf",
                    "C:/Windows/Fonts/calibri.ttf", 
                    "C:/Windows/Fonts/tahoma.ttf",
                    "C:/Windows/Fonts/verdana.ttf",
                    # Fontes do Linux com suporte a Unicode
                    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
                    "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
                    "/System/Library/Fonts/Arial.ttf",  # macOS
                    # Fallback para fontes básicas
                    "arial.ttf",
                    "calibri.ttf",
                    "tahoma.ttf"
                ]
                
                for font_path in font_paths:
                    try:
                        font = ImageFont.truetype(font_path, size)
                        print(f"[DEBUG PREVIEW] Fonte carregada com sucesso: {font_path}")
                        return font
                    except (OSError, IOError):
                        continue
                
                # Se nenhuma fonte TrueType funcionar, usar fonte padrão
                print("[DEBUG PREVIEW] Usando fonte padrão - algumas fontes TrueType não disponíveis")
                return ImageFont.load_default()

            # Lógica de fonte dinâmica com suporte a Unicode
            if total_chars <= 20 and num_lines <= 2:
                # Texto curto - fonte maior
                font_size = 16
                font = load_unicode_font_preview(font_size)
                line_spacing = 20
            elif total_chars <= 40 and num_lines <= 3:
                # Texto médio - fonte normal
                font_size = 14
                font = load_unicode_font_preview(font_size)
                line_spacing = 18
            else:
                # Texto longo - fonte menor
                font_size = 12
                font = load_unicode_font_preview(font_size)
                line_spacing = 15

            print(
                f"[DEBUG] Texto: {total_chars} chars, {num_lines} linhas, espaçamento: {line_spacing}px"
            )

            # Desenhar cada linha do texto com fonte dinâmica
            for i, line in enumerate(lines):
                line_y = text_y + (i * line_spacing)
                # Garantir que o texto está corretamente codificado para renderização
                try:
                    # Normalizar o texto para garantir compatibilidade com a fonte
                    import unicodedata
                    normalized_line = unicodedata.normalize('NFC', str(line))
                    print(f"[DEBUG PREVIEW] Renderizando linha {i+1}: '{normalized_line}'")
                    draw.text((150, line_y), normalized_line, fill="black", font=font, anchor="mm")
                except Exception as e:
                    print(f"[ERROR PREVIEW] Erro ao renderizar linha '{line}': {e}")
                    # Fallback: tentar renderizar sem normalização
                    draw.text((150, line_y), str(line), fill="black", font=font, anchor="mm")

            # Se houver logo do cliente - posicionada no rodapé do card
            if "client_logo" in request.FILES:
                client_logo = Image.open(request.FILES["client_logo"])
                client_logo.thumbnail((logo_size, logo_size))

                # Posicionar no rodapé, centralizada horizontalmente
                # Garantir margem confortável do rodapé (20px em vez de 10px)
                max_logo_height = min(
                    logo_size, 80
                )  # Limitar altura para não sobrepor o texto
                logo_y = max(
                    400 - logo_size - 20, 320
                )  # Margem mais confortável do rodapé
                logo_x = (300 - logo_size) // 2

                logo_pos = (logo_x, logo_y)
                print(
                    f"[DEBUG] Logo do cliente posicionada em: {logo_pos}, tamanho: {logo_size}px"
                )

                final_image.paste(
                    client_logo,
                    logo_pos,
                    client_logo if client_logo.mode == "RGBA" else None,
                )

            # Converter para base64
            buffer = BytesIO()
            final_image.save(buffer, format="PNG")
            qr_base64 = base64.b64encode(buffer.getvalue()).decode()

            qr_codes.append(
                {
                    "data": qr_data_content,  # Usar o conteúdo determinado (link ou dados estruturados)
                    "image": qr_base64,
                    "custom_description": custom_description,
                }
            )

        return JsonResponse({"success": True, "qr_codes": qr_codes})

    except Exception as e:
        import traceback

        print(f"Erro ao gerar QR Code: {str(e)}")
        print(f"Traceback: {traceback.format_exc()}")
        return JsonResponse({"error": f"Erro ao gerar QR Code: {str(e)}"}, status=500)


@login_required
def get_locations(request):
    print(f"[DEBUG] get_locations called with CR: {request.GET.get('cr')}")
    try:
        # OTIMIZADO: Usar cache para CRs consultados
        from django.core.cache import cache

        # Obter o CR da query string
        cr = request.GET.get("cr")
        print(f"Buscando locais para CR: {cr}")

        # Query base - excluir registros com status = 4 (excluídos)
        query = Estrutura.objects.exclude(status=4)

        # Buscar informações do CR
        cr_info = None
        if cr:
            # Limpar o CR de possíveis espaços e caracteres especiais
            cr_cleaned = cr.strip()
            print(f"CR limpo: {cr_cleaned}")

            # OTIMIZADO: Tentar pegar do cache primeiro
            cache_key = f"cr_info_{cr_cleaned}"
            cr_info = cache.get(cache_key)

            if not cr_info:
                # Buscar estrutura (sem only() devido a managed=False)
                # CORREÇÃO: Usar mesma lógica do qr_generator para encontrar CRs como 18416
                from django.db.models import Q
                estrutura = Estrutura.objects.filter(
                    Q(cr=cr_cleaned)
                    | Q(cr__iexact=cr_cleaned)  # Exato
                    | Q(cr__contains=cr_cleaned)  # Contém o número
                ).first()

                if estrutura:
                    cr_info = {
                        "cr": estrutura.cr,
                        "descricao": estrutura.descricao,
                        "hierarquia": estrutura.hierarquiadescricao,
                        "grupo": estrutura.grupo,
                        "diretor": estrutura.diretor,
                        "gr": estrutura.gr,
                        "gc": estrutura.gc,
                    }
                    # OTIMIZADO: Armazenar no cache por 1 hora
                    cache.set(cache_key, cr_info, 3600)

            # OTIMIZADO: Filtrar locais
            # Nota: Não usando only() pois Estrutura tem managed=False e pode causar problemas
            # CORREÇÃO: Usar startswith para corresponder ao comportamento LIKE 'prefix%' do código Python
            base_query = query.filter(cr__startswith=cr_cleaned, nivel_4__isnull=False).exclude(nivel_4="")

            # CORREÇÃO: Retornar TODOS os registros sem filtro de priorização
            # Isso garante que registros sem nivel_5 também sejam incluídos
            locations = list(
                base_query.order_by("nivel_4", "nivel_5", "nivel_6")[:500]
            )  # OTIMIZADO: Converter para lista
            print(
                f"[DEBUG] Retornando {len(locations)} locais para CR {cr_cleaned}"
            )
        else:
            locations = []
            print("[DEBUG] Nenhum CR fornecido")

        print(f"[DEBUG] Total de locais encontrados: {len(locations)}")

        # OTIMIZADO: Construir hierarquia sem loops extras
        print(f"[DEBUG] Construindo hierarquia para {len(locations)} locais...")

        # OTIMIZADO: Detecção de CR complexo simplificada (sem loop extra)
        # Apenas verificar os primeiros 10 registros para decidir o tipo
        sample_size = min(10, len(locations))
        if sample_size > 0:
            complex_sample = 0
            for loc in locations[:sample_size]:
                try:
                    if (
                        getattr(loc, "nivel_4", None)
                        and getattr(loc, "nivel_5", None)
                        and getattr(loc, "nivel_6", None)
                        and getattr(loc, "nivel_7", None)
                    ):
                        complex_sample += 1
                except AttributeError:
                    continue
            is_complex_cr = complex_sample > sample_size * 0.5
        else:
            is_complex_cr = False

        min_levels = 4 if is_complex_cr else 1  # CORREÇÃO: Permitir registros com 1+ níveis
        print(f"[DEBUG] CR complexo: {is_complex_cr}, mínimo de níveis: {min_levels}")

        hierarchy = {}

        # Função auxiliar para gerar IDs HTML seguros e evitar falhas na interface
        import re, unicodedata, uuid
        def make_safe_id(prefix, text):
            if not text: return f"{prefix}_{uuid.uuid4().hex[:8]}"
            text_norm = ''.join(c for c in unicodedata.normalize('NFD', str(text)) if unicodedata.category(c) != 'Mn')
            safe_text = re.sub(r'[^a-zA-Z0-9]', '_', text_norm)
            safe_text = re.sub(r'_+', '_', safe_text).strip('_')
            return f"{prefix}_{safe_text}_{uuid.uuid4().hex[:6]}"

        # PRIMEIRA PASSAGEM: Identificar quais registros são containers (têm filhos)
        container_signatures = set()
        try:
            for loc in locations:
                try:
                    # Se tem nivel_6 preenchido, então nivel_5 é um container
                    if getattr(loc, "nivel_6", None):
                        nivel_4 = getattr(loc, "nivel_4", None)
                        nivel_5 = getattr(loc, "nivel_5", None)
                        if nivel_4 and nivel_5:
                            container_signatures.add((nivel_4, nivel_5))
                            print(f"[DEBUG] Container nivel_5 detectado: {nivel_4} -> {nivel_5}")

                    # Se tem nivel_7 preenchido, então nivel_6 é um container
                    if getattr(loc, "nivel_7", None):
                        nivel_4 = getattr(loc, "nivel_4", None)
                        nivel_5 = getattr(loc, "nivel_5", None)
                        nivel_6 = getattr(loc, "nivel_6", None)
                        if nivel_4 and nivel_5 and nivel_6:
                            container_signatures.add((nivel_4, nivel_5, nivel_6))
                            print(f"[DEBUG] Container nivel_6 detectado: {nivel_4} -> {nivel_5} -> {nivel_6}")
                except Exception as e:
                    print(f"[DEBUG] Erro ao identificar container individual: {e}")
                    continue

            print(f"[DEBUG] Total containers identificados: {len(container_signatures)}")
        except Exception as e:
            print(f"[ERROR] Erro fatal ao identificar containers: {e}")
            import traceback
            print(traceback.format_exc())

        # OTIMIZADO: Loop principal mais eficiente
        for loc in locations:
            # OTIMIZADO: Extrair níveis de forma mais eficiente e segura
            try:
                levels = [
                    nivel
                    for nivel in [
                        getattr(loc, "nivel_4", None),
                        getattr(loc, "nivel_5", None),
                        getattr(loc, "nivel_6", None),
                        getattr(loc, "nivel_7", None),
                        getattr(loc, "nivel_8", None),
                        getattr(loc, "nivel_9", None),
                        getattr(loc, "nivel_10", None),
                        getattr(loc, "nivel_11", None),
                    ]
                    if nivel
                ]
            except AttributeError as e:
                print(f"[ERROR] Erro ao acessar níveis do loc {loc.id}: {e}")
                continue

            if not levels:
                continue

            # CORREÇÃO ESPECIAL: Se a descrição contém "/" no nome,
            # usar a descrição completa como local único ao invés da hierarquia de níveis
            descricao = getattr(loc, "descricao", "")
            if (len(levels) >= 3 and descricao and "/" in descricao):
                # Verificar se o nivel_6 é parte do nome (ex: OPY/RICART)
                nivel_5 = getattr(loc, "nivel_5", "")
                nivel_6 = getattr(loc, "nivel_6", "")

                # Se descrição contém a barra e nivel_6 parece ser continuação do nivel_5
                if nivel_5 and nivel_6 and f"{nivel_5}/{nivel_6}" in descricao:
                    print(f"[DEBUG] Detectado local com '/' na descrição: {descricao}")
                    print(f"[DEBUG]   - Antes: {' -> '.join(levels)}")
                    # Usar apenas nivel_4 (building) + descrição completa como nivel_5
                    levels = [levels[0], descricao]
                    print(f"[DEBUG]   - Depois: {' -> '.join(levels)}")
                elif nivel_6 and nivel_6.strip() and "/" in nivel_5:
                    # Caso alternativo: nivel_5 já tem a barra (ex: "OPY/RICART")
                    print(f"[DEBUG] Detectado barra no nivel_5: {nivel_5}")
                    print(f"[DEBUG]   - Antes: {' -> '.join(levels)}")
                    levels = [levels[0], descricao]
                    print(f"[DEBUG]   - Depois: {' -> '.join(levels)} (usando descrição completa)")

            # CORREÇÃO: Verificar se este registro é um container intermediário
            is_container = tuple(levels) in container_signatures

            # FILTRO OTIMIZADO: Aplicar filtro baseado no tipo de CR detectado
            # CORREÇÃO: Para CRs complexos, aceitar registros com menos níveis se forem registros raiz
            if is_complex_cr and len(levels) < min_levels:
                # Permitir registros com apenas nivel_4 se puderem ser registros raiz importantes
                if len(levels) == 1:
                    print(f"[DEBUG] Incluindo registro raiz com 1 nível: {levels[0]} (ID: {loc.id})")
                    # Não fazer continue, deixa processar
                else:
                    continue
            elif not is_complex_cr and len(levels) < min_levels:
                # EXCEÇÃO: Permitir se for um container intermediário
                if not is_container:
                    continue

            # Navegar pela hierarquia criando containers
            current = hierarchy

            # LÓGICA ADAPTATIVA baseada no número de níveis
            if len(levels) == 1:
                # Apenas 1 nível - criar local selecionável direto no building
                building_name = levels[0]

                # CORREÇÃO: Verificar se este é o registro raiz (building principal)
                is_root_building = (
                    loc.descricao and
                    loc.descricao.strip().lower() == building_name.strip().lower()
                )

                if building_name not in current:
                    # Criar building (pode usar o ID real se for o registro raiz)
                    building_id = str(loc.id) if is_root_building else make_safe_id("bldg", building_name)
                    current[building_name] = {
                        "id": building_id,
                        "name": building_name,
                        "type": "building",
                        "isContainer": True,  # Building é container, não selecionável
                        "isSelectableContainer": is_root_building,  # Selecionável se for registro real
                        "displayName": building_name if is_root_building else None,  # Adicionar displayName se selecionável
                        "children": {},  # CORREÇÃO: Usar dict consistentemente
                    }
                    print(f"[DEBUG] Building criado: {building_name} (ID: {building_id}, is_root: {is_root_building})")

                # CORREÇÃO: Se for o registro raiz, NÃO adicionar como local filho
                if is_root_building:
                    print(f"[DEBUG] Ignorando registro raiz como local filho: {building_name} (ID: {loc.id})")
                    continue

                # Criar lista de _locations se não existir
                if "_locations" not in current[building_name]["children"]:
                    current[building_name]["children"]["_locations"] = []

                # Adicionar como local selecionável
                display_name = building_name
                location_name = loc.descricao if loc.descricao else building_name
                if loc.descricao and loc.descricao != building_name:
                    display_name += f" - {loc.descricao}"

                current[building_name]["children"]["_locations"].append(
                    {
                        "id": str(loc.id),
                        "name": decode_unicode_string(location_name),
                        "displayName": decode_unicode_string(display_name),
                        "type": "room",
                        "descricao_completa": decode_unicode_string(loc.descricao)
                        if loc.descricao
                        else None,
                        "cr_completo": loc.cr,
                        "qrcode": loc.qrcode if loc.qrcode else None,
                    }
                )
                print(f"[DEBUG] Local final 1-nível criado: {building_name} (ID: {loc.id})")

            elif len(levels) == 2:
                # 2 níveis - building + (container OU local final)
                building_name = levels[0]
                local_name = levels[1]
                print(f"[DEBUG] Processando 2 níveis: {building_name} -> {local_name} (ID: {loc.id}, is_container: {is_container})")

                # Criar building se não existe
                if building_name not in current:
                    current[building_name] = {
                        "id": make_safe_id("bldg", building_name),
                        "name": building_name,
                        "type": "building",
                        "isContainer": True,  # Building é container, não selecionável
                        "children": {},  # CORREÇÃO: Usar dict para permitir containers aninhados
                    }

                # CORREÇÃO: Verificar se este registro é um container intermediário
                if is_container:
                    # Este é um container (ex: GERAL, PRIMEIRO ANDAR, SUBSOLO)
                    # Criar como container dentro do building
                    if local_name not in current[building_name]["children"]:
                        current[building_name]["children"][local_name] = {
                            "id": str(loc.id),  # ID real do container
                            "name": local_name,
                            "displayName": " -> ".join(levels),  # Torná-lo selecionável
                            "type": "floor",
                            "isSelectableContainer": True,  # Container que também é local
                            "children": {},  # Este container pode ter filhos
                        }
                        print(f"[DEBUG] Container intermediário criado: {local_name} (ID: {loc.id})")
                else:
                    # Este é um local final (sem filhos)
                    # Garantir que children seja uma lista para locais finais
                    if "_locations" not in current[building_name]["children"]:
                        current[building_name]["children"]["_locations"] = []

                    display_name = " -> ".join(levels)
                    # CORREÇÃO: Se local_name JÁ É a descrição completa (detectado pela barra),
                    # usar ela diretamente sem adicionar novamente
                    location_name = local_name  # Já é a descrição se foi modificado acima

                    # Não adicionar descrição extra se local_name já é a descrição
                    descricao_original = getattr(loc, "descricao", "")
                    if descricao_original and "/" not in descricao_original and descricao_original != local_name:
                        display_name += f" - {descricao_original}"

                    current[building_name]["children"]["_locations"].append(
                        {
                            "id": str(loc.id),
                            "name": decode_unicode_string(location_name),
                            "displayName": decode_unicode_string(display_name),
                            "type": "room",
                            "descricao_completa": decode_unicode_string(descricao_original)
                            if descricao_original
                            else None,
                            "cr_completo": loc.cr,
                            "qrcode": loc.qrcode if loc.qrcode else None,
                        }
                    )
                    print(f"[DEBUG] Local final 2-níveis criado: {display_name} (ID: {loc.id})")

            else:
                # 3+ níveis - usar lógica hierárquica completa
                print(f"[DEBUG] Processando 3+ níveis: {' -> '.join(levels)} (ID: {loc.id})")
                # Todos os níveis exceto o último são containers
                for i, level_name in enumerate(levels[:-1]):
                    # Determinar tipo
                    if i == 0:
                        level_type = "building"
                    elif i == 1:
                        level_type = "floor"
                    else:
                        level_type = "section"

                    # Verificar se current é dict antes de usar
                    if not isinstance(current, dict):
                        continue

                    # Criar container se não existe
                    if level_name not in current:
                        # CORREÇÃO: Verificar se este container também é um local real no banco
                        container_path = levels[: i + 1]
                        container_displayName = " -> ".join(container_path)

                        # Buscar se existe um registro no banco para este container
                        container_record = None
                        try:
                            # ESTRATÉGIA 1: Busca exata por descrição
                            for potential_loc in locations:
                                if (
                                    hasattr(potential_loc, "descricao")
                                    and potential_loc.descricao
                                    and potential_loc.descricao.strip().lower()
                                    == level_name.strip().lower()
                                ):
                                    container_record = potential_loc
                                    print(
                                        f"[DEBUG] Encontrado por descrição exata: {level_name} -> {potential_loc.id}"
                                    )
                                    break

                            # ESTRATÉGIA 2: Se não encontrou, buscar por conteúdo na descrição
                            if not container_record:
                                for potential_loc in locations:
                                    if (
                                        hasattr(potential_loc, "descricao")
                                        and potential_loc.descricao
                                        and level_name.lower()
                                        in potential_loc.descricao.lower()
                                    ):
                                        container_record = potential_loc
                                        print(
                                            f"[DEBUG] Encontrado por conteúdo: {level_name} -> {potential_loc.id}"
                                        )
                                        break

                            # ESTRATÉGIA 3: Para estruturas mestres, buscar por níveis hierárquicos
                            if (
                                not container_record and i == 0
                            ):  # Primeiro nível (estrutura mestre)
                                for potential_loc in locations:
                                    # Verificar se é uma estrutura mestre comparando níveis
                                    if (
                                        hasattr(potential_loc, "nivel_4")
                                        and potential_loc.nivel_4
                                        and potential_loc.nivel_4.strip().lower()
                                        == level_name.strip().lower()
                                    ):
                                        container_record = potential_loc
                                        print(
                                            f"[DEBUG] Encontrado estrutura mestre por nivel_4: {level_name} -> {potential_loc.id}"
                                        )
                                        break

                            # ESTRATÉGIA 4: Buscar por hierarquiadescricao (para estruturas mestres)
                            if not container_record and i == 0:
                                for potential_loc in locations:
                                    if (
                                        hasattr(potential_loc, "hierarquiadescricao")
                                        and potential_loc.hierarquiadescricao
                                        and level_name.lower()
                                        in potential_loc.hierarquiadescricao.lower()
                                    ):
                                        container_record = potential_loc
                                        print(
                                            f"[DEBUG] Encontrado por hierarquiadescricao: {level_name} -> {potential_loc.id}"
                                        )
                                        break

                            # ESTRATÉGIA 5: Para estruturas mestres, pegar o primeiro registro do CR (fallback)
                            if not container_record and i == 0 and locations:
                                # Pegar o primeiro registro como representante da estrutura mestre
                                container_record = locations[0]
                                print(
                                    f"[DEBUG] Usando primeiro registro como estrutura mestre: {level_name} -> {container_record.id}"
                                )
                                print(
                                    f"[DEBUG] Registro usado: {container_record.descricao} (CR: {container_record.cr})"
                                )

                        except Exception as e:
                            print(f"[DEBUG] Erro ao buscar container record: {e}")

                        # Se encontrou registro, é um local selecionável; senão, é apenas container
                        if container_record:
                            current[level_name] = {
                                "id": str(container_record.id),  # ID real do banco
                                "name": level_name,
                                "type": level_type,
                                "displayName": container_displayName,  # Selecionável
                                "isSelectableContainer": True,  # Container que também é local
                                "children": {},
                            }
                            print(
                                f"[DEBUG] Container selecionável criado: {level_name} (ID: {container_record.id})"
                            )
                        else:
                            current[level_name] = {
                                "id": make_safe_id(level_type, level_name),
                                "name": level_name,
                                "type": level_type,
                                "isContainer": True,  # Apenas container organizacional
                                "children": {},
                            }
                            print(
                                f"[DEBUG] Container organizacional criado: {level_name}"
                            )

                    # Continuar navegando - garantir que children seja dict
                    if not isinstance(current[level_name]["children"], dict):
                        current[level_name]["children"] = {}

                    current = current[level_name]["children"]

                # Adicionar o local final diretamente no nível anterior (sem container intermediário)
                # Verificar se current é dict antes de usar
                if not isinstance(current, dict):
                    continue

                # Criar uma lista especial para itens finais se não existir
                if "_locations" not in current:
                    current["_locations"] = []

                # Adicionar o local final
                display_name = " -> ".join(levels)
                location_name = loc.descricao if loc.descricao else levels[-1]

                current["_locations"].append(
                    {
                        "id": str(loc.id),
                        "name": decode_unicode_string(location_name),
                        "displayName": decode_unicode_string(display_name),
                        "type": "room",
                        "descricao_completa": decode_unicode_string(loc.descricao)
                        if loc.descricao
                        else None,
                        "cr_completo": loc.cr,
                        "qrcode": loc.qrcode if loc.qrcode else None,
                    }
                )
                print(f"[DEBUG] Local final 3+ níveis adicionado: {display_name} (ID: {loc.id})")

        print(f"[DEBUG] Hierarquia construída com {len(hierarchy)} edifícios")

        # DEBUG: Mostrar estrutura detalhada da hierarquia
        import json
        for building_name, building_data in hierarchy.items():
            print(f"[DEBUG HIERARCHY] Building: {building_name}")
            print(f"[DEBUG HIERARCHY]   - ID: {building_data.get('id')}")
            print(f"[DEBUG HIERARCHY]   - Type: {building_data.get('type')}")
            print(f"[DEBUG HIERARCHY]   - isContainer: {building_data.get('isContainer')}")
            print(f"[DEBUG HIERARCHY]   - isSelectableContainer: {building_data.get('isSelectableContainer')}")
            children = building_data.get('children')
            if isinstance(children, dict):
                print(f"[DEBUG HIERARCHY]   - Children (dict): {len(children)} keys")
                for child_name in children.keys():
                    if not child_name.startswith('_'):
                        print(f"[DEBUG HIERARCHY]     * {child_name}")
            elif isinstance(children, list):
                print(f"[DEBUG HIERARCHY]   - Children (list): {len(children)} items")
            else:
                print(f"[DEBUG HIERARCHY]   - Children: {type(children)}")

        # Converter para formato do frontend (recursivo para suportar hierarquia dinâmica)
        def convert_hierarchy_to_frontend(hierarchy_dict):
            result = []
            for key, node in hierarchy_dict.items():
                # Pular chaves especiais
                if key.startswith("_"):
                    continue

                node_data = {
                    "id": node["id"],
                    "name": node["name"],
                    "type": node["type"],
                }

                # CORREÇÃO: Adicionar displayName para locais selecionáveis
                if "displayName" in node:
                    # Incluir displayName se for local final OU container selecionável
                    if not node.get("isContainer", False) or node.get(
                        "isSelectableContainer", False
                    ):
                        node_data["displayName"] = node["displayName"]

                # Propagar flags para o frontend
                if node.get("isContainer", False):
                    node_data["isContainer"] = True
                if node.get("isSelectableContainer", False):
                    node_data["isSelectableContainer"] = True

                # Verificar se tem children e qual o tipo
                children = node.get("children")
                locations = node.get("_locations", [])

                if isinstance(children, dict):
                    # É um container - tem subníveis como dict
                    converted_children = convert_hierarchy_to_frontend(children)

                    # Se também tem _locations, adicionar à lista de children
                    child_locations = children.get("_locations", [])
                    if child_locations:
                        converted_children.extend(child_locations)

                    node_data["children"] = converted_children

                elif isinstance(children, list):
                    # É um nível final - tem locais selecionáveis como lista
                    node_data["children"] = children + locations
                else:
                    # Só tem locations ou não tem nada
                    node_data["children"] = locations

                result.append(node_data)
            return result

        formatted_locations = convert_hierarchy_to_frontend(hierarchy)

        print(f"[DEBUG] Total de edifícios formatados: {len(formatted_locations)}")
        print(
            f"[DEBUG] Hierarchy keys: {list(hierarchy.keys()) if hierarchy else 'empty'}"
        )

        # DEBUG: Mostrar estrutura do JSON que será enviado ao frontend
        if formatted_locations:
            print(f"[DEBUG FINAL] Primeira localização formatada:")
            first_loc = formatted_locations[0]
            print(f"[DEBUG FINAL]   - ID: {first_loc.get('id')}")
            print(f"[DEBUG FINAL]   - Name: {first_loc.get('name')}")
            print(f"[DEBUG FINAL]   - Type: {first_loc.get('type')}")
            print(f"[DEBUG FINAL]   - isContainer: {first_loc.get('isContainer')}")
            print(f"[DEBUG FINAL]   - isSelectableContainer: {first_loc.get('isSelectableContainer')}")
            children = first_loc.get('children', [])
            print(f"[DEBUG FINAL]   - Children: {len(children)} items")
            if children:
                for idx, child in enumerate(children[:5]):  # Mostrar até 5 children
                    print(f"[DEBUG FINAL]     [{idx}] {child.get('name')} (ID: {child.get('id')})")

        return JsonResponse({"locations": formatted_locations, "cr_info": cr_info})
    except Exception as e:
        import traceback

        error_details = {
            "error": str(e),
            "type": type(e).__name__,
            "traceback": traceback.format_exc(),
        }

        print(f"[ERROR] Erro ao buscar locais: {str(e)}")
        print(f"[ERROR] Tipo do erro: {type(e).__name__}")
        print(f"[ERROR] Traceback completo:\n{traceback.format_exc()}")

        # Tentar identificar onde o erro ocorreu
        if "hierarchy" in locals():
            print(f"[DEBUG] Estado da hierarchy quando erro ocorreu: {type(hierarchy)}")
            print(
                f"[DEBUG] Keys da hierarchy: {list(hierarchy.keys()) if isinstance(hierarchy, dict) else 'não é dict'}"
            )

        if "locations" in locals():
            print(
                f"[DEBUG] Número de locations: {len(locations) if hasattr(locations, '__len__') else 'N/A'}"
            )

        return JsonResponse(error_details, status=500)


@login_required
def get_service_logo(request):
    """
    API endpoint para buscar logo de um serviço específico
    """
    try:
        service_name = request.GET.get("service")
        print(f"[DEBUG] Requisição para logo do serviço: {service_name}")

        if not service_name:
            print("[DEBUG] Nome do serviço não fornecido")
            return JsonResponse({"error": "Nome do serviço é obrigatório"}, status=400)

        # Buscar logo do serviço na tabela logo_servico
        logo_servico = LogoServico.objects.filter(nome=service_name).first()
        print(f"[DEBUG] Logo encontrada: {logo_servico is not None}")

        if logo_servico:
            # Verificar se a imagem base64 tem formato correto
            img_data = logo_servico.img_base64
            print(f"[DEBUG] Tamanho da imagem base64: {len(img_data)}")
            print(f"[DEBUG] Início da imagem: {img_data[:50]}...")

            return JsonResponse(
                {"success": True, "service": service_name, "logo_base64": img_data}
            )
        else:
            print(f"[DEBUG] Logo não encontrada para: {service_name}")
            # Listar todos os logos disponíveis para debug
            available_logos = list(LogoServico.objects.values_list("nome", flat=True))
            print(f"[DEBUG] Logos disponíveis: {available_logos}")

            return JsonResponse(
                {
                    "success": False,
                    "error": f"Logo não encontrada para o serviço: {service_name}",
                    "available_services": available_logos,
                },
                status=404,
            )

    except Exception as e:
        import traceback

        print(f"Erro ao buscar logo do serviço: {str(e)}")
        print(f"Traceback: {traceback.format_exc()}")
        return JsonResponse({"error": str(e)}, status=500)


@login_required
def desativacao_cr(request):
    if request.method == "GET" and request.GET.get("id"):
        # Buscar um registro específico para edição
        try:
            desativacao = get_object_or_404(DesativacaoCR, id=request.GET.get("id"))
            return JsonResponse(
                {
                    "id": str(desativacao.id),
                    "contrato": desativacao.contrato,
                    "solicitante": desativacao.solicitante,
                    "status": desativacao.status,
                }
            )
        except Exception as e:
            return JsonResponse({"message": str(e)}, status=400)

    if request.method == "POST":
        data = json.loads(request.body)
        try:
            desativacao = DesativacaoCR.objects.create(
                contrato=data["contrato"],
                solicitante=data["solicitante"],
                status=data["status"],
            )
            return JsonResponse(
                {
                    "id": str(desativacao.id),
                    "contrato": desativacao.contrato,
                    "solicitante": desativacao.solicitante,
                    "status": desativacao.status,
                    "datasolicitacao": desativacao.datasolicitacao.strftime("%Y-%m-%d"),
                }
            )
        except Exception as e:
            return JsonResponse({"message": str(e)}, status=400)

    elif request.method == "PUT":
        data = json.loads(request.body)
        try:
            desativacao = get_object_or_404(DesativacaoCR, id=data["id"])
            desativacao.contrato = data["contrato"]
            desativacao.solicitante = data["solicitante"]
            desativacao.status = data["status"]
            desativacao.save()
            return JsonResponse(
                {
                    "id": str(desativacao.id),
                    "contrato": desativacao.contrato,
                    "solicitante": desativacao.solicitante,
                    "status": desativacao.status,
                    "datasolicitacao": desativacao.datasolicitacao.strftime("%Y-%m-%d"),
                }
            )
        except Exception as e:
            return JsonResponse({"message": str(e)}, status=400)

    elif request.method == "DELETE":
        data = json.loads(request.body)
        try:
            desativacao = get_object_or_404(DesativacaoCR, id=data["id"])
            desativacao.delete()
            return JsonResponse({"message": "Solicitação excluída com sucesso"})
        except Exception as e:
            return JsonResponse({"message": str(e)}, status=400)

    elif request.method == "PATCH":
        try:
            data = json.loads(request.body)
            ids = data.get('ids', [])
            novo_status = data.get('status')

            # Validação de dados obrigatórios
            if not ids or not novo_status:
                return JsonResponse({
                    'success': False,
                    'message': 'IDs e status são obrigatórios'
                }, status=400)

            # Validar se o status é válido
            status_validos = [choice[0] for choice in DesativacaoCR.STATUS_CHOICES]
            if novo_status not in status_validos:
                return JsonResponse({
                    'success': False,
                    'message': f'Status inválido. Valores permitidos: {", ".join(status_validos)}'
                }, status=400)

            # Atualizar registros em massa
            registros_atualizados = DesativacaoCR.objects.filter(
                id__in=ids
            ).update(status=novo_status)

            return JsonResponse({
                'success': True,
                'message': f'{registros_atualizados} registro(s) atualizado(s) com sucesso',
                'count': registros_atualizados
            })

        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'message': 'Dados JSON inválidos'
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Erro ao atualizar registros: {str(e)}'
            }, status=500)

    # Filtros

    cr_filter = request.GET.get("cr", "")
    status_filter = request.GET.get("status", "todos")

    # Paginação
    por_pagina = int(request.GET.get("por_pagina", 10))
    pagina = request.GET.get("pagina", 1)

    # Query base
    desativacoes = DesativacaoCR.objects.all()

    # Aplicar filtros
    if cr_filter:
        desativacoes = desativacoes.filter(contrato__icontains=cr_filter)
    if status_filter != "todos":
        desativacoes = desativacoes.filter(status=status_filter)

    # Ordenar por data de solicitação (mais recentes primeiro)
    desativacoes = desativacoes.order_by("-datasolicitacao")

    # Aplicar paginação
    paginator = Paginator(desativacoes, por_pagina)
    try:
        desativacoes_paginadas = paginator.page(pagina)
    except PageNotAnInteger:
        desativacoes_paginadas = paginator.page(1)
    except EmptyPage:
        desativacoes_paginadas = paginator.page(paginator.num_pages)

    context = {
        "desativacoes": desativacoes_paginadas,
        "filtros": {
            "cr": cr_filter,
            "status": status_filter,
        },
        "status_choices": DesativacaoCR.STATUS_CHOICES,
        "por_pagina": por_pagina,
    }

    return render(request, "desativacao_cr.html", context)


@login_required
def controle_chips(request):
    """View para a página de controle de chips"""
    if request.method == "GET" and request.GET.get("id"):
        # Buscar um registro específico para edição
        try:
            registro = get_object_or_404(ControleChip, id=request.GET.get("id"))
            return JsonResponse(
                {
                    "id": str(registro.id),
                    "data": registro.data.strftime("%Y-%m-%d"),
                    "id_portal": registro.id_portal,
                    "solicitante": registro.solicitante,
                    "cr": registro.cr,
                    "operadora": registro.operadora,
                    "numero_telefone": registro.numero_telefone,
                    "responsavel_retirada": registro.responsavel_retirada,
                    "status": registro.status,
                    "observacoes": registro.observacoes,
                }
            )
        except Exception as e:
            return JsonResponse(
                {"success": False, "message": f"Erro ao buscar registro: {str(e)}"},
                status=400,
            )

    elif request.method == "POST":
        try:
            data = json.loads(request.body)
            # Converter a string da data para objeto date
            data_obj = datetime.strptime(data["data"], "%Y-%m-%d").date()

            registro = ControleChip.objects.create(
                data=data_obj,  # Usar o objeto date convertido
                id_portal=data["id_portal"],
                solicitante=data["solicitante"],
                cr=data["cr"],
                operadora=data["operadora"],
                numero_telefone=data.get("numero_telefone", ""),
                responsavel_retirada=data["responsavel_retirada"],
                status=data["status"],
                observacoes=data.get("observacoes", ""),
            )
            return JsonResponse(
                {
                    "success": True,
                    "message": "Registro criado com sucesso",
                    "data": {
                        "id": str(registro.id),
                        "data": registro.data.strftime("%Y-%m-%d"),
                        "id_portal": registro.id_portal,
                        "solicitante": registro.solicitante,
                        "cr": registro.cr,
                        "operadora": registro.get_operadora_display(),
                        "numero_telefone": registro.numero_telefone,
                        "responsavel_retirada": registro.responsavel_retirada,
                        "status": registro.get_status_display(),
                        "observacoes": registro.observacoes,
                    },
                }
            )
        except ValueError as e:
            return JsonResponse(
                {"success": False, "message": f"Erro de validação: {str(e)}"},
                status=400,
            )
        except Exception as e:
            import traceback

            print(f"Erro ao criar registro: {str(e)}")
            print(f"Traceback: {traceback.format_exc()}")
            return JsonResponse(
                {"success": False, "message": f"Erro ao criar registro: {str(e)}"},
                status=400,
            )

    elif request.method == "PUT":
        try:
            data = json.loads(request.body)
            registro = get_object_or_404(ControleChip, id=data["id"])

            # Converter a string da data para objeto date
            data_obj = datetime.strptime(data["data"], "%Y-%m-%d").date()

            # Atualizar os campos
            registro.data = data_obj
            registro.id_portal = data["id_portal"]
            registro.solicitante = data["solicitante"]
            registro.cr = data["cr"]
            registro.operadora = data["operadora"]
            registro.numero_telefone = data.get("numero_telefone", "")
            registro.responsavel_retirada = data["responsavel_retirada"]
            registro.status = data["status"]
            registro.observacoes = data.get("observacoes", "")
            registro.save()

            return JsonResponse(
                {
                    "success": True,
                    "message": "Registro atualizado com sucesso",
                    "data": {
                        "id": str(registro.id),
                        "data": registro.data.strftime("%Y-%m-%d"),
                        "id_portal": registro.id_portal,
                        "solicitante": registro.solicitante,
                        "cr": registro.cr,
                        "operadora": registro.get_operadora_display(),
                        "numero_telefone": registro.numero_telefone,
                        "responsavel_retirada": registro.responsavel_retirada,
                        "status": registro.get_status_display(),
                        "observacoes": registro.observacoes,
                    },
                }
            )
        except ControleChip.DoesNotExist:
            return JsonResponse(
                {"success": False, "message": "Registro não encontrado"}, status=404
            )
        except ValueError as e:
            return JsonResponse(
                {"success": False, "message": f"Erro de validação: {str(e)}"},
                status=400,
            )
        except Exception as e:
            import traceback

            print(f"Erro ao atualizar registro: {str(e)}")
            print(f"Traceback: {traceback.format_exc()}")
            return JsonResponse(
                {"success": False, "message": f"Erro ao atualizar registro: {str(e)}"},
                status=400,
            )

    elif request.method == "DELETE":
        try:
            data = json.loads(request.body)
            registro = get_object_or_404(ControleChip, id=data["id"])
            registro.delete()
            return JsonResponse(
                {"success": True, "message": "Registro excluído com sucesso"}
            )
        except ControleChip.DoesNotExist:
            return JsonResponse(
                {"success": False, "message": "Registro não encontrado"}, status=404
            )
        except Exception as e:
            import traceback

            print(f"Erro ao excluir registro: {str(e)}")
            print(f"Traceback: {traceback.format_exc()}")
            return JsonResponse(
                {"success": False, "message": f"Erro ao excluir registro: {str(e)}"},
                status=400,
            )

    # GET request sem ID - listar registros

    # Filtros
    cr_filter = request.GET.get("cr", "").strip()
    telefone_filter = request.GET.get("telefone", "").strip()
    id_portal_filter = request.GET.get("id_portal", "").strip()
    status_filter = request.GET.get("status", "todos")

    # Ordenação
    ordenar = request.GET.get("ordenar", "-data")
    if ordenar.endswith("_desc"):
        ordenar = ordenar.replace("_desc", "")
    else:
        if not ordenar.startswith("-"):
            ordenar = "-" + ordenar

    # Paginação
    por_pagina = int(request.GET.get("por_pagina", 10))
    pagina = request.GET.get("pagina", 1)

    # Query base
    registros = ControleChip.objects.all()

    # Aplicar filtros
    if cr_filter:
        registros = registros.filter(cr__icontains=cr_filter)
    if telefone_filter:
        registros = registros.filter(numero_telefone__icontains=telefone_filter)
    if id_portal_filter:
        registros = registros.filter(id_portal__icontains=id_portal_filter)
    if status_filter != "todos":
        registros = registros.filter(status=status_filter)

    # Aplicar ordenação
    registros = registros.order_by(ordenar, "-created_at")

    # Aplicar paginação
    paginator = Paginator(registros, por_pagina)
    try:
        registros_paginados = paginator.page(pagina)
    except PageNotAnInteger:
        registros_paginados = paginator.page(1)
    except EmptyPage:
        registros_paginados = paginator.page(paginator.num_pages)

    

    context = {
        "registros": registros_paginados,
        "status_choices": ControleChip.STATUS_CHOICES,
        "operadora_choices": ControleChip.OPERADORA_CHOICES,
        "filtros": {
            "cr": cr_filter,
            "telefone": telefone_filter,
            "id_portal": id_portal_filter,
            "status": status_filter,
            
        },
        
        "ordenacao": request.GET.get("ordenar", "data"),
        "por_pagina": por_pagina,
    }

    return render(request, "controle_chips.html", context)


@login_required
def controle_acessos(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            action = data.get("action")

            if action == "get_user_permissions":
                try:
                    # Nova ação para buscar permissões do usuário
                    user_id = data.get("user_id")

                    if not user_id:
                        return JsonResponse(
                            {"error": "ID do usuário não fornecido"}, status=400
                        )

                    user = get_object_or_404(CustomUser, id=user_id)

                    # Se for administrador, usar as permissões padrão completas
                    if user.role == "administrador":
                        permissions = user.get_default_permissions()
                    else:
                        # Começar com permissões padrão do role
                        permissions = user.get_default_permissions()

                        # Se tiver permissões customizadas, sobrescrever
                        if user.page_permissions:
                            permissions.update(user.page_permissions)

                    return JsonResponse(
                        {
                            "permissions": permissions,
                            "role": user.role,
                            "name": user.name,
                            "username": user.username,
                            "email": user.email,
                            "is_regulatory": user.is_regulatory,
                            "is_general": user.is_general,
                            "crs": user.crs,
                            "setor": user.setor or "",
                            "regional": user.regional_id,
                            "notificar_livro_ata": user.notificar_livro_ata,
                            "whatsapp_notificacao": user.whatsapp_notificacao or "",
                            "is_global_admin": user.is_global_admin,
                            "success": True,
                        }
                    )

                except CustomUser.DoesNotExist:
                    return JsonResponse({"error": "Usuário não encontrado"}, status=404)
                except Exception as e:
                    print(f"Erro ao buscar permissões do usuário: {e}")
                    return JsonResponse(
                        {"error": f"Erro interno: {str(e)}"}, status=500
                    )

            elif action == "create":
                try:
                    # Criar novo usuário
                    # Se não for adm supremo (global), força a regional do criador
                    if getattr(request.user, "is_global_admin", False):
                        regional_id = data.get("regional")
                        # Somente um admin global pode criar outro admin global
                        is_global_admin_value = bool(data.get("is_global_admin", False))
                    else:
                        regional_id = request.user.regional_id
                        is_global_admin_value = False

                    user = CustomUser.objects.create(
                        username=data["username"],
                        name=data["name"],
                        password=make_password(data["password"]),
                        role=data["role"],
                        email=data.get("email", ""),
                        is_regulatory=data.get("is_regulatory", False),
                        is_general=data.get("is_general", False),
                        crs=data.get("crs", ""),
                        setor=data.get("setor", ""),
                        regional_id=regional_id,
                        notificar_livro_ata=data.get("notificar_livro_ata", False),
                        whatsapp_notificacao=data.get("whatsapp_notificacao", ""),
                        is_global_admin=is_global_admin_value,
                    )

                    # Se não for administrador, definir permissões específicas
                    if data["role"] != "administrador":
                        user.page_permissions = data.get("page_permissions", {})
                        user.save()

                    return JsonResponse(
                        {
                            "id": str(user.id),  # Converter UUID para string
                            "name": user.name,
                            "username": user.username,
                            "role": user.role,
                            "status": "offline",
                            "success": True,
                        }
                    )
                except Exception as e:
                    print(f"Erro ao criar usuário: {e}")
                    return JsonResponse(
                        {"error": f"Erro ao criar usuário: {str(e)}"}, status=500
                    )

            elif action == "update":
                try:
                    # Atualizar usuário existente
                    user_id = data.get("user_id")
                    if not user_id:
                        return JsonResponse(
                            {"error": "ID do usuário não fornecido"}, status=400
                        )

                    user = get_object_or_404(CustomUser, id=user_id)
                    user.name = data.get("name", user.name)
                    user.username = data.get("username", user.username)
                    user.role = data.get("role", user.role)
                    user.email = data.get("email", user.email)
                    user.is_regulatory = data.get("is_regulatory", user.is_regulatory)
                    user.is_general = data.get("is_general", user.is_general)
                    user.crs = data.get("crs", user.crs)
                    user.setor = data.get("setor", user.setor)
                    user.notificar_livro_ata = data.get("notificar_livro_ata", user.notificar_livro_ata)
                    user.whatsapp_notificacao = data.get("whatsapp_notificacao", user.whatsapp_notificacao)

                    # Apenas adm supremo pode editar a regional/flag de admin global de um usuário
                    if getattr(request.user, "is_global_admin", False):
                        if "regional" in data:
                            user.regional_id = data["regional"] if data["regional"] else None
                        if "is_global_admin" in data:
                            new_is_global_admin = bool(data["is_global_admin"])
                            # Impede que o próprio admin global remova essa flag de si mesmo
                            # (evitaria lockout: ninguém mais conseguiria promover admins globais)
                            if not (str(user.id) == str(request.user.id) and not new_is_global_admin):
                                user.is_global_admin = new_is_global_admin

                    # Se não for administrador, atualizar permissões específicas
                    if data["role"] != "administrador":
                        user.page_permissions = data.get("page_permissions", {})
                    else:
                        user.page_permissions = user.get_default_permissions()

                    user.save()
                    return JsonResponse({"status": "success", "success": True})
                except Exception as e:
                    print(f"Erro ao atualizar usuário: {e}")
                    return JsonResponse(
                        {"error": f"Erro ao atualizar usuário: {str(e)}"}, status=500
                    )

            elif action == "delete":
                try:
                    # Excluir usuário
                    user_id = data.get("user_id")
                    if not user_id:
                        return JsonResponse(
                            {"error": "ID do usuário não fornecido"}, status=400
                        )

                    user = get_object_or_404(CustomUser, id=user_id)
                    user.delete()
                    return JsonResponse({"status": "success", "success": True})
                except Exception as e:
                    print(f"Erro ao excluir usuário: {e}")
                    return JsonResponse(
                        {"error": f"Erro ao excluir usuário: {str(e)}"}, status=500
                    )

            elif action == "reset_password":
                try:
                    # Resetar senha
                    user_id = data.get("user_id")
                    if not user_id:
                        return JsonResponse(
                            {"error": "ID do usuário não fornecido"}, status=400
                        )

                    user = get_object_or_404(CustomUser, id=user_id)
                    new_password = "senha123"  # Senha padrão para reset
                    user.password = make_password(new_password)
                    user.save()
                    return JsonResponse({"status": "success", "success": True})
                except Exception as e:
                    print(f"Erro ao resetar senha: {e}")
                    return JsonResponse(
                        {"error": f"Erro ao resetar senha: {str(e)}"}, status=500
                    )

            else:
                return JsonResponse({"error": "Ação não reconhecida"}, status=400)

        except json.JSONDecodeError:
            return JsonResponse({"error": "Dados JSON inválidos"}, status=400)
        except Exception as e:
            print(f"Erro geral no controle de acessos: {e}")
            return JsonResponse(
                {"error": f"Erro interno do servidor: {str(e)}"}, status=500
            )

    # Filtros para GET request
    search_term = request.GET.get("search", "")
    role_filter = request.GET.get("role", "todos")

    # Query base - com tratamento de erro para IDs inválidos
    try:
        users = CustomUser.objects.only(
            'id', 'name', 'username', 'role', 'is_online', 'is_active', 'regional'
        ).all()

        # Se não for global_admin, filtra apenas os da regional dele
        if not getattr(request.user, "is_global_admin", False):
            if request.user.regional:
                users = users.filter(regional=request.user.regional)
            else:
                users = CustomUser.objects.none()
        else:
            # Admin Global: respeita o filtro de regional ativa no topo da
            # tela (Filtro Regional). Sem filtro selecionado = ve todo mundo.
            active_regional_slug = request.session.get("active_regional")
            if active_regional_slug:
                users = users.filter(regional__db_slug=active_regional_slug)

        # Aplicar filtros
        if search_term:
            users = users.filter(
                Q(name__icontains=search_term) | Q(username__icontains=search_term)
            )
        if role_filter != "todos":
            users = users.filter(role=role_filter)

        # Tentar executar a query para detectar problemas de UUID
        list(users[:1])  # Força a execução da query

    except Exception as e:
        # Se houver erro de UUID, limpar a query e exibir usuários vazios
        print(f"Erro ao buscar usuários (provavelmente IDs antigos): {e}")
        users = CustomUser.objects.none()

    # Filtrar permissões baseado no usuário logado
    user_permissions = (
        request.user.page_permissions
        if hasattr(request.user, "page_permissions")
        else {}
    )

    # Se for administrador, mostrar todas as permissões
    if request.user.role == "administrador":
        available_permissions = CustomUser.PAGE_PERMISSIONS
    else:
        # Filtrar apenas as permissões que o usuário logado tem acesso
        available_permissions = []
        for perm_code, perm_name in CustomUser.PAGE_PERMISSIONS:
            # Mostrar a permissão se o usuário logado tem acesso a ela
            if request.user.has_page_permission(perm_code):
                available_permissions.append((perm_code, perm_name))

    solicitacoes_pendentes = SolicitacaoCadastro.objects.filter(status='pendente')
    if not getattr(request.user, "is_global_admin", False):
        if request.user.regional_id:
            solicitacoes_pendentes = solicitacoes_pendentes.filter(regional_id=request.user.regional_id)
        else:
            solicitacoes_pendentes = SolicitacaoCadastro.objects.none()
    else:
        active_regional_slug = request.session.get("active_regional")
        if active_regional_slug:
            solicitacoes_pendentes = solicitacoes_pendentes.filter(regional__db_slug=active_regional_slug)
    solicitacoes_count = solicitacoes_pendentes.count()
    
    from .models import Regional
    regionais = Regional.objects.all().order_by('nome')

    context = {
        "users": users,
        "role_choices": CustomUser.ROLE_CHOICES,
        "page_permissions": available_permissions,
        "filters": {"search": search_term, "role": role_filter},
        'solicitacoes_count': solicitacoes_count,
        'regionais': regionais,
        'is_global_admin': getattr(request.user, "is_global_admin", False),
    }

    return render(request, "controle_acessos.html", context)


@login_required
def logout_view(request):
    if request.user.is_authenticated:
        # Marcar offline sem bloquear: update direto em vez de save() no objeto inteiro
        CustomUser.objects.filter(pk=request.user.pk).update(is_online=False)
        logout(request)
    return redirect("gestao_a_vista:login")


@login_required
def download_qr(request, qr_data):
    try:
        # Decodificar os dados do QR
        decoded_data = base64.b64decode(qr_data).decode()
        qr_info = json.loads(decoded_data)

        # Gerar QR code
        qr = qrcode.QRCode(version=1, box_size=10, border=5)
        qr.add_data(json.dumps(qr_info))
        qr.make(fit=True)
        qr_image = qr.make_image(fill_color="black", back_color="white")

        # Converter para PNG
        buffer = BytesIO()
        qr_image.save(buffer, format="PNG")
        buffer.seek(0)

        # Retornar como download
        response = HttpResponse(buffer, content_type="image/png")
        response[
            "Content-Disposition"
        ] = f'attachment; filename="qr-code-{qr_info["cr"]}.png"'
        return response
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@login_required
def portaria_base(request):
    """View para a página de portaria base"""
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            action = data.get("action")

            if action == "create":
                from datetime import datetime

                # Converter CPF para integer (remover formatação)
                cpf_numeric = int("".join(filter(str.isdigit, data["cpf"])))

                # Converter data_nascimento se for string
                data_nascimento = data["data_nascimento"]
                if isinstance(data_nascimento, str):
                    try:
                        # Tentar formato YYYY-MM-DD
                        data_nascimento = datetime.strptime(
                            data_nascimento, "%Y-%m-%d"
                        ).date()
                    except ValueError:
                        try:
                            # Tentar formato DD/MM/YYYY
                            data_nascimento = datetime.strptime(
                                data_nascimento, "%d/%m/%Y"
                            ).date()
                        except ValueError:
                            raise ValueError("Formato de data inválido")

                # Buscar área responsável
                try:
                    area_responsavel = AreaResponsavel.objects.get(
                        id=data["com_quem_veio_falar"]
                    )
                except AreaResponsavel.DoesNotExist:
                    raise ValueError("Área responsável não encontrada")

                # Criar nova entrada
                entrada = PortariaBase.objects.create(
                    nome=data["nome"],
                    cpf=cpf_numeric,
                    data_nascimento=data_nascimento,
                    motivo_entrada=data["motivo"],
                    area_responsavel=area_responsavel,
                    user_cadastro=request.user,
                )

                # Formatar CPF para exibição - lidar com string ou integer
                if isinstance(entrada.cpf, str):
                    cpf_numbers = "".join(filter(str.isdigit, entrada.cpf))
                    cpf_padded = cpf_numbers.zfill(11)  # Pad com zeros à esquerda
                else:
                    cpf_padded = f"{entrada.cpf:011d}"
                cpf_formatted = f"{cpf_padded[:3]}.{cpf_padded[3:6]}.{cpf_padded[6:9]}-{cpf_padded[9:]}"

                # Converter para timezone do Brasil
                data_brasilia = convert_to_brasilia_timezone(entrada.data)

                # Formatar datas para retorno
                data_nascimento_str = entrada.data_nascimento.strftime("%Y-%m-%d")
                data_registro_str = data_brasilia.strftime("%Y-%m-%d %H:%M:%S")

                return JsonResponse(
                    {
                        "success": True,
                        "message": "Entrada registrada com sucesso!",
                        "entrada": {
                            "id": str(entrada.id),  # Converter UUID para string
                            "nome": entrada.nome,
                            "cpf": cpf_formatted,
                            "data_nascimento": data_nascimento_str,
                            "motivo": entrada.motivo_entrada,
                            "com_quem_veio_falar": entrada.area_responsavel.nome,
                            "liberado_por": entrada.liberado_por,
                            "data_hora_registro": data_registro_str,
                        },
                    }
                )

        except ValueError as e:
            return JsonResponse(
                {"success": False, "message": f"Erro de validação: {str(e)}"},
                status=400,
            )
        except Exception as e:
            import traceback

            print(f"Erro ao registrar entrada: {str(e)}")
            print(f"Traceback: {traceback.format_exc()}")
            return JsonResponse(
                {"success": False, "message": f"Erro ao registrar entrada: {str(e)}"},
                status=400,
            )

    # GET request - buscar entradas
    search_term = request.GET.get("search", "")
    date_filter = request.GET.get("date", "")

    # OTIMIZADO: Query base com select_related para evitar N+1
    entradas = PortariaBase.objects.select_related(
        "area_responsavel", "user_cadastro"
    ).all()

    # Aplicar filtros
    if search_term:
        # Se o termo de busca for numérico, buscar por CPF também
        q_filter = Q(nome__icontains=search_term) | Q(
            area_responsavel__nome__icontains=search_term
        )
        if search_term.replace(".", "").replace("-", "").isdigit():
            cpf_numeric = int("".join(filter(str.isdigit, search_term)))
            q_filter |= Q(cpf=cpf_numeric)
        entradas = entradas.filter(q_filter)

    if date_filter:
        entradas = entradas.filter(data__date=date_filter)

    # OTIMIZADO: Obter entradas de hoje com select_related
    today_brasilia = convert_to_brasilia_timezone(timezone.now()).date()
    entradas_hoje = PortariaBase.objects.select_related(
        "area_responsavel", "user_cadastro"
    ).filter(data__date=today_brasilia)

    # OTIMIZADO: Cache para áreas responsáveis (dados que mudam pouco)
    from django.core.cache import cache

    areas_responsaveis = cache.get_or_set(
        "areas_responsaveis_ativas",
        lambda: list(
            AreaResponsavel.objects.filter(ativa=True).order_by("ordem", "nome")
        ),
        1800,  # Cache por 30 minutos
    )

    # Serializar entradas para uso no JavaScript
    entradas_data = []
    for entrada in entradas:
        try:
            # Formatar CPF para exibição - lidar com string ou integer
            if isinstance(entrada.cpf, str):
                cpf_numbers = "".join(filter(str.isdigit, entrada.cpf))
                cpf_padded = cpf_numbers.zfill(11)  # Pad com zeros à esquerda
            else:
                cpf_padded = f"{entrada.cpf:011d}"
            cpf_formatted = (
                f"{cpf_padded[:3]}.{cpf_padded[3:6]}.{cpf_padded[6:9]}-{cpf_padded[9:]}"
            )

            # Converter para timezone do Brasil
            data_brasilia = convert_to_brasilia_timezone(entrada.data)

            entradas_data.append(
                {
                    "id": str(entrada.id),  # Converter UUID para string
                    "nome": entrada.nome,
                    "cpf": cpf_formatted,
                    "data_nascimento": entrada.data_nascimento.strftime("%d/%m/%Y"),
                    "motivo": entrada.motivo_entrada,
                    "com_quem_veio_falar": entrada.area_responsavel.nome,
                    "data_hora_registro": data_brasilia.strftime("%d/%m/%Y às %H:%M"),
                    "data_hora_registro_iso": data_brasilia.isoformat(),
                    "liberado_por": entrada.liberado_por,
                }
            )
        except Exception as e:
            print(f"Erro ao serializar entrada {entrada.id}: {e}")
            continue

    entradas_hoje_data = []
    for entrada in entradas_hoje:
        try:
            # Formatar CPF para exibição - lidar com string ou integer
            if isinstance(entrada.cpf, str):
                cpf_numbers = "".join(filter(str.isdigit, entrada.cpf))
                cpf_padded = cpf_numbers.zfill(11)  # Pad com zeros à esquerda
            else:
                cpf_padded = f"{entrada.cpf:011d}"
            cpf_formatted = (
                f"{cpf_padded[:3]}.{cpf_padded[3:6]}.{cpf_padded[6:9]}-{cpf_padded[9:]}"
            )

            # Converter para timezone do Brasil
            data_brasilia = convert_to_brasilia_timezone(entrada.data)

            entradas_hoje_data.append(
                {
                    "id": str(entrada.id),  # Converter UUID para string
                    "nome": entrada.nome,
                    "cpf": cpf_formatted,
                    "data_nascimento": entrada.data_nascimento.strftime("%d/%m/%Y"),
                    "motivo": entrada.motivo_entrada,
                    "com_quem_veio_falar": entrada.area_responsavel.nome,
                    "data_hora_registro": data_brasilia.strftime("%H:%M"),
                    "data_hora_registro_iso": data_brasilia.isoformat(),
                    "liberado_por": entrada.liberado_por,
                }
            )
        except Exception as e:
            print(f"Erro ao serializar entrada hoje {entrada.id}: {e}")
            continue

    context = {
        "entradas": entradas,
        "entradas_hoje": entradas_hoje,
        "entradas_json": json.dumps(entradas_data),
        "entradas_hoje_json": json.dumps(entradas_hoje_data),
        "search_term": search_term,
        "date_filter": date_filter,
        "total_entradas": entradas.count(),
        "total_entradas_hoje": entradas_hoje.count(),
        "areas_responsaveis": areas_responsaveis,
    }

    return render(request, "portaria_base.html", context)


class GestaoSalasView(LoginRequiredMixin, ListView):
    model = GestaoSala
    template_name = "gestao_salas.html"
    context_object_name = "salas"

    def get_queryset(self):
        unidade_id = self.request.GET.get("unidade", 1)
        return GestaoSala.objects.filter(unidade_id=unidade_id).order_by("nome")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["unidades"] = Unidade.objects.filter(ativa=True)
        unidade_id = self.request.GET.get("unidade", 1)
        user = self.request.user
        from django.core.cache import cache
        from django.db.models import Count, Case, When, IntegerField, Q

        # 1. Capturar filtros da URL (GET)
        filtro_nome = self.request.GET.get('filter_nome', '')
        filtro_responsavel = self.request.GET.get('filter_responsavel', '')
        filtro_tipo = self.request.GET.get('filter_tipo', 'all')
        filtro_prioridade = self.request.GET.get('filter_prioridade', 'all')
        
        # 2. Query Base
        projetos_queryset = PlannerProject.objects.select_related(
            'tipo_servico'
        ).prefetch_related(
            'projeto_responsaveis__responsavel',
            'checklist_items'
        ).all()

        # 3. Aplicar Filtros
        if filtro_nome:
            projetos_queryset = projetos_queryset.filter(nome__icontains=filtro_nome)
        
        filtro_responsavel = self.request.GET.get('filter_responsavel')
        if filtro_responsavel:
            # Verificação robusta para UUID ou ID numérico
            is_valid_id = False
            if filtro_responsavel.isdigit():
                is_valid_id = True
            else:
                try:
                    uuid.UUID(str(filtro_responsavel))
                    is_valid_id = True
                except (ValueError, AttributeError):
                  
                  is_valid_id = False

            if is_valid_id:
                projetos_queryset = projetos_queryset.filter(
                    projeto_responsaveis__responsavel__id=filtro_responsavel
                )
            else:
                projetos_queryset = projetos_queryset.filter(
                    projeto_responsaveis__responsavel__first_name__icontains=filtro_responsavel
                )

            # Usar distinct() para evitar duplicatas ao filtrar Many-to-Many
            projetos_queryset = projetos_queryset.distinct()

        if filtro_tipo and filtro_tipo != 'all':
            projetos_queryset = projetos_queryset.filter(tipo_servico__nome=filtro_tipo)

        if filtro_prioridade and filtro_prioridade != 'all':
            projetos_queryset = projetos_queryset.filter(prioridade=filtro_prioridade)

        # Ordenação padrão
        projetos_queryset = projetos_queryset.order_by('-prioridade', 'data_conclusao')

        # 4. Calcular Contadores (Baseado no filtro atual ou total?)
        # Se quiser que os cards do topo mudem com o filtro, use o queryset filtrado.
        # Se quiser totais fixos, mantenha o cache antigo. Vamos atualizar com o filtro:
        status_counts = projetos_queryset.aggregate(
            ativo_count=Count(Case(When(status='Ativo', then=1), output_field=IntegerField())),
            andamento_count=Count(Case(When(status='Em andamento', then=1), output_field=IntegerField())),
            pausado_count=Count(Case(When(status='Pausado', then=1), output_field=IntegerField())),
            concluido_count=Count(Case(When(status='Concluído', then=1), output_field=IntegerField())),
        )

        # 5. Preparar dados para o Template
        class StatusGroup:
            def __init__(self, count):
                self.count = count

        projetos_por_status = {
            "Ativo": StatusGroup(status_counts['ativo_count']),
            "Em_andamento": StatusGroup(status_counts['andamento_count']),
            "Pausado": StatusGroup(status_counts['pausado_count']),
            "Concluido": StatusGroup(status_counts['concluido_count']),
        }

        # Carregar listas auxiliares (Consulta direta para evitar cache obsoleto)
        tipos_servico_cached = list(TipoServico.objects.filter(ativo=True).values('id', 'nome'))
        usuarios_cached = list(CustomUser.objects.filter(is_active=True).values('id', 'name', 'username'))

        import time
        context.update({
            "projetos": list(projetos_queryset), # Converte o queryset filtrado para lista
            "projetos_por_status": projetos_por_status,
            "tipos_servico": tipos_servico_cached,
            "usuarios": usuarios_cached,
            "timestamp": int(time.time()),
            # Passar filtros atuais para manter o form preenchido
            "filtros": {
                "nome": filtro_nome,
                "responsavel": filtro_responsavel,
                "tipo": filtro_tipo,
                "prioridade": filtro_prioridade
            }
        })

        return context


class SalaCriarView(LoginRequiredMixin, CreateView):
    model = GestaoSala
    fields = [
        "nome",
        "capacidade",
        "hora_inicio",
        "hora_fim",
        "quantidade_m",
        "foto",
        "unidade",
    ]
    template_name = "sala_form.html"

    def get_success_url(self):
        return (
            reverse_lazy("gestao_a_vista:gestao_salas")
            + f"?unidade={self.object.unidade.id}"
        )

    def get_initial(self):
        initial = super().get_initial()
        unidade_id = self.request.GET.get("unidade", 1)
        initial["unidade"] = unidade_id
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["unidades"] = Unidade.objects.filter(ativa=True)
        return context

    def form_valid(self, form):
        """
        Chama a validação customizada antes de salvar
        """
        try:
            form.instance.clean()
            return super().form_valid(form)
        except ValidationError as e:
            form.add_error(None, e)
            return self.form_invalid(form)


class SalaEditarView(LoginRequiredMixin, UpdateView):
    model = GestaoSala
    fields = [
        "nome",
        "capacidade",
        "hora_inicio",
        "hora_fim",
        "quantidade_m",
        "foto",
        "unidade",
    ]
    template_name = "sala_editar.html"

    def get_success_url(self):
        return (
            reverse_lazy("gestao_a_vista:gestao_salas")
            + f"?unidade={self.object.unidade.id}"
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["unidades"] = Unidade.objects.filter(ativa=True)
        return context

    def form_valid(self, form):
        """
        Chama a validação customizada antes de salvar
        """
        try:
            form.instance.clean()
            return super().form_valid(form)
        except ValidationError as e:
            form.add_error(None, e)
            return self.form_invalid(form)


class SalaDeletarView(LoginRequiredMixin, DeleteView):
    model = GestaoSala
    template_name = "sala_confirm_delete.html"

    def get_success_url(self):
        unidade_id = self.object.unidade.id
        return reverse_lazy("gestao_a_vista:gestao_salas") + f"?unidade={unidade_id}"
    
    def delete(self, request, *args, **kwargs):
        """
        Sobrescreve o método delete para remover reservas relacionadas
        """
        try:
            from .audit_utils import registrar_acao_auditoria
            
            self.object = self.get_object()
            sala_nome = self.object.nome
            sala_id = str(self.object.id_sala)
            
            # Registrar auditoria antes da exclusão
            registrar_acao_auditoria(
                usuario=request.user,
                acao="excluiu",
                tipo_item="sala",
                item_id=sala_id,
                detalhes=f"Sala '{sala_nome}' foi excluída pelo usuário {request.user.username}",
                status_anterior="ativa",
                status_novo="excluída"
            )
            
            # Remove reservas relacionadas à sala
            reservas_removidas = ReservaSala.objects.filter(sala_id=self.object.id_sala).count()
            ReservaSala.objects.filter(sala_id=self.object.id_sala).delete()
            
            # Remove mesas relacionadas à sala
            mesas_removidas = GestaoMesa.objects.filter(id_sala=self.object).count()
            GestaoMesa.objects.filter(id_sala=self.object).delete()
            
            success_url = self.get_success_url()
            self.object.delete()
            
            # Registrar detalhes adicionais da exclusão
            if reservas_removidas > 0 or mesas_removidas > 0:
                registrar_acao_auditoria(
                    usuario=request.user,
                    acao="limpeza",
                    tipo_item="sala",
                    item_id=sala_id,
                    detalhes=f"Removidas {reservas_removidas} reservas e {mesas_removidas} mesas relacionadas à sala '{sala_nome}'"
                )
            
            messages.success(request, f'Sala "{sala_nome}" foi excluída com sucesso.')
            return HttpResponseRedirect(success_url)
            
        except Exception as e:
            messages.error(request, f'Erro ao excluir a sala: {str(e)}')
            return HttpResponseRedirect(self.get_success_url())



# Endpoints da API para gerenciamento de mesas
def adicionar_mesa(request, sala_id):
    if request.method == "POST":
        sala = get_object_or_404(GestaoSala, id_sala=sala_id)
        mesa_nome = request.POST.get("nome")
        mesa = GestaoMesa.objects.create(mesa=mesa_nome, id_sala=sala)
        return JsonResponse({"id": mesa.id_mesa, "nome": mesa.mesa})
    return JsonResponse({"erro": "Requisição inválida"}, status=400)


def remover_mesa(request, mesa_id):
    if request.method == "POST":
        mesa = get_object_or_404(GestaoMesa, id_mesa=mesa_id)
        mesa.delete()
        return JsonResponse({"sucesso": True})
    return JsonResponse({"erro": "Requisição inválida"}, status=400)


@method_decorator(xframe_options_sameorigin, name='dispatch')
class LivroAtaView(TemplateView):
    template_name = "livro_ata.html"

    def get(self, request, *args, **kwargs):
        action = kwargs.get("action")

        if action == "get_shifts":
            return self.get_shifts(request)
        elif action == "get_shift_details":
            shift_id = kwargs.get("shift_id")
            return self.get_shift_details(request, shift_id)
        elif action == "get_relatorio":
            return self.get_relatorio(request)
        elif action == "get_gerentes":
            return self.get_gerentes(request)
        elif action == "get_relatorio_mensal":
            return self.get_relatorio_mensal(request)
        elif action == "get_relatorio_consolidado":
            return self.get_relatorio_consolidado(request)

        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Verificar se há parâmetro qrcode na URL (GET param ou URL param)
        qrcode_id = self.request.GET.get('qrcode') or kwargs.get('qrcode_id')
        if qrcode_id:
            # LIMPEZA: Remove a barra final se ela vier colada
            context['qrcode_id'] = str(qrcode_id).replace('/', '').strip()
        return context

    def get_shifts(self, request):
        cr_number = request.GET.get("cr_number")
        qrcode_id = request.GET.get("qrcode")

        # Se veio de QR Code, buscar o cr_id correspondente
        if qrcode_id:
            try:
                # LIMPEZA: Remove a barra final para não quebrar a busca do UUID no banco
                qrcode_cleaned = str(qrcode_id).replace('/', '').strip()
                
                from .models import LivroAtaQRCode
                
                cr_id = None
                
                # 1. Tentar buscar pelo ID do LivroAtaQRCode (padrão atual)
                try:
                    livro_ata_qr = LivroAtaQRCode.objects.using('default').get(id=qrcode_cleaned)
                    cr_id = livro_ata_qr.cr_id
                except LivroAtaQRCode.DoesNotExist:
                    # 2. Tentar buscar onde o cr_id seja igual ao qrcode (QR codes gerados com erro anteriormente)
                    livro_ata_qr = LivroAtaQRCode.objects.using('default').filter(cr_id=qrcode_cleaned).first()
                    if livro_ata_qr:
                        cr_id = livro_ata_qr.cr_id
                    else:
                        # 3. Tentar buscar direto na tabela Estrutura (compatibilidade máxima)
                        if Estrutura.objects.using('default').filter(id=qrcode_cleaned).exists():
                            cr_id = qrcode_cleaned
                            
                if cr_id:
                    return self.get_shifts_from_dw_vista(request, cr_id)
                else:
                    return JsonResponse({"error": "QR Code não encontrado no banco de dados."}, status=404)
                    
            except Exception as e:
                return JsonResponse({"error": f"Erro ao processar QR Code: {str(e)}"}, status=500)

        # Fluxo por CR number - buscar o cr_id da Estrutura
        if not cr_number:
            return JsonResponse({"error": "Número do CR é obrigatório"}, status=400)

        try:
            # Buscar TODAS as estruturas pelo campo 'cr' (que armazena o número do CR com formatação)
            # O campo cr contém algo como "12981 - GO - SEG - ORION BUSINESS - GOIANIA - (PRG)"
            # Então usamos startswith para buscar apenas pelo número
            estruturas = list(Estrutura.objects.using('default').filter(cr__startswith=cr_number.upper()).values_list('id', flat=True))

            if not estruturas:
                return JsonResponse({"error": f"CR '{cr_number.upper()}' não encontrado no sistema"}, status=404)

            # Buscar plantões de TODAS as estruturas usando a mesma lógica do QR Code
            all_shifts = []

            for cr_id in estruturas:
                # Usar a mesma função que funciona para QR Code
                shifts_response = self.get_shifts_from_dw_vista(request, cr_id)

                # Extrair os dados da resposta JsonResponse
                import json
                shifts_data = json.loads(shifts_response.content.decode())

                if 'shifts' in shifts_data:
                    all_shifts.extend(shifts_data['shifts'])

            # Ordenar por data/hora mais recente
            all_shifts.sort(
                key=lambda x: (x.get('date', ''), x.get('time', '')),
                reverse=True
            )

            return JsonResponse({"shifts": all_shifts})
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Erro ao buscar estrutura para CR {cr_number}: {e}")
            return JsonResponse({"error": f"Erro ao processar CR: {str(e)}"}, status=500)

    def get_shifts_from_dw_vista(self, request, cr_id):
        """
        Consulta dados do DW_Vista usando o cr_id da estrutura
        """
        from django.db import connections
        from django.core.cache import cache

        # Cache de 5 minutos para dados do DW_Vista (VPN) por CR
        cache_key = f"dw_vista_shifts_{cr_id}"
        cached = cache.get(cache_key)
        if cached is not None:
            return JsonResponse({"shifts": cached})

        cursor = None
        try:
            # Usar conexão readonly (DW_Vista)
            db_conn_name = 'dw_vpn' if 'dw_vpn' in connections else ('readonly' if 'readonly' in connections else 'default')
            cursor = connections[db_conn_name].cursor()
            
            # Query atualizada com nome do colaborador e coluna expirada
            query = """
            WITH ultimas_tarefas AS (
              SELECT t.id
              FROM dbo.tarefa t
              INNER JOIN dbo.checklist c ON c.id = t.checklistid
              WHERE c.id = '6687b862-10d0-4144-ae30-8bdc55f22ee3'
                AND t.status = 85
                AND t.estruturaid = %s 
              ORDER BY t.terminoreal DESC
              LIMIT 4
            )
            SELECT
              t.id,
              t.nome,
              t.numero,
              t.inicio,
              t.terminoreal,
              t.estruturaqrcode,
              t.expirada,
              e.descricao AS estrutura_descricao,
              r.nome AS colaborador,
              ex.perguntadescricao,
              ex.conteudo
            FROM dbo.tarefa t
            INNER JOIN ultimas_tarefas u ON u.id = t.id
            LEFT JOIN dbo.execucao ex ON ex.tarefaid = t.id
            INNER JOIN dbo.checklist c ON c.id = t.checklistid
            INNER JOIN dbo.estrutura e ON e.id = t.estruturaid
            INNER JOIN dbo.recurso r ON r.codigohash = t.finalizadoporhash
            """
            
            cursor.execute(query, [cr_id])
            results = cursor.fetchall()
            
            # Processar resultados
            shifts_data = self.process_dw_vista_results(results)
            
            cache.set(cache_key, shifts_data, 300)  # Cache 5 minutos
            return JsonResponse({
                "shifts": shifts_data,
                "debug": {
                    "cr_id": cr_id,
                    "total_records": len(results),
                    "found_shifts": len(shifts_data)
                }
            })
            
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Erro ao consultar DW_Vista: {e}")
            
            # Se a conexão com DW_Vista falhar, retornar dados vazios em vez de erro
            # Isso permite que a interface funcione mesmo sem acesso ao DW_Vista
            return JsonResponse({
                "shifts": [],
                "error": f"Serviço temporariamente indisponível. Tente novamente em alguns minutos.",
                "debug": {
                    "cr_id": cr_id,
                    "connection_error": str(e)
                }
            })
        finally:
            # Fechar cursor apenas se foi inicializado
            if cursor is not None:
                try:
                    cursor.close()
                except:
                    pass  # Ignorar erros ao fechar cursor

    def get_relatorio(self, request):
        """
        Gera um PDF com os plantões dos últimos 30 dias, seguindo o mesmo
        critério de busca usado na tela (número do CR ou QR Code).
        """
        cr_number = request.GET.get("cr_number")
        qrcode_id = request.GET.get("qrcode")

        try:
            cr_ids = []
            titulo_busca = ""

            if qrcode_id:
                qrcode_cleaned = str(qrcode_id).replace('/', '').strip()
                from .models import LivroAtaQRCode

                cr_id = None
                try:
                    livro_ata_qr = LivroAtaQRCode.objects.using('default').get(id=qrcode_cleaned)
                    cr_id = livro_ata_qr.cr_id
                except LivroAtaQRCode.DoesNotExist:
                    livro_ata_qr = LivroAtaQRCode.objects.using('default').filter(cr_id=qrcode_cleaned).first()
                    if livro_ata_qr:
                        cr_id = livro_ata_qr.cr_id
                    elif Estrutura.objects.using('default').filter(id=qrcode_cleaned).exists():
                        cr_id = qrcode_cleaned

                if not cr_id:
                    return JsonResponse({"error": "QR Code não encontrado no banco de dados."}, status=404)

                cr_ids = [cr_id]
                titulo_busca = f"QR Code: {qrcode_cleaned}"
            else:
                if not cr_number:
                    return JsonResponse({"error": "Número do CR é obrigatório"}, status=400)

                cr_ids = list(Estrutura.objects.using('default').filter(cr__startswith=cr_number.upper()).values_list('id', flat=True))
                if not cr_ids:
                    return JsonResponse({"error": f"CR '{cr_number.upper()}' não encontrado no sistema"}, status=404)

                titulo_busca = f"CR: {cr_number.upper()}"
        except Exception as e:
            return JsonResponse({"error": f"Erro ao localizar CR: {str(e)}"}, status=500)

        all_shifts = []
        for cr_id in cr_ids:
            all_shifts.extend(self.get_shifts_relatorio_from_dw_vista(cr_id, dias=30))

        all_shifts.sort(key=lambda x: (x.get('date') or '', x.get('time') or ''), reverse=True)

        return self.gerar_pdf_relatorio(titulo_busca, all_shifts)

    def get_shifts_relatorio_from_dw_vista(self, cr_id, dias=30):
        """
        Consulta o DW_Vista buscando TODOS os plantões de uma estrutura
        (sem limite de 4) dentro da janela de dias informada.
        """
        from django.db import connections
        from datetime import datetime, timedelta
        import logging

        cursor = None
        try:
            db_conn_name = 'dw_vpn' if 'dw_vpn' in connections else ('readonly' if 'readonly' in connections else 'default')
            cursor = connections[db_conn_name].cursor()

            data_limite = (datetime.now() - timedelta(days=dias)).strftime('%Y-%m-%d %H:%M:%S')

            query = """
            SELECT
              t.id,
              t.nome,
              t.numero,
              t.inicio,
              t.terminoreal,
              t.estruturaqrcode,
              t.expirada,
              e.descricao AS estrutura_descricao,
              r.nome AS colaborador,
              ex.perguntadescricao,
              ex.conteudo
            FROM dbo.tarefa t
            LEFT JOIN dbo.execucao ex ON ex.tarefaid = t.id
            INNER JOIN dbo.checklist c ON c.id = t.checklistid
            INNER JOIN dbo.estrutura e ON e.id = t.estruturaid
            INNER JOIN dbo.recurso r ON r.codigohash = t.finalizadoporhash
            WHERE c.id = '6687b862-10d0-4144-ae30-8bdc55f22ee3'
              AND t.status = 85
              AND t.estruturaid = %s
              AND (CASE WHEN t.expirada THEN t.inicio ELSE t.terminoreal END) >= %s
            ORDER BY t.terminoreal DESC
            """

            cursor.execute(query, [cr_id, data_limite])
            results = cursor.fetchall()

            return self.process_dw_vista_results(results)
        except Exception as e:
            logging.getLogger(__name__).error(f"Erro ao consultar relatório do DW_Vista para CR {cr_id}: {e}")
            return []
        finally:
            if cursor is not None:
                try:
                    cursor.close()
                except:
                    pass

    def gerar_pdf_relatorio(self, titulo_busca, shifts):
        """
        Monta o PDF do relatório de 30 dias com os plantões encontrados.
        """
        from reportlab.lib import colors
        from django.utils import timezone
        from datetime import datetime

        agora = timezone.now()
        response = HttpResponse(content_type='application/pdf')
        filename = f"Relatorio_Livro_Ata_{agora.strftime('%Y%m%d_%H%M')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        c = canvas.Canvas(response, pagesize=A4)
        largura, altura = A4

        def desenhar_header():
            c.setFillColorRGB(15/255, 23/255, 42/255)  # #0f172a
            c.rect(0, altura - 80, largura, 80, stroke=0, fill=1)
            c.setFillColor(colors.white)
            c.setFont("Helvetica-Bold", 16)
            c.drawString(40, altura - 40, "GRUPO EXEMPLO - RELATORIO DO LIVRO ATA")
            c.setFont("Helvetica", 10)
            c.drawString(40, altura - 58, f"{titulo_busca}  |  Periodo: ultimos 30 dias")
            c.drawString(40, altura - 72, f"Gerado em: {agora.strftime('%d/%m/%Y %H:%M')}")

        def desenhar_colunas(y):
            c.setFillColorRGB(15/255, 23/255, 42/255)
            c.setFont("Helvetica-Bold", 9)
            c.drawString(40, y, "DATA")
            c.drawString(90, y, "HORA")
            c.drawString(130, y, "COLABORADOR")
            c.drawString(340, y, "TIPO")
            c.drawString(420, y, "STATUS")
            y -= 8
            c.setStrokeColorRGB(15/255, 23/255, 42/255)
            c.line(40, y, largura - 40, y)
            return y

        desenhar_header()
        y = altura - 110
        c.setFillColorRGB(30/255, 64/255, 175/255)  # #1e40af
        c.setFont("Helvetica-Bold", 12)
        c.drawString(40, y, f"PLANTOES REGISTRADOS ({len(shifts)})")
        y -= 25
        y = desenhar_colunas(y)
        y -= 15

        c.setFont("Helvetica", 8)
        c.setFillColor(colors.black)

        if not shifts:
            c.drawString(40, y, "Nenhum plantao encontrado no periodo.")
        else:
            for shift in shifts:
                data_fmt = shift.get('date') or ''
                if data_fmt:
                    try:
                        data_fmt = datetime.strptime(data_fmt, '%Y-%m-%d').strftime('%d/%m/%Y')
                    except Exception:
                        pass
                hora_fmt = (shift.get('time') or '')[:5]
                expirada = bool(shift.get('expirada'))

                c.setFillColor(colors.black)
                c.drawString(40, y, data_fmt)
                c.drawString(90, y, hora_fmt)
                c.drawString(130, y, (shift.get('guard_name') or '')[:45])
                c.drawString(340, y, (shift.get('shift_type') or '')[:14])
                c.setFillColor(colors.red if expirada else colors.green)
                c.drawString(420, y, "NAO REALIZADO" if expirada else "REALIZADO")

                y -= 16
                if y < 60:
                    c.showPage()
                    desenhar_header()
                    y = altura - 110
                    y = desenhar_colunas(y)
                    y -= 15
                    c.setFont("Helvetica", 8)
                    c.setFillColor(colors.black)

        c.setFont("Helvetica-Oblique", 8)
        c.setFillColor(colors.gray)
        c.drawString(40, 30, "Documento gerado pelo sistema de Gestao a Vista - Grupo Exemplo.")

        c.showPage()
        c.save()
        return response

    def get_gerentes(self, request):
        """Lista os gerentes ativos, usada para popular o filtro de busca por Gerente."""
        gerentes = (
            CustomUser.objects.using('default')
            .filter(role='gerente', is_active=True)
            .order_by('name')
        )
        return JsonResponse({
            "gerentes": [{"id": str(g.id), "name": g.name or g.username} for g in gerentes]
        })

    def _crs_do_gerente(self, gerente):
        """
        Retorna [(cr_id, cr_nome)] das estruturas com Livro Ata cadastrado que o
        gerente cobre. Fonte principal: campo Estrutura.gc (Gerente do Contrato),
        que já vem preenchido pela base da Estrutura e não depende de cadastro
        manual. Também aceita o campo CustomUser.crs como complemento, para
        gerentes cujo vínculo não esteja refletido em gc.
        """
        import re

        from .models import Estrutura, LivroAtaQRCode

        livro_atas = list(LivroAtaQRCode.objects.using('default').all())
        if gerente.is_general:
            return [(la.cr_id, la.cr_descricao or la.cr_id) for la in livro_atas]

        nome_gerente = (gerente.name or '').strip()
        cr_ids_por_gc = set()
        if nome_gerente:
            cr_ids_por_gc = {
                str(cr_id) for cr_id in
                Estrutura.objects.using('default')
                .filter(id__in=[la.cr_id for la in livro_atas], gc__iexact=nome_gerente)
                .values_list('id', flat=True)
            }

        crs_nums, crs_texto = [], []
        if gerente.crs:
            crs_nums = [m.group() for c in gerente.crs.split(',') if (m := re.search(r'\d+', c))]
            crs_texto = [c.strip() for c in gerente.crs.split(',') if c.strip()]

        resultado = []
        for livro_ata in livro_atas:
            nome = livro_ata.cr_descricao or livro_ata.cr_id
            match = livro_ata.cr_id in cr_ids_por_gc
            if not match and crs_nums:
                num_match = re.search(r'\d+', str(nome))
                num = num_match.group() if num_match else None
                match = bool(num and num in crs_nums)
            if not match and crs_texto:
                match = nome.strip() in crs_texto
            if match:
                resultado.append((livro_ata.cr_id, nome))
        return resultado

    def _turnos_feitos_no_periodo(self, cr_id, inicio, fim):
        """
        Retorna um set de (data, turno) com os turnos concluídos (não expirados)
        de um CR entre [inicio, fim). 'turno' é 'diurno' (06h-18h) ou 'noturno'
        (18h de um dia até 06h do dia seguinte, ancorado no dia em que começa).
        """
        from datetime import time as time_type, timedelta as td
        from django.db import connections

        db_conn_name = 'dw_vpn' if 'dw_vpn' in connections else ('readonly' if 'readonly' in connections else 'default')
        cursor = connections[db_conn_name].cursor()
        try:
            query = """
                SELECT t.terminoreal
                FROM dbo.tarefa t
                INNER JOIN dbo.checklist c ON c.id = t.checklistid
                WHERE c.id = %s
                  AND t.status = 85
                  AND t.estruturaid = %s
                  AND (t.expirada = false OR t.expirada IS NULL)
                  AND t.terminoreal >= %s
                  AND t.terminoreal < %s
            """
            cursor.execute(query, [
                '6687b862-10d0-4144-ae30-8bdc55f22ee3',
                cr_id,
                inicio.strftime('%Y-%m-%d %H:%M:%S'),
                fim.strftime('%Y-%m-%d %H:%M:%S'),
            ])
            feitos = set()
            for (terminoreal,) in cursor.fetchall():
                if terminoreal is None:
                    continue
                hora = terminoreal.time()
                if time_type(6, 0) <= hora < time_type(18, 0):
                    feitos.add((terminoreal.date(), 'diurno'))
                elif hora >= time_type(18, 0):
                    feitos.add((terminoreal.date(), 'noturno'))
                else:
                    feitos.add((terminoreal.date() - td(days=1), 'noturno'))
            return feitos
        finally:
            cursor.close()

    def get_relatorio_mensal(self, request):
        """
        Retorna, para os CRs cobertos por um gerente, a situação (feito/não
        feito) de cada turno diurno/noturno ao longo de um mês - usado pelo
        filtro de busca por Gerente na tela do Livro Ata.
        """
        from datetime import datetime as dt, timedelta as td

        gerente_id = request.GET.get('gerente_id')
        mes_str = request.GET.get('mes')  # formato YYYY-MM, opcional

        if not gerente_id:
            return JsonResponse({"error": "gerente_id é obrigatório"}, status=400)

        try:
            gerente = CustomUser.objects.using('default').get(id=gerente_id, role='gerente')
        except CustomUser.DoesNotExist:
            return JsonResponse({"error": "Gerente não encontrado"}, status=404)

        hoje = dt.now()
        if mes_str:
            try:
                ano, mes = map(int, mes_str.split('-'))
            except ValueError:
                return JsonResponse({"error": "Parâmetro 'mes' inválido. Use o formato YYYY-MM"}, status=400)
        else:
            ano, mes = hoje.year, hoje.month

        inicio_mes = dt(ano, mes, 1)
        fim_mes = dt(ano + 1, 1, 1) if mes == 12 else dt(ano, mes + 1, 1)
        hoje_00h = hoje.replace(hour=0, minute=0, second=0, microsecond=0)
        limite = min(fim_mes, hoje_00h + td(days=1))

        crs = self._crs_do_gerente(gerente)
        if not crs or limite <= inicio_mes:
            return JsonResponse({"gerente": gerente.name, "mes": f"{ano:04d}-{mes:02d}", "crs": []})

        dias_no_periodo = (limite - inicio_mes).days

        resultado_crs = []
        for cr_id, cr_nome in crs:
            feitos = self._turnos_feitos_no_periodo(cr_id, inicio_mes, fim_mes)
            dias = []
            for i in range(dias_no_periodo):
                dia_data = (inicio_mes + td(days=i)).date()
                dias.append({
                    "dia": dia_data.day,
                    "diurno": (dia_data, 'diurno') in feitos,
                    "noturno": (dia_data, 'noturno') in feitos,
                })
            resultado_crs.append({"cr_id": cr_id, "cr_nome": cr_nome, "dias": dias})

        return JsonResponse({
            "gerente": gerente.name,
            "mes": f"{ano:04d}-{mes:02d}",
            "crs": resultado_crs,
        })

    def get_relatorio_consolidado(self, request):
        """
        PDF consolidado dos últimos 30 dias: uma linha por CR com o total de
        turnos diurnos/noturnos feitos, em vez do detalhe plantão a plantão.
        Se gerente_id não for informado, consolida todos os CRs com Livro Ata cadastrado.
        """
        from datetime import datetime as dt, timedelta as td

        from .models import LivroAtaQRCode

        gerente_id = request.GET.get('gerente_id')
        gerente_nome = "Todos os CRs"

        if gerente_id:
            try:
                gerente = CustomUser.objects.using('default').get(id=gerente_id, role='gerente')
            except CustomUser.DoesNotExist:
                return JsonResponse({"error": "Gerente não encontrado"}, status=404)
            crs = self._crs_do_gerente(gerente)
            gerente_nome = gerente.name or gerente.username
        else:
            crs = [
                (la.cr_id, la.cr_descricao or la.cr_id)
                for la in LivroAtaQRCode.objects.using('default').all()
            ]

        if not crs:
            return JsonResponse({"error": "Nenhum CR encontrado para este gerente."}, status=404)

        fim = dt.now()
        inicio = fim - td(days=30)
        dias_corridos = 30
        esperado_por_turno = dias_corridos

        linhas = []
        for cr_id, cr_nome in crs:
            feitos = self._turnos_feitos_no_periodo(cr_id, inicio, fim)
            diurnos_feitos = sum(1 for (_, turno) in feitos if turno == 'diurno')
            noturnos_feitos = sum(1 for (_, turno) in feitos if turno == 'noturno')
            total_feito = diurnos_feitos + noturnos_feitos
            total_esperado = esperado_por_turno * 2
            percentual = (total_feito / total_esperado * 100) if total_esperado else 0
            linhas.append({
                "cr_nome": cr_nome,
                "diurnos_feitos": diurnos_feitos,
                "noturnos_feitos": noturnos_feitos,
                "total_feito": total_feito,
                "total_esperado": total_esperado,
                "percentual": percentual,
            })

        return self.gerar_pdf_relatorio_consolidado(gerente_nome, linhas)

    def gerar_pdf_relatorio_consolidado(self, titulo, linhas):
        """Monta o PDF consolidado (um CR por linha) dos últimos 30 dias."""
        from reportlab.lib import colors
        from django.utils import timezone

        agora = timezone.now()
        response = HttpResponse(content_type='application/pdf')
        filename = f"Relatorio_Consolidado_Livro_Ata_{agora.strftime('%Y%m%d_%H%M')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        c = canvas.Canvas(response, pagesize=A4)
        largura, altura = A4

        def desenhar_header():
            c.setFillColorRGB(15/255, 23/255, 42/255)
            c.rect(0, altura - 80, largura, 80, stroke=0, fill=1)
            c.setFillColor(colors.white)
            c.setFont("Helvetica-Bold", 16)
            c.drawString(40, altura - 40, "GRUPO EXEMPLO - RELATORIO CONSOLIDADO DO LIVRO ATA")
            c.setFont("Helvetica", 10)
            c.drawString(40, altura - 58, f"{titulo}  |  Periodo: ultimos 30 dias")
            c.drawString(40, altura - 72, f"Gerado em: {agora.strftime('%d/%m/%Y %H:%M')}")

        def desenhar_colunas(y):
            c.setFillColorRGB(15/255, 23/255, 42/255)
            c.setFont("Helvetica-Bold", 9)
            c.drawString(40, y, "CR")
            c.drawString(300, y, "DIURNOS")
            c.drawString(360, y, "NOTURNOS")
            c.drawString(430, y, "TOTAL")
            c.drawString(490, y, "CONFORMIDADE")
            y -= 8
            c.setStrokeColorRGB(15/255, 23/255, 42/255)
            c.line(40, y, largura - 40, y)
            return y

        desenhar_header()
        y = altura - 110
        c.setFillColorRGB(30/255, 64/255, 175/255)
        c.setFont("Helvetica-Bold", 12)
        c.drawString(40, y, f"CRs NO RELATORIO ({len(linhas)})")
        y -= 25
        y = desenhar_colunas(y)
        y -= 15

        c.setFont("Helvetica", 8)
        c.setFillColor(colors.black)

        for linha in linhas:
            c.setFillColor(colors.black)
            c.drawString(40, y, str(linha["cr_nome"])[:55])
            c.drawString(300, y, str(linha["diurnos_feitos"]))
            c.drawString(360, y, str(linha["noturnos_feitos"]))
            c.drawString(430, y, f'{linha["total_feito"]}/{linha["total_esperado"]}')
            percentual = linha["percentual"]
            c.setFillColor(colors.green if percentual >= 80 else (colors.orange if percentual >= 50 else colors.red))
            c.drawString(490, y, f'{percentual:.0f}%')

            y -= 16
            if y < 60:
                c.showPage()
                desenhar_header()
                y = altura - 110
                y = desenhar_colunas(y)
                y -= 15
                c.setFont("Helvetica", 8)
                c.setFillColor(colors.black)

        c.setFont("Helvetica-Oblique", 8)
        c.setFillColor(colors.gray)
        c.drawString(40, 30, "Documento gerado pelo sistema de Gestao a Vista - Grupo Exemplo.")

        c.showPage()
        c.save()
        return response

    def process_dw_vista_results(self, results):
        """
        Processa os resultados da consulta DW_Vista e organiza por plantão
        """
        from collections import defaultdict
        import re
        
        # Organizar dados por tarefa
        tasks_data = defaultdict(lambda: {
            'task_info': None,
            'executions': []
        })
        
        for row in results:
            task_id, nome, numero, inicio, terminoreal, estruturaqrcode, expirada, estrutura_descricao, colaborador, perguntadescricao, conteudo = row
            
            # Informações da tarefa
            if not tasks_data[task_id]['task_info']:
                tasks_data[task_id]['task_info'] = {
                    'id': task_id,
                    'nome': nome,
                    'numero': numero,
                    'inicio': inicio,
                    'terminoreal': terminoreal,
                    'estruturaqrcode': estruturaqrcode,
                    'expirada': expirada,
                    'estrutura_descricao': estrutura_descricao,
                    'colaborador': colaborador
                }
            
            # Execuções (perguntas e respostas)
            if perguntadescricao and conteudo:
                tasks_data[task_id]['executions'].append({
                    'pergunta': perguntadescricao,
                    'conteudo': conteudo
                })
        
        # Converter para formato de plantões
        shifts = []
        for task_id, task_data in tasks_data.items():
            task_info = task_data['task_info']
            executions = task_data['executions']
            
            # Extrair tipo de plantão do nome (após o hífen)
            shift_type = "DIURNO"  # default
            if task_info['nome']:
                match = re.search(r'-\s*(.+)$', task_info['nome'])
                if match:
                    shift_type = match.group(1).strip()
            
            # Buscar descrição do plantão
            description = ""
            for exec_item in executions:
                if exec_item['pergunta'] == "RELATE INFORMAÇÕES DO PLANTÃO:":
                    description = exec_item['conteudo']
                    break
            
            # Processar data/hora - formato: "2025-10-21 18:51:49.074"
            # Para registros expirados, usar data de início; para realizados, usar terminoreal
            date_str = None
            time_str = None
            
            # Escolher qual data usar baseado no status expirado
            date_source = task_info['inicio'] if task_info.get('expirada') else task_info['terminoreal']
            
            if date_source:
                from datetime import datetime
                try:
                    # Converter string para datetime
                    date_source_str = str(date_source)
                    
                    # Remover microsegundos se existirem
                    if '.' in date_source_str:
                        date_source_str = date_source_str.split('.')[0]
                    
                    # Parse da data/hora
                    dt = datetime.strptime(date_source_str, '%Y-%m-%d %H:%M:%S')
                    
                    # Formato ISO para JavaScript
                    date_str = dt.date().isoformat()  # YYYY-MM-DD
                    time_str = dt.time().isoformat()  # HH:MM:SS
                    
                except Exception as e:
                    # Fallback para formato original
                    print(f"Erro ao processar data/hora {date_source}: {e}")
                    date_str = str(date_source)[:10]
                    time_str = str(date_source)[11:19]

            # Criar dados do plantão
            shift_data = {
                "id": str(task_id),
                "cr_number": task_info['estruturaqrcode'] or "N/A",
                "guard_name": task_info['colaborador'] or "Colaborador não identificado",
                "guard_number": task_info['numero'],
                "date": date_str,
                "time": time_str,
                "shift_type": shift_type,
                "location": task_info['estrutura_descricao'],
                "description": description,
                "expirada": task_info['expirada'],
                "executions": executions  # Dados adicionais para detalhes
            }
            
            shifts.append(shift_data)
        
        return shifts

    def get_example_shift_details(self, request, shift_id):
        """
        Retorna detalhes de exemplo para testes
        """
        from datetime import date, time
        
        example_data = {
            "example-1": {
                "id": "example-1",
                "cr_number": "CR-001",
                "guard_name": "João Silva",
                "guard_number": "12345",
                "date": date.today().isoformat(),
                "time": time(8, 0).isoformat(),
                "shift_type": "diurno",
                "location": "Portaria Principal",
                "description": "Plantão de exemplo para testes",
                "evidences": [
                    {
                        "id": "evidence-1",
                        "image_url": None,
                        "description": "Evidência de exemplo 1"
                    }
                ],
                "compliance_items": [
                    {
                        "id": "compliance-1",
                        "description": "Item de compliance de exemplo",
                        "status": "completed"
                    }
                ]
            },
            "example-2": {
                "id": "example-2",
                "cr_number": "CR-001",
                "guard_name": "Maria Santos",
                "guard_number": "67890",
                "date": date.today().isoformat(),
                "time": time(20, 0).isoformat(),
                "shift_type": "noturno",
                "location": "Portaria Principal",
                "description": "Plantão noturno de exemplo",
                "evidences": [],
                "compliance_items": []
            }
        }
        
        if shift_id in example_data:
            return JsonResponse({"shift": example_data[shift_id]})
        else:
            return JsonResponse({"error": "Plantão de exemplo não encontrado"}, status=404)

    def get_shift_details(self, request, shift_id):
        # Verificar se é um ID de exemplo (para testes)
        if str(shift_id).startswith('example-'):
            return self.get_example_shift_details(request, shift_id)
        
        # Verificar se é um ID de tarefa do DW_Vista (UUID format)
        try:
            import uuid
            # Se shift_id já é UUID, converter para string
            if isinstance(shift_id, uuid.UUID):
                shift_id = str(shift_id)
            # Tentar validar como UUID
            uuid.UUID(shift_id)
            # É um UUID, buscar no DW_Vista
            return self.get_shift_details_from_dw_vista(request, shift_id)
        except (ValueError, TypeError):
            # Não é UUID, buscar no banco local
            pass
        
        try:
            shift = ShiftRecord.objects.get(id=shift_id)
        except ShiftRecord.DoesNotExist:
            return JsonResponse({"error": "Plantão não encontrado"}, status=404)

        evidences = shift.evidences.all()
        compliance_items = shift.compliance_items.all()

        evidence_data = [
            {
                "id": str(evidence.id),
                "image_url": evidence.image.url if evidence.image else None,
                "description": evidence.description,
            }
            for evidence in evidences
        ]

        compliance_data = [
            {
                "id": str(item.id),
                "item_description": item.item_description,
                "status": item.status,
                "observations": item.observations,
            }
            for item in compliance_items
        ]

        data = {
            "shift": {
                "id": str(shift.id),
                "cr_number": shift.cr_number,
                "guard_name": shift.guard_name,
                "guard_number": shift.guard_number,
                "date": shift.shift_date.isoformat() if shift.shift_date else None,
                "time": shift.start_time.isoformat() if shift.start_time else None,
                "shift_type": shift.shift_type,
                "location": shift.location,
                "description": shift.description,
            },
            "evidences": evidence_data,
            "compliance_items": compliance_data,
        }

        return JsonResponse(data)

    def get_shift_details_from_dw_vista(self, request, task_id):
        """
        Busca detalhes de um plantão específico no DW_Vista
        """
        from django.db import connections
        
        try:
            db_conn_name = 'dw_vpn' if 'dw_vpn' in connections else ('readonly' if 'readonly' in connections else 'default')
            cursor = connections[db_conn_name].cursor()
            
            # Query para buscar detalhes de uma tarefa específica (incluindo expirada e colaborador)
            query = """
            SELECT
              t.id,
              t.nome,
              t.numero,
              t.inicio,
              t.terminoreal,
              t.estruturaqrcode,
              t.expirada,
              e.descricao AS estrutura_descricao,
              r.nome AS colaborador,
              ex.perguntadescricao,
              ex.conteudo
            FROM dbo.tarefa t
            LEFT JOIN dbo.execucao ex ON ex.tarefaid = t.id
            INNER JOIN dbo.checklist c ON c.id = t.checklistid
            INNER JOIN dbo.estrutura e ON e.id = t.estruturaid
            INNER JOIN dbo.recurso r ON r.codigohash = t.finalizadoporhash
            WHERE t.id = %s
              AND c.id = '6687b862-10d0-4144-ae30-8bdc55f22ee3'
              AND t.status = 85
            """
            
            cursor.execute(query, [task_id])
            results = cursor.fetchall()
            
            if not results:
                return JsonResponse({"error": "Plantão não encontrado"}, status=404)
            
            # Processar dados da tarefa
            task_info = None
            executions = []
            
            for row in results:
                task_id_db, nome, numero, inicio, terminoreal, estruturaqrcode, expirada, estrutura_descricao, colaborador, perguntadescricao, conteudo = row
                
                if not task_info:
                    task_info = {
                        'id': task_id_db,
                        'nome': nome,
                        'numero': numero,
                        'inicio': inicio,
                        'terminoreal': terminoreal,
                        'estruturaqrcode': estruturaqrcode,
                        'expirada': expirada,
                        'estrutura_descricao': estrutura_descricao,
                        'colaborador': colaborador
                    }
                
                if perguntadescricao and conteudo:
                    executions.append({
                        'pergunta': perguntadescricao,
                        'conteudo': conteudo
                    })
            
            # Processar dados para o formato esperado
            shift_data, evidences_data, compliance_data, occurrences_data = self.process_task_details(task_info, executions)
            
            return JsonResponse({
                "shift": shift_data,
                "evidences": evidences_data,
                "compliance_items": compliance_data,
                "occurrences": occurrences_data
            })
            
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Erro ao buscar detalhes no DW_Vista: {e}")
            return JsonResponse({"error": f"Erro ao buscar detalhes: {str(e)}"}, status=500)
        finally:
            cursor.close()

    def process_task_details(self, task_info, executions):
        """
        Processa os detalhes de uma tarefa do DW_Vista para o formato esperado
        """
        import re
        import logging
        logger = logging.getLogger(__name__)

        # Extrair tipo de plantão
        shift_type = "DIURNO"
        if task_info['nome']:
            match = re.search(r'-\s*(.+)$', task_info['nome'])
            if match:
                shift_type = match.group(1).strip()

        # Buscar descrição do plantão
        description = ""
        evidences = []
        compliance_items = []
        occurrences = []  # Nova lista para ocorrências

        logger.info(f"Processando {len(executions)} itens de execução")

        for exec_item in executions:
            pergunta = exec_item['pergunta']
            conteudo = exec_item['conteudo']

            # Log de debug
            logger.debug(f"Pergunta: {pergunta[:80] if pergunta else 'NULL'}")
            
            # Descrição do plantão
            if pergunta == "RELATE INFORMAÇÕES DO PLANTÃO:":
                description = conteudo
            
            # Ocorrências e tratativas
            elif pergunta == "REGISTRE A NÃO CONFORMIDADE ENCONTRADA.":
                occurrences.append({
                    "type": "non_conformity",
                    "title": "Não Conformidade Encontrada",
                    "content": conteudo
                })
            elif pergunta == "QUAL FOI A TRATATIVA TOMADA PARA SANAR OS PROBLEMAS IDENTIFICADOS?":
                occurrences.append({
                    "type": "treatment",
                    "title": "Tratativa Aplicada",
                    "content": conteudo
                })
            
            # Evidências (URLs que começam com https://api.opsvista.example.com/api/armazenamento)
            elif conteudo and conteudo.startswith("https://api.opsvista.example.com/api/armazenamento"):
                evidences.append({
                    "id": f"evidence_{len(evidences)}",
                    "image_url": conteudo,
                    "description": pergunta
                })
            
            # Conformidades (perguntas que contêm "conforme" - case insensitive)
            elif pergunta and "conforme" in pergunta.lower():
                # Remover espaços extras e converter para maiúsculas para comparação
                conteudo_limpo = conteudo.strip().upper() if conteudo else ""
                if conteudo_limpo in ["CONFORME", "NÃO CONFORME"]:
                    status = "conforme" if conteudo_limpo == "CONFORME" else "nao_conforme"
                    # Remover a palavra "conforme?" do final/início da descrição
                    item_description = pergunta.replace("Conforme?", "").replace("conforme?", "").replace("CONFORME?", "").strip()
                    # Se ficou vazio, usar a pergunta inteira
                    if not item_description:
                        item_description = pergunta

                    compliance_items.append({
                        "id": f"compliance_{len(compliance_items)}",
                        "item_description": item_description,
                        "status": status,
                        "observations": ""
                    })
                    logger.info(f"Conformidade adicionada: {status} - {item_description[:50]}")
                else:
                    # Log para debug: registrar quando o conteúdo não corresponde ao esperado
                    logger.debug(f"Conteúdo inesperado para pergunta '{pergunta}': '{conteudo}' (limpo: '{conteudo_limpo}')")

        logger.info(f"Total de conformidades encontradas: {len(compliance_items)}")
        logger.info(f"Total de evidências encontradas: {len(evidences)}")
        
        # Processar data/hora - formato: "2025-10-21 18:51:49.074"
        # Para registros expirados, usar data de início; para realizados, usar terminoreal
        date_str = None
        time_str = None
        
        # Escolher qual data usar baseado no status expirado
        date_source = task_info.get('inicio') if task_info.get('expirada') else task_info.get('terminoreal')
        
        if date_source:
            from datetime import datetime
            try:
                # Converter string para datetime
                date_source_str = str(date_source)
                
                # Remover microsegundos se existirem
                if '.' in date_source_str:
                    date_source_str = date_source_str.split('.')[0]
                
                # Parse da data/hora
                dt = datetime.strptime(date_source_str, '%Y-%m-%d %H:%M:%S')
                
                # Formato ISO para JavaScript
                date_str = dt.date().isoformat()  # YYYY-MM-DD
                time_str = dt.time().isoformat()  # HH:MM:SS
                
            except Exception as e:
                # Fallback para formato original
                print(f"Erro ao processar data/hora {date_source}: {e}")
                date_str = str(date_source)[:10]
                time_str = str(date_source)[11:19]

        # Dados do plantão
        shift_data = {
            "id": str(task_info['id']),
            "cr_number": task_info['estruturaqrcode'] or "N/A",
            "guard_name": task_info.get('colaborador', "Colaborador não identificado") or "Colaborador não identificado",
            "guard_number": task_info['numero'],
            "date": date_str,
            "time": time_str,
            "shift_type": shift_type,
            "location": task_info['estrutura_descricao'],
            "description": description,
            "expirada": task_info.get('expirada', False)
        }
        
        return shift_data, evidences, compliance_items, occurrences


class PlannerView(LoginRequiredMixin, TemplateView):
    """
    View para a página do Planner (Kanban)
    """
    template_name = "planner.html"

    def get(self, request, *args, **kwargs):
        """
        Manipula as requisições GET (AJAX e Normal)
        """
        action = request.GET.get("action")

        if action == "download_report":
            try:
                project_id = request.GET.get("project_id")
                projeto = PlannerProject.objects.select_related('tipo_servico').prefetch_related(
                    'projeto_responsaveis__responsavel',
                    'checklist_items',
                    'anexos',
                    'comentarios__autor',
                ).get(id=project_id)

                if projeto.status != "Concluído":
                    return JsonResponse({"status": "error", "message": "Relatório disponível apenas para demandas concluídas."}, status=400)

                from io import BytesIO
                from django.http import HttpResponse
                from django.utils.text import slugify
                from reportlab.lib import colors
                from reportlab.lib.pagesizes import A4
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.units import cm
                from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak

                buffer = BytesIO()
                doc = SimpleDocTemplate(
                    buffer,
                    pagesize=A4,
                    rightMargin=1.5 * cm,
                    leftMargin=1.5 * cm,
                    topMargin=1.5 * cm,
                    bottomMargin=1.5 * cm,
                    title=f"Relatório - {projeto.nome}",
                )
                styles = getSampleStyleSheet()
                styles.add(ParagraphStyle(name="SectionTitle", parent=styles["Heading2"], textColor=colors.HexColor("#1d4ed8"), spaceBefore=12, spaceAfter=8))
                styles.add(ParagraphStyle(name="SmallText", parent=styles["BodyText"], fontSize=8, leading=10))

                def safe(value):
                    if value is None or value == "":
                        return "-"
                    return str(value).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace("\n", "<br/>")

                def date_br(value):
                    return value.strftime("%d/%m/%Y") if value else "-"

                def money(value):
                    return f"R$ {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".") if value is not None else "-"

                def section(title):
                    story.append(Paragraph(title, styles["SectionTitle"]))

                def kv_table(rows):
                    table = Table(
                        [[Paragraph(f"<b>{safe(k)}</b>", styles["BodyText"]), Paragraph(safe(v), styles["BodyText"])] for k, v in rows],
                        colWidths=[5.0 * cm, 11.0 * cm],
                    )
                    table.setStyle(TableStyle([
                        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
                        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f1f5f9")),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 6),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                        ("TOPPADDING", (0, 0), (-1, -1), 5),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ]))
                    story.append(table)
                    story.append(Spacer(1, 0.25 * cm))

                story = []
                story.append(Paragraph("Relatório de Conclusão", styles["Title"]))
                story.append(Paragraph(safe(projeto.nome), styles["Heading2"]))
                story.append(Paragraph(f"Gerado em {timezone.now().strftime('%d/%m/%Y %H:%M')} por {safe(getattr(request.user, 'name', '') or request.user.username)}", styles["SmallText"]))
                story.append(Spacer(1, 0.4 * cm))

                section("Resumo da Demanda")
                kv_table([
                    ("Status", projeto.status),
                    ("Tipo da demanda", projeto.get_tipo_demanda_display()),
                    ("Categoria", projeto.tipo_servico.nome if projeto.tipo_servico else "-"),
                    ("Cliente / Unidade", projeto.cliente),
                    ("Solicitante", projeto.solicitante),
                    ("Contato", projeto.contato),
                    ("Telefone", projeto.telefone),
                    ("E-mail", projeto.email),
                    ("Prioridade", projeto.prioridade),
                    ("Impacto", projeto.impacto),
                    ("SLA", projeto.sla),
                    ("Progresso", f"{projeto.percentual_progresso}%"),
                    ("Início", date_br(projeto.data_inicial)),
                    ("Entrega / Conclusão prevista", date_br(projeto.data_conclusao)),
                    ("Valor / esforço estimado", money(projeto.valor_estimado)),
                ])

                section("CRM 360º, Operacional e Colaborativo")
                kv_table([
                    ("Tipo de CRM", projeto.get_crm_tipo_display()),
                    ("Visão 360º", projeto.visao_360),
                    ("Saúde do relacionamento", projeto.saude_relacionamento),
                    ("Canal de relacionamento", projeto.get_canal_relacionamento_display() if projeto.canal_relacionamento else "-"),
                    ("Última interação", date_br(projeto.ultima_interacao)),
                    ("Próxima interação", date_br(projeto.proxima_interacao)),
                    ("Departamentos envolvidos", projeto.departamentos_envolvidos),
                    ("Stakeholders e papéis", projeto.stakeholders),
                    ("Resumo colaborativo", projeto.resumo_colaborativo),
                ])

                section("Responsáveis")
                responsaveis = ", ".join([r.responsavel.name or r.responsavel.username for r in projeto.projeto_responsaveis.all()])
                kv_table([
                    ("Responsáveis do card", responsaveis or "-"),
                    ("Responsável técnico", projeto.responsavel_tecnico),
                    ("Validador / aprovador", projeto.validador),
                ])

                section("Informações por Etapa")
                kv_table([
                    ("Triagem - diagnóstico", projeto.triagem_diagnostico),
                    ("Triagem - priorização / impacto", projeto.triagem_priorizacao),
                    ("Planejamento - entregáveis", projeto.planejamento_entregaveis),
                    ("Planejamento - riscos / dependências", projeto.planejamento_riscos),
                    ("Execução - andamento", projeto.execucao_andamento),
                    ("Validação - resultado", projeto.validacao_resultado),
                ])

                section("5W2H")
                kv_table([
                    ("What / O quê", projeto.w2h_what),
                    ("Why / Por quê", projeto.w2h_why),
                    ("Where / Onde", projeto.w2h_where),
                    ("When / Quando", date_br(projeto.w2h_when)),
                    ("Who / Quem", projeto.w2h_who),
                    ("How / Como", projeto.w2h_how),
                    ("How much / Quanto custa", money(projeto.w2h_how_much)),
                ])

                section("Dados Técnicos / Implantação / Ocorrência")
                kv_table([
                    ("Módulo / Sistema", projeto.modulo_sistema),
                    ("Ambiente", projeto.get_ambiente_display() if projeto.ambiente else "-"),
                    ("Etapa da implantação", projeto.etapa_implantacao),
                    ("Go-live previsto", date_br(projeto.go_live_previsto)),
                    ("Treinamento realizado", "Sim" if projeto.treinamento_realizado else "Não"),
                    ("Severidade", projeto.severidade),
                    ("Causa raiz", projeto.causa_raiz),
                    ("Ação corretiva", projeto.acao_corretiva),
                    ("Ação preventiva", projeto.acao_preventiva),
                    ("Link / PR / documentação", projeto.link_referencia),
                    ("Critério de aceite", projeto.criterio_aceite),
                ])

                section("Observações e Próxima Ação")
                kv_table([
                    ("Observações", projeto.observacao),
                    ("Próxima ação", projeto.proxima_acao),
                ])

                section("Checklist")
                checklist_rows = [("Item", "Status")]
                for item in projeto.checklist_items.all():
                    checklist_rows.append((item.texto, "Concluído" if item.concluido else "Pendente"))
                if len(checklist_rows) == 1:
                    checklist_rows.append(("Nenhum item cadastrado", "-"))
                table = Table([[Paragraph(safe(a), styles["BodyText"]), Paragraph(safe(b), styles["BodyText"])] for a, b in checklist_rows], colWidths=[11 * cm, 5 * cm])
                table.setStyle(TableStyle([
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dbeafe")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]))
                story.append(table)
                story.append(Spacer(1, 0.25 * cm))

                section("Anexos / Evidências")
                anexos = list(projeto.anexos.all())
                if anexos:
                    kv_table([(anexo.nome, request.build_absolute_uri(anexo.arquivo.url)) for anexo in anexos])
                else:
                    kv_table([("Anexos", "Nenhum anexo cadastrado")])

                section("Comentários")
                comentarios = list(projeto.comentarios.all()[:20])
                if comentarios:
                    kv_table([(f"{c.autor.name or c.autor.username} - {c.created_at.strftime('%d/%m/%Y %H:%M')}", c.conteudo) for c in comentarios])
                else:
                    kv_table([("Comentários", "Nenhum comentário cadastrado")])

                doc.build(story)
                filename = f"relatorio-{slugify(projeto.nome) or projeto.id}.pdf"
                response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
                response["Content-Disposition"] = f'attachment; filename="{filename}"'
                return response
            except PlannerProject.DoesNotExist:
                return JsonResponse({"status": "error", "message": "Projeto não encontrado"}, status=404)
            except Exception as e:
                return JsonResponse({"status": "error", "message": str(e)}, status=500)

        if action == "get_project":
            try:
                project_id = request.GET.get("project_id")
                projeto = PlannerProject.objects.select_related(
                    'tipo_servico'
                ).prefetch_related(
                    'projeto_responsaveis__responsavel',
                    'checklist_items',
                    'anexos'
                ).get(id=project_id)

                return JsonResponse({
                    "status": "success",
                    "project": {
                        "nome": projeto.nome,
                        "cliente": projeto.cliente or "",
                        "contato": projeto.contato or "",
                        "telefone": projeto.telefone or "",
                        "email": projeto.email or "",
                        "origem_lead": projeto.origem_lead or "",
                        "valor_estimado": str(projeto.valor_estimado or ""),
                        "probabilidade": projeto.probabilidade or 0,
                        "proxima_acao": projeto.proxima_acao or "",
                        "crm_tipo": projeto.crm_tipo or "operacional",
                        "visao_360": projeto.visao_360 or "",
                        "saude_relacionamento": projeto.saude_relacionamento or "",
                        "canal_relacionamento": projeto.canal_relacionamento or "",
                        "ultima_interacao": projeto.ultima_interacao.strftime("%Y-%m-%d") if projeto.ultima_interacao else "",
                        "proxima_interacao": projeto.proxima_interacao.strftime("%Y-%m-%d") if projeto.proxima_interacao else "",
                        "departamentos_envolvidos": projeto.departamentos_envolvidos or "",
                        "stakeholders": projeto.stakeholders or "",
                        "resumo_colaborativo": projeto.resumo_colaborativo or "",
                        "tipo_demanda": projeto.tipo_demanda or "implantacao",
                        "impacto": projeto.impacto or "Médio",
                        "solicitante": projeto.solicitante or "",
                        "responsavel_tecnico": projeto.responsavel_tecnico or "",
                        "validador": projeto.validador or "",
                        "sla": projeto.sla or "",
                        "percentual_progresso": projeto.percentual_progresso or 0,
                        "modulo_sistema": projeto.modulo_sistema or "",
                        "ambiente": projeto.ambiente or "",
                        "etapa_implantacao": projeto.etapa_implantacao or "",
                        "go_live_previsto": projeto.go_live_previsto.strftime("%Y-%m-%d") if projeto.go_live_previsto else "",
                        "treinamento_realizado": projeto.treinamento_realizado,
                        "severidade": projeto.severidade or "",
                        "causa_raiz": projeto.causa_raiz or "",
                        "acao_corretiva": projeto.acao_corretiva or "",
                        "acao_preventiva": projeto.acao_preventiva or "",
                        "link_referencia": projeto.link_referencia or "",
                        "criterio_aceite": projeto.criterio_aceite or "",
                        "triagem_diagnostico": projeto.triagem_diagnostico or "",
                        "triagem_priorizacao": projeto.triagem_priorizacao or "",
                        "planejamento_entregaveis": projeto.planejamento_entregaveis or "",
                        "planejamento_riscos": projeto.planejamento_riscos or "",
                        "execucao_andamento": projeto.execucao_andamento or "",
                        "validacao_resultado": projeto.validacao_resultado or "",
                        "w2h_what": projeto.w2h_what or "",
                        "w2h_why": projeto.w2h_why or "",
                        "w2h_where": projeto.w2h_where or "",
                        "w2h_when": projeto.w2h_when.strftime("%Y-%m-%d") if projeto.w2h_when else "",
                        "w2h_who": projeto.w2h_who or "",
                        "w2h_how": projeto.w2h_how or "",
                        "w2h_how_much": str(projeto.w2h_how_much or ""),
                        "responsaveis": [
                            {"id": str(resp.responsavel.id), "name": resp.responsavel.name or resp.responsavel.username} 
                            for resp in projeto.projeto_responsaveis.all()
                        ],
                        "data_inicial": projeto.data_inicial.strftime("%Y-%m-%d"),
                        "data_conclusao": projeto.data_conclusao.strftime("%Y-%m-%d"),
                        "prioridade": projeto.prioridade,
                        "observacao": projeto.observacao,
                        "status": projeto.status,
                        "tipo": projeto.tipo,
                        "tipo_servico": str(projeto.tipo_servico.id) if projeto.tipo_servico else "",
                        "tipo_servico_id": str(projeto.tipo_servico.id) if projeto.tipo_servico else "",
                        "checklist_items": [
                            {"id": str(item.id), "text": item.texto, "completed": item.concluido} 
                            for item in projeto.checklist_items.all()
                        ],
                        "attachments": [
                            {"id": str(anexo.id), "nome": anexo.nome, "url": anexo.arquivo.url}
                            for anexo in projeto.anexos.all()
                        ],
                    },
                })
            except PlannerProject.DoesNotExist:
                return JsonResponse({"status": "error", "message": "Projeto não encontrado"})
            except Exception as e:
                return JsonResponse({"status": "error", "message": str(e)})

        elif action == "check_service_type":
            try:
                nome = request.GET.get("nome", "").strip()
                if nome:
                    exists = TipoServico.objects.filter(nome__iexact=nome).exists()
                    return JsonResponse({"exists": exists})
                return JsonResponse({"exists": False})
            except Exception as e:
                return JsonResponse({"exists": False, "error": str(e)})
        
        elif action == "get_counters":
            # Retorna contadores via AJAX se necessário (mas o template já carrega via context)
            try:
                from django.db.models import Count, Case, When, IntegerField
                status_counts = PlannerProject.objects.aggregate(
                    entrada=Count(Case(When(status='Entrada / Backlog', then=1), output_field=IntegerField())),
                    triagem=Count(Case(When(status='Triagem', then=1), output_field=IntegerField())),
                    planejamento=Count(Case(When(status='Planejamento 5W2H', then=1), output_field=IntegerField())),
                    execucao=Count(Case(When(status='Em Execução', then=1), output_field=IntegerField())),
                    aguardando=Count(Case(When(status='Aguardando Terceiros / Cliente', then=1), output_field=IntegerField())),
                    validacao=Count(Case(When(status='Validação / Testes', then=1), output_field=IntegerField())),
                    concluido=Count(Case(When(status='Concluído', then=1), output_field=IntegerField())),
                    cancelado=Count(Case(When(status='Cancelado / Suspenso', then=1), output_field=IntegerField())),
                    total=Count('id')
                )
                return JsonResponse({"status": "success", "counters": status_counts})
            except Exception as e:
                return JsonResponse({"status": "error", "message": str(e)})

        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from django.core.cache import cache
        from django.core.paginator import Paginator
        from django.db.models import Count, Case, When, IntegerField, Q
        import time

        # 1. CAPTURAR FILTROS DA URL
        filtro_nome = self.request.GET.get('filter_nome', '').strip()
        filtro_responsavel = self.request.GET.get('filter_responsavel', '').strip()
        filtro_tipo = self.request.GET.get('filter_tipo', 'all')
        filtro_prioridade = self.request.GET.get('filter_prioridade', 'all')
        filtro_tipo_demanda = self.request.GET.get('filter_tipo_demanda', 'all')
        view_mode = self.request.GET.get('view', 'kanban')
        if view_mode not in ('kanban', 'grid'):
            view_mode = 'kanban'

        # 2. QUERYSET BASE (Otimizada)
        # 'etapa' entra no select_related/only pois o template acessa
        # projeto.etapa.nome/.cor em todo card (Kanban) e linha (Grid) --
        # sem isso, cada card dispara 2 queries extras (campo etapa_id
        # adiado + FK relacionada), virando um N+1 de centenas/milhares de
        # queries por request.
        projetos_queryset = PlannerProject.objects.select_related(
            'tipo_servico', 'etapa'
        ).prefetch_related(
            'projeto_responsaveis__responsavel',
            'checklist_items',
            'anexos'
        ).only(
            'id', 'nome', 'cliente', 'contato', 'telefone', 'email', 'valor_estimado', 'probabilidade', 'proxima_acao', 'crm_tipo', 'visao_360', 'saude_relacionamento', 'canal_relacionamento', 'ultima_interacao', 'proxima_interacao', 'departamentos_envolvidos', 'stakeholders', 'resumo_colaborativo', 'status', 'prioridade', 'tipo_demanda', 'impacto', 'solicitante', 'responsavel_tecnico', 'percentual_progresso', 'modulo_sistema', 'ambiente', 'sla', 'triagem_diagnostico', 'triagem_priorizacao', 'planejamento_entregaveis', 'planejamento_riscos', 'execucao_andamento', 'validacao_resultado', 'data_inicial', 'data_conclusao', 'observacao', 'tipo_servico__nome', 'w2h_when', 'etapa__nome', 'etapa__cor'
        ).order_by('-prioridade', 'data_conclusao')

        # 3. APLICAR FILTROS (Se existirem)
        # Se tiver filtro, não usamos cache de contagem para refletir a busca real
        has_filters = any([filtro_nome, filtro_responsavel, filtro_tipo != 'all', filtro_prioridade != 'all', filtro_tipo_demanda != 'all'])

        if filtro_responsavel:
            # Tenta verificar se é um ID válido (UUID ou Inteiro)
            is_valid_id = False
            if filtro_responsavel.isdigit():
                is_valid_id = True
            else:
                try:
                    uuid.UUID(str(filtro_responsavel))
                    is_valid_id = True
                except (ValueError, AttributeError):
                    is_valid_id = False

            if is_valid_id:
                # Filtra pelo ID exato do usuário na tabela de relacionamento
                projetos_queryset = projetos_queryset.filter(
                    projeto_responsaveis__responsavel__id=filtro_responsavel
                )
            else:
                # Filtra pelo nome se não for um ID
                projetos_queryset = projetos_queryset.filter(
                    projeto_responsaveis__responsavel__name__icontains=filtro_responsavel
                )

        if filtro_tipo and filtro_tipo != 'all':
            projetos_queryset = projetos_queryset.filter(tipo_servico__nome=filtro_tipo)

        if filtro_prioridade and filtro_prioridade != 'all':
            projetos_queryset = projetos_queryset.filter(prioridade=filtro_prioridade)

        if filtro_tipo_demanda and filtro_tipo_demanda != 'all':
            projetos_queryset = projetos_queryset.filter(tipo_demanda=filtro_tipo_demanda)

        if filtro_nome:
            projetos_queryset = projetos_queryset.filter(
                Q(nome__icontains=filtro_nome) |
                Q(cliente__icontains=filtro_nome) |
                Q(contato__icontains=filtro_nome) |
                Q(email__icontains=filtro_nome) |
                Q(telefone__icontains=filtro_nome) |
                Q(solicitante__icontains=filtro_nome) |
                Q(modulo_sistema__icontains=filtro_nome) |
                Q(departamentos_envolvidos__icontains=filtro_nome) |
                Q(stakeholders__icontains=filtro_nome) |
                Q(visao_360__icontains=filtro_nome) |
                Q(triagem_diagnostico__icontains=filtro_nome) |
                Q(planejamento_entregaveis__icontains=filtro_nome)
            )

        # 4. CALCULAR CONTADORES (Baseado no resultado filtrado)
        # Isso garante que se você filtrar "Alta", os cards mostrem apenas a contagem de "Alta"
        # Sem filtro ativo, os contadores representam o board padrão e podem
        # ser cacheados por alguns segundos -- antes essa infraestrutura de
        # cache existia (invalidate_planner_cache já limpava a chave), mas
        # nada nunca escrevia/lia dela, então todo GET refazia o aggregate.
        def _calcular_status_counts():
            return projetos_queryset.aggregate(
                entrada_count=Count(Case(When(status='Entrada / Backlog', then=1), output_field=IntegerField())),
                triagem_count=Count(Case(When(status='Triagem', then=1), output_field=IntegerField())),
                planejamento_count=Count(Case(When(status='Planejamento 5W2H', then=1), output_field=IntegerField())),
                execucao_count=Count(Case(When(status='Em Execução', then=1), output_field=IntegerField())),
                aguardando_count=Count(Case(When(status='Aguardando Terceiros / Cliente', then=1), output_field=IntegerField())),
                validacao_count=Count(Case(When(status='Validação / Testes', then=1), output_field=IntegerField())),
                concluido_count=Count(Case(When(status='Concluído', then=1), output_field=IntegerField())),
                cancelado_count=Count(Case(When(status='Cancelado / Suspenso', then=1), output_field=IntegerField())),
            )

        if has_filters:
            status_counts = _calcular_status_counts()
        else:
            status_counts = cache.get('planner_status_counts')
            if status_counts is None:
                status_counts = _calcular_status_counts()
                cache.set('planner_status_counts', status_counts, 60)

        class StatusGroup:
            def __init__(self, count):
                self.count = count or 0

        projetos_por_status = {
            "Entrada": StatusGroup(status_counts['entrada_count']),
            "Triagem": StatusGroup(status_counts['triagem_count']),
            "Planejamento": StatusGroup(status_counts['planejamento_count']),
            "Execucao": StatusGroup(status_counts['execucao_count']),
            "Aguardando": StatusGroup(status_counts['aguardando_count']),
            "Validacao": StatusGroup(status_counts['validacao_count']),
            "Concluido": StatusGroup(status_counts['concluido_count']),
            "Cancelado": StatusGroup(status_counts['cancelado_count']),
        }

        # 5. CARREGAR DADOS AUXILIARES (mudam raramente -- cache curto em vez
        # de recarregar do banco em todo GET/POST da página)
        tipos_servico_cached = cache.get('planner_tipos_servico')
        if tipos_servico_cached is None:
            tipos_servico_cached = list(TipoServico.objects.filter(ativo=True).values('id', 'nome'))
            cache.set('planner_tipos_servico', tipos_servico_cached, 300)

        usuarios_cached = cache.get('planner_usuarios')
        if usuarios_cached is None:
            usuarios_cached = list(CustomUser.objects.filter(is_active=True).values('id', 'name', 'username'))
            cache.set('planner_usuarios', usuarios_cached, 300)

        # 6. PAGINAÇÃO REAL (Paginator -- antes carregava até 500 objetos
        # completos de uma vez e fatiava em Python em 8 list comprehensions)
        page_size = 300 if has_filters else 150
        try:
            page_number = int(self.request.GET.get('page', 1))
        except (TypeError, ValueError):
            page_number = 1
        paginator = Paginator(projetos_queryset, page_size)
        page_obj = paginator.get_page(page_number)
        projetos_lista = list(page_obj.object_list)

        # 7. MONTAR APENAS A VISÃO ATIVA (Kanban ou Grid)
        # As duas visões (view_mode) usam a mesma URL/template, mas antes o
        # servidor sempre montava e renderizava as DUAS ao mesmo tempo (só
        # escondidas via CSS no browser), dobrando o HTML/DOM gerado pra
        # nada -- o toggle agora recarrega a página com ?view=kanban|grid.
        status_key_map = {
            "Entrada / Backlog": "entrada",
            "Triagem": "triagem",
            "Planejamento 5W2H": "planejamento",
            "Em Execução": "execucao",
            "Aguardando Terceiros / Cliente": "aguardando",
            "Validação / Testes": "validacao",
            "Concluído": "concluido",
            "Cancelado / Suspenso": "cancelado",
        }
        status_meta = [
            ("Entrada / Backlog", "fa-inbox", "#2563eb"),
            ("Triagem", "fa-filter-circle-dollar", "#7c3aed"),
            ("Planejamento 5W2H", "fa-list-check", "#0891b2"),
            ("Em Execução", "fa-gears", "#f59e0b"),
            ("Aguardando Terceiros / Cliente", "fa-clock", "#ea580c"),
            ("Validação / Testes", "fa-vial-circle-check", "#4f46e5"),
            ("Concluído", "fa-circle-check", "#16a34a"),
            ("Cancelado / Suspenso", "fa-ban", "#64748b"),
        ]
        crm_columns = []
        for status_label, icon, color in status_meta:
            key = status_key_map[status_label]
            crm_columns.append({
                "status": status_label,
                "key": key,
                "icon": icon,
                "color": color,
                # Contagem real (total filtrado, não limitada pela página atual)
                "count": status_counts.get(f"{key}_count", 0) or 0,
                "projects": [p for p in projetos_lista if p.status == status_label] if view_mode == "kanban" else [],
            })

        projetos_grid = projetos_lista if view_mode == "grid" else []

        context.update({
            "projetos": projetos_grid,
            "crm_columns": crm_columns,
            "projetos_por_status": projetos_por_status,
            "tipos_servico": tipos_servico_cached,
            "usuarios": usuarios_cached,
            "timestamp": int(time.time()),
            "view_mode": view_mode,
            "page_obj": page_obj,
            "paginator": paginator,
            # Passar filtros para o template popular os campos
            "filtros": {
                "nome": filtro_nome,
                "responsavel": filtro_responsavel,
                "tipo": filtro_tipo,
                "prioridade": filtro_prioridade,
                "tipo_demanda": filtro_tipo_demanda
            },
            "tipo_demanda_choices": PlannerProject.TIPO_DEMANDA_CHOICES,
            "impacto_choices": PlannerProject.IMPACTO_CHOICES,
            "ambiente_choices": PlannerProject.AMBIENTE_CHOICES,
            "crm_tipo_choices": PlannerProject.CRM_TIPO_CHOICES,
            "canal_relacionamento_choices": PlannerProject.CANAL_RELACIONAMENTO_CHOICES,
            "status_choices": PlannerProject.STATUS_CHOICES
        })
        return context

    def invalidate_planner_cache(self):
        """Invalida o cache do planner"""
        from django.core.cache import cache
        cache.delete('planner_status_counts')
        cache.delete('planner_data_v2')
        # Não deletamos usuarios/tipos aqui pois mudam pouco

    def post(self, request, *args, **kwargs):
        """
        Manipula as requisições POST para criar/editar/excluir projetos
        """
        action = request.POST.get("action")
        response_data = {"status": "error", "message": "Ação inválida"}

        # === ATUALIZAR STATUS (DRAG & DROP) ===
        if action == "update_project_status":
            try:
                project_id = request.POST.get('project_id')
                new_status = request.POST.get('status')
                
                if not project_id or not new_status:
                    return JsonResponse({'status': 'error', 'message': 'Dados incompletos'})

                valid_statuses = [choice[0] for choice in PlannerProject.STATUS_CHOICES]
                if new_status not in valid_statuses:
                    return JsonResponse({'status': 'error', 'message': 'Status inválido'})

                try:
                    projeto = PlannerProject.objects.get(id=project_id)
                except (PlannerProject.DoesNotExist, ValueError, ValidationError):
                    return JsonResponse({'status': 'error', 'message': 'Projeto não encontrado'})
                
                old_status = projeto.status
                
                if old_status != new_status:
                    projeto.status = new_status
                    projeto.save()
                    
                    # Registrar histórico
                    PlannerProjectChangeHistory.objects.create(
                        projeto=projeto,
                        usuario=request.user,
                        tipo_alteracao='status_alterado',
                        valor_anterior=old_status,
                        valor_novo=new_status,
                        descricao=f"Arrastado no Kanban de {old_status} para {new_status}"
                    )
                    
                    self.invalidate_planner_cache()
                    
                return JsonResponse({'status': 'success', 'report_url': f'{request.path}?action=download_report&project_id={projeto.id}' if projeto.status == 'Concluído' else ''})
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': str(e)})

        # === CRIAR PROJETO ===
        elif action == "create_project":
            try:
                nome = request.POST.get('nome', '').strip()
                if not nome:
                    return JsonResponse({'status': 'error', 'message': 'O nome do projeto é obrigatório'})

                tipo_servico_id = request.POST.get('tipo_servico')
                if not tipo_servico_id or tipo_servico_id in ("None", "null", ""):
                    return JsonResponse({'status': 'error', 'message': 'O tipo de serviço é obrigatório'})

                data_inicial = request.POST.get('data_inicial')
                data_conclusao = request.POST.get('data_conclusao')
                if not data_inicial or not data_conclusao:
                    return JsonResponse({'status': 'error', 'message': 'Datas obrigatórias'})

                try:
                    tipo_servico = TipoServico.objects.get(id=tipo_servico_id, ativo=True)
                except (TipoServico.DoesNotExist, ValueError, TypeError):
                    return JsonResponse({'status': 'error', 'message': 'Tipo de serviço inválido'})

                # Parse responsáveis e checklist
                try:
                    responsaveis_data = json.loads(request.POST.get('responsaveis', '[]'))
                    if not responsaveis_data:
                        responsaveis_data = [{'id': str(request.user.id)}]
                except json.JSONDecodeError:
                    responsaveis_data = []

                try:
                    checklist_items = json.loads(request.POST.get('checklist_items', '[]'))
                except json.JSONDecodeError:
                    checklist_items = []

                with transaction.atomic():
                    projeto = PlannerProject.objects.create(
                        nome=nome,
                        data_inicial=data_inicial,
                        data_conclusao=data_conclusao,
                        prioridade=request.POST.get('prioridade', 'Médio'),
                        observacao=request.POST.get('observacao', '').strip(),
                        status=request.POST.get('status', 'Entrada / Backlog'),
                        tipo_servico=tipo_servico,
                        cliente=request.POST.get('cliente', '').strip(),
                        contato=request.POST.get('contato', '').strip(),
                        telefone=request.POST.get('telefone', '').strip(),
                        email=request.POST.get('email', '').strip(),
                        origem_lead=request.POST.get('origem_lead', '').strip(),
                        valor_estimado=request.POST.get('valor_estimado') or None,
                        probabilidade=request.POST.get('probabilidade') or 0,
                        proxima_acao=request.POST.get('proxima_acao', '').strip(),
                        crm_tipo=request.POST.get('crm_tipo', 'operacional'),
                        visao_360=request.POST.get('visao_360', '').strip(),
                        saude_relacionamento=request.POST.get('saude_relacionamento', '').strip(),
                        canal_relacionamento=request.POST.get('canal_relacionamento', ''),
                        ultima_interacao=request.POST.get('ultima_interacao') or None,
                        proxima_interacao=request.POST.get('proxima_interacao') or None,
                        departamentos_envolvidos=request.POST.get('departamentos_envolvidos', '').strip(),
                        stakeholders=request.POST.get('stakeholders', '').strip(),
                        resumo_colaborativo=request.POST.get('resumo_colaborativo', '').strip(),
                        tipo_demanda=request.POST.get('tipo_demanda', 'implantacao'),
                        impacto=request.POST.get('impacto', 'Médio'),
                        solicitante=request.POST.get('solicitante', '').strip(),
                        responsavel_tecnico=request.POST.get('responsavel_tecnico', '').strip(),
                        validador=request.POST.get('validador', '').strip(),
                        sla=request.POST.get('sla', '').strip(),
                        percentual_progresso=request.POST.get('percentual_progresso') or 0,
                        modulo_sistema=request.POST.get('modulo_sistema', '').strip(),
                        ambiente=request.POST.get('ambiente', ''),
                        etapa_implantacao=request.POST.get('etapa_implantacao', '').strip(),
                        go_live_previsto=request.POST.get('go_live_previsto') or None,
                        treinamento_realizado=request.POST.get('treinamento_realizado') == 'on',
                        severidade=request.POST.get('severidade') or None,
                        causa_raiz=request.POST.get('causa_raiz', '').strip(),
                        acao_corretiva=request.POST.get('acao_corretiva', '').strip(),
                        acao_preventiva=request.POST.get('acao_preventiva', '').strip(),
                        link_referencia=request.POST.get('link_referencia', '').strip(),
                        criterio_aceite=request.POST.get('criterio_aceite', '').strip(),
                        triagem_diagnostico=request.POST.get('triagem_diagnostico', '').strip(),
                        triagem_priorizacao=request.POST.get('triagem_priorizacao', '').strip(),
                        planejamento_entregaveis=request.POST.get('planejamento_entregaveis', '').strip(),
                        planejamento_riscos=request.POST.get('planejamento_riscos', '').strip(),
                        execucao_andamento=request.POST.get('execucao_andamento', '').strip(),
                        validacao_resultado=request.POST.get('validacao_resultado', '').strip(),
                        w2h_what=request.POST.get('w2h_what', '').strip(),
                        w2h_why=request.POST.get('w2h_why', '').strip(),
                        w2h_where=request.POST.get('w2h_where', '').strip(),
                        w2h_when=request.POST.get('w2h_when') or None,
                        w2h_who=request.POST.get('w2h_who', '').strip(),
                        w2h_how=request.POST.get('w2h_how', '').strip(),
                        w2h_how_much=request.POST.get('w2h_how_much') or None,
                    )
                    
                    PlannerProjectChangeHistory.objects.create(
                        projeto=projeto,
                        usuario=request.user,
                        tipo_alteracao='criado',
                        valor_novo=nome,
                        descricao="Projeto criado"
                    )

                    for resp in responsaveis_data:
                        resp_id = resp.get('id')
                        if resp_id and resp_id not in ("None", "null", ""):
                            try:
                                if CustomUser.objects.filter(id=resp_id).exists():
                                    PlannerProjectResponsavel.objects.create(projeto=projeto, responsavel_id=resp_id)
                            except (ValueError, ValidationError):
                                pass

                    for i, item in enumerate(checklist_items):
                        text = item.get('text', '').strip()
                        if text:
                            PlannerChecklistItem.objects.create(
                                projeto=projeto,
                                texto=text,
                                concluido=item.get('completed', False),
                                ordem=i,
                            )

                    # Salvar anexos
                    arquivos = request.FILES.getlist('arquivos')
                    for f in arquivos:
                        PlannerAttachment.objects.create(
                            projeto=projeto,
                            nome=f.name,
                            arquivo=f,
                            uploaded_by=request.user
                        )

                    self.invalidate_planner_cache()
                
                return JsonResponse({'status': 'success', 'message': 'Projeto criado com sucesso'})

            except Exception as e:
                return JsonResponse({'status': 'error', 'message': f'Erro: {str(e)}'})

        # === CRIAR TIPO DE SERVIÇO ===
        elif action == "create_service_type":
            try:
                nome = request.POST.get("nome")
                if not nome:
                    raise ValueError("Nome obrigatório")
                
                if TipoServico.objects.filter(nome=nome).exists():
                    raise ValueError("Este tipo já existe")

                novo_tipo = TipoServico.objects.create(nome=nome, ativo=True)
                self.invalidate_planner_cache()

                return JsonResponse({
                    "status": "success", 
                    "message": "Tipo criado",
                    "tipo_id": novo_tipo.id,
                    "tipo_nome": novo_tipo.nome
                })
            except Exception as e:
                return JsonResponse({"status": "error", "message": str(e)})

        # === ATUALIZAR PROJETO ===
        elif action == "update_project":
            try:
                project_id = request.POST.get('project_id')
                try:
                    projeto = PlannerProject.objects.get(id=project_id)
                except (PlannerProject.DoesNotExist, ValueError, ValidationError):
                    return JsonResponse({'status': 'error', 'message': 'Projeto não encontrado'})
                
                projeto.nome = request.POST.get('nome', projeto.nome)
                projeto.status = request.POST.get('status', projeto.status)
                projeto.prioridade = request.POST.get('prioridade', projeto.prioridade)
                projeto.data_inicial = request.POST.get('data_inicial', projeto.data_inicial)
                projeto.data_conclusao = request.POST.get('data_conclusao', projeto.data_conclusao)
                projeto.observacao = request.POST.get('observacao', projeto.observacao)
                projeto.cliente = request.POST.get('cliente', '').strip()
                projeto.contato = request.POST.get('contato', '').strip()
                projeto.telefone = request.POST.get('telefone', '').strip()
                projeto.email = request.POST.get('email', '').strip()
                projeto.origem_lead = request.POST.get('origem_lead', '').strip()
                projeto.valor_estimado = request.POST.get('valor_estimado') or None
                projeto.probabilidade = request.POST.get('probabilidade') or 0
                projeto.proxima_acao = request.POST.get('proxima_acao', '').strip()
                projeto.crm_tipo = request.POST.get('crm_tipo', projeto.crm_tipo)
                projeto.visao_360 = request.POST.get('visao_360', '').strip()
                projeto.saude_relacionamento = request.POST.get('saude_relacionamento', '').strip()
                projeto.canal_relacionamento = request.POST.get('canal_relacionamento', '')
                projeto.ultima_interacao = request.POST.get('ultima_interacao') or None
                projeto.proxima_interacao = request.POST.get('proxima_interacao') or None
                projeto.departamentos_envolvidos = request.POST.get('departamentos_envolvidos', '').strip()
                projeto.stakeholders = request.POST.get('stakeholders', '').strip()
                projeto.resumo_colaborativo = request.POST.get('resumo_colaborativo', '').strip()
                projeto.tipo_demanda = request.POST.get('tipo_demanda', projeto.tipo_demanda)
                projeto.impacto = request.POST.get('impacto', projeto.impacto)
                projeto.solicitante = request.POST.get('solicitante', '').strip()
                projeto.responsavel_tecnico = request.POST.get('responsavel_tecnico', '').strip()
                projeto.validador = request.POST.get('validador', '').strip()
                projeto.sla = request.POST.get('sla', '').strip()
                projeto.percentual_progresso = request.POST.get('percentual_progresso') or 0
                projeto.modulo_sistema = request.POST.get('modulo_sistema', '').strip()
                projeto.ambiente = request.POST.get('ambiente', '')
                projeto.etapa_implantacao = request.POST.get('etapa_implantacao', '').strip()
                projeto.go_live_previsto = request.POST.get('go_live_previsto') or None
                projeto.treinamento_realizado = request.POST.get('treinamento_realizado') == 'on'
                projeto.severidade = request.POST.get('severidade') or None
                projeto.causa_raiz = request.POST.get('causa_raiz', '').strip()
                projeto.acao_corretiva = request.POST.get('acao_corretiva', '').strip()
                projeto.acao_preventiva = request.POST.get('acao_preventiva', '').strip()
                projeto.link_referencia = request.POST.get('link_referencia', '').strip()
                projeto.criterio_aceite = request.POST.get('criterio_aceite', '').strip()
                projeto.triagem_diagnostico = request.POST.get('triagem_diagnostico', '').strip()
                projeto.triagem_priorizacao = request.POST.get('triagem_priorizacao', '').strip()
                projeto.planejamento_entregaveis = request.POST.get('planejamento_entregaveis', '').strip()
                projeto.planejamento_riscos = request.POST.get('planejamento_riscos', '').strip()
                projeto.execucao_andamento = request.POST.get('execucao_andamento', '').strip()
                projeto.validacao_resultado = request.POST.get('validacao_resultado', '').strip()
                projeto.w2h_what = request.POST.get('w2h_what', '').strip()
                projeto.w2h_why = request.POST.get('w2h_why', '').strip()
                projeto.w2h_where = request.POST.get('w2h_where', '').strip()
                projeto.w2h_when = request.POST.get('w2h_when') or None
                projeto.w2h_who = request.POST.get('w2h_who', '').strip()
                projeto.w2h_how = request.POST.get('w2h_how', '').strip()
                projeto.w2h_how_much = request.POST.get('w2h_how_much') or None
                
                tipo_id = request.POST.get('tipo_servico')
                if tipo_id:
                    if tipo_id in ("None", "null", ""):
                        projeto.tipo_servico = None
                    else:
                        try:
                            projeto.tipo_servico = TipoServico.objects.get(id=tipo_id)
                        except (TipoServico.DoesNotExist, ValueError, TypeError):
                            return JsonResponse({'status': 'error', 'message': 'Tipo de serviço inválido'})

                with transaction.atomic():
                    projeto.save()
                    
                    # Atualizar responsáveis
                    responsaveis_data = json.loads(request.POST.get('responsaveis', '[]'))
                    if responsaveis_data:
                        projeto.projeto_responsaveis.all().delete()
                        for resp in responsaveis_data:
                            resp_id = resp.get('id')
                            if resp_id and resp_id not in ("None", "null", ""):
                                try:
                                    if CustomUser.objects.filter(id=resp_id).exists():
                                        PlannerProjectResponsavel.objects.create(projeto=projeto, responsavel_id=resp_id)
                                except (ValueError, ValidationError):
                                    pass
                    
                    # Atualizar checklist
                    checklist_items = json.loads(request.POST.get('checklist_items', '[]'))
                    if checklist_items:
                        projeto.checklist_items.all().delete()
                        for i, item in enumerate(checklist_items):
                            text = item.get('text', '').strip()
                            if text:
                                PlannerChecklistItem.objects.create(
                                    projeto=projeto,
                                    texto=text,
                                    concluido=item.get('completed', False),
                                    ordem=i
                                )

                    # Remover anexos antigos deletados no frontend
                    try:
                        removed_attachments = json.loads(request.POST.get('removed_attachments', '[]'))
                    except json.JSONDecodeError:
                        removed_attachments = []
                    
                    for att_id in removed_attachments:
                        try:
                            attachment = PlannerAttachment.objects.get(id=att_id, projeto=projeto)
                            attachment.arquivo.delete(save=False)
                            attachment.delete()
                        except PlannerAttachment.DoesNotExist:
                            pass

                    # Salvar novos anexos
                    arquivos = request.FILES.getlist('arquivos')
                    for f in arquivos:
                        PlannerAttachment.objects.create(
                            projeto=projeto,
                            nome=f.name,
                            arquivo=f,
                            uploaded_by=request.user
                        )
                    
                    PlannerProjectChangeHistory.objects.create(
                        projeto=projeto,
                        usuario=request.user,
                        tipo_alteracao='atualizado',
                        descricao="Projeto editado via formulário"
                    )

                    self.invalidate_planner_cache()

                return JsonResponse({'status': 'success', 'message': 'Projeto atualizado'})

            except Exception as e:
                return JsonResponse({"status": "error", "message": str(e)})

        # === ATUALIZAR CHECKLIST (ITEM INDIVIDUAL) ===
        elif action == "update_checklist_item":
            try:
                item_id = request.POST.get("item_id")
                completed = request.POST.get("completed") == "true"
                item = PlannerChecklistItem.objects.get(id=item_id)
                item.concluido = completed
                item.save()
                return JsonResponse({"status": "success"})
            except Exception as e:
                return JsonResponse({"status": "error", "message": str(e)})

        # === EXCLUIR PROJETO ===
        elif action == "delete_project":
            try:
                project_id = request.POST.get("project_id")
                try:
                    projeto = PlannerProject.objects.get(id=project_id)
                except (PlannerProject.DoesNotExist, ValueError, ValidationError):
                    return JsonResponse({'status': 'error', 'message': 'Projeto não encontrado'})
                projeto.delete()
                self.invalidate_planner_cache()
                return JsonResponse({"status": "success", "message": "Projeto excluído"})
            except Exception as e:
                return JsonResponse({"status": "error", "message": str(e)})

        return JsonResponse(response_data)

           
# Em Gestao_a_Vista/views.py

from django.utils.decorators import method_decorator
from .models import OcorrenciaPlanoAcao, ReincidenciaOcorrencia

@method_decorator(never_cache, name='dispatch')

class LivroOcorrenciasView(LoginRequiredMixin, TemplateView):
    template_name = "livro_ocorrencias.html"
    def get_context_data(self, **kwargs):
        import datetime  # Import local: garante que 'datetime' seja o módulo (não a classe) nesta função
        context = super().get_context_data(**kwargs)

        # --- 1. LÓGICA DINÂMICA DE FILTRO (30, 60 ou 90 dias) ---
        dias_filtro = self.request.GET.get('dias', '30')
        
        if dias_filtro == '60':
            limite_linhas = 900000
            intervalo_sql = "90 days"
        elif dias_filtro == '90':
            limite_linhas = 1000000
            intervalo_sql = "120 days"
        else:
            dias_filtro = '30' # Padrão
            limite_linhas = 500000
            intervalo_sql = "90 days"
            
        context['dias_atuais'] = dias_filtro
        
        # =========================================================
        # NOVA LÓGICA: LENDO DIRETO DO BANCO LOCAL (MODELO DJANGO)
        # =========================================================
        from .models import LivroOcorrencia
        from datetime import timedelta
        from django.utils import timezone
        
        try:
            dias_int = int(dias_filtro) if dias_filtro in ['30', '60', '90'] else 30
        except ValueError:
            dias_int = 30
            
        data_limite = timezone.now() - timedelta(days=dias_int)
        
        # Mapeamento do status local (NC, AND, C) para os textos esperados pelo restante da view
        mapa_status_reverso = {
            'NC': 'NÃO CONFORME',
            'AND': 'EM ANDAMENTO',
            'C': 'REALIZADO'
        }

        try:
            # Filtramos as ocorrências da base local pelo modelo
            ocorrencias = LivroOcorrencia.objects.filter(
                data_criacao__gte=data_limite
            ).order_by('-data_criacao')[:limite_linhas]
            
            # Filtro de acesso aos CRs (NCs do painel)
            usuario = self.request.user
            if usuario.role != 'administrador' and not getattr(usuario, 'is_superuser', False) and not getattr(usuario, 'is_general', False) and getattr(usuario, 'crs', '').strip():
                import re
                crs_usuario_nums = [re.search(r'\d+', c).group() for c in usuario.crs.split(',') if re.search(r'\d+', c)]
                
                ocorrencias_filtradas = []
                for oc in ocorrencias:
                    cr_plano = str(oc.cr) if oc.cr else ""
                    cr_num_map = re.search(r'\d+', cr_plano)
                    
                    if cr_num_map and cr_num_map.group() in crs_usuario_nums:
                        ocorrencias_filtradas.append(oc)
                    elif any(c.strip() in cr_plano for c in usuario.crs.split(',') if c.strip()):
                        ocorrencias_filtradas.append(oc)
                ocorrencias = ocorrencias_filtradas
            
            rows = []
            ocorrencia_numeros = {}
            for oc in ocorrencias:
                cr_str = str(oc.cr).strip() if oc.cr else ""
                item_str = str(oc.item).strip() if oc.item else ""
                if cr_str and item_str and oc.numero:
                    key = (normalizar_texto_global(cr_str), normalizar_texto_global(item_str))
                    if key not in ocorrencia_numeros:
                        ocorrencia_numeros[key] = oc.numero

                texto_resp = str(oc.observacao).strip() if oc.observacao else ""
                if not texto_resp or texto_resp.startswith('http'):
                    texto_resp = 'Sem relato registrado'

                rows.append({
                    'tarefa_id': str(oc.id),
                    'numero': oc.numero,
                    'estruturaid': oc.cr, 
                    'terminoreal': oc.data_criacao,
                    'item_nome': oc.item,
                    'status_resposta': mapa_status_reverso.get(oc.status, 'REALIZADO'),
                    'colaborador': oc.solicitante,
                    'estrutura_descricao': oc.cr,
                    'resposta': texto_resp,
                })
        except Exception as e:
            print(f"Erro ao buscar no LivroOcorrencia local: {e}")
            rows = []

        # --- PREPARAÇÃO DO HISTÓRICO E PLANOS ATUAIS ---
        # UMA única query com select_related substitui a query de .only() separada
        # e as 8+ queries .filter()/.count() posteriores — todos os planos são
        # carregados uma vez e particionados em Python.
        todos_planos = list(
            OcorrenciaPlanoAcao.objects
            .select_related('criador_plano', 'aprovador', 'comprador', 'retirante')
            .order_by('-data_criacao')
        )

        planos_set = set()
        historico_set = set()
        regulatorio_set = set()
        planos_concluidos_recentes = {}  # Para o timer de 24h
        agora = timezone.now()

        for p in todos_planos:
            cr_str = p.cr_colaborador.strip() if p.cr_colaborador else ""
            item_str = p.item_em_falta.strip() if p.item_em_falta else ""
            if cr_str and item_str:
                key = (normalizar_texto_global(cr_str), normalizar_texto_global(item_str))

                # Sempre adicionar ao histórico para manter rastreabilidade
                historico_set.add(key)
                if getattr(p, 'is_regulatory', False):
                    regulatorio_set.add(key)
                
                # Associa o numero do chamado do banco de dados ao plano
                p.numero_chamado = ocorrencia_numeros.get(key)

                if p.status not in ('rejeitado', 'concluido', 'excluido'):
                    planos_set.add(key)

                # Guarda a data do plano concluído para cálculo das 24h
                if p.status == 'concluido':
                    data_ref = p.data_criacao
                    if data_ref:
                        if key not in planos_concluidos_recentes or data_ref > planos_concluidos_recentes[key]['data']:
                            planos_concluidos_recentes[key] = {
                                'data': data_ref,
                                'plano': p,
                            }

        # Pré-carregar IDs de planos com reincidência pendente em 1 única query (evita N+1)
        planos_com_reincidencia_ids = set(
            ReincidenciaOcorrencia.objects.values_list('plano_original_id', flat=True)
        )

        rows_limpos = []
        for r in rows:
            if r['item_nome'] is not None and r['status_resposta'] is not None:
                resp_upper = str(r['status_resposta']).upper()
                if 'NÃO APLICADO' in resp_upper or 'NAO APLICADO' in resp_upper or 'N/A' in resp_upper or 'NÃO APLICÁVEL' in resp_upper or 'NAO APLICAVEL' in resp_upper:
                    continue
                if ' - ' in r['item_nome']:
                    r['item_nome'] = r['item_nome'].split(' - ')[0].strip()
                rows_limpos.append(r)

        ocorrencias_agrupadas = []
        stats_status = {'nao_conforme': 0, 'em_andamento': 0, 'conforme': 0}
        stats_itens = {}
        stats_cr = {}

        # Frozensets para lookups O(1) — definidos uma vez fora do loop
        TOPICOS_IGNORADOS = frozenset([
            "HA NAO CONFORMIDADES", "HA VEICULOS NO LOCAL",
            "HA POSTO ARMADO", "HA OCORRENCIAS", "HOUVE ALGUMA NAO CONFORMIDADE",
            "SISTEMA", "ALGEMA", "EPI'S BASICOS", "EPIS BASICOS", "EPIS BÁSICOS",
            "IDENTIFIQUE AS PLACAS DOS CARROS E O POSTO RESPONSÁVEL:",
            "IDENTIFIQUE AS PLACAS DOS CARROS E O POSTO RESPONSAVEL:"
        ])
        RESPOSTAS_NC = frozenset(['NÃO CONFORME', 'NAO CONFORME'])
        
        # Referência às constantes globais
        ITENS_REGULATORIOS = ITENS_REGULATORIOS_GLOBAL
        def normalizar_texto(t): return normalizar_texto_global(t)

        rows_limpos.sort(key=lambda x: (str(x['estruturaid']), x['item_nome']))

        for key, group in groupby(rows_limpos, key=lambda x: (str(x['estruturaid']), x['item_nome'])):
            est_id, item_name = key
            items_list = list(group)

            item_limpo = normalizar_texto(item_name)
            # Filtro por colaborador 'Sistema' (conforme solicitação)
            if "SISTEMA" in str(items_list[0].get('colaborador', '')).upper():
                continue

            if any(frase in item_limpo for frase in TOPICOS_IGNORADOS):
                continue

            total_items = len(items_list)

            # Calcula upper() uma vez por item — reutilizado para nc_count e status_final
            conteudos_upper = [str(i['status_resposta']).upper() for i in items_list]
            nc_count = sum(1 for c in conteudos_upper if c in RESPOSTAS_NC)

            # Buscar o item que gerou a não conformidade para pegar o relato e criador corretos
            first_item = next((i for i in items_list if str(i['status_resposta']).upper() in RESPOSTAS_NC), items_list[0])
            
            cr_label = first_item.get('estrutura_descricao', f"CR {est_id}").strip()
            item_name = item_name.strip()

            # Formatação Data
            primeira_data = first_item.get('terminoreal')
            data_iso, data_display = "", "-"
            if primeira_data:
                if isinstance(primeira_data, (datetime.datetime, datetime.date)):
                    data_iso = primeira_data.strftime('%Y-%m-%d')
                    data_display = primeira_data.strftime('%d/%m/%Y')
                else:
                    data_str = str(primeira_data)[:10]
                    if len(data_str) == 10:
                        data_iso, data_display = data_str, f"{data_str[8:10]}/{data_str[5:7]}/{data_str[0:4]}"

            status_final = 'conforme'
            item_key = (normalizar_texto_global(cr_label), normalizar_texto_global(item_name))

            if any(c in RESPOSTAS_NC for c in conteudos_upper):
                status_final = 'nao_conforme'
            elif 'EM ANDAMENTO' in conteudos_upper or 'PROBLEMA' in conteudos_upper:
                status_final = 'em_andamento'
            
            if item_key in planos_set:
                status_final = 'em_andamento'

            # --- LÓGICA DO TIMER DE 24H E REINCIDÊNCIA ---
            pular_ocorrencia = False
            
            if status_final == 'nao_conforme' and item_key in planos_concluidos_recentes:
                dados_conclusao = planos_concluidos_recentes[item_key]
                data_conclusao = dados_conclusao['data']
                plano_original = dados_conclusao['plano']
                
                # Se não passou 24h, oculta da tela
                prazo_carencia = timedelta(days=15) if getattr(plano_original, 'is_regulatory', False) else timedelta(hours=24)
                
                if agora <= data_conclusao + prazo_carencia:
                    pular_ocorrencia = True
                else:
                    # Passou 24h e voltou a ficar reprovado -> Gerar Reincidência
                    # Usa set pré-carregado para evitar N+1 (1 query para todos os planos)
                    reincidencia_existe = plano_original.id in planos_com_reincidencia_ids

                    if not reincidencia_existe:
                        nova_reincidencia = ReincidenciaOcorrencia.objects.create(
                            plano_original=plano_original,
                            cr_colaborador=cr_label,
                            item_reincidente=item_name
                        )
                        # Atualiza o set para evitar duplicatas na mesma requisição
                        planos_com_reincidencia_ids.add(plano_original.id)
                        # Engatilha envio de email passando os objetos
                        self.enviar_email_reincidencia(nova_reincidencia, plano_original)

            # Adiciona apenas se não estiver nas 24h de carência
            if not pular_ocorrencia:
                numero_str = str(first_item.get('numero', ''))
                tarefa_id_str = str(first_item.get('tarefa_id', ''))
                
                # Se tiver numero, usa o numero do BD, senao usa fallback pro id da tarefa cortado
                if numero_str and numero_str != 'None':
                    chamado_valor = numero_str
                else:
                    chamado_valor = tarefa_id_str[:8].upper() if tarefa_id_str else '-'

                ocorrencias_agrupadas.append({
                    'id_hash': f"{est_id}|{item_name}", 
                    'chamado': chamado_valor,
                    'cr': cr_label,
                    'estrutura_id': est_id,
                    'item': item_name,
                    'quantidade': total_items,
                    'quantidade_nc': nc_count, 
                    'status': status_final,
                    'colaborador': first_item.get('colaborador', 'Não identificado'),
                    'resposta': first_item.get('resposta', 'Sem relato'),
                    'data_iso': data_iso,
                    'data_display': data_display,
                    'tem_historico': item_key in historico_set,
                    'is_regulatory': any(reg_key in item_limpo for reg_key in ITENS_REGULATORIOS)
                })

                stats_status[status_final] += 1
                if nc_count > 0:
                    stats_itens[item_name] = stats_itens.get(item_name, 0) + nc_count
                    stats_cr[cr_label] = stats_cr.get(cr_label, 0) + nc_count

        context['kanban'] = {
            'nao_conforme': [o for o in ocorrencias_agrupadas if o['status'] == 'nao_conforme'],
            'em_andamento': [o for o in ocorrencias_agrupadas if o['status'] == 'em_andamento'],
            'conforme': [o for o in ocorrencias_agrupadas if o['status'] == 'conforme'],
        }

        itens_sorted = dict(sorted(stats_itens.items(), key=lambda item: item[1], reverse=True))
        crs_sorted = dict(sorted(stats_cr.items(), key=lambda item: item[1], reverse=True))
        context['charts'] = json.dumps({
            'status_labels': ['Não Conforme', 'Em Andamento', 'Conforme'],
            'status_values': [stats_status['nao_conforme'], stats_status['em_andamento'], stats_status['conforme']],
            'itens_labels': list(itens_sorted.keys())[:15], 'itens_values': list(itens_sorted.values())[:15],
            'crs_labels': list(crs_sorted.keys())[:15], 'crs_values': list(crs_sorted.values())[:15],
        })

        # Particionamento em Python usando a lista já carregada acima.
        # Elimina as 8+ queries .filter()/.count() que antes acessavam o banco
        # separadamente para cada status.
        planos_ativos = []
        for p in todos_planos:
            if p.status == 'excluido':
                continue
            item_limpo = normalizar_texto_global(p.item_em_falta) if p.item_em_falta else ""
            if any(frase in item_limpo for frase in TOPICOS_IGNORADOS):
                continue
            planos_ativos.append(p)

        # Filtro de acesso aos CRs (Aprovação)
        usuario = self.request.user
        if usuario.role != 'administrador' and getattr(usuario, 'crs', '').strip():
            import re
            # Extrai os números dos CRs cadastrados no perfil do gestor
            crs_usuario_nums = [re.search(r'\d+', c).group() for c in usuario.crs.split(',') if re.search(r'\d+', c)]
            planos_filtrados = []
            for p in planos_ativos:
                adicionar = False
                cr_plano = str(p.cr_colaborador) if p.cr_colaborador else ""
                
                cr_num_map = re.search(r'\d+', cr_plano)
                if cr_num_map and cr_num_map.group() in crs_usuario_nums:
                    adicionar = True
                elif any(c.strip() in cr_plano for c in usuario.crs.split(',') if c.strip()):
                    adicionar = True
                    
                if adicionar:
                    planos_filtrados.append(p)
            
            planos_ativos = planos_filtrados

        # OTIMIZAÇÃO: Garante que is_regulatory seja exibido se o nome do item for regulatório,
        # mesmo para registros antigos que não tinham a flag setada no banco.
        for p in planos_ativos:
            item_limpo = normalizar_texto_global(p.item_em_falta)
            if not getattr(p, 'is_regulatory', False):
                if any(reg_key in item_limpo for reg_key in ITENS_REGULATORIOS_GLOBAL):
                    p.is_regulatory = True

        context['planos_em_aprovacao']     = [p for p in planos_ativos if p.status == 'em_aprovacao']
        context['planos_compra_cadastrar'] = [p for p in planos_ativos if p.status == 'compra_cadastrar']
        context['planos_compra_pedido']    = [p for p in planos_ativos if p.status == 'compra_pedido']
        context['planos_compra_entregue']  = [p for p in planos_ativos if p.status == 'compra_entregue']
        context['planos_concluidos']       = [p for p in planos_ativos if p.status == 'concluido']
        context['total_compras']           = (
            len(context['planos_compra_cadastrar']) + len(context['planos_compra_pedido'])
        )
        context['historico_planos']        = planos_ativos[:100]
        
        return context

    def enviar_email_reincidencia(self, reincidencia, plano_original):
        if reincidencia.email_enviado: return
        
        link_aprovar = self.request.build_absolute_uri(reverse('gestao_a_vista:aprovar_reincidencia', args=[reincidencia.id]))
        
        gerente_nome = "Gerente"
        coordenador_nome = "Coordenador"
        destinatarios = [] 
        
        try:
            from .models import CustomUser
            from django.db import models
            
            import re
            
            cr = getattr(plano_original, 'cr_colaborador', None)
            is_reg = getattr(plano_original, 'is_regulatory', False)
            
            # Extrai apenas números do CR para facilitar o match
            cr_num = re.search(r'\d+', str(cr)).group() if cr and re.search(r'\d+', str(cr)) else None
            
            # Buscar na base de Gestores/Coordenadores ativos
            gestores = CustomUser.objects.filter(role__in=['gerente', 'coordenador'], is_active=True)

            for gestor in gestores:
                # TRAVA ABSOLUTA: Regulatório só recebe regulatório, Normal só recebe normal
                gestor_is_reg = getattr(gestor, 'is_regulatory', False)
                if is_reg and not gestor_is_reg:
                    continue
                if not is_reg and gestor_is_reg:
                    continue

                match = False
                if gestor.is_general:
                    match = True
                elif cr and gestor.crs:
                    # Extrai os números dos CRs cadastrados no perfil do gestor
                    gestor_crs_nums = [re.search(r'\d+', c).group() for c in gestor.crs.split(',') if re.search(r'\d+', c)]
                    if cr_num and cr_num in gestor_crs_nums:
                        match = True
                    # Fallback para string normal caso não haja números
                    elif cr.strip() in [c.strip() for c in gestor.crs.split(',')]:
                        match = True
                
                if match:
                    if gestor.email:
                        destinatarios.append(gestor.email)
                        
            # O fallback global foi removido. A lista final depende unicamente dos `CustomUser`s configurados.
                
        except Exception as e:
            import traceback
            traceback.print_exc()
            print(f"Erro ao buscar configuração de e-mail nos usuários: {e}")

        if not destinatarios:
            print(f"Nenhum e-mail de gerente ou coordenador encontrado nas lógicas!")
            return
        
        url_foto = None
        if hasattr(plano_original, 'foto_retirada') and plano_original.foto_retirada:
            url_foto = self.request.build_absolute_uri(plano_original.foto_retirada.url)

        context = {
            'plano': plano_original,
            'nome_gerente': f"{coordenador_nome} / {gerente_nome}",
            'cr': reincidencia.cr_colaborador,
            'item_falta': reincidencia.item_reincidente,
            'link_aprovacao': link_aprovar,
            'url_foto': url_foto,
            'is_reincidencia': True,
            'aviso_extra': "ATENÇÃO: Este item apresentou reincidência de falha APÓS a tratativa de 24h."
        }

        from django.template.loader import render_to_string
        from django.core.mail import EmailMessage

        html_content = render_to_string('emails/aprovacao_plano_acao.html', context)
        assunto = f"ALERTA DE AUDITORIA E REINCIDÊNCIA - CR {reincidencia.cr_colaborador}"
        
        email_msg = EmailMessage(
            subject=assunto,
            body=html_content,
            to=destinatarios,
        )
        email_msg.content_subtype = "html"
        
        try:
            email_msg.send(fail_silently=False)
            reincidencia.email_enviado = True
            reincidencia.save()
        except Exception as e:
            print(f"Erro no envio do email de reincidência: {e}")

# =============================================================================
# VIEW DE DETALHES (MODAL IFRAME) - SOMENTE ÚLTIMOS 30 DIAS
# =============================================================================
@method_decorator(xframe_options_exempt, name='dispatch')
@method_decorator(never_cache, name='dispatch')
class LivroOcorrenciasDetalheView(LoginRequiredMixin, TemplateView):
    template_name = "livro_ocorrencias_detalhes.html"

    def get_context_data(self, **kwargs):
        from django.utils import timezone
        from datetime import timedelta
        
        context = super().get_context_data(**kwargs)
        logger = logging.getLogger(__name__)
        
        estrutura_id = self.kwargs.get('pk')
        item_nome = self.request.GET.get('item', '')
        
        if not estrutura_id:
            context['erro_critico'] = "ID da estrutura inválido ou não fornecido."
            return context

        dados_cr, dados_gerente = {}, {}
        sql_info_local = """
            SELECT 
                e.descricao as cr_nome, e.cr as cr_codigo, e.nivel_4 as regional,
                e.nivel_5 as filial, e.gc as gerente_nome, u.email as gerente_email
            FROM public.estrutura e
            LEFT JOIN "Gestao_a_Vista_customuser" u ON u.name = e.gc 
            WHERE e.id::text = %s OR e.cr = %s OR e.descricao = %s LIMIT 1
        """
        try:
            with connections['default'].cursor() as cursor:
                cursor.execute(sql_info_local, [str(estrutura_id), str(estrutura_id), str(estrutura_id)])
                row = cursor.fetchone()
                if row:
                    cols = [col[0] for col in cursor.description]
                    res = dict(zip(cols, row))
                    dados_cr = {'cr': res.get('cr_codigo'), 'descricao': res.get('cr_nome'), 'nivel_4': res.get('regional'), 'nivel_5': res.get('filial')}
                    dados_gerente = {'nome': res.get('gerente_nome') or 'Não Identificado', 'email': res.get('gerente_email'), 'telefone': 'Não disponível'}
        except Exception as e:
            pass

        # === FILTRO DE 30 DIAS APLICADO AQUI ===
        detalhes_itens = []
        data_limite = timezone.now() - timedelta(days=30)

        filtro_status = self.request.GET.get('status')
        cache_key = f"livro_det_{estrutura_id}_{hash(item_nome)}_{filtro_status or 'all'}"
        cached_detalhes = cache.get(cache_key)
        
        if cached_detalhes is not None:
            context['ocorrencia_grupo'] = {'item': item_nome, 'cr': dados_cr.get('cr', ''), 'status_geral': cached_detalhes[0]['status'] if cached_detalhes else 'conforme'}
            context['itens'] = cached_detalhes
            context['estrutura'] = dados_cr
            context['gerente'] = dados_gerente
            return context

        # =========================================================
        # NOVA LÓGICA: LENDO DIRETO DO BANCO LOCAL (MODELO DJANGO)
        # =========================================================
        try:
            from .models import LivroOcorrencia
            
            # O nome do CR que o nosso script guardou (COALESCE(cr, descricao))
            nome_cr_busca = dados_cr.get('cr') or dados_cr.get('descricao') or estrutura_id
            
            # Mapeamento do status do banco (NC, AND, C) para o padrão HTML/CSS do site
            mapa_status_html = {
                'NC': 'nao_conforme',
                'AND': 'pendente',
                'C': 'conforme'
            }

            # Faz a busca no ORM local (Milissegundos, pois não tem JOIN)
            ocorrencias = LivroOcorrencia.objects.filter(
                cr=nome_cr_busca,
                item__startswith=item_nome,
                data_criacao__gte=data_limite
            ).order_by('-data_criacao')

            for oc in ocorrencias:
                status_convertido = mapa_status_html.get(oc.status, 'conforme')
                
                # Aplica o filtro de status vindo da URL (se existir)
                if filtro_status and status_convertido != filtro_status:
                    continue
                
                # Verifica se a observação é um link (foto)
                texto_resp = str(oc.observacao).strip() if oc.observacao else ""
                e_foto = texto_resp.startswith('http')
                
                detalhes_itens.append({
                    'tarefa_id': str(oc.id),
                    'created_at': oc.data_criacao,
                    'colaborador': oc.solicitante,
                    'status': status_convertido,
                    'resposta': "Ver anexo" if e_foto else texto_resp,
                    'quantidade': oc.quantidade if oc.quantidade and oc.quantidade > 1 else None,
                    'fotos': [texto_resp] if e_foto else []
                })

            status_geral = detalhes_itens[0]['status'] if detalhes_itens else 'conforme'
            cache.set(cache_key, detalhes_itens, 300)

        except Exception as e:
            logger.exception(f"Erro ao buscar no LivroOcorrencia local: {e}")
            detalhes_itens = []
            status_geral = 'erro'

        is_reg = any(reg_key in normalizar_texto_global(item_nome) for reg_key in ITENS_REGULATORIOS_GLOBAL)
        context['ocorrencia_grupo'] = {
            'item': item_nome, 
            'cr': dados_cr.get('cr', ''), 
            'status_geral': status_geral,
            'is_regulatory': is_reg
        }
        context['itens'] = detalhes_itens
        context['estrutura'] = dados_cr
        context['gerente'] = dados_gerente
        return context
    

@require_POST
@login_required
def criar_plano_ocorrencia(request):
    import traceback
    from django.core.mail import EmailMessage
    from django.template.loader import render_to_string
    from django.conf import settings
    from django.urls import reverse
    from .models import OcorrenciaPlanoAcao, CustomUser

    try:
        data = json.loads(request.body)
        item = data.get('item_em_falta')
        colaborador = data.get('colaborador_nc')
        cr = data.get('cr_colaborador')
        criador_id = data.get('criador_plano_id')
        tem_estoque = data.get('tem_estoque', False)

        is_regulatory = data.get('is_regulatory', False)

        criador = CustomUser.objects.get(id=criador_id) if criador_id else request.user
        criador_nome = criador.name if criador.name else criador.username

        status_inicial = 'compra_entregue' if tem_estoque else 'em_aprovacao'

        plano = OcorrenciaPlanoAcao.objects.create(
            item_em_falta=item, colaborador_nc=colaborador, cr_colaborador=cr,
            criador_plano=criador, tem_estoque=tem_estoque, status=status_inicial,
            is_regulatory=is_regulatory
        )

        from .models import HistoricoPlanoAcao
        HistoricoPlanoAcao.objects.create(
            plano=plano, status_anterior=None, status_novo=status_inicial,
            usuario=criador, observacao="Plano criado"
        )

        if status_inicial == 'em_aprovacao':
            destinatarios = []
            try:
                from django.db import models
                import re
                
                # Extrai apenas números do CR para facilitar o match (exemplo: "88390 - GO" -> "88390")
                cr_num = re.search(r'\d+', str(cr)).group() if cr and re.search(r'\d+', str(cr)) else None
                
                # Buscar na base de Gestores/Coordenadores ativos
                gestores = CustomUser.objects.filter(role__in=['gerente', 'coordenador'], is_active=True)

                for gestor in gestores:
                    # TRAVA ABSOLUTA: Regulatório só recebe regulatório, Normal só recebe normal
                    gestor_is_reg = getattr(gestor, 'is_regulatory', False)
                    if is_regulatory and not gestor_is_reg:
                        continue
                    if not is_regulatory and gestor_is_reg:
                        continue

                    match = False
                    if gestor.is_general:
                        match = True
                    elif cr and gestor.crs:
                        # Identifica todos os números de CRs que o gestor atende
                        gestor_crs_nums = [re.search(r'\d+', c).group() for c in gestor.crs.split(',') if re.search(r'\d+', c)]
                        if cr_num and cr_num in gestor_crs_nums:
                            match = True
                        elif cr.strip() in [c.strip() for c in gestor.crs.split(',')]:
                            match = True
                    
                    if match:
                        if gestor.email:
                            destinatarios.append(gestor.email)
                
                # Fallback antigo removido. Destinatários agora vêm exclusivamente de CustomUser.
            except Exception as e:
                import traceback
                traceback.print_exc()
                print(f"Erro ao buscar os emails no CustomUser: {e}")
            
            if destinatarios:
                # GERA OS LINKS SEGUROS PARA O E-MAIL DE APROVAÇÃO
                link_aprovar = request.build_absolute_uri(reverse('gestao_a_vista:acao_plano_email', args=[plano.id, 'aprovar']))
                link_rejeitar = request.build_absolute_uri(reverse('gestao_a_vista:acao_plano_email', args=[plano.id, 'rejeitar']))
                link_painel = request.build_absolute_uri(reverse('gestao_a_vista:torre_controle'))
                
                contexto_email = {
                    'cr': cr, 'item': item, 'colaborador': colaborador, 'criador_nome': criador_nome,
                    'link_aprovar': link_aprovar, 'link_rejeitar': link_rejeitar, 'link_painel': link_painel
                }
                html_content = render_to_string("emails/aprovacao_plano_acao.html", contexto_email)

                email_msg = EmailMessage(
                    subject=f"Aprovação de Plano de Ação Pendente - {cr}",
                    body=html_content, to=destinatarios
                )
                email_msg.content_subtype = "html"
                email_msg.send(fail_silently=False)
            else:
                print(f"Nenhum destinatário encontrado com CR correspondente a '{cr_num}' ou 'Geral' nas regras regulatórias={is_regulatory}!")

        return JsonResponse({'success': True, 'message': 'Plano criado com sucesso!'})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'message': f"Erro: {str(e)}"}, status=500)


@require_POST
@login_required
def excluir_plano_ocorrencia(request, pk):
    try:
        plano = OcorrenciaPlanoAcao.objects.get(id=pk)
        
        # CHECAGEM DE PERMISSÃO: Admin pode deletar qualquer um. Outros só os próprios.
        if request.user.role != 'administrador' and not request.user.is_superuser:
            if plano.criador_plano != request.user:
                return JsonResponse({'success': False, 'message': 'Operação restrita. Você só tem permissão para excluir planos criados por você.'}, status=403)

        status_anterior = plano.status
        plano.status = 'excluido'
        try:
            plano.excluido_por = request.user
        except:
            pass
        plano.save()
        
        from .models import HistoricoPlanoAcao
        HistoricoPlanoAcao.objects.create(
            plano=plano, status_anterior=status_anterior, status_novo='excluido',
            usuario=request.user, observacao="Plano excluído via interface"
        )
        return JsonResponse({'success': True, 'message': 'Plano de ação excluído com sucesso.'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)

@require_POST
@login_required
def aprovar_plano_ocorrencia(request, pk):
    # Libera coordenadores além de gerentes (e variações de regulatório)
    roles_permitidas = ['administrador', 'gerente', 'coordenador', 'gerente_regulatorio', 'coordenador_regulatorio']
    if request.user.role not in roles_permitidas:
        return JsonResponse({'success': False, 'message': 'Apenas gerentes e coordenadores podem aprovar.'}, status=403)

    try:
        data = json.loads(request.body)
        acao = data.get('acao') # 'aprovar' ou 'rejeitar'
        justificativa = data.get('justificativa', '')
        plano = OcorrenciaPlanoAcao.objects.get(id=pk)

        # OTIMIZAÇÃO: Força leitura dinâmica da flag para planos com dados legados
        is_plano_regulatory = bool(getattr(plano, 'is_regulatory', False))
        if not is_plano_regulatory:
            item_limpo = normalizar_texto_global(plano.item_em_falta)
            if any(reg_key in item_limpo for reg_key in ITENS_REGULATORIOS_GLOBAL):
                is_plano_regulatory = True

        user_is_reg_by_role = 'regulatorio' in request.user.role.lower()
        is_user_regulatory = bool(getattr(request.user, 'is_regulatory', False)) or user_is_reg_by_role

        # Trava mutuamente exclusiva de is_regulatory (ignora para administradores)
        if request.user.role != 'administrador':
            if is_plano_regulatory != is_user_regulatory:
                tipo_plano = "Regulatório" if is_plano_regulatory else "Normal"
                tipo_user = "Regulatório" if is_user_regulatory else "Normal"
                return JsonResponse({
                    'success': False,
                    'message': f'Acesso Negado. Este plano é {tipo_plano}, mas o seu perfil de aprovação é {tipo_user}.'
                }, status=403)

        status_anterior = plano.status
        plano.aprovador = request.user
        plano.status = 'compra_cadastrar' if acao == 'aprovar' else 'rejeitado'
        plano.justificativa_aprovacao = justificativa
        
        # Atualiza a flag no banco de dados para evitar problemas futuros
        plano.is_regulatory = is_plano_regulatory
        plano.save()

        from .models import HistoricoPlanoAcao
        HistoricoPlanoAcao.objects.create(
            plano=plano,
            status_anterior=status_anterior,
            status_novo=plano.status,
            usuario=request.user,
            observacao=f"Plano {acao} com justificativa: {justificativa}"
        )
        
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)
        
        
@require_POST
@login_required
def enviar_para_auditoria(request):
    try:
        import json
        data = json.loads(request.body)
        cr = data.get('cr_colaborador')
        item = data.get('item_em_falta')
        observacao = data.get('observacao')
        
        if not cr or not item:
            return JsonResponse({'success': False, 'message': 'CR e Item são obrigatórios.'}, status=400)
            
        from .models import Estrutura, ReincidenciaOcorrencia, OcorrenciaPlanoAcao
        from django.db.models import Q
        
        estrutura = Estrutura.objects.using('default').filter(Q(descricao__icontains=cr) | Q(cr=cr)).first()
        gerente_nome = estrutura.gc if (estrutura and estrutura.gc) else 'Não Atribuído'
        
        plano = OcorrenciaPlanoAcao.objects.create(
            cr_colaborador=cr,
            item_em_falta=item,
            criador_plano=request.user,
            status='auditoria',
            tem_estoque=False,
            is_regulatory=False
        )
        
        ReincidenciaOcorrencia.objects.create(
            plano_original=plano,
            cr_colaborador=cr,
            item_reincidente=item,
            coordenador='Sistema', 
            gerente=gerente_nome,
            status_aprovacao='pendente',
            observacao=observacao
        )
        
        return JsonResponse({'success': True, 'message': 'Item enviado para auditoria com sucesso.'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)

@require_POST
@login_required
def cadastrar_compra_ocorrencia(request, pk):
    if request.user.role not in ['administrador', 'gerente', 'coordenador']:
        return JsonResponse({'success': False, 'message': 'Apenas gerentes e coordenadores podem cadastrar compras.'}, status=403)

    try:
        if 'nota_fiscal' not in request.FILES:
            return JsonResponse({'success': False, 'message': 'O anexo da Nota Fiscal (ou comprovante) é obrigatório!'}, status=400)

        plano = OcorrenciaPlanoAcao.objects.get(id=pk)

        # OTIMIZAÇÃO: Força leitura dinâmica da flag para planos com dados legados
        is_plano_regulatory = getattr(plano, 'is_regulatory', False)
        if not is_plano_regulatory:
            item_limpo = normalizar_texto_global(plano.item_em_falta)
            if any(reg_key in item_limpo for reg_key in ITENS_REGULATORIOS_GLOBAL):
                is_plano_regulatory = True

        is_user_regulatory = getattr(request.user, 'is_regulatory', False)

        # Trava mutuamente exclusiva de is_regulatory (ignora para administradores)
        if request.user.role != 'administrador':
            if is_plano_regulatory != is_user_regulatory:
                tipo_plano = "Regulatório" if is_plano_regulatory else "Normal"
                tipo_user = "Regulatório" if is_user_regulatory else "Normal"
                return JsonResponse({
                    'success': False,
                    'message': f'Acesso Negado. Este plano é {tipo_plano}, mas o seu perfil é {tipo_user}.'
                }, status=403)

        plano.data_compra = request.POST.get('data_compra')
        previsao = request.POST.get('previsao_entrega')
        if previsao:
            plano.previsao_entrega = previsao
            
        plano.itens_compra = json.loads(request.POST.get('itens_compra', '[]'))
        
        comprador_id = request.POST.get('comprador_id')
        plano.comprador = CustomUser.objects.get(id=comprador_id) if comprador_id else request.user
        
        plano.nota_fiscal = request.FILES['nota_fiscal']
        
        if 'guia_de_trafego' in request.FILES:
            plano.guia_de_trafego = request.FILES['guia_de_trafego']
            
        status_anterior = plano.status
        plano.status = 'compra_pedido'
        
        # Atualiza a flag no banco de dados
        plano.is_regulatory = is_plano_regulatory
        plano.save()

        from .models import HistoricoPlanoAcao
        HistoricoPlanoAcao.objects.create(
            plano=plano,
            status_anterior=status_anterior,
            status_novo='compra_pedido',
            usuario=request.user,
            observacao="Compra/Pedido cadastrado com anexos."
        )
        
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)
    

@require_POST
@login_required
def atualizar_status_compra(request, pk):
    try:
        data = json.loads(request.body)
        plano = OcorrenciaPlanoAcao.objects.get(id=pk)
        status_anterior = plano.status
        plano.status = data.get('status')
        
        # Salva a data de entrega se enviada (do novo modal)
        if data.get('data'):
            plano.data_compra = data.get('data')

        plano.save()
        
        from .models import HistoricoPlanoAcao
        HistoricoPlanoAcao.objects.create(
            plano=plano, status_anterior=status_anterior, status_novo=plano.status,
            usuario=request.user, observacao="Status da compra atualizado manualmente."
        )
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)

@require_POST
@login_required
def registrar_retirada_ocorrencia(request):
    try:
        data = request.POST
        plano = OcorrenciaPlanoAcao.objects.get(id=data.get('plano_id'))
        status_anterior = plano.status
        retirante_id = data.get('retirante_id')
        
        plano.retirante = CustomUser.objects.get(id=retirante_id) if retirante_id else request.user
        plano.recebedor = data.get('recebedor')
        plano.item_retirado = data.get('item_retirado')
        
        if 'foto_retirada' in request.FILES:
            plano.foto_retirada = request.FILES['foto_retirada']

        from django.utils import timezone
        plano.data_retirada = timezone.now()
        plano.status = 'concluido'
        plano.save()
        
        from .models import HistoricoPlanoAcao
        HistoricoPlanoAcao.objects.create(
            plano=plano, status_anterior=status_anterior, status_novo='concluido',
            usuario=request.user, observacao=f"Item recebido por {plano.recebedor}."
        )
        
        return JsonResponse({'success': True})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'message': str(e)}, status=400)


@login_required
def download_qr_pdf(request):
    """
    Gera e baixa um PDF com todos os QR Codes selecionados
    """
    if request.method != "POST":
        return JsonResponse({"error": "Método não permitido"}, status=405)

    try:
        data = json.loads(request.POST.get("data", "{}"))
        print(f"[DEBUG] Dados recebidos para PDF: {data}")

        service = data.get("service")
        cr_number = data.get("cr_number")
        locations = data.get("locations", [])
        description_config = data.get("description_config", {"mode": "auto"})

        print(f"[DEBUG PDF] Service: {service}")
        print(f"[DEBUG PDF] CR Number: {cr_number}")
        print(f"[DEBUG PDF] Total locations recebidas: {len(locations)}")
        for i, loc in enumerate(locations):
            print(
                f"[DEBUG PDF] Location {i+1}: {loc.get('displayName', 'NO DISPLAY NAME')} (ID: {loc.get('id', 'NO ID')})"
            )

        # Converter para int com tratamento de erro
        try:
            logo_size = int(data.get("logo_size", 80))
        except (ValueError, TypeError):
            logo_size = 80

        try:
            service_logo_size = int(data.get("service_logo_size", 60))
        except (ValueError, TypeError):
            service_logo_size = 60

        if not service or not cr_number or not locations:
            return JsonResponse({"error": "Dados incompletos"}, status=400)

        # ==========================================
        # 1. PRÉ-CARREGAR DADOS E ASSETS (OTIMIZAÇÃO)
        # ==========================================
        
        # 1.1 Buscar todas as estruturas de uma vez (Bulk Fetch)
        location_ids = []
        for loc in locations:
            lid = loc.get("id")
            if lid:
                if isinstance(lid, str) and len(lid) > 10:
                    location_ids.append(lid)
                elif str(lid).replace("-", "").replace("_", "").isdigit():
                    location_ids.append(lid)
        
        estruturas_dict = {}
        if location_ids:
            estruturas_list = Estrutura.objects.filter(id__in=location_ids)
            for est in estruturas_list:
                estruturas_dict[str(est.id)] = est

        # 1.2 Carregar Logo OpsVista apenas uma vez
        app_logo_resized = None
        import os
        possible_paths = [
            os.path.join("Gestao_a_Vista", "templates", "image", "visa.png"),
            os.path.join("Gestao_a_Vista", "templates", "image", "logo.png"),
            os.path.join(os.path.dirname(__file__), "templates", "image", "visa.png"),
            os.path.join(os.path.dirname(__file__), "templates", "image", "logo.png"),
        ]
        for logo_path in possible_paths:
            if os.path.exists(logo_path):
                try:
                    ops_vista_logo = Image.open(logo_path)
                    logo_height = 80
                    logo_width = 250
                    app_logo_resized = ops_vista_logo.copy()
                    if hasattr(Image, 'Resampling'):
                        app_logo_resized.thumbnail((logo_width, logo_height), Image.Resampling.LANCZOS)
                    else:
                        app_logo_resized.thumbnail((logo_width, logo_height), Image.LANCZOS)
                    break
                except Exception:
                    pass

        # 1.3 Carregar Logo do Serviço apenas uma vez
        service_logo_resized = None
        try:
            logo_servico = LogoServico.objects.filter(nome=service).first()
            if logo_servico and logo_servico.img_base64:
                img_data = logo_servico.img_base64
                if img_data.startswith("data:image/"):
                    img_data = img_data.split(",")[1]
                img_bytes = base64.b64decode(img_data)
                service_logo = Image.open(BytesIO(img_bytes))
                service_logo_resized = service_logo.copy()
                if hasattr(Image, 'Resampling'):
                    service_logo_resized.thumbnail((service_logo_size, service_logo_size), Image.Resampling.LANCZOS)
                else:
                    service_logo_resized.thumbnail((service_logo_size, service_logo_size), Image.LANCZOS)
        except Exception as e:
            print(f"Erro ao carregar logo do serviço em bulk: {e}")

        # 1.4 Carregar Logo do Cliente apenas uma vez
        client_logo_resized = None
        if "client_logo" in request.FILES:
            try:
                client_logo = Image.open(request.FILES["client_logo"])
                client_logo_resized = client_logo.copy()
                if hasattr(Image, 'Resampling'):
                    client_logo_resized.thumbnail((logo_size, logo_size), Image.Resampling.LANCZOS)
                else:
                    client_logo_resized.thumbnail((logo_size, logo_size), Image.LANCZOS)
            except Exception as e:
                print(f"Erro ao processar logo do cliente: {e}")

        # 1.5 Carregar Fontes uma vez
        def get_unicode_font(size):
            font_paths = [
                "C:/Windows/Fonts/arial.ttf", "C:/Windows/Fonts/calibri.ttf", 
                "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
                "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
                "/System/Library/Fonts/Arial.ttf", "arial.ttf"
            ]
            for font_path in font_paths:
                try:
                    return ImageFont.truetype(font_path, size)
                except (OSError, IOError):
                    continue
            return ImageFont.load_default()
            
        fonts = {
            16: get_unicode_font(16),
            14: get_unicode_font(14),
            12: get_unicode_font(12)
        }
        fallback_font = ImageFont.load_default()

        # ==========================================
        # 2. GERAR IMAGENS
        # ==========================================
        qr_images = []
        import unicodedata

        for location in locations:
            location_id = str(location.get("id", ""))
            estrutura = estruturas_dict.get(location_id)
            
            # Fallback rápido se não achou no bulk
            if not estrutura and location.get("displayName"):
                estrutura = Estrutura.objects.filter(descricao__iexact=location["displayName"]).first()

            qr_link = None
            if estrutura and estrutura.qrcode and estrutura.qrcode.strip():
                qr_link = estrutura.qrcode.strip()

            # Gerar descrição personalizada baseada na configuração
            custom_description = generate_custom_description(
                estrutura, location["displayName"], description_config
            )

            if qr_link:
                qr_data_content = qr_link
            else:
                location_name = custom_description
                words = location_name.split(" ")
                lines = []
                current_line = ""

                for word in words:
                    if len(current_line + " " + word) <= 30:
                        current_line += (" " + word) if current_line else word
                    else:
                        if current_line:
                            lines.append(current_line)
                        current_line = word

                if current_line:
                    lines.append(current_line)

                formatted_location = "\n".join(lines)

                qr_data_fallback = {
                    "cr": cr_number,
                    "service": service,
                    "location": formatted_location,
                }
                # Usar ensure_ascii=False para não codificar caracteres especiais
                qr_data_content = json.dumps(qr_data_fallback, ensure_ascii=False)

            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=10,
                border=4,
            )
            qr.add_data(qr_data_content)
            qr.make(fit=True)
            qr_image = qr.make_image(fill_color="black", back_color="white")

            final_image = Image.new("RGB", (300, 400), "white")
            draw = ImageDraw.Draw(final_image)
            draw.rectangle([(5, 5), (295, 395)], outline="black", width=2)

            if app_logo_resized:
                logo_x = (300 - app_logo_resized.width) // 2
                logo_y = 120 - app_logo_resized.height - 5
                if app_logo_resized.mode == "RGBA":
                    final_image.paste(app_logo_resized, (logo_x, logo_y), app_logo_resized)
                else:
                    final_image.paste(app_logo_resized, (logo_x, logo_y))
            else:
                draw.text((150, 35), "OPSVISTA", fill="black", font=fallback_font, anchor="mm")

            if service_logo_resized:
                logo_x = max(300 - service_logo_size - 15, 10)
                logo_y = 15
                if service_logo_resized.mode == "RGBA":
                    final_image.paste(service_logo_resized, (logo_x, logo_y), service_logo_resized)
                else:
                    final_image.paste(service_logo_resized, (logo_x, logo_y))
            else:
                draw.text((260, 20), service.upper(), fill="black", font=fallback_font, anchor="mm")

            qr_size = 140
            qr_pos = ((300 - qr_size) // 2, 120)
            final_image.paste(qr_image.resize((qr_size, qr_size)), qr_pos)

            text_y = qr_pos[1] + qr_size + 15
            location_text = custom_description.upper()
            words = location_text.split(" ")
            lines = []
            current_line = ""

            for word in words:
                test_line = current_line + (" " + word if current_line else word)
                if len(test_line) <= 25:
                    current_line = test_line
                else:
                    if current_line:
                        lines.append(current_line)
                    current_line = word

            if current_line:
                lines.append(current_line)

            total_chars = len(location_text)
            num_lines = len(lines)

            if total_chars <= 20 and num_lines <= 2:
                font = fonts[16]
                line_spacing = 20
            elif total_chars <= 40 and num_lines <= 3:
                font = fonts[14]
                line_spacing = 18
            else:
                font = fonts[12]
                line_spacing = 15

            for i, line in enumerate(lines):
                line_y = text_y + (i * line_spacing)
                try:
                    normalized_line = unicodedata.normalize('NFC', str(line))
                    draw.text((150, line_y), normalized_line, fill="black", font=font, anchor="mm")
                except Exception:
                    draw.text((150, line_y), str(line), fill="black", font=font, anchor="mm")

            if client_logo_resized:
                logo_y = max(400 - client_logo_resized.height - 20, 320)
                logo_x = (300 - client_logo_resized.width) // 2
                if client_logo_resized.mode == "RGBA":
                    final_image.paste(client_logo_resized, (logo_x, logo_y), client_logo_resized)
                else:
                    final_image.paste(client_logo_resized, (logo_x, logo_y))

            # Converter para bytes para o PDF
            buffer = BytesIO()
            final_image.save(buffer, format="PNG")
            buffer.seek(0)

            qr_images.append({"image": buffer, "location": custom_description})

            # Criar PDF
            response = HttpResponse(content_type="application/pdf")
            response[
                "Content-Disposition"
            ] = f'attachment; filename="qr_codes_{cr_number}.pdf"'

            pdf_buffer = BytesIO()
            p = canvas.Canvas(pdf_buffer, pagesize=A4)

            width, height = A4

            cols = 3
            rows = 4
            qr_codes_per_page = cols * rows

            qr_width = 140
            qr_height = 187

            margin_top = 30
            margin_bottom = 30
            margin_left = 40
            margin_right = 40

            spacing_x = 15
            spacing_y = 15

            available_width = width - margin_left - margin_right
            available_height = height - margin_top - margin_bottom

            total_qr_width = cols * qr_width + (cols - 1) * spacing_x
            total_qr_height = rows * qr_height + (rows - 1) * spacing_y

            start_x = margin_left + (available_width - total_qr_width) / 2
            start_y = height - margin_top - (available_height - total_qr_height) / 2

            total_pages = (len(qr_images) + qr_codes_per_page - 1) // qr_codes_per_page

            for page_num in range(total_pages):
                if page_num > 0:
                    p.showPage()

                page_start = page_num * qr_codes_per_page
                page_end = min(page_start + qr_codes_per_page, len(qr_images))
                page_qrs = qr_images[page_start:page_end]

                for idx, qr_data in enumerate(page_qrs):
                    row = idx // cols
                    col = idx % cols

                    x = start_x + col * (qr_width + spacing_x)
                    y = start_y - row * (qr_height + spacing_y) - qr_height

                    qr_data["image"].seek(0)
                    p.drawImage(
                        ImageReader(qr_data["image"]),
                        x,
                        y,
                        width=qr_width,
                        height=qr_height,
                    )

                    p.setStrokeColorRGB(0.8, 0.8, 0.8)
                    p.setDash(2, 2)

                    if row > 0:
                        cut_y = y + qr_height + spacing_y / 2
                        p.line(
                            x - spacing_x / 4,
                            cut_y,
                            x + qr_width + spacing_x / 4,
                            cut_y,
                        )

                    if col > 0:
                        cut_x = x - spacing_x / 2
                        p.line(
                            cut_x,
                            y - spacing_y / 4,
                            cut_x,
                            y + qr_height + spacing_y / 4,
                        )

                    p.setDash()
                    p.setStrokeColorRGB(0, 0, 0)

        p.save()

        pdf_buffer.seek(0)
        response.write(pdf_buffer.getvalue())
        pdf_buffer.close()

        return response

    except Exception as e:
        import traceback

        print(f"[ERROR] Erro ao gerar PDF: {str(e)}")
        print(f"[ERROR] Traceback: {traceback.format_exc()}")
        return JsonResponse({"error": f"Erro ao gerar PDF: {str(e)}"}, status=500)

        return super().get(request, *args, **kwargs)


@csrf_exempt
@login_required
@login_required
@csrf_exempt
def download_livro_ata_pdf(request):
    """
    Gera e baixa um PDF com o QR Code do Livro Ata
    """
    if request.method != "POST":
        return JsonResponse({"error": "Método não permitido"}, status=405)

    try:
        data = json.loads(request.POST.get("data", "{}"))
        livro_ata_id = data.get("livro_ata_id")
        logo_size = int(data.get("logo_size", 80))
        service_logo_size = int(data.get("service_logo_size", 60))
        
        if not livro_ata_id:
            return JsonResponse({"error": "ID do Livro Ata não fornecido"}, status=400)

        # Buscar o registro do Livro Ata
        from .models import LivroAtaQRCode
        try:
            livro_ata = LivroAtaQRCode.objects.using('default').get(id=livro_ata_id)
        except LivroAtaQRCode.DoesNotExist:
            return JsonResponse({"error": "Livro Ata não encontrado"}, status=404)

        # Gerar o QR Code usando a mesma lógica da função generate_livro_ata_qr
        # Buscar a estrutura pelo cr_id (que é o ID da estrutura como UUID)
        try:
            cr_estrutura = Estrutura.objects.using('default').get(id=livro_ata.cr_id)
        except Estrutura.DoesNotExist:
            return JsonResponse({"error": "Estrutura não encontrada"}, status=404)

        # Forçar atualização para o novo formato correto
        nova_url = f"{settings.SITE_URL}/livroata/qrcode={livro_ata.id}/"
        if livro_ata.qr_code_url != nova_url:
            livro_ata.qr_code_url = nova_url
            livro_ata.save(using='default')

        # Gerar QR Code
        qr_data = livro_ata.qr_code_url
        
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(qr_data)
        qr.make(fit=True)

        # Criar imagem do QR code
        qr_image = qr.make_image(fill_color="black", back_color="white")
        
        # Criar imagem final com layout IGUAL AOS LOCAIS (300x400)
        final_img = Image.new("RGB", (300, 400), "white")
        draw = ImageDraw.Draw(final_img)

        # Adicionar borda IGUAL AOS LOCAIS
        draw.rectangle([(5, 5), (295, 395)], outline="black", width=2)

        # Redimensionar QR Code para o tamanho padrão dos locais
        qr_size = 140
        qr_image = qr_image.resize((qr_size, qr_size))
        
        # Carregar logo fixa do OpsVista (IGUAL AOS LOCAIS)
        ops_vista_logo = None
        try:
            import os
            from django.conf import settings
            
            possible_paths = [
                os.path.join("Gestao_a_Vista", "templates", "image", "visa.png"),
                os.path.join("Gestao_a_Vista", "templates", "image", "logo.png"),
                os.path.join(os.path.dirname(__file__), "templates", "image", "visa.png"),
                os.path.join(os.path.dirname(__file__), "templates", "image", "logo.png"),
            ]
            
            for logo_path in possible_paths:
                if os.path.exists(logo_path):
                    ops_vista_logo = Image.open(logo_path)
                    break
        except Exception as e:
            print(f"Erro ao carregar logo OpsVista: {e}")
            ops_vista_logo = None
        
        # Posicionar QR Code no centro (igual aos locais)
        qr_x = (300 - qr_size) // 2
        qr_y = 120  # Mesma posição dos locais
        final_img.paste(qr_image, (qr_x, qr_y))
        
        # Adicionar logo fixa do OpsVista (IGUAL AOS LOCAIS)
        if ops_vista_logo:
            try:
                logo_height = service_logo_size
                logo_width = int(service_logo_size * 3.125)  # Proporção 250/80
                app_logo_resized = ops_vista_logo.copy()
                
                try:
                    app_logo_resized.thumbnail((logo_width, logo_height), Image.Resampling.LANCZOS)
                except AttributeError:
                    app_logo_resized.thumbnail((logo_width, logo_height), Image.LANCZOS)
                
                logo_x = (300 - app_logo_resized.width) // 2
                logo_y = qr_y - logo_height - 5
                
                if app_logo_resized.mode == "RGBA":
                    final_img.paste(app_logo_resized, (logo_x, logo_y), app_logo_resized)
                else:
                    final_img.paste(app_logo_resized, (logo_x, logo_y))
            except Exception as e:
                print(f"Erro ao adicionar logo OpsVista: {e}")

        # ✅ ADICIONAR LOGO DO SERVIÇO (canto superior direito)
        service_name = data.get("service")
        print(f"[DEBUG] service_name: {service_name}")
        
        if service_name:
            try:
                from .models import LogoServico
                logo_servico = LogoServico.objects.filter(nome=service_name).first()
                print(f"[DEBUG] logo_servico encontrado: {logo_servico is not None}")
                
                if logo_servico and logo_servico.img_base64:
                    print(f"[DEBUG] Processando logo do serviço: {service_name}")
                    
                    # Extrair dados base64 (remover prefixo data:image/png;base64, se existir)
                    img_data = logo_servico.img_base64
                    if img_data.startswith("data:image/"):
                        img_data = img_data.split(",")[1]

                    # Decodificar base64 e criar imagem
                    import base64
                    img_bytes = base64.b64decode(img_data)
                    service_logo = Image.open(BytesIO(img_bytes))
                    
                    print(f"[DEBUG] Logo do serviço aberta. Mode: {service_logo.mode}, Size: {service_logo.size}")

                    # Redimensionar conforme tamanho selecionado pelo usuário
                    service_logo_resized = service_logo.copy()
                    try:
                        service_logo_resized.thumbnail((service_logo_size, service_logo_size), Image.Resampling.LANCZOS)
                    except AttributeError:
                        service_logo_resized.thumbnail((service_logo_size, service_logo_size), Image.LANCZOS)
                    
                    print(f"[DEBUG] Logo do serviço redimensionada para: {service_logo_resized.size}")

                    # Posicionar no canto superior direito com margens confortáveis (IGUAL AOS LOCAIS)
                    service_logo_x = 300 - service_logo_resized.width - 15  # 15px da direita
                    service_logo_y = 15  # 15px do topo
                    
                    print(f"[DEBUG] Posição da logo do serviço: x={service_logo_x}, y={service_logo_y}")

                    if service_logo_resized.mode == "RGBA":
                        final_img.paste(service_logo_resized, (service_logo_x, service_logo_y), service_logo_resized)
                    else:
                        final_img.paste(service_logo_resized, (service_logo_x, service_logo_y))
                    
                    print(f"[INFO] [OK] Logo do serviço '{service_name}' adicionada ao PDF (tamanho: {service_logo_size}px)")
                else:
                    print(f"[WARNING] Logo do serviço '{service_name}' não encontrada no banco de dados")
            except Exception as e:
                print(f"[ERROR] Erro ao carregar logo do serviço: {e}")
                import traceback
                print(f"[ERROR] Traceback: {traceback.format_exc()}")

        # Adicionar texto "LIVRO DE OCORRÊNCIA"
        def load_unicode_font_pdf(size):
            font_paths = [
                "C:/Windows/Fonts/arial.ttf",
                "C:/Windows/Fonts/calibri.ttf", 
                "arial.ttf",
                "calibri.ttf",
            ]
            
            for font_path in font_paths:
                try:
                    font = ImageFont.truetype(font_path, size)
                    return font
                except (OSError, IOError):
                    continue
            
            return ImageFont.load_default()
        
        text = "LIVRO DE OCORRÊNCIA"
        font_size = 16
        font = load_unicode_font_pdf(font_size)
        
        text_y = qr_y + qr_size + 15
        bbox = draw.textbbox((0, 0), text, font=font)
        text_width = bbox[2] - bbox[0]
        text_x = (300 - text_width) // 2
        
        draw.text((text_x, text_y), text, fill="black", font=font)
        
        # ✅ ADICIONAR LOGO DO CLIENTE NO RODAPÉ (se fornecida)
        client_logo_file = request.FILES.get("client_logo")
        print(f"[DEBUG] request.FILES: {request.FILES}")
        print(f"[DEBUG] client_logo_file: {client_logo_file}")
        
        if client_logo_file:
            try:
                print(f"[DEBUG] Processando logo do cliente: {client_logo_file.name}, size: {client_logo_file.size}")
                client_logo = Image.open(client_logo_file)
                print(f"[DEBUG] Logo aberta com sucesso. Mode: {client_logo.mode}, Size: {client_logo.size}")
                
                # Converter para RGBA se necessário
                if client_logo.mode not in ('RGB', 'RGBA'):
                    client_logo = client_logo.convert('RGBA')
                
                # Redimensionar logo do cliente
                client_logo_resized = client_logo.copy()
                try:
                    client_logo_resized.thumbnail((logo_size, logo_size), Image.Resampling.LANCZOS)
                except AttributeError:
                    client_logo_resized.thumbnail((logo_size, logo_size), Image.LANCZOS)
                
                print(f"[DEBUG] Logo redimensionada para: {client_logo_resized.size}")
                
                # Posicionar no rodapé (centralizado)
                client_logo_x = (300 - client_logo_resized.width) // 2
                client_logo_y = 400 - client_logo_resized.height - 10
                
                print(f"[DEBUG] Posição da logo: x={client_logo_x}, y={client_logo_y}")
                
                # Colar logo com transparência se disponível
                if client_logo_resized.mode == "RGBA":
                    final_img.paste(client_logo_resized, (client_logo_x, client_logo_y), client_logo_resized)
                else:
                    final_img.paste(client_logo_resized, (client_logo_x, client_logo_y))
                    
                print(f"[INFO] [OK] Logo do cliente adicionada ao PDF do Livro Ata (tamanho: {logo_size}px)")
            except Exception as e:
                print(f"[ERROR] Erro ao adicionar logo do cliente ao PDF: {e}")
                import traceback
                print(f"[ERROR] Traceback: {traceback.format_exc()}")
        else:
            print(f"[WARNING] Nenhuma logo do cliente foi enviada no request.FILES")

        # Converter para buffer
        buffer = BytesIO()
        final_img.save(buffer, format='PNG')
        buffer.seek(0)

        # Criar PDF
        response = HttpResponse(content_type="application/pdf")
        response['Content-Disposition'] = 'attachment; filename="livro_ata_qr_code.pdf"'

        pdf_buffer = BytesIO()
        p = canvas.Canvas(pdf_buffer, pagesize=A4)

        # Dimensões A4 em pontos
        width, height = A4
        
        # Centralizar QR Code na página
        qr_width = 200  # Tamanho do QR Code no PDF
        qr_height = 267  # Altura proporcional (200 * 400/300)
        
        x = (width - qr_width) / 2
        y = (height - qr_height) / 2

        # Adicionar QR Code ao PDF
        buffer.seek(0)
        p.drawImage(ImageReader(buffer), x, y, width=qr_width, height=qr_height)

        p.save()
        pdf_buffer.seek(0)
        response.write(pdf_buffer.getvalue())
        pdf_buffer.close()

        return response

    except Exception as e:
        import traceback
        print(f"[ERROR] Erro ao gerar PDF do Livro Ata: {str(e)}")
        print(f"[ERROR] Traceback: {traceback.format_exc()}")
        return JsonResponse({"error": f"Erro ao gerar PDF: {str(e)}"}, status=500)


class ReservaSalasView( TemplateView):
    template_name = "reserva_salas.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        unidade_slug = self.kwargs.get("unidade_slug")
        unidade = get_object_or_404(Unidade, slug=unidade_slug, ativa=True)

        context["unidade"] = unidade
        context["today"] = timezone.now().date()
        return context

    def get(self, request, *args, **kwargs):
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            # Verificar se é uma requisição de horários baseada na URL ou parâmetros
            if ("sala" in request.GET and "data" in request.GET) or "horarios" in request.GET:
                return self.get_horarios_disponiveis(request)
                
            unidade_slug = kwargs.get("unidade_slug")
            unidade = get_object_or_404(Unidade, slug=unidade_slug, ativa=True)

            salas = GestaoSala.objects.filter(unidade=unidade)

            salas_data = []
            for sala in salas:
                salas_data.append(
                    {
                        "id": sala.id_sala,
                        "nome": sala.nome,
                        "capacidade": sala.capacidade,
                        "hora_inicio": sala.hora_inicio.strftime("%H:%M"),
                        "hora_fim": sala.hora_fim.strftime("%H:%M"),
                        "quantidade_m": sala.quantidade_m,
                        'foto': sala.foto.url if sala.foto and getattr(sala.foto, 'url', None) else None,
                        "mesas": [
                            {"id": i + 1, "nome": f"Mesa {str(i+1).zfill(2)}"}
                            for i in range(sala.quantidade_m)
                        ]
                        if sala.quantidade_m > 1
                        else [],
                    }
                )

            return JsonResponse({"salas": salas_data})
        return super().get(request, *args, **kwargs)

    def get_horarios_disponiveis(self, request):
        # Forçamos a importação exata das classes necessárias para evitar conflitos de módulo
        from datetime import datetime, timedelta
        
        try:
            sala_id = request.GET.get("sala")
            data = request.GET.get("data")
            mesa = request.GET.get("mesa")
            
            if not sala_id or not data:
                return JsonResponse({"success": False, "message": "Parâmetros inválidos"}, status=400)

            sala = get_object_or_404(GestaoSala, id_sala=sala_id)
            
            # strptime funcionando perfeitamente devido ao import forçado no topo da função
            data_obj = datetime.strptime(data, "%Y-%m-%d").date()

            # Buscar reservas existentes
            reservas = ReservaSala.objects.filter(sala_id=sala.id_sala, data=data_obj, status='ativa')
            
            if mesa:
                reservas = reservas.filter(mesa_id=mesa)
            elif sala.quantidade_m > 1:
                reservas = reservas.filter(mesa_id__isnull=True)

            horarios_reservados = set()
            for reserva in reservas:
                hora_inicio = None
                hora_fim = None
                
                # Extrai corretamente o horário da string da reserva
                if reserva.horario and " - " in reserva.horario:
                    try:
                        inicio_str, fim_str = reserva.horario.split(" - ")
                        hora_inicio = datetime.strptime(inicio_str.strip(), "%H:%M").time()
                        hora_fim = datetime.strptime(fim_str.strip(), "%H:%M").time()
                    except ValueError:
                        # Se a string do horário não estiver no formato esperado, ignora
                        continue 

                if hora_inicio and hora_fim:
                    # Usa data_obj (que já é tipo date) ao invés de datetime.today()
                    hora_atual_dt = datetime.combine(data_obj, hora_inicio)
                    hora_fim_dt = datetime.combine(data_obj, hora_fim)
                    
                    while hora_atual_dt < hora_fim_dt:
                        horarios_reservados.add(hora_atual_dt.strftime("%H:%M"))
                        hora_atual_dt += timedelta(minutes=30)

            return JsonResponse({"success": True, "horarios_reservados": list(horarios_reservados)})

        except Exception as e:
            import traceback
            print(f"Erro na API de horários disponíveis: {traceback.format_exc()}")
            return JsonResponse({"success": False, "message": f"Erro ao carregar horários: {str(e)}"}, status=500)

    def post(self, request, *args, **kwargs):
        import json
        import logging
        import traceback
        from django.http import JsonResponse
        from django.core.exceptions import ValidationError
        from django.template.loader import render_to_string
        from django.utils.html import strip_tags
        from django.core.mail import EmailMultiAlternatives
        from django.conf import settings
        from django.utils import timezone
        from datetime import datetime, timedelta
        from .models import PrestadorServico, Unidade, GestaoSala, ReservaSala

        logger = logging.getLogger(__name__)
        unidade_slug = self.kwargs.get("unidade_slug")

        try:
            # 1. Lê os dados enviados via JS
            if request.content_type == 'application/json':
                data_post = json.loads(request.body)
            else:
                data_post = request.POST

            # 2. Busca unidade e sala
            unidade = Unidade.objects.get(slug=unidade_slug, ativa=True)
            sala_id = data_post.get("sala_id")
            
            if not sala_id:
                return JsonResponse({"success": False, "message": "O ID da sala não foi enviado."}, status=400)
                
            sala = GestaoSala.objects.get(id_sala=sala_id)

            # 3. Captura e tratamento de dados do Payload JSON
            data_str = data_post.get("data")
            data_reserva = datetime.strptime(data_str, "%Y-%m-%d").date() if data_str else None
            
            horarios = data_post.get("horarios", [])
            dia_todo = data_post.get("dia_todo", False)
            
            if not horarios and not dia_todo:
                return JsonResponse({"success": False, "message": "Selecione pelo menos um horário ou marque o Dia Todo."}, status=400)

            if dia_todo:
                hora_inicio_sala = sala.hora_inicio.strftime("%H:%M") if hasattr(sala.hora_inicio, 'strftime') else str(sala.hora_inicio)[:5]
                hora_fim_sala = sala.hora_fim.strftime("%H:%M") if hasattr(sala.hora_fim, 'strftime') else str(sala.hora_fim)[:5]
                horario_str = f"{hora_inicio_sala} - {hora_fim_sala}"
            else:
                hora_inicio = horarios[0]
                # Corrige o bug do "10:30 - 10:30" adicionando 30 minutos ao último bloco
                hora_ultimo_bloco = datetime.strptime(horarios[-1], "%H:%M")
                hora_fim_real = (hora_ultimo_bloco + timedelta(minutes=30)).strftime("%H:%M")
                horario_str = f"{hora_inicio} - {hora_fim_real}"

            mesa_id = data_post.get("mesa_numero") or data_post.get("mesa")
            
            titulo = data_post.get("titulo", "")
            nome = data_post.get("nome", titulo)
            if not nome:
                nome = "Nova Reserva"
                
            descricao = data_post.get("descricao", "")
            observacoes = data_post.get("observacoes", descricao)
            
            email = data_post.get("email", "")
            telefone = data_post.get("telefone", "")

            # Serviços extras
            servico_limpeza = data_post.get("servico_limpeza", False)
            servico_entreposto = data_post.get("servico_entreposto", False)
            servico_coffe = data_post.get("servico_coffe", False)

            # 4. Instancia e Salva a reserva
            reserva = ReservaSala(
                sala_id=sala.id_sala,
                data=data_reserva,
                horario=horario_str,
                solicitante=nome,
                observacoes=observacoes,
                status='ativa',
                email=email,
                telefone=telefone,
                mesa_id=mesa_id if mesa_id else None,
                servico_limpeza=servico_limpeza,
                servico_entreposto=servico_entreposto,
                servico_coffe=servico_coffe,
                created_at=timezone.now(),
                updated_at=timezone.now(),
            )
            reserva.clean()
            reserva.save()

            # 5. Auditoria
            try:
                from .audit_utils import registrar_acao_auditoria
                
                # Modificação que realizamos para suportar acesso público
                usuario_auditoria = request.user if request.user.is_authenticated else None
                
                registrar_acao_auditoria(
                    usuario=usuario_auditoria,
                    acao="criou",
                    tipo_item="reserva",
                    item_id=str(reserva.pk),
                    detalhes=f"Reserva criada para {nome} na sala {sala.nome}. ID: {reserva.pk}",
                    status_anterior="N/A",
                    status_novo="ativa"
                )
            except Exception as e:
                logger.error(f"Erro na auditoria de reserva: {e}")

            # =========================================================
            # 6. ENVIO DE E-MAILS (Alterado para não omitir erros críticos)
            # =========================================================
            
            # 6.1 Envio para o Solicitante
            if reserva.email:
                subject = f"Confirmação de Reserva - {sala.nome}"
                email_context = {"reserva": reserva, "sala": sala}
                
                html_content = render_to_string("emails/reserva_comprador.html", email_context)
                text_content = strip_tags(html_content)
                
                email_msg = EmailMultiAlternatives(subject, text_content, settings.DEFAULT_FROM_EMAIL, [reserva.email])
                email_msg.attach_alternative(html_content, "text/html")
                
                # Se falhar aqui, o código desce direto pro 'except Exception as e' lá embaixo
                email_msg.send()

            # 6.2 Envio para os Prestadores de Serviço
            areas_solicitadas = []
            if servico_limpeza: areas_solicitadas.append('limpeza')
            if servico_entreposto: areas_solicitadas.append('entreposto')
            if servico_coffe: areas_solicitadas.append('coffe')

            if areas_solicitadas:
                prestadores = PrestadorServico.objects.filter(ativo=True, area_servico__in=areas_solicitadas)
                
                for prestador in prestadores:
                    area_nome = prestador.get_area_servico_display() if hasattr(prestador, 'get_area_servico_display') else prestador.area_servico
                    subject_prestador = f"Nova Solicitação de Serviço ({area_nome}) - Sala {sala.nome}"
                    
                    context_prestador = {
                        "reserva": reserva,
                        "sala": sala,
                        "prestador": prestador,
                        "nome_prestador": prestador.nome,
                        "servico_nome": area_nome
                    }
                    
                    html_content_prestador = render_to_string("emails/reserva_prestador.html", context_prestador)
                    text_content_prestador = strip_tags(html_content_prestador)
                    
                    email_prestador_msg = EmailMultiAlternatives(
                        subject_prestador, text_content_prestador, settings.DEFAULT_FROM_EMAIL, [prestador.email]
                    )
                    email_prestador_msg.attach_alternative(html_content_prestador, "text/html")
                    
                    # Envio direto para capturar o erro caso exista
                    email_prestador_msg.send()

            # 7. Resposta JSON de Sucesso
            return JsonResponse({
                "success": True, 
                "message": "Reserva realizada com sucesso!", 
                "reserva_id": str(reserva.pk)
            })

        except Unidade.DoesNotExist:
            return JsonResponse({"success": False, "message": "Unidade não encontrada."}, status=404)
        except GestaoSala.DoesNotExist:
            return JsonResponse({"success": False, "message": "Sala não encontrada."}, status=404)
        except ValidationError as e:
            msgs = "; ".join(e.messages) if hasattr(e, 'messages') else str(e)
            return JsonResponse({"success": False, "message": f"Erro de validação: {msgs}"}, status=400)
        except Exception as e:
            # AGORA SIM! Se der erro de SMTP ou de credencial do Google, ele vai estourar aqui no terminal
            print("==================================================")
            print("ERRO CRÍTICO NO ENVIO DA RESERVA / E-MAIL:")
            print(traceback.format_exc())
            print("==================================================")
            return JsonResponse({
                "success": False, 
                "message": f"Erro do sistema (Possível falha no E-mail): {str(e)}"
            }, status=500)
        
        


class CalendarioReservasView(TemplateView):
    template_name = "calendario_reservas.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Buscar unidades ativas (sem cache para debug)
        unidades = Unidade.objects.filter(ativa=True)
        context["unidades"] = unidades
        return context

    def get(self, request, *args, **kwargs):
        # Verificar se é requisição AJAX ou API
        is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"
        is_api_path = "/api/reservas/" in request.path
        
        if is_ajax or is_api_path:
            try:
                data_inicio = request.GET.get("data_inicio")
                data_fim = request.GET.get("data_fim")
                unidade_id = request.GET.get("unidade")
                sala_id = request.GET.get("sala")

                # Query base com filtros
                reservas = ReservaSala.objects.all()

                # Filtros
                if data_inicio:
                    reservas = reservas.filter(data__gte=data_inicio)
                if data_fim:
                    reservas = reservas.filter(data__lte=data_fim)
                if sala_id:
                    reservas = reservas.filter(sala_id=sala_id)

                # Filtro por unidade requer join manual
                if unidade_id:
                    salas_da_unidade = GestaoSala.objects.filter(unidade_id=unidade_id).values_list('id_sala', flat=True)
                    reservas = reservas.filter(sala_id__in=salas_da_unidade)

                # Ordenar por data e horario
                reservas = reservas.order_by("data", "horario")

                # Pré-carregar dados das salas para evitar N+1 queries
                salas_dict = {}
                if reservas.exists():
                    sala_ids = list(reservas.values_list('sala_id', flat=True).distinct())
                    salas = GestaoSala.objects.filter(id_sala__in=sala_ids).select_related('unidade')
                    salas_dict = {str(sala.id_sala): sala for sala in salas}

                # Serializar dados
                reservas_data = []
                for reserva in reservas:
                    try:
                        # Usar dados pré-carregados
                        sala_obj = salas_dict.get(str(reserva.sala_id))
                        sala_nome = sala_obj.nome if sala_obj else "Sala não encontrada"
                        unidade_nome = sala_obj.unidade.nome if sala_obj and sala_obj.unidade else "Unidade não encontrada"
                        unidade_id_val = sala_obj.unidade.id if sala_obj and sala_obj.unidade else None
                        
                        # Extrair hora_inicio e hora_fim do campo horario
                        horario_str = reserva.horario or ""
                        hora_inicio = ""
                        hora_fim = ""
                        
                        if " - " in horario_str:
                            partes = horario_str.split(" - ")
                            hora_inicio = partes[0].strip()
                            hora_fim = partes[1].strip() if len(partes) > 1 else ""
                        else:
                            hora_inicio = horario_str
                        
                        reservas_data.append({
                            "id": str(reserva.id_reserva),
                            "data": reserva.data.strftime("%Y-%m-%d") if reserva.data else "",
                            "hora_inicio": hora_inicio,
                            "hora_fim": hora_fim,
                            "horario": reserva.horario or "",
                            "nome": reserva.nome or "",
                            "solicitante": reserva.solicitante or "",
                            "email": reserva.email or "",
                            "telefone": reserva.telefone or "",
                            "observacoes": reserva.observacoes or "",
                            "sala": sala_nome,
                            "unidade": unidade_nome,
                            "unidade_id": unidade_id_val,
                            "mesa_numero": reserva.mesa_id,
                            "status": reserva.status or "ativa",
                        })
                    except Exception as e:
                        # Log do erro mas continua processando outras reservas
                        print(f"Erro ao serializar reserva {reserva.id_reserva}: {e}")
                        continue

                return JsonResponse({"reservas": reservas_data})

            except Exception as e:
                import traceback
                error_details = traceback.format_exc()
                print(f"ERRO na API de reservas: {error_details}")
                return JsonResponse(
                    {
                        "success": False,
                        "message": f"Erro ao carregar reservas: {str(e)}",
                        "error_details": error_details if settings.DEBUG else None,
                    },
                    status=500,
                )

        return super().get(request, *args, **kwargs)

    def patch(self, request, reserva_id, *args, **kwargs):
        """Atualizar uma reserva (editar ou cancelar)"""
        try:
            import json
            data = json.loads(request.body)
            
            reserva = get_object_or_404(ReservaSala, id_reserva=reserva_id)
            
            # Atualizar campos se fornecidos
            if 'status' in data:
                reserva.status = data['status']
            if 'nome' in data:
                reserva.nome = data['nome']
                reserva.solicitante = data['nome']  # Manter compatibilidade
            if 'email' in data:
                reserva.email = data['email']
            if 'telefone' in data:
                reserva.telefone = data['telefone']
            if 'data' in data:
                from datetime import datetime
                reserva.data = datetime.strptime(data['data'], '%Y-%m-%d').date()
            if 'horario' in data:
                reserva.horario = data['horario']
            if 'observacoes' in data:
                reserva.observacoes = data['observacoes']
            
            reserva.save()
            
            return JsonResponse({
                "success": True, 
                "message": "Reserva atualizada com sucesso"
            })
            
        except Exception as e:
            return JsonResponse(
                {
                    "success": False,
                    "message": f"Erro ao atualizar reserva: {str(e)}",
                },
                status=400,
            )

    def delete(self, request, reserva_id, *args, **kwargs):
        try:
            reserva = get_object_or_404(ReservaSala, id_reserva=reserva_id)
            reserva.delete()
            return JsonResponse(
                {"success": True, "message": "Reserva excluída com sucesso"}
            )
        except Exception as e:
            return JsonResponse(
                {"success": False, "message": f"Erro ao excluir reserva: {str(e)}"},
                status=400,
            )


@login_required
def api_historico_reservas(request):
    """API para listar histórico completo de reservas com paginação"""
    try:
        # Parâmetros de paginação
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 50))  # 50 reservas por página
        
        # Buscar todas as reservas ordenadas
        reservas_query = ReservaSala.objects.all().order_by('-data', '-created_at')

        # Aplicar paginação
        paginator = Paginator(reservas_query, page_size)
        try:
            reservas_page = paginator.page(page)
        except PageNotAnInteger:
            reservas_page = paginator.page(1)
        except EmptyPage:
            reservas_page = paginator.page(paginator.num_pages)
        
        # Pré-carregar dados das salas para evitar N+1 queries
        salas_dict = {}
        if reservas_page:
            # Extrair sala_ids da página atual sem usar distinct() após slice
            sala_ids = list(set(reserva.sala_id for reserva in reservas_page))
            salas = GestaoSala.objects.filter(id_sala__in=sala_ids).select_related('unidade')
            salas_dict = {str(sala.id_sala): sala for sala in salas}

        reservas_data = []
        for reserva in reservas_page:
            try:
                # Usar dados pré-carregados
                sala_obj = salas_dict.get(str(reserva.sala_id))
                sala_nome = sala_obj.nome if sala_obj else "Sala não encontrada"
                unidade_nome = sala_obj.unidade.nome if sala_obj and sala_obj.unidade else "Unidade não encontrada"
                unidade_id = sala_obj.unidade.id if sala_obj and sala_obj.unidade else None
                
                # Extrair hora_inicio e hora_fim do campo horario
                horario_str = reserva.horario or ""
                hora_inicio = ""
                hora_fim = ""
                
                if " - " in horario_str:
                    partes = horario_str.split(" - ")
                    hora_inicio = partes[0].strip()
                    hora_fim = partes[1].strip() if len(partes) > 1 else ""
                else:
                    hora_inicio = horario_str
                
                reservas_data.append({
                    "id": str(reserva.id_reserva),
                    "data": reserva.data.strftime("%Y-%m-%d") if reserva.data else "",
                    "hora_inicio": hora_inicio,
                    "hora_fim": hora_fim,
                    "horario": reserva.horario or "",
                    "nome": reserva.nome or "",
                    "solicitante": reserva.solicitante or "",
                    "email": reserva.email or "",
                    "telefone": reserva.telefone or "",
                    "observacoes": reserva.observacoes or "",
                    "sala": sala_nome,
                    "unidade": unidade_nome,
                    "unidade_id": unidade_id,
                    "mesa_numero": reserva.mesa_id,
                    "status": reserva.status or "ativa",
                })
            except Exception as e:
                # Log do erro mas continua processando outras reservas
                print(f"Erro ao serializar reserva {reserva.id_reserva}: {e}")
                continue
        
        return JsonResponse({
            "reservas": reservas_data,
            "pagination": {
                "current_page": reservas_page.number,
                "total_pages": paginator.num_pages,
                "total_items": paginator.count,
                "page_size": page_size,
                "has_next": reservas_page.has_next(),
                "has_previous": reservas_page.has_previous(),
            }
        })
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"ERRO na API de histórico: {error_details}")
        return JsonResponse({
            'success': False,
            'message': f'Erro ao carregar histórico: {str(e)}',
            'error_details': error_details if settings.DEBUG else None,
        }, status=500)

class SelecionarUnidadeView(TemplateView):
    template_name = "selecionar_unidade.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        unidades = Unidade.objects.filter(ativa=True)
        context["unidades"] = unidades.prefetch_related("salas_contrato")
        return context

@login_required
def api_salas(request):
    """API para listar salas por unidade"""
    try:
        unidade_id = request.GET.get('unidade')
        if not unidade_id:
            return JsonResponse({'salas': []})

        salas = GestaoSala.objects.filter(unidade_id=unidade_id).order_by('nome')
        salas_data = []
        for sala in salas:
            salas_data.append({
                'id': str(sala.id_sala),
                'nome': sala.nome,
                'capacidade': sala.capacidade,
                'foto': sala.foto.url if sala.foto and getattr(sala.foto, 'url', None) else None,
            })
        return JsonResponse({'salas': salas_data})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erro ao carregar salas: {str(e)}'}, status=400)


def api_salas_por_unidade(request, unidade_slug):
    """API para listar salas por slug da unidade (compatibilidade com reserva_salas)"""
    try:
        unidade = get_object_or_404(Unidade, slug=unidade_slug, ativa=True)

        salas = GestaoSala.objects.filter(unidade=unidade).order_by('nome')
        salas_data = []
        for sala in salas:
            salas_data.append({
                'id': str(sala.id_sala),
                'nome': sala.nome,
                'capacidade': sala.capacidade,
                'hora_inicio': sala.hora_inicio.strftime('%H:%M') if sala.hora_inicio else '',
                'hora_fim': sala.hora_fim.strftime('%H:%M') if sala.hora_fim else '',
                'quantidade_m': sala.quantidade_m,
                'foto': sala.foto.url if sala.foto and getattr(sala.foto, 'url', None) else None,
            })
        return JsonResponse({'salas': salas_data})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erro ao carregar salas: {str(e)}'}, status=400)

@login_required
@check_page_permission("implantacoes_opsvista")
def implantacoes_opsvista(request):
    """View para a página de Implantações OpsVista"""
    if request.method == "GET" and request.GET.get("id"):
        # Buscar um registro específico para edição
        try:
            registro = get_object_or_404(ImplantacoesOpsVista, id=request.GET.get("id"))
            return JsonResponse(
                {
                    "id": str(registro.id),
                    "cr": str(registro.cr_id),
                    "cr_descricao": "",  # Não buscar descrição para evitar erro
                    "sistema": registro.sistema,
                    "implantacoes": registro.implantacoes,
                    "dashboards": registro.dashboards,
                    "servico": registro.servico,
                    "status": registro.status,
                    "observacoes": registro.observacoes,
                }
            )
        except Exception as e:
            return JsonResponse(
                {"success": False, "message": f"Erro ao buscar registro: {str(e)}"},
                status=400,
            )

    elif request.method == "POST":
        try:
            data = json.loads(request.body)

            # Validar se o CR foi fornecido
            cr_id = data.get("cr")
            if not cr_id:
                return JsonResponse(
                    {"success": False, "message": "CR é obrigatório"}, status=400
                )

            # Processar implantações (converter string para lista se necessário)
            implantacoes = data.get("implantacoes", [])
            if isinstance(implantacoes, str):
                implantacoes = [
                    item.strip() for item in implantacoes.split(",") if item.strip()
                ]

            registro = ImplantacoesOpsVista.objects.create(
                cr_id=cr_id,
                sistema=data["sistema"],
                implantacoes=implantacoes,
                dashboards=data.get("dashboards", "nao_possui"),
                servico=data["servico"],
                status=data["status"],
                observacoes=data.get("observacoes", ""),
            )

            return JsonResponse(
                {
                    "success": True,
                    "message": "Registro criado com sucesso",
                    "data": {
                        "id": str(registro.id),
                        "cr": str(registro.cr_id),
                        "cr_descricao": "",  # Não buscar descrição para evitar erro
                        "sistema": registro.sistema,
                        "implantacoes": registro.implantacoes,
                        "dashboards": registro.get_dashboards_display(),
                        "servico": registro.get_servico_display(),
                        "status": registro.get_status_display(),
                        "observacoes": registro.observacoes,
                        "created_at": registro.created_at.strftime("%d/%m/%Y %H:%M"),
                    },
                }
            )
        except ValueError as e:
            return JsonResponse(
                {"success": False, "message": f"Erro de validação: {str(e)}"},
                status=400,
            )
        except Exception as e:
            import traceback

            print(f"Erro ao criar registro: {str(e)}")
            print(f"Traceback: {traceback.format_exc()}")
            return JsonResponse(
                {"success": False, "message": f"Erro ao criar registro: {str(e)}"},
                status=400,
            )

    elif request.method == "PUT":
        try:
            data = json.loads(request.body)
            registro = get_object_or_404(ImplantacoesOpsVista, id=data["id"])

            # Validar se o CR foi fornecido
            cr_id = data.get("cr")
            if not cr_id:
                return JsonResponse(
                    {"success": False, "message": "CR é obrigatório"}, status=400
                )

            # Processar implantações
            implantacoes = data.get("implantacoes", [])
            if isinstance(implantacoes, str):
                implantacoes = [
                    item.strip() for item in implantacoes.split(",") if item.strip()
                ]

            # Atualizar os campos
            registro.cr_id = cr_id
            registro.sistema = data["sistema"]
            registro.implantacoes = implantacoes
            registro.dashboards = data.get("dashboards", "nao_possui")
            registro.servico = data["servico"]
            registro.status = data["status"]
            registro.observacoes = data.get("observacoes", "")
            registro.save()

            return JsonResponse(
                {
                    "success": True,
                    "message": "Registro atualizado com sucesso",
                    "data": {
                        "id": str(registro.id),
                        "cr": str(registro.cr_id),
                        "cr_descricao": "",  # Não buscar descrição para evitar erro
                        "sistema": registro.sistema,
                        "implantacoes": registro.implantacoes,
                        "dashboards": registro.get_dashboards_display(),
                        "servico": registro.get_servico_display(),
                        "status": registro.get_status_display(),
                        "observacoes": registro.observacoes,
                        "updated_at": registro.updated_at.strftime("%d/%m/%Y %H:%M"),
                    },
                }
            )
        except ImplantacoesOpsVista.DoesNotExist:
            return JsonResponse(
                {"success": False, "message": "Registro não encontrado"}, status=404
            )
        except ValueError as e:
            return JsonResponse(
                {"success": False, "message": f"Erro de validação: {str(e)}"},
                status=400,
            )
        except Exception as e:
            import traceback

            print(f"Erro ao atualizar registro: {str(e)}")
            print(f"Traceback: {traceback.format_exc()}")
            return JsonResponse(
                {"success": False, "message": f"Erro ao atualizar registro: {str(e)}"},
                status=400,
            )

    elif request.method == "DELETE":
        try:
            data = json.loads(request.body)
            registro = get_object_or_404(ImplantacoesOpsVista, id=data["id"])
            registro.delete()
            return JsonResponse(
                {"success": True, "message": "Registro excluído com sucesso"}
            )
        except ImplantacoesOpsVista.DoesNotExist:
            return JsonResponse(
                {"success": False, "message": "Registro não encontrado"}, status=404
            )
        except Exception as e:
            import traceback

            print(f"Erro ao excluir registro: {str(e)}")
            print(f"Traceback: {traceback.format_exc()}")
            return JsonResponse(
                {"success": False, "message": f"Erro ao excluir registro: {str(e)}"},
                status=400,
            )

    # GET request sem ID - listar registros

    # Filtros
    cr_filter = request.GET.get("cr", "").strip()
    sistema_filter = request.GET.get("sistema", "").strip()
    servico_filter = request.GET.get("servico", "todos")
    status_filter = request.GET.get("status", "todos")

    # Ordenação
    ordenar = request.GET.get("ordenar", "-created_at")
    if ordenar.endswith("_desc"):
        ordenar = ordenar.replace("_desc", "")
    else:
        if not ordenar.startswith("-"):
            ordenar = "-" + ordenar

    # Paginação
    por_pagina = int(request.GET.get("por_pagina", 10))
    pagina = request.GET.get("pagina", 1)

    # Query base
    registros = ImplantacoesOpsVista.objects.all()

    # Aplicar filtros
    if cr_filter:
        # Buscar estruturas que contenham o filtro no campo CR (otimizado)
        cr_ids = Estrutura.objects.filter(
            cr__icontains=cr_filter
        ).values_list('id', flat=True)[:100]  # Limitar resultados
        registros = registros.filter(cr_id__in=list(cr_ids))
    if sistema_filter:
        registros = registros.filter(sistema__icontains=sistema_filter)
    if servico_filter != "todos":
        registros = registros.filter(servico=servico_filter)
    if status_filter != "todos":
        registros = registros.filter(status=status_filter)

    # Aplicar ordenação
    registros = registros.order_by(ordenar, "-created_at")

    # Aplicar paginação
    paginator = Paginator(registros, por_pagina)
    try:
        registros_paginados = paginator.page(pagina)
    except PageNotAnInteger:
        registros_paginados = paginator.page(1)
    except EmptyPage:
        registros_paginados = paginator.page(paginator.num_pages)

    # Para AJAX requests de busca de estruturas
    if request.GET.get('search_estruturas'):
        search_term = request.GET.get('search_estruturas', '').strip()
        estruturas_data = []
        if len(search_term) >= 2:  # Só buscar com 2+ caracteres
            try:
                # Otimização 1: Usar startswith em vez de icontains para melhor performance
                # Otimização 2: Limitar drasticamente o número de registros
                # Otimização 3: Usar distinct no banco quando possível
                estruturas_raw = Estrutura.objects.only('id', 'cr').filter(
                    cr__startswith=search_term,
                    cr__isnull=False
                ).exclude(cr='').order_by("cr")[:20]  # Reduzir para 20 registros
                
                # Filtrar CRs únicos manualmente (mais rápido com menos registros)
                crs_vistos = set()
                for e in estruturas_raw:
                    if e.cr not in crs_vistos and len(estruturas_data) < 10:  # Máximo 10 resultados
                        crs_vistos.add(e.cr)
                        estruturas_data.append({'id': e.id, 'cr': e.cr})
                
                # Se não encontrou com startswith, tentar icontains mas com limite ainda menor
                if not estruturas_data and len(search_term) >= 3:
                    estruturas_raw = Estrutura.objects.only('id', 'cr').filter(
                        cr__icontains=search_term,
                        cr__isnull=False
                    ).exclude(cr='').order_by("cr")[:10]  # Apenas 10 registros
                    
                    for e in estruturas_raw:
                        if e.cr not in crs_vistos and len(estruturas_data) < 5:  # Máximo 5 resultados
                            crs_vistos.add(e.cr)
                            estruturas_data.append({'id': e.id, 'cr': e.cr})
                            
            except Exception as e:
                # Em caso de erro, retornar lista vazia
                print(f"Erro na busca de estruturas: {e}")
                estruturas_data = []
        
        return JsonResponse({'estruturas': estruturas_data})
    
    # Buscar apenas algumas estruturas para inicialização (muito otimizado)
    # Reduzir drasticamente para melhorar performance inicial
    try:
        estruturas_raw = Estrutura.objects.only('id', 'cr').filter(
            cr__isnull=False
        ).exclude(cr='').order_by("cr")[:50]  # Reduzir de 500 para 50
        # Filtrar CRs únicos manualmente
        crs_vistos = set()
        estruturas = []
        for e in estruturas_raw:
            if e.cr not in crs_vistos and len(estruturas) < 20:  # Reduzir de 100 para 20
                crs_vistos.add(e.cr)
                estruturas.append(e)
    except Exception as e:
        print(f"Erro ao carregar estruturas iniciais: {e}")
        estruturas = []

    # --- Dashboard (Visão Geral) ---
    # Agregados sobre TODOS os registros (não sobre a página/filtros atuais),
    # para os KPIs e gráficos de andamento no topo da página.
    from django.db.models.functions import TruncMonth

    todos = ImplantacoesOpsVista.objects.all()
    dash_total = todos.count()
    dash_ativos = todos.filter(status='ativo').count()
    dash_desmobilizados = todos.filter(status='desmobilizado').count()
    # CRs em atendimento: só conta CRs com implantação ATIVA (mesma leitura
    # do KPI "Em Andamento", agrupada por CR)
    dash_crs = todos.filter(status='ativo').values('cr_id').distinct().count()

    servico_labels = dict(ImplantacoesOpsVista.SERVICO_CHOICES)
    por_servico = {value: {'ativo': 0, 'desmobilizado': 0} for value, _label in ImplantacoesOpsVista.SERVICO_CHOICES}
    for row in todos.values('servico', 'status').annotate(qtd=Count('id')):
        if row['servico'] in por_servico and row['status'] in ('ativo', 'desmobilizado'):
            por_servico[row['servico']][row['status']] = row['qtd']

    dashboards_labels = dict(ImplantacoesOpsVista.DASHBOARDS_CHOICES)
    por_dashboards = {value: 0 for value, _label in ImplantacoesOpsVista.DASHBOARDS_CHOICES}
    for row in todos.values('dashboards').annotate(qtd=Count('id')):
        if row['dashboards'] in por_dashboards:
            por_dashboards[row['dashboards']] = row['qtd']

    # Novas implantações por mês (últimos 6 meses, incluindo meses zerados)
    meses_pt = ['jan', 'fev', 'mar', 'abr', 'mai', 'jun', 'jul', 'ago', 'set', 'out', 'nov', 'dez']
    agora = timezone.now()
    meses_janela = []
    ano, mes = agora.year, agora.month
    for _ in range(6):
        meses_janela.append((ano, mes))
        mes -= 1
        if mes == 0:
            mes, ano = 12, ano - 1
    meses_janela.reverse()

    por_mes_raw = {}
    inicio_janela = timezone.datetime(meses_janela[0][0], meses_janela[0][1], 1, tzinfo=agora.tzinfo)
    for row in (todos.filter(created_at__gte=inicio_janela)
                .annotate(mes_ref=TruncMonth('created_at'))
                .values('mes_ref').annotate(qtd=Count('id'))):
        if row['mes_ref']:
            por_mes_raw[(row['mes_ref'].year, row['mes_ref'].month)] = row['qtd']

    dashboard_stats = {
        'kpis': {
            'total': dash_total,
            'ativos': dash_ativos,
            'desmobilizados': dash_desmobilizados,
            'crs': dash_crs,
        },
        'status': {
            'labels': ['Ativo', 'Desmobilizado'],
            'valores': [dash_ativos, dash_desmobilizados],
        },
        'servico': {
            'labels': [servico_labels[v] for v in por_servico],
            'ativos': [por_servico[v]['ativo'] for v in por_servico],
            'desmobilizados': [por_servico[v]['desmobilizado'] for v in por_servico],
        },
        'dashboards': {
            'labels': [dashboards_labels[v] for v in por_dashboards],
            'valores': [por_dashboards[v] for v in por_dashboards],
        },
        'mensal': {
            'labels': [f"{meses_pt[m - 1]}/{str(a)[2:]}" for a, m in meses_janela],
            'valores': [por_mes_raw.get((a, m), 0) for a, m in meses_janela],
        },
    }

    context = {
        'registros': registros_paginados,
        'status_choices': ImplantacoesOpsVista.STATUS_CHOICES,
        'servico_choices': ImplantacoesOpsVista.SERVICO_CHOICES,
        'dashboards_choices': ImplantacoesOpsVista.DASHBOARDS_CHOICES,
        'estruturas': estruturas,
        'dashboard_stats': dashboard_stats,
        'filtros': {
            'cr': cr_filter,
            'sistema': sistema_filter,
            'servico': servico_filter,
            'status': status_filter,
        },
        'ordenacao': request.GET.get('ordenar', 'created_at'),
        'por_pagina': por_pagina,
    }

    return render(request, 'implantacoes_opsvista.html', context)


@login_required
@check_page_permission("implantacoes_fluxo")
def implantacoes_fluxo(request):
    """View para o Kanban de Fluxo de Implantações (Planner de Implantações)"""
    from django.utils import timezone
    from .models import CardImplantacao
    from django.views.decorators.csrf import csrf_exempt
    import json

    # 1. Requisição GET com ID - Retornar dados do card para o modal
    if request.method == "GET" and request.GET.get("id"):
        try:
            card = get_object_or_404(CardImplantacao.objects.select_related('created_by'), id=request.GET.get("id"))
            
            # Calcular dias restantes ou de atraso para o BI
            dias_restantes = None
            if card.etapa_atual == 9 and card.bi_inicio_data:
                dias_decorridos = (timezone.now() - card.bi_inicio_data).days
                dias_restantes = 20 - dias_decorridos

            return JsonResponse({
                "success": True,
                "id": str(card.id),
                "nome": card.nome,
                "status": card.status,
                "etapa_atual": card.etapa_atual,
                "tipo_implantacao": card.tipo_implantacao,
                "mapeamento_locais": card.mapeamento_locais,
                "anexo_mapeamento_url": card.anexo_mapeamento.url if card.anexo_mapeamento else None,
                "anexo_mapeamento_nome": card.anexo_mapeamento.name.split('/')[-1] if card.anexo_mapeamento else None,
                "anexo_checklist_url": card.anexo_checklist.url if card.anexo_checklist else None,
                "anexo_checklist_nome": card.anexo_checklist.name.split('/')[-1] if card.anexo_checklist else None,
                "rotinas_criadas": card.rotinas_criadas,
                "anexo_rotinas_url": card.anexo_rotinas.url if card.anexo_rotinas else None,
                "anexo_rotinas_nome": card.anexo_rotinas.name.split('/')[-1] if card.anexo_rotinas else None,
                "anexo_qrcodes_url": card.anexo_qrcodes.url if card.anexo_qrcodes else None,
                "anexo_qrcodes_nome": card.anexo_qrcodes.name.split('/')[-1] if card.anexo_qrcodes else None,
                "anexo_treinamento_url": card.anexo_treinamento.url if card.anexo_treinamento else None,
                "anexo_treinamento_nome": card.anexo_treinamento.name.split('/')[-1] if card.anexo_treinamento else None,
                "anexo_entrega_url": card.anexo_entrega.url if card.anexo_entrega else None,
                "anexo_entrega_nome": card.anexo_entrega.name.split('/')[-1] if card.anexo_entrega else None,
                "link_bi": card.link_bi,
                "anexo_bi_url": card.anexo_bi.url if card.anexo_bi else None,
                "anexo_bi_nome": card.anexo_bi.name.split('/')[-1] if card.anexo_bi else None,
                "bi_inicio_data": card.bi_inicio_data.strftime("%d/%m/%Y %H:%M") if card.bi_inicio_data else None,
                "dias_restantes": dias_restantes,
                "created_at": card.created_at.strftime("%d/%m/%Y %H:%M"),
                "created_by_name": card.created_by.name or card.created_by.username if card.created_by else "Sistema",
                "step2_concluido_em": card.step2_concluido_em.strftime("%d/%m/%Y %H:%M") if card.step2_concluido_em else None,
                "step3_concluido_em": card.step3_concluido_em.strftime("%d/%m/%Y %H:%M") if card.step3_concluido_em else None,
                "step4_concluido_em": card.step4_concluido_em.strftime("%d/%m/%Y %H:%M") if card.step4_concluido_em else None,
                "step5_concluido_em": card.step5_concluido_em.strftime("%d/%m/%Y %H:%M") if card.step5_concluido_em else None,
                "step6_concluido_em": card.step6_concluido_em.strftime("%d/%m/%Y %H:%M") if card.step6_concluido_em else None,
                "step7_concluido_em": card.step7_concluido_em.strftime("%d/%m/%Y %H:%M") if card.step7_concluido_em else None,
                "step8_concluido_em": card.step8_concluido_em.strftime("%d/%m/%Y %H:%M") if card.step8_concluido_em else None,
                "step9_concluido_em": card.step9_concluido_em.strftime("%d/%m/%Y %H:%M") if card.step9_concluido_em else None,
            })
        except Exception as e:
            return JsonResponse({"success": False, "message": f"Erro ao buscar card: {str(e)}"}, status=400)

    # 2. Requisição POST - Criar novo card (Coordenador - Abertura)
    elif request.method == "POST" and request.POST.get("action") == "create":
        try:
            nome = request.POST.get("nome", "").strip()
            if not nome:
                return JsonResponse({"success": False, "message": "Nome do fluxo é obrigatório"}, status=400)
            
            card = CardImplantacao.objects.create(
                nome=nome,
                status="em_andamento",
                etapa_atual=1,
                created_by=request.user
            )
            try:
                enviar_email_implantacao(card, 1)
            except Exception as mail_err:
                print(f"Erro ao enviar email de implantacao na criacao: {mail_err}")
            return JsonResponse({
                "success": True,
                "message": "Fluxo de implantação aberto com sucesso!",
                "id": str(card.id),
                "nome": card.nome,
                "status": card.status,
                "etapa_atual": card.etapa_atual
            })
        except Exception as e:
            return JsonResponse({"success": False, "message": f"Erro ao criar fluxo: {str(e)}"}, status=400)

    # Requisição POST - Excluir card (Somente administradores)
    elif request.method == "POST" and request.POST.get("action") == "delete":
        if request.user.role != 'administrador':
            return JsonResponse({"success": False, "message": "Apenas administradores podem excluir fluxos."}, status=403)
        try:
            card_id = request.POST.get("id")
            card = get_object_or_404(CardImplantacao, id=card_id)
            card.delete()
            return JsonResponse({"success": True, "message": "Fluxo de implantação excluído com sucesso!"})
        except Exception as e:
            return JsonResponse({"success": False, "message": f"Erro ao excluir fluxo: {str(e)}"}, status=400)

    # 3. Requisição POST - Atualizar coluna (Mudar status Kanban)
    elif request.method == "POST" and request.POST.get("action") == "update_status":
        try:
            card_id = request.POST.get("id")
            novo_status = request.POST.get("status")
            if novo_status not in ['em_andamento', 'pausada', 'concluida']:
                return JsonResponse({"success": False, "message": "Status inválido"}, status=400)
            
            card = get_object_or_404(CardImplantacao, id=card_id)
            card.status = novo_status
            
            # Se foi movido manualmente para concluído e estava em etapa inferior, pode finalizar ou manter
            if novo_status == 'concluida' and card.etapa_atual < 10:
                card.etapa_atual = 10
                
            card.save()
            return JsonResponse({"success": True, "message": "Status do fluxo atualizado com sucesso!"})
        except Exception as e:
            return JsonResponse({"success": False, "message": f"Erro ao atualizar status: {str(e)}"}, status=400)

    # 4. Requisição POST - Avançar/Atualizar Etapa (Wizard)
    elif request.method == "POST" and request.POST.get("action") == "update_step":
        try:
            card_id = request.POST.get("id")
            card = get_object_or_404(CardImplantacao, id=card_id)
            etapa = card.etapa_atual
            
            if etapa == 1:
                # Transiciona para etapa 2 automaticamente na abertura
                card.etapa_atual = 2
                
            elif etapa == 2:
                # Definir serviço (Coordenador)
                tipo = request.POST.get("tipo_implantacao")
                if tipo not in ['seguranca', 'limpeza', 'engenharia']:
                    return JsonResponse({"success": False, "message": "Tipo de implantação inválido. Selecione Segurança, Limpeza ou Engenharia."}, status=400)
                card.tipo_implantacao = tipo
                card.step2_concluido_em = timezone.now()
                card.etapa_atual = 3 # vai para Supervisor
                
            elif etapa == 3:
                # Mapeamento de locais (Supervisor)
                locais = request.POST.get("mapeamento_locais", "").strip()
                if not locais:
                    return JsonResponse({"success": False, "message": "Mapeamento dos locais é obrigatório."}, status=400)
                
                if 'anexo_mapeamento' in request.FILES:
                    card.anexo_mapeamento = request.FILES['anexo_mapeamento']
                
                card.mapeamento_locais = locais
                card.step3_concluido_em = timezone.now()
                card.etapa_atual = 4 # vai para Projetos - Checklist
                
            elif etapa == 4:
                # Criação do Checklist (Projetos)
                if 'anexo_checklist' not in request.FILES and not card.anexo_checklist:
                    return JsonResponse({"success": False, "message": "O print de comprovação do checklist é obrigatório."}, status=400)
                
                if 'anexo_checklist' in request.FILES:
                    card.anexo_checklist = request.FILES['anexo_checklist']
                
                card.step4_concluido_em = timezone.now()
                card.etapa_atual = 5 # vai para Projetos - Rotinas
                
            elif etapa == 5:
                # Criação de Rotinas (Projetos)
                rotinas = request.POST.get("rotinas_criadas") == "true"
                if not rotinas:
                    return JsonResponse({"success": False, "message": "Você deve confirmar que as rotinas de RONDA e LIVRO foram criadas."}, status=400)
                
                if 'anexo_rotinas' not in request.FILES and not card.anexo_rotinas:
                    return JsonResponse({"success": False, "message": "O print de comprovação das rotinas é obrigatório."}, status=400)
                
                if 'anexo_rotinas' in request.FILES:
                    card.anexo_rotinas = request.FILES['anexo_rotinas']
                
                card.rotinas_criadas = True
                card.step5_concluido_em = timezone.now()
                card.etapa_atual = 6 # vai para QR codes
                
            elif etapa == 6:
                # Criação de QR codes (Projetos)
                if 'anexo_qrcodes' not in request.FILES and not card.anexo_qrcodes:
                    return JsonResponse({"success": False, "message": "O documento em PDF contendo os QR Codes é obrigatório."}, status=400)
                
                if 'anexo_qrcodes' in request.FILES:
                    card.anexo_qrcodes = request.FILES['anexo_qrcodes']
                
                card.step6_concluido_em = timezone.now()
                card.etapa_atual = 7 # vai para Treinamento
                
            elif etapa == 7:
                # Treinamento (Projetos)
                if 'anexo_treinamento' not in request.FILES and not card.anexo_treinamento:
                    return JsonResponse({"success": False, "message": "A foto/print de comprovação do treinamento realizado é obrigatória."}, status=400)
                
                if 'anexo_treinamento' in request.FILES:
                    card.anexo_treinamento = request.FILES['anexo_treinamento']
                
                card.step7_concluido_em = timezone.now()
                card.etapa_atual = 8 # vai para Entrega
                
            elif etapa == 8:
                # Entrega de Projeto (Projetos)
                if 'anexo_entrega' not in request.FILES and not card.anexo_entrega:
                    return JsonResponse({"success": False, "message": "O anexo geral de entrega de projetos é obrigatório."}, status=400)
                
                if 'anexo_entrega' in request.FILES:
                    card.anexo_entrega = request.FILES['anexo_entrega']
                
                card.step8_concluido_em = timezone.now()
                card.etapa_atual = 9 # vai para BI (Inicia prazo de 20 dias)
                card.bi_inicio_data = timezone.now()
                
            elif etapa == 9:
                # BI (BI)
                link = request.POST.get("link_bi", "").strip()
                if not link:
                    return JsonResponse({"success": False, "message": "O link do dashboard do Power BI é obrigatório."}, status=400)
                
                if 'anexo_bi' not in request.FILES and not card.anexo_bi:
                    return JsonResponse({"success": False, "message": "O print de comprovação do dashboard do Power BI é obrigatório."}, status=400)
                
                if 'anexo_bi' in request.FILES:
                    card.anexo_bi = request.FILES['anexo_bi']
                
                card.link_bi = link
                card.step9_concluido_em = timezone.now()
                card.etapa_atual = 10 # Concluído!
                card.status = "concluida" # Vai automaticamente para a coluna de concluídos
                
            card.save()
            try:
                enviar_email_implantacao(card, card.etapa_atual)
            except Exception as mail_err:
                print(f"Erro ao enviar email de implantacao no avanco: {mail_err}")
            return JsonResponse({
                "success": True,
                "message": f"Etapa {etapa} concluída com sucesso! O fluxo avançou.",
                "etapa_atual": card.etapa_atual,
                "status": card.status
            })
        except Exception as e:
            return JsonResponse({"success": False, "message": f"Erro ao salvar etapa: {str(e)}"}, status=400)

    # 5. Listagem de Cards (Página Inicial)
    cards = CardImplantacao.objects.all().select_related('created_by')
    
    # Calcular contadores
    total_cards = cards.count()
    em_andamento_count = cards.filter(status='em_andamento').count()
    pausada_count = cards.filter(status='pausada').count()
    concluida_count = cards.filter(status='concluida').count()
    
    # Calcular atrasos do BI (cards na etapa 9 há mais de 20 dias)
    atrasados_bi_count = 0
    for card in cards.filter(etapa_atual=9, bi_inicio_data__isnull=False):
        if (timezone.now() - card.bi_inicio_data).days > 20:
            atrasados_bi_count += 1
            
    # Mapeamento de descrições e responsáveis das etapas para o card Kanban
    etapas_info = {
        1: {"label": "Solicitação IMPLANTAÇÃO", "cargo": "Coordenador", "esperando": "Coordenadores!"},
        2: {"label": "Tipo de implantação", "cargo": "Coordenador", "esperando": "Coordenadores!"},
        3: {"label": "Mapeamento dos locais", "cargo": "Supervisor", "esperando": "Supervisor!"},
        4: {"label": "Criação do Checklist", "cargo": "Projetos", "esperando": "Projetos (Checklist)!"},
        5: {"label": "Criação das rotinas (RONDA, LIVRO)", "cargo": "Projetos", "esperando": "Projetos (Rotinas)!"},
        6: {"label": "Criação dos QR codes", "cargo": "Projetos", "esperando": "Projetos (QR Codes)!"},
        7: {"label": "Treinamento", "cargo": "Projetos", "esperando": "Projetos (Treinamento)!"},
        8: {"label": "Entrega do projeto", "cargo": "Projetos", "esperando": "Projetos (Entrega)!"},
        9: {"label": "Link e painel do BI", "cargo": "BI", "esperando": "BI!"},
        10: {"label": "Finalizado", "cargo": "Concluído", "esperando": "Finalizado"}
    }

    # Enriquecer os cards com dados dinâmicos de renderização
    cards_list = []
    for card in cards:
        info = etapas_info.get(card.etapa_atual, {"label": "Finalizado", "cargo": "Concluído", "esperando": "Finalizado"})
        
        # Calcular dias restantes do BI se estiver na etapa 9
        dias_restantes = None
        atrasado = False
        if card.etapa_atual == 9 and card.bi_inicio_data:
            decorrido = (timezone.now() - card.bi_inicio_data).days
            dias_restantes = 20 - decorrido
            if dias_restantes < 0:
                atrasado = True
                
        cards_list.append({
            "id": card.id,
            "nome": card.nome,
            "status": card.status,
            "etapa_atual": card.etapa_atual,
            "label_etapa": info["label"],
            "esperando_texto": info["esperando"],
            "tipo_implantacao": card.tipo_implantacao,
            "dias_restantes_bi": dias_restantes,
            "atrasado_bi": atrasado,
            "created_at": card.created_at,
            "created_by_name": card.created_by.name or card.created_by.username if card.created_by else "Sistema",
        })

    # Dividir nas raias
    cards_andamento = [c for c in cards_list if c["status"] == 'em_andamento']
    cards_pausada = [c for c in cards_list if c["status"] == 'pausada']
    cards_concluida = [c for c in cards_list if c["status"] == 'concluida']

    context = {
        "cards_andamento": cards_andamento,
        "cards_pausada": cards_pausada,
        "cards_concluida": cards_concluida,
        "total_cards": total_cards,
        "em_andamento_count": em_andamento_count,
        "pausada_count": pausada_count,
        "concluida_count": concluida_count,
        "atrasados_bi_count": atrasados_bi_count,
    }
    
    return render(request, "implantacoes_fluxo.html", context)


@login_required
def relatorios(request):
    """View para a página de relatórios de EPI e APR"""
    from .models_relatorios import RelatorioView
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    from django.http import JsonResponse, HttpResponse
    import requests
    import logging
    
    logger = logging.getLogger(__name__)
    
    # Verificar se é uma ação de download
    if request.GET.get('action') == 'download':
        tarefa_id = request.GET.get('id')
        if not tarefa_id:
            return JsonResponse({'error': 'ID da tarefa não fornecido'}, status=400)
        
        try:
            # Fazer proxy do download da API externa
            api_url = f"https://api.opsvista.example.com/api/relatorio/tarefa?tarefaId={tarefa_id}"
            
            # Fazer requisição para a API externa
            response = requests.get(api_url, timeout=30)
            response.raise_for_status()
            
            # Retornar o arquivo como resposta
            http_response = HttpResponse(
                response.content,
                content_type=response.headers.get('content-type', 'application/pdf')
            )
            
            # Definir nome do arquivo
            filename = f"relatorio_{tarefa_id}.pdf"
            http_response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            return http_response
            
        except requests.exceptions.RequestException as e:
            logger.error(f"Erro ao baixar relatório {tarefa_id}: {e}")
            return JsonResponse({
                'error': 'Erro ao baixar relatório',
                'message': 'Não foi possível conectar com o servidor de relatórios'
            }, status=500)
        except Exception as e:
            logger.error(f"Erro inesperado ao baixar relatório {tarefa_id}: {e}")
            return JsonResponse({
                'error': 'Erro interno',
                'message': 'Erro inesperado ao processar download'
            }, status=500)
    
    # Determinar tipo ativo (EPI ou APR)
    active_tab = request.GET.get('tipo', 'EPI')
    
    # Definir checklist ID baseado no tipo
    if active_tab == 'APR':
        checklist_id = RelatorioView.CHECKLIST_APR_VIAGEM_SEGURA
    else:
        checklist_id = RelatorioView.CHECKLIST_INSPECAO_EPI
    
    # Coletar filtros
    filtros = {
        'search': request.GET.get('search', '').strip(),
        'responsavel': request.GET.get('responsavel', 'all'),
        'cr': request.GET.get('cr', 'all'),
        'data_inicial': request.GET.get('data_inicial', ''),
        'data_final': request.GET.get('data_final', ''),
    }
    
    # Remover filtros vazios
    filtros = {k: v for k, v in filtros.items() if v and v != 'all'}
    
    try:
        # Buscar relatórios
        relatorios = RelatorioView.get_relatorios(checklist_id, filtros, request.user)
        
        # Buscar listas para filtros (sempre buscar para ambos os tipos)
        responsaveis_list_epi = RelatorioView.get_responsaveis_list(RelatorioView.CHECKLIST_INSPECAO_EPI, request.user)
        crs_list_epi = RelatorioView.get_crs_list(RelatorioView.CHECKLIST_INSPECAO_EPI, request.user)
        responsaveis_list_apr = RelatorioView.get_responsaveis_list(RelatorioView.CHECKLIST_APR_VIAGEM_SEGURA, request.user)
        crs_list_apr = RelatorioView.get_crs_list(RelatorioView.CHECKLIST_APR_VIAGEM_SEGURA, request.user)
        
        # Implementar paginação
        page = request.GET.get('page', 1)
        paginator = Paginator(relatorios, 15)  # 15 registros por página
        
        try:
            relatorios_paginados = paginator.page(page)
        except PageNotAnInteger:
            relatorios_paginados = paginator.page(1)
        except EmptyPage:
            relatorios_paginados = paginator.page(paginator.num_pages)
        
        # Separar relatórios por tipo para o template
        if active_tab == 'APR':
            relatorios_apr = relatorios_paginados
            relatorios_epi = []
            responsaveis_list = responsaveis_list_apr
            crs_list = crs_list_apr
        else:
            relatorios_epi = relatorios_paginados
            relatorios_apr = []
            responsaveis_list = responsaveis_list_epi
            crs_list = crs_list_epi
            
    except Exception as e:
        # Em caso de erro, mostrar listas vazias
        relatorios_epi = []
        relatorios_apr = []
        responsaveis_list = []
        crs_list = []
        
        # Log do erro
        logger.error(f"Erro ao buscar relatórios: {e}")
    
    context = {
        'active_tab': active_tab,
        'relatorios_epi': relatorios_epi,
        'relatorios_apr': relatorios_apr,
        'responsaveis_list': responsaveis_list,
        'crs_list': crs_list,
        'search_term': request.GET.get('search', ''),
        'filter_responsavel': request.GET.get('responsavel', 'all'),
        'filter_cr': request.GET.get('cr', 'all'),
        'filter_data_inicial': request.GET.get('data_inicial', ''),
        'filter_data_final': request.GET.get('data_final', ''),
    }
    
    return render(request, 'relatorios.html', context)


@csrf_exempt
@login_required
def livro_ata_status(request):
    """API para verificar status do QR Code do Livro Ata para um CR"""
    from .models import LivroAtaQRCode
    from django.http import JsonResponse
    
    cr = request.GET.get('cr', '').strip()
    if not cr:
        return JsonResponse({'error': 'CR não informado'}, status=400)
    
    try:
        # Primeiro, buscar a estrutura pelo CR para obter o ID
        from .models import Estrutura
        cr_estrutura = Estrutura.objects.using('default').filter(
            cr__startswith=cr,
            nivel_4__isnull=True
        ).exclude(status=4).first()

        if not cr_estrutura:
            return JsonResponse({'error': 'CR não encontrado'}, status=404)

        # Verificar se já existe QR Code para este CR usando o ID da estrutura
        livro_ata = LivroAtaQRCode.objects.using('default').filter(cr_id=str(cr_estrutura.id)).first()
        
        if livro_ata:
            # Forçar a atualização para o formato correto
            nova_url = f"{settings.SITE_URL}/livroata/qrcode={livro_ata.id}/"
            if livro_ata.qr_code_url != nova_url:
                livro_ata.qr_code_url = nova_url
                livro_ata.save(using='default')

            return JsonResponse({
                'exists': True,
                'qr_code_url': livro_ata.qr_code_url,
                'created_at': livro_ata.created_at.isoformat(),
                'cr_descricao': livro_ata.cr_descricao
            })
        else:
            return JsonResponse({
                'exists': False,
                'message': 'CR válido, QR Code pode ser criado'
            })
            
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Erro ao verificar status do Livro Ata: {e}")
        return JsonResponse({'error': 'Erro interno do servidor'}, status=500)


@csrf_exempt
@login_required
def generate_livro_ata_qr(request):
    """API para gerar QR Code do Livro Ata"""
    from .models import LivroAtaQRCode
    from django.http import JsonResponse
    import json
    import qrcode
    from io import BytesIO
    import base64
    from PIL import Image, ImageDraw, ImageFont
    import os
    
    if request.method != "POST":
        return JsonResponse({"error": "Método não permitido"}, status=405)

    try:
        data = json.loads(request.POST.get("data", "{}"))
        service = data.get("service")
        cr_number = data.get("cr_number")
        logo_size = int(data.get("logo_size", 80))
        service_logo_size = int(data.get("service_logo_size", 60))

        if not service:
            return JsonResponse({"error": "Serviço não selecionado"}, status=400)
        if not cr_number:
            return JsonResponse({"error": "Número do CR não informado"}, status=400)

        # Verificar se o CR existe na tabela estrutura (nivel_4 is null, status <> 4)
        try:
            cr_estrutura = Estrutura.objects.using('default').filter(
                cr__startswith=cr_number,
                nivel_4__isnull=True
            ).exclude(status=4).first()

            if not cr_estrutura:
                return JsonResponse({"error": "CR não encontrado"}, status=404)

        except Exception as e:
            return JsonResponse({"error": f"Erro ao buscar CR: {str(e)}"}, status=500)

        # Verificar se já existe QR Code para este CR
        # CORREÇÃO: Usar o ID da estrutura ao invés do CR
        livro_ata, created = LivroAtaQRCode.objects.using('default').get_or_create(
            cr_id=str(cr_estrutura.id),  # Usar o ID da estrutura
            defaults={
                'cr_descricao': cr_estrutura.descricao or cr_estrutura.cr
            }
        )
        
        # Forçar a atualização da URL para garantir que todos fiquem no formato correto
        nova_url = f"{settings.SITE_URL}/livroata/qrcode={livro_ata.id}/"
        if livro_ata.qr_code_url != nova_url:
            livro_ata.qr_code_url = nova_url
            livro_ata.save(using='default')

        # Gerar QR Code - REPLICANDO EXATAMENTE A LÓGICA DOS LOCAIS
        qr_data = livro_ata.qr_code_url
        
        # Criar QR Code
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(qr_data)
        qr.make(fit=True)

        # Criar imagem do QR code
        qr_image = qr.make_image(fill_color="black", back_color="white")
        
        # Criar imagem final com layout IGUAL AOS LOCAIS (300x400)
        final_img = Image.new("RGB", (300, 400), "white")
        draw = ImageDraw.Draw(final_img)

        # Adicionar borda IGUAL AOS LOCAIS
        draw.rectangle([(5, 5), (295, 395)], outline="black", width=2)

        # Redimensionar QR Code para o tamanho padrão dos locais
        qr_size = 140
        qr_image = qr_image.resize((qr_size, qr_size))
        
        # CORREÇÃO: Carregar logo fixa do OpsVista (IGUAL AOS LOCAIS)
        ops_vista_logo = None
        try:
            import os
            from django.conf import settings
            
            # Caminhos possíveis para a logo OpsVista
            possible_paths = [
                os.path.join("Gestao_a_Vista", "templates", "image", "visa.png"),
                os.path.join("Gestao_a_Vista", "templates", "image", "logo.png"),
                os.path.join(os.path.dirname(__file__), "templates", "image", "visa.png"),
                os.path.join(os.path.dirname(__file__), "templates", "image", "logo.png"),
            ]
            
            for logo_path in possible_paths:
                if os.path.exists(logo_path):
                    ops_vista_logo = Image.open(logo_path)
                    break
        except Exception as e:
            print(f"Erro ao carregar logo OpsVista: {e}")
            ops_vista_logo = None
        
        # Posicionar QR Code no centro (igual aos locais)
        qr_x = (300 - qr_size) // 2
        qr_y = 120  # Mesma posição dos locais
        final_img.paste(qr_image, (qr_x, qr_y))
        
        # CORREÇÃO: Adicionar logo fixa do OpsVista (IGUAL AOS LOCAIS)
        if ops_vista_logo:
            try:
                # Redimensionar logo para caber no topo
                logo_height = 80
                logo_width = 250
                app_logo_resized = ops_vista_logo.copy()
                
                try:
                    app_logo_resized.thumbnail((logo_width, logo_height), Image.Resampling.LANCZOS)
                except AttributeError:
                    app_logo_resized.thumbnail((logo_width, logo_height), Image.LANCZOS)
                
                # Centralizar horizontalmente e posicionar próxima ao QR Code
                logo_x = (300 - app_logo_resized.width) // 2
                logo_y = qr_y - logo_height - 5  # 5px de margem entre logo e QR Code
                
                if app_logo_resized.mode == "RGBA":
                    final_img.paste(app_logo_resized, (logo_x, logo_y), app_logo_resized)
                else:
                    final_img.paste(app_logo_resized, (logo_x, logo_y))
            except Exception as e:
                print(f"Erro ao adicionar logo OpsVista: {e}")

        # CORREÇÃO: Carregar logo do serviço do banco de dados (IGUAL AOS LOCAIS)
        try:
            from .models import LogoServico
            logo_servico = LogoServico.objects.filter(nome=service).first()
            if logo_servico and logo_servico.img_base64:
                # Extrair dados base64 (remover prefixo data:image/png;base64, se existir)
                img_data = logo_servico.img_base64
                if img_data.startswith("data:image/"):
                    img_data = img_data.split(",")[1]

                # Decodificar base64 e criar imagem
                import base64
                img_bytes = base64.b64decode(img_data)
                service_logo = Image.open(BytesIO(img_bytes))

                # Redimensionar conforme tamanho selecionado pelo usuário
                service_logo.thumbnail((service_logo_size, service_logo_size))

                # Posicionar no canto superior direito com margens confortáveis (IGUAL AOS LOCAIS)
                logo_x = max(300 - service_logo_size - 15, 10)  # Margem confortável de 15px da direita
                logo_y = 15  # Margem confortável de 15px do topo

                if service_logo.mode == "RGBA":
                    final_img.paste(service_logo, (logo_x, logo_y), service_logo)
                else:
                    final_img.paste(service_logo, (logo_x, logo_y))
            else:
                # Fallback para texto se não houver logo (posição no topo direito com margens)
                draw.text(
                    (260, 20),
                    service.upper(),
                    fill="black",
                    font=ImageFont.load_default(),
                    anchor="mm",
                )
        except Exception as e:
            print(f"Erro ao carregar logo do serviço: {e}")
            draw.text(
                (260, 20),
                service.upper(),
                fill="black",
                font=ImageFont.load_default(),
                anchor="mm",
            )

        # CORREÇÃO: Adicionar apenas texto "LIVRO DE OCORRÊNCIA" (posicionado igual aos locais)
        # Função para carregar fonte com suporte a Unicode
        def load_unicode_font_livro_ata(size):
            font_paths = [
                "C:/Windows/Fonts/arial.ttf",
                "C:/Windows/Fonts/calibri.ttf", 
                "C:/Windows/Fonts/tahoma.ttf",
                "C:/Windows/Fonts/verdana.ttf",
                "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
                "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
                "/System/Library/Fonts/Arial.ttf",
                "arial.ttf",
                "calibri.ttf",
                "tahoma.ttf"
            ]
            
            for font_path in font_paths:
                try:
                    font = ImageFont.truetype(font_path, size)
                    return font
                except (OSError, IOError):
                    continue
            
            return ImageFont.load_default()
        
        # Texto principal - apenas "LIVRO DE OCORRÊNCIA"
        text = "LIVRO DE OCORRÊNCIA"
        font_size = 16
        font = load_unicode_font_livro_ata(font_size)
        
        # Posicionar texto abaixo do QR code (igual aos locais)
        text_y = qr_y + qr_size + 15
        
        # Centralizar texto
        bbox = draw.textbbox((0, 0), text, font=font)
        text_width = bbox[2] - bbox[0]
        text_x = (300 - text_width) // 2
        
        draw.text((text_x, text_y), text, fill="black", font=font)

        # Processar logos se fornecidas
        client_logo = request.FILES.get('client_logo')
        if client_logo:
            try:
                logo_img = Image.open(client_logo)
                logo_img = logo_img.convert('RGBA')
                
                # Redimensionar logo do cliente
                logo_img.thumbnail((logo_size, logo_size), Image.Resampling.LANCZOS)
                
                # Posicionar logo do cliente no rodapé (IGUAL AOS LOCAIS)
                logo_x = (300 - logo_img.width) // 2
                logo_y = 400 - logo_img.height - 15
                
                # Criar máscara para transparência
                if logo_img.mode == 'RGBA':
                    final_img.paste(logo_img, (logo_x, logo_y), logo_img)
                else:
                    final_img.paste(logo_img, (logo_x, logo_y))
            except Exception as e:
                print(f"Erro ao processar logo do cliente: {e}")


        # Converter para base64
        buffer = BytesIO()
        final_img.save(buffer, format='PNG')
        img_base64 = base64.b64encode(buffer.getvalue()).decode()

        return JsonResponse({
            "success": True,
            "is_existing": not created,
            "livro_ata_id": str(livro_ata.id),
            "qr_code_url": livro_ata.qr_code_url,
            "qr_code": {
                "image": img_base64,
                "data": qr_data
            }
        })

    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Erro ao gerar QR Code do Livro Ata: {e}")
        return JsonResponse({"error": f"Erro interno: {str(e)}"}, status=500)


@login_required
@require_http_methods(["POST"])
def enviar_notificacao_teste_whatsapp(request):
    """
    Envia uma mensagem de teste via WhatsApp (uazapi) para todos os
    usuários com a notificação de Livro Ata habilitada no Controle de
    Usuários, usado para validar a integração de ponta a ponta.
    """
    if request.user.role != 'administrador':
        return JsonResponse({"error": "Apenas administradores podem enviar notificações de teste."}, status=403)

    from .uazapi_client import enviar_whatsapp

    destinatarios = list(
        CustomUser.objects.using('default')
        .filter(notificar_livro_ata=True)
        .exclude(whatsapp_notificacao='')
        .values_list('whatsapp_notificacao', flat=True)
    )

    if not destinatarios:
        numero_teste = getattr(settings, 'LIVRO_ATA_NOTIFICACAO_NUMERO_TESTE', None)
        if not numero_teste:
            return JsonResponse({"error": "Nenhum usuário com notificação de Livro Ata habilitada no Controle de Usuários."}, status=400)
        destinatarios = [numero_teste]

    mensagem = (
        "🔔 *Teste de Notificação - Livro Ata*\n\n"
        "Se você recebeu esta mensagem, a integração com o WhatsApp (uazapi) "
        "está funcionando corretamente no sistema Gestão à Vista."
    )

    falhas = []
    for numero in destinatarios:
        sucesso, detalhe = enviar_whatsapp(numero, mensagem)
        if not sucesso:
            falhas.append(f"{numero}: {detalhe}")

    if falhas and len(falhas) == len(destinatarios):
        return JsonResponse({"error": "Falha ao enviar para todos os destinatários: " + "; ".join(falhas)}, status=502)

    mensagem_resultado = f"Notificação de teste enviada para {len(destinatarios) - len(falhas)} de {len(destinatarios)} destinatário(s)."
    if falhas:
        mensagem_resultado += " Falhas: " + "; ".join(falhas)

    return JsonResponse({"success": True, "message": mensagem_resultado})


@login_required
@require_http_methods(["GET"])
def livro_ata_whatsapp_status(request):
    """
    Status da conexão WhatsApp (uazapi) usada pelo Livro Ata, para a aba de
    Configurações visível somente para administradores. Se a instância não
    estiver conectada, também aciona a reconexão (/instance/connect) para
    devolver um QR Code/código de pareamento atualizado.
    """
    if request.user.role != 'administrador' and not request.user.is_superuser:
        return JsonResponse({"error": "Apenas administradores podem ver esta configuração."}, status=403)

    from .uazapi_client import conectar_instancia, obter_status_instancia

    sucesso, dados = obter_status_instancia()
    if not sucesso:
        return JsonResponse({"error": dados}, status=502)

    instance = dados.get("instance") or {}
    status = instance.get("status", "disconnected")

    resultado = {
        "status": status,
        "profile_name": instance.get("profileName"),
    }

    if status != "connected":
        sucesso_connect, dados_connect = conectar_instancia()
        if sucesso_connect:
            instance_connect = dados_connect.get("instance") or {}
            resultado["status"] = instance_connect.get("status", status)
            resultado["qrcode"] = instance_connect.get("qrcode")
            resultado["paircode"] = instance_connect.get("paircode")
        else:
            resultado["error"] = dados_connect

    return JsonResponse(resultado)


# API Endpoints para Histórico de Mudanças do Projeto
@login_required
@require_http_methods(["GET"])
def get_project_change_history(request, project_id):
    """
    Retorna o histórico de mudanças de um projeto específico

    Query Parameters:
    - limit: número máximo de registros (padrão: 50)
    - offset: número de registros a pular (padrão: 0)
    - type: tipo de alteração para filtrar (opcional)
    """
    try:
        # Verificar se o projeto existe
        projeto = PlannerProject.objects.get(id=project_id)

        # Verificar permissão (usuário criador ou responsável)
        is_responsavel = projeto.responsaveis.filter(id=request.user.id).exists()
        if projeto.created_by != request.user and not is_responsavel:
            return JsonResponse(
                {"error": "Você não tem permissão para visualizar este histórico"},
                status=403
            )

        # Parâmetros de paginação
        limit = int(request.GET.get('limit', 50))
        offset = int(request.GET.get('offset', 0))
        change_type = request.GET.get('type', '')

        # Validar limites
        limit = min(limit, 100)  # Máximo 100
        offset = max(offset, 0)

        # Query base
        history_query = PlannerProjectChangeHistory.objects.filter(
            projeto_id=project_id
        ).select_related('usuario')

        # Filtrar por tipo se fornecido
        if change_type:
            history_query = history_query.filter(tipo_alteracao=change_type)

        # Contar total
        total = history_query.count()

        # Aplicar paginação
        history = history_query[offset:offset + limit]

        # Serializar dados
        historia_data = []
        for h in history:
            historia_data.append({
                'id': str(h.id),
                'tipo_alteracao': h.get_tipo_alteracao_display(),
                'tipo_alteracao_key': h.tipo_alteracao,
                'campo': h.campo,
                'valor_anterior': h.valor_anterior,
                'valor_novo': h.valor_novo,
                'descricao': h.descricao,
                'created_at': h.created_at.isoformat(),
                'usuario': {
                    'id': str(h.usuario.id) if h.usuario else None,
                    'nome': h.usuario.name if h.usuario else 'Sistema',
                    'email': h.usuario.email if h.usuario else None,
                },
                'ip_address': h.ip_address,
            })

        return JsonResponse({
            'success': True,
            'total': total,
            'limit': limit,
            'offset': offset,
            'count': len(historia_data),
            'history': historia_data,
        })

    except PlannerProject.DoesNotExist:
        return JsonResponse(
            {"error": "Projeto não encontrado"},
            status=404
        )
    except ValueError:
        return JsonResponse(
            {"error": "Parâmetros de paginação inválidos"},
            status=400
        )
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Erro ao recuperar histórico do projeto: {e}")
        return JsonResponse(
            {"error": f"Erro ao recuperar histórico: {str(e)}"},
            status=500
        )


@login_required
@require_http_methods(["GET"])
def get_project_history_summary(request, project_id):
    """
    Retorna um resumo estatístico do histórico de mudanças
    """
    try:
        # Verificar se o projeto existe
        projeto = PlannerProject.objects.get(id=project_id)

        # Verificar permissão
        is_responsavel = projeto.responsaveis.filter(id=request.user.id).exists()
        if projeto.created_by != request.user and not is_responsavel:
            return JsonResponse(
                {"error": "Você não tem permissão para visualizar este resumo"},
                status=403
            )

        # Buscar histórico
        history = PlannerProjectChangeHistory.objects.filter(projeto_id=project_id)

        # Contar por tipo
        from django.db.models import Count
        by_type = history.values('tipo_alteracao').annotate(count=Count('id'))
        by_user = history.values('usuario__name').annotate(count=Count('id'))

        # Última mudança
        last_change = history.order_by('-created_at').first()

        # Preparar resumo
        summary = {
            'total_changes': history.count(),
            'first_change': history.order_by('created_at').first().created_at.isoformat() if history.exists() else None,
            'last_change': last_change.created_at.isoformat() if last_change else None,
            'changes_by_type': [
                {
                    'type': item['tipo_alteracao'],
                    'type_display': dict(PlannerProjectChangeHistory.CHANGE_TYPE_CHOICES).get(item['tipo_alteracao']),
                    'count': item['count']
                }
                for item in by_type
            ],
            'changes_by_user': [
                {
                    'user': item['usuario__name'] or 'Sistema',
                    'count': item['count']
                }
                for item in by_user
            ]
        }

        return JsonResponse({
            'success': True,
            'summary': summary,
        })

    except PlannerProject.DoesNotExist:
        return JsonResponse(
            {"error": "Projeto não encontrado"},
            status=404
        )
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Erro ao gerar resumo do histórico: {e}")
        return JsonResponse(
            {"error": f"Erro ao gerar resumo: {str(e)}"},
            status=500
        )


# ==================== GESTÃO DA QUALIDADE ====================

class GestaoQualidadeView(LoginRequiredMixin, TemplateView):
    """
    View principal da Gestão da Qualidade
    """
    template_name = 'gestao_qualidade.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from datetime import date
        from django.db.models import Q

        # Isolamento de dados: mostrar apenas dados do usuário logado
        user = self.request.user

        # Estatísticas gerais - cache por 60s para evitar 5 queries COUNT via VPN
        from django.core.cache import cache
        cache_key = f'qualidade_counts_{user.id}'
        counts = cache.get(cache_key)

        if counts is None:
            counts = {
                'total_treinamentos': Treinamento.objects.filter(created_by=user).count(),
                'total_visitas': VisitaTecnica.objects.filter(created_by=user).count(),
                'total_nao_conformidades': NaoConformidade.objects.filter(created_by=user).count(),
                'total_planos_acao': PlanoAcao.objects.filter(created_by=user).count(),
                'total_planos_acao_pendentes': PlanoAcao.objects.filter(
                    created_by=user, 
                    status__in=['pendente', 'em_andamento']
                ).count(),
            }
            cache.set(cache_key, counts, 60)

        context.update(counts)

        # Treinamentos recentes
        context['treinamentos_recentes'] = Treinamento.objects.filter(
            created_by=user
        ).select_related('created_by').order_by('-data')[:5]

        # Visitas recentes - Correção do erro de GROUP BY substituindo annotate por prefetch_related + len()
        visitas = list(VisitaTecnica.objects.filter(
            created_by=user
        ).select_related('created_by').prefetch_related('evidencias').order_by('-data')[:5])
        
        for v in visitas:
            v.evidencias_count = len(v.evidencias.all())
        context['visitas_recentes'] = visitas

        # Não conformidades pendentes
        context['nao_conformidades_pendentes'] = NaoConformidade.objects.filter(
            created_by=user,
            status__in=['pendente', 'em_andamento']
        ).select_related('visita_tecnica').order_by('prazo')[:5]

        # Planos de ação recentes - Corrigido o campo de ordenação de '-data_inicio' para '-created_at'
        planos = list(PlanoAcao.objects.filter(
            created_by=user
        ).select_related('nao_conformidade').prefetch_related('evidencias').order_by('-created_at')[:5])
        
        for p in planos:
            p.evidencias_count = len(p.evidencias.all())
        context['planos_acao'] = planos

        return context


@login_required
@require_http_methods(["POST"])
def criar_treinamento(request):
    """Cria um novo treinamento"""
    try:
        data = json.loads(request.body)

        treinamento = Treinamento.objects.create(
            data=data.get('data'),
            tema=data.get('tema'),
            local=data.get('local'),
            responsavel=data.get('responsavel'),
            status=data.get('status', 'pendente'),
            observacoes=data.get('observacoes', ''),
            created_by=request.user
        )

        return JsonResponse({
            'success': True,
            'message': 'Treinamento criado com sucesso!',
            'id': str(treinamento.id)
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(["POST"])
def criar_visita_tecnica(request):
    """Cria uma nova visita técnica com upload de evidências"""
    import logging
    logger = logging.getLogger(__name__)

    try:
        # Log dos dados recebidos para debug
        logger.info(f'=== CRIAR VISITA TÉCNICA DEBUG ===')
        logger.info(f'User: {request.user.username if request.user.is_authenticated else "Anonymous"}')
        logger.info(f'POST data keys: {list(request.POST.keys())}')
        logger.info(f'FILES data keys: {list(request.FILES.keys())}')
        logger.info(f'Content-Type: {request.content_type}')

        # Processar dados JSON do formulário
        data_json = request.POST.get('data')
        if not data_json:
            logger.error('Campo "data" não encontrado no POST')
            return JsonResponse({
                'success': False,
                'error': 'Campo "data" é obrigatório'
            }, status=400)

        try:
            data = json.loads(data_json)
            logger.info(f'Dados JSON parseados: {data}')
        except json.JSONDecodeError as e:
            logger.error(f'Erro ao decodificar JSON: {str(e)}')
            return JsonResponse({
                'success': False,
                'error': f'Erro ao processar dados JSON: {str(e)}'
            }, status=400)

        # Obter resultado_conformidade do campo separado ou do JSON
        resultado_conformidade = request.POST.get('resultado') or data.get('resultado_conformidade')
        logger.info(f'Resultado conformidade: {resultado_conformidade}')

        # Validar resultado_conformidade se fornecido
        if resultado_conformidade:
            resultados_validos = ['conforme', 'nao_conforme']
            if resultado_conformidade not in resultados_validos:
                logger.error(f'Resultado de conformidade inválido: {resultado_conformidade}')
                return JsonResponse({
                    'success': False,
                    'error': f'Resultado de conformidade inválido: "{resultado_conformidade}". Valores permitidos: {", ".join(resultados_validos)}'
                }, status=400)

        # Validar campos obrigatórios
        campos_obrigatorios = ['data', 'tipo', 'local', 'responsavel']
        campos_faltantes = []

        for campo in campos_obrigatorios:
            valor = data.get(campo)
            if not valor or (isinstance(valor, str) and not valor.strip()):
                campos_faltantes.append(campo)

        if campos_faltantes:
            logger.error(f'Campos obrigatórios faltando ou vazios: {campos_faltantes}')
            return JsonResponse({
                'success': False,
                'error': f'Campos obrigatórios faltando ou vazios: {", ".join(campos_faltantes)}'
            }, status=400)

        # Validar tipo
        tipos_validos = ['inspecao', 'auditoria', 'visita_tecnica']
        tipo = data.get('tipo')
        if tipo not in tipos_validos:
            logger.error(f'Tipo inválido: {tipo}')
            return JsonResponse({
                'success': False,
                'error': f'Tipo inválido: "{tipo}". Valores permitidos: {", ".join(tipos_validos)}'
            }, status=400)

        # Converter data de string para objeto date
        from datetime import datetime
        data_str = data.get('data')

        if not data_str or not isinstance(data_str, str):
            logger.error(f'Campo data inválido ou vazio: {data_str}')
            return JsonResponse({
                'success': False,
                'error': 'Campo "data" é obrigatório e deve ser uma string válida'
            }, status=400)

        try:
            # Tentar formato YYYY-MM-DD primeiro (padrão do input date HTML5)
            if '-' in data_str:
                data_obj = datetime.strptime(data_str, '%Y-%m-%d').date()
            # Tentar formato DD/MM/YYYY
            elif '/' in data_str:
                data_obj = datetime.strptime(data_str, '%d/%m/%Y').date()
            else:
                raise ValueError('Formato de data não contém separadores válidos (- ou /)')
        except (ValueError, TypeError) as e:
            logger.error(f'Erro ao converter data "{data_str}": {str(e)}')
            return JsonResponse({
                'success': False,
                'error': f'Formato de data inválido: "{data_str}". Use YYYY-MM-DD ou DD/MM/YYYY'
            }, status=400)

        # Validar status se fornecido
        status = data.get('status', 'pendente')
        status_validos = ['pendente', 'em_andamento', 'concluido']
        if status not in status_validos:
            logger.error(f'Status inválido: {status}')
            return JsonResponse({
                'success': False,
                'error': f'Status inválido: "{status}". Valores permitidos: {", ".join(status_validos)}'
            }, status=400)

        # Validar checklist se fornecido
        checklist = data.get('checklist', {})
        if checklist is not None and not isinstance(checklist, dict):
            logger.error(f'Checklist deve ser um objeto JSON/dicionário, recebido: {type(checklist)}')
            return JsonResponse({
                'success': False,
                'error': 'Campo "checklist" deve ser um objeto JSON válido'
            }, status=400)

        # Validar tamanho dos campos
        local = data.get('local').strip()
        responsavel = data.get('responsavel').strip()

        if len(local) > 255:
            logger.error(f'Campo "local" excede tamanho máximo: {len(local)} caracteres')
            return JsonResponse({
                'success': False,
                'error': f'Campo "local" excede o tamanho máximo de 255 caracteres (atual: {len(local)})'
            }, status=400)

        if len(responsavel) > 255:
            logger.error(f'Campo "responsavel" excede tamanho máximo: {len(responsavel)} caracteres')
            return JsonResponse({
                'success': False,
                'error': f'Campo "responsavel" excede o tamanho máximo de 255 caracteres (atual: {len(responsavel)})'
            }, status=400)

        # Criar visita técnica
        try:
            visita = VisitaTecnica.objects.create(
                data=data_obj,
                tipo=data.get('tipo'),
                local=local,
                responsavel=responsavel,
                checklist=checklist,
                observacoes=data.get('observacoes', '').strip() if data.get('observacoes') else '',
                status=status,
                resultado_conformidade=resultado_conformidade,
                created_by=request.user
            )
            logger.info(f'Visita técnica criada com ID: {visita.id}')
        except Exception as e:
            logger.error(f'Erro ao criar visita técnica no banco: {str(e)}')
            return JsonResponse({
                'success': False,
                'error': f'Erro ao salvar visita técnica: {str(e)}'
            }, status=400)

        # Se resultado é não conforme, criar automaticamente uma NaoConformidade
        nao_conformidade_criada = None
        if resultado_conformidade == 'nao_conforme':
            from datetime import timedelta
            data_visita = visita.data

            # Obter campos opcionais para NC (prazo e criticidade customizados)
            prazo_nc = request.POST.get('prazo_nc') or data.get('prazo_nc')
            criticidade_nc = request.POST.get('criticidade_nc') or data.get('criticidade_nc')

            # Se prazo_nc foi fornecido, converter para objeto date
            if prazo_nc:
                try:
                    if '-' in prazo_nc:
                        prazo = datetime.strptime(prazo_nc, '%Y-%m-%d').date()
                    elif '/' in prazo_nc:
                        prazo = datetime.strptime(prazo_nc, '%d/%m/%Y').date()
                    else:
                        prazo = data_visita + timedelta(days=30)
                except Exception as e:
                    logger.warning(f'Erro ao converter prazo_nc: {str(e)}. Usando prazo padrão.')
                    prazo = data_visita + timedelta(days=30)
            else:
                # Prazo padrão: data da visita + 30 dias
                prazo = data_visita + timedelta(days=30)

            # Validar criticidade ou usar padrão
            classificacoes_validas = ['baixa', 'media', 'alta', 'critica']
            if criticidade_nc and criticidade_nc in classificacoes_validas:
                classificacao = criticidade_nc
            else:
                classificacao = 'media'

            descricao_nc = data.get('observacoes', '') or f'Não conformidade identificada em {visita.get_tipo_display()} no local {visita.local}'

            nao_conformidade = NaoConformidade.objects.create(
                visita_tecnica=visita,
                descricao=descricao_nc,
                data_identificacao=data_visita,
                classificacao=classificacao,
                responsavel=visita.responsavel,
                prazo=prazo,
                status='pendente',
                created_by=request.user
            )
            nao_conformidade_criada = {
                'id': str(nao_conformidade.id),
                'descricao': descricao_nc,
                'prazo': prazo.strftime('%Y-%m-%d'),
                'criticidade': classificacao
            }
            logger.info(f'Não conformidade criada automaticamente: {nao_conformidade.id} - Prazo: {prazo} - Criticidade: {classificacao}')

        # Processar upload de evidências (múltiplos arquivos numerados)
        evidencias_criadas = []

        # Primeiro tentar pegar arquivos no formato evidencia_0, evidencia_1, etc
        import os
        i = 0
        while True:
            arquivo_key = f'evidencia_{i}'
            descricao_key = f'evidencia_{i}_descricao'

            if arquivo_key not in request.FILES:
                break

            arquivo = request.FILES[arquivo_key]
            descricao = request.POST.get(descricao_key, f'Evidência {i+1} da visita técnica')

            logger.info(f'Processando evidência {i}: {arquivo.name} ({arquivo.size} bytes)')

            # Validar tipo de arquivo
            extensao = os.path.splitext(arquivo.name)[1].lower()
            extensoes_permitidas = ['.jpg', '.jpeg', '.png', '.gif', '.pdf']

            if extensao not in extensoes_permitidas:
                # Se houver arquivo inválido, deletar a visita criada e retornar erro
                visita.delete()
                logger.error(f'Tipo de arquivo não permitido: {arquivo.name}')
                return JsonResponse({
                    'success': False,
                    'error': f'Tipo de arquivo não permitido: {arquivo.name}. Apenas imagens (jpg, jpeg, png, gif) e PDF são aceitos.'
                }, status=400)

            # Validar tamanho do arquivo (máximo 5MB)
            if arquivo.size > 5 * 1024 * 1024:  # 5MB em bytes
                visita.delete()
                logger.error(f'Arquivo muito grande: {arquivo.name} ({arquivo.size} bytes)')
                return JsonResponse({
                    'success': False,
                    'error': f'Arquivo muito grande: {arquivo.name}. Tamanho máximo: 5MB.'
                }, status=400)

            # Criar evidência
            evidencia = EvidenciaQualidade.objects.create(
                tipo='visita',
                descricao=descricao,
                arquivo=arquivo,
                visita_tecnica=visita,
                uploaded_by=request.user
            )
            evidencias_criadas.append({
                'id': str(evidencia.id),
                'nome': arquivo.name,
                'descricao': descricao,
                'tamanho': arquivo.size
            })
            logger.info(f'Evidência criada: {evidencia.id}')

            i += 1

        # Se não encontrou no formato numerado, tentar formato antigo (evidencias)
        if not evidencias_criadas:
            arquivos = request.FILES.getlist('evidencias')
            logger.info(f'Tentando formato antigo: {len(arquivos)} arquivos')

            for arquivo in arquivos:
                # Validar tipo de arquivo
                extensao = os.path.splitext(arquivo.name)[1].lower()
                extensoes_permitidas = ['.jpg', '.jpeg', '.png', '.gif', '.pdf']

                if extensao not in extensoes_permitidas:
                    visita.delete()
                    return JsonResponse({
                        'success': False,
                        'error': f'Tipo de arquivo não permitido: {arquivo.name}. Apenas imagens (jpg, jpeg, png, gif) e PDF são aceitos.'
                    }, status=400)

                if arquivo.size > 5 * 1024 * 1024:
                    visita.delete()
                    return JsonResponse({
                        'success': False,
                        'error': f'Arquivo muito grande: {arquivo.name}. Tamanho máximo: 5MB.'
                    }, status=400)

                evidencia = EvidenciaQualidade.objects.create(
                    tipo='visita',
                    descricao=f'Evidência da visita técnica - {arquivo.name}',
                    arquivo=arquivo,
                    visita_tecnica=visita,
                    uploaded_by=request.user
                )
                evidencias_criadas.append({
                    'id': str(evidencia.id),
                    'nome': arquivo.name,
                    'tamanho': arquivo.size
                })

        logger.info(f'Total de evidências criadas: {len(evidencias_criadas)}')

        response_data = {
            'success': True,
            'message': 'Visita técnica criada com sucesso!',
            'id': str(visita.id),
            'evidencias_count': len(evidencias_criadas),
            'evidencias': evidencias_criadas
        }

        if nao_conformidade_criada:
            response_data['nao_conformidade_criada'] = nao_conformidade_criada
            response_data['message'] = 'Visita técnica criada com sucesso! Uma não conformidade foi criada automaticamente.'

        logger.info(f'=== VISITA TÉCNICA CRIADA COM SUCESSO: {visita.id} ===')
        return JsonResponse(response_data)

    except ValueError as e:
        logger.error(f'Erro de validação ao criar visita técnica: {str(e)}')
        return JsonResponse({
            'success': False,
            'error': f'Erro de validação: {str(e)}'
        }, status=400)
    except KeyError as e:
        logger.error(f'Campo obrigatório ausente: {str(e)}')
        return JsonResponse({
            'success': False,
            'error': f'Campo obrigatório ausente: {str(e)}'
        }, status=400)
    except Exception as e:
        logger.exception(f'Erro inesperado ao criar visita técnica: {str(e)}')
        import traceback
        return JsonResponse({
            'success': False,
            'error': f'Erro ao criar visita técnica: {str(e)}',
            'details': traceback.format_exc() if logger.level <= 10 else None  # Incluir traceback apenas em DEBUG
        }, status=500)


@login_required
@require_http_methods(["POST"])
def criar_nao_conformidade(request):
    """Cria uma nova não conformidade"""
    try:
        data = json.loads(request.body)

        nao_conformidade = NaoConformidade.objects.create(
            data_identificacao=data.get('data_identificacao'),
            descricao=data.get('descricao'),
            referencia_normativa=data.get('referencia_normativa', ''),
            classificacao=data.get('classificacao'),
            responsavel=data.get('responsavel'),
            prazo=data.get('prazo'),
            status=data.get('status', 'pendente'),
            visita_tecnica_id=data.get('visita_tecnica_id') if data.get('visita_tecnica_id') else None,
            created_by=request.user
        )

        return JsonResponse({
            'success': True,
            'message': 'Não conformidade criada com sucesso!',
            'id': str(nao_conformidade.id)
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(["POST"])
def criar_plano_acao(request):
    """Cria um novo plano de ação associado a uma não conformidade"""
    try:
        # Processar dados - aceitar FormData ou JSON
        if request.content_type and 'multipart/form-data' in request.content_type:
            # FormData com possível upload de arquivos
            data_json = request.POST.get('data')
            if data_json:
                data = json.loads(data_json)
            else:
                # Extrair dados do POST
                data = {}
                for key in request.POST.keys():
                    if key not in ['csrfmiddlewaretoken', 'data']:
                        data[key] = request.POST.get(key)
        else:
            # JSON puro
            data = json.loads(request.body) if request.body else {}

        logger.info(f'[CRIAR-PLANO-ACAO] Dados recebidos: {data}')
        logger.info(f'[CRIAR-PLANO-ACAO] Arquivos recebidos: {request.FILES.keys()}')

        # Validar que a não conformidade existe e pertence ao usuário
        nao_conformidade_id = data.get('nao_conformidade_id')
        if not nao_conformidade_id:
            return JsonResponse({
                'success': False,
                'error': 'O ID da não conformidade é obrigatório.'
            }, status=400)

        # Verificar se a não conformidade existe e foi criada pelo usuário (isolamento de dados)
        try:
            nao_conformidade = NaoConformidade.objects.get(
                id=nao_conformidade_id,
                created_by=request.user
            )
        except NaoConformidade.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Não conformidade não encontrada ou você não tem permissão para acessá-la.'
            }, status=404)

        # Criar plano de ação
        plano = PlanoAcao.objects.create(
            nao_conformidade=nao_conformidade,
            tipo_acao=data.get('tipo_acao', 'corretiva'),
            descricao=data.get('descricao'),
            responsavel=data.get('responsavel'),
            prazo=data.get('prazo'),
            status=data.get('status', 'pendente'),
            observacoes=data.get('observacoes', ''),
            created_by=request.user
        )

        logger.info(f'[CRIAR-PLANO-ACAO] Plano de ação criado: {plano.id}')

        # Processar upload de evidências se houver
        evidencias_criadas = 0
        if request.FILES:
            for file_key in request.FILES.keys():
                if file_key.startswith('evidencia_'):
                    arquivo = request.FILES[file_key]
                    # Obter descrição correspondente
                    descricao_key = file_key.replace('evidencia_', 'descricao_evidencia_')
                    descricao = request.POST.get(descricao_key, f'Evidência - {arquivo.name}')

                    # Validar tamanho do arquivo (max 5MB)
                    if arquivo.size > 5 * 1024 * 1024:
                        logger.warning(f'Arquivo {arquivo.name} muito grande: {arquivo.size} bytes')
                        continue

                    # Criar evidência
                    evidencia = EvidenciaQualidade.objects.create(
                        plano_acao=plano,
                        tipo='plano_acao',
                        descricao=descricao,
                        arquivo=arquivo,
                        uploaded_by=request.user
                    )
                    evidencias_criadas += 1
                    logger.info(f'[CRIAR-PLANO-ACAO] Evidência criada: {evidencia.id}')

        return JsonResponse({
            'success': True,
            'message': f'Plano de ação criado com sucesso! {evidencias_criadas} evidência(s) anexada(s).',
            'id': str(plano.id)
        })

    except Exception as e:
        logger.error(f'[CRIAR-PLANO-ACAO] Erro: {str(e)}')
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(['GET'])
def listar_evidencias_visita(request, visita_id):
    """Lista todas as evidências de uma visita técnica específica"""
    logger.info(f'Listando evidências para visita: {visita_id}')
    logger.info(f'Usuário requisitante: {request.user.username}')

    try:
        # Verificar se a visita técnica existe e foi criada pelo usuário
        visita = get_object_or_404(VisitaTecnica, id=visita_id, created_by=request.user)
        logger.info(f'Visita encontrada: {visita.local} - ID: {visita.id}')

        # Buscar evidências relacionadas
        evidencias = EvidenciaQualidade.objects.filter(
            visita_tecnica=visita
        ).select_related('uploaded_by').order_by('-created_at')

        logger.info(f'Evidências encontradas: {evidencias.count()}')

        # Serializar dados
        evidencias_list = []
        for evidencia in evidencias:
            evidencia_data = {
                'id': str(evidencia.id),
                'tipo': evidencia.tipo,
                'descricao': evidencia.descricao,
                'arquivo_url': evidencia.arquivo.url if evidencia.arquivo else None,
                'arquivo_nome': evidencia.arquivo.name.split('/')[-1] if evidencia.arquivo else None,
                'uploaded_by': evidencia.uploaded_by.username if evidencia.uploaded_by else 'Sistema',
                'created_at': evidencia.created_at.strftime('%d/%m/%Y %H:%M')
            }
            evidencias_list.append(evidencia_data)
            logger.debug(f'Evidência processada: ID={evidencia.id}, Tipo={evidencia.tipo}')

        response_data = {
            'success': True,
            'visita_id': str(visita.id),
            'visita_local': visita.local,
            'evidencias': evidencias_list,
            'total': len(evidencias_list)
        }

        logger.info(f'Retornando {len(evidencias_list)} evidências para visita {visita_id}')
        return JsonResponse(response_data)

    except VisitaTecnica.DoesNotExist:
        logger.warning(f'Visita técnica não encontrada ou sem permissão: {visita_id} - Usuário: {request.user.username}')
        return JsonResponse({
            'success': False,
            'error': 'Visita técnica não encontrada ou você não tem permissão para acessá-la.'
        }, status=404)
    except Exception as e:
        logger.error(f'Erro ao listar evidências da visita {visita_id}: {str(e)}', exc_info=True)
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(['GET'])
def visualizar_evidencia(request, evidencia_id):
    """Visualiza ou faz download de uma evidência específica"""
    try:
        # Buscar evidência e verificar permissão
        evidencia = get_object_or_404(
            EvidenciaQualidade,
            id=evidencia_id,
            uploaded_by=request.user
        )

        # Verificar se o arquivo existe
        if not evidencia.arquivo:
            return JsonResponse({
                'success': False,
                'error': 'Arquivo não encontrado.'
            }, status=404)

        # Obter caminho do arquivo
        arquivo_path = evidencia.arquivo.path

        # Verificar se arquivo existe no sistema de arquivos
        import os
        if not os.path.exists(arquivo_path):
            return JsonResponse({
                'success': False,
                'error': 'Arquivo não encontrado no sistema.'
            }, status=404)

        # Determinar tipo MIME
        import mimetypes
        content_type, _ = mimetypes.guess_type(arquivo_path)
        if not content_type:
            content_type = 'application/octet-stream'

        # Abrir e retornar arquivo
        with open(arquivo_path, 'rb') as arquivo:
            response = HttpResponse(arquivo.read(), content_type=content_type)

            # Verificar se é para download ou visualização
            modo = request.GET.get('modo', 'visualizar')
            if modo == 'download':
                nome_arquivo = evidencia.arquivo.name.split('/')[-1]
                response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
            else:
                # Para imagens, permitir visualização inline
                if content_type.startswith('image/'):
                    response['Content-Disposition'] = 'inline'
                else:
                    nome_arquivo = evidencia.arquivo.name.split('/')[-1]
                    response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'

            return response

    except EvidenciaQualidade.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Evidência não encontrada ou você não tem permissão para acessá-la.'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)
# ================== CRUD ENDPOINTS - GESTÃO DA QUALIDADE ==================

# ========== TREINAMENTO ==========

@login_required
@require_http_methods(['GET'])
def obter_treinamento(request, treinamento_id):
    '''Obtém detalhes de um treinamento específico'''
    import logging
    logger = logging.getLogger(__name__)

    try:
        # Verificar se o treinamento existe e foi criado pelo usuário
        treinamento = get_object_or_404(Treinamento, id=treinamento_id, created_by=request.user)

        # Buscar evidências relacionadas
        evidencias = EvidenciaQualidade.objects.filter(
            treinamento=treinamento
        ).values('id', 'descricao', 'arquivo', 'created_at')

        # Serializar dados
        data = {
            'id': str(treinamento.id),
            'data': treinamento.data.strftime('%Y-%m-%d'),
            'tema': treinamento.tema,
            'local': treinamento.local,
            'responsavel': treinamento.responsavel,
            'status': treinamento.status,
            'observacoes': treinamento.observacoes or '',
            'created_at': treinamento.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'updated_at': treinamento.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
            'evidencias': [
                {
                    'id': str(e['id']),
                    'descricao': e['descricao'],
                    'arquivo': e['arquivo'],
                    'created_at': e['created_at'].strftime('%Y-%m-%d %H:%M:%S')
                } for e in evidencias
            ]
        }

        return JsonResponse({
            'success': True,
            'data': data
        })

    except Exception as e:
        logger.exception(f'Erro ao obter treinamento: {str(e)}')
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(['PUT'])
def atualizar_treinamento(request, treinamento_id):
    '''Atualiza um treinamento existente'''
    import logging
    logger = logging.getLogger(__name__)

    try:
        # Verificar se o treinamento existe e foi criado pelo usuário
        treinamento = get_object_or_404(Treinamento, id=treinamento_id, created_by=request.user)

        # Processar dados - aceitar FormData ou JSON
        if request.content_type and 'multipart/form-data' in request.content_type:
            # FormData - extrair JSON do campo 'data'
            data_json = request.POST.get('data')
            if data_json:
                data = json.loads(data_json)
            else:
                data = {}
            # Também pode ter campos diretos do FormData
            for key in request.POST.keys():
                if key not in ['csrfmiddlewaretoken', 'data']:
                    data[key] = request.POST.get(key)
        else:
            # JSON puro
            data = json.loads(request.body) if request.body else {}

        # Atualizar campos se fornecidos
        if 'data' in data:
            treinamento.data = data['data']
        if 'tema' in data:
            treinamento.tema = data['tema']
        if 'local' in data:
            treinamento.local = data['local']
        if 'responsavel' in data:
            treinamento.responsavel = data['responsavel']
        if 'status' in data:
            treinamento.status = data['status']
        if 'observacoes' in data:
            treinamento.observacoes = data['observacoes']

        treinamento.save()

        logger.info(f'Treinamento atualizado: {treinamento.id}')

        return JsonResponse({
            'success': True,
            'message': 'Treinamento atualizado com sucesso!',
            'id': str(treinamento.id)
        })

    except Exception as e:
        logger.exception(f'Erro ao atualizar treinamento: {str(e)}')
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(['DELETE'])
def deletar_treinamento(request, treinamento_id):
    '''Deleta um treinamento existente'''
    import logging
    logger = logging.getLogger(__name__)

    try:
        # Verificar se o treinamento existe e foi criado pelo usuário
        treinamento = get_object_or_404(Treinamento, id=treinamento_id, created_by=request.user)

        tema = treinamento.tema
        treinamento.delete()

        logger.info(f'Treinamento deletado: {treinamento_id} - {tema}')

        return JsonResponse({
            'success': True,
            'message': 'Treinamento deletado com sucesso!'
        })

    except Exception as e:
        logger.exception(f'Erro ao deletar treinamento: {str(e)}')
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


# ========== VISITA TÉCNICA ==========

@login_required
@require_http_methods(['GET'])
def obter_visita_tecnica(request, visita_id):
    '''Obtém detalhes de uma visita técnica específica'''
    import logging
    logger = logging.getLogger(__name__)

    try:
        # Verificar se a visita existe e foi criada pelo usuário
        visita = get_object_or_404(VisitaTecnica, id=visita_id, created_by=request.user)

        # Buscar evidências relacionadas
        evidencias = EvidenciaQualidade.objects.filter(
            visita_tecnica=visita
        ).values('id', 'descricao', 'arquivo', 'created_at')

        # Buscar não conformidades relacionadas
        nao_conformidades = NaoConformidade.objects.filter(
            visita_tecnica=visita
        ).values('id', 'descricao', 'classificacao', 'status', 'prazo')

        # Serializar dados
        data = {
            'id': str(visita.id),
            'data': visita.data.strftime('%Y-%m-%d'),
            'tipo': visita.tipo,
            'local': visita.local,
            'responsavel': visita.responsavel,
            'checklist': visita.checklist,
            'observacoes': visita.observacoes or '',
            'status': visita.status,
            'resultado_conformidade': visita.resultado_conformidade,
            'created_at': visita.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'updated_at': visita.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
            'evidencias': [
                {
                    'id': str(e['id']),
                    'descricao': e['descricao'],
                    'arquivo': e['arquivo'],
                    'created_at': e['created_at'].strftime('%Y-%m-%d %H:%M:%S')
                } for e in evidencias
            ],
            'nao_conformidades': [
                {
                    'id': str(nc['id']),
                    'descricao': nc['descricao'],
                    'classificacao': nc['classificacao'],
                    'status': nc['status'],
                    'prazo': nc['prazo'].strftime('%Y-%m-%d')
                } for nc in nao_conformidades
            ]
        }

        return JsonResponse({
            'success': True,
            'data': data
        })

    except Exception as e:
        logger.exception(f'Erro ao obter visita técnica: {str(e)}')
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(['PUT'])
def atualizar_visita_tecnica(request, visita_id):
    '''Atualiza uma visita técnica existente'''
    import logging
    logger = logging.getLogger(__name__)

    try:
        # Verificar se a visita existe e foi criada pelo usuário
        visita = get_object_or_404(VisitaTecnica, id=visita_id, created_by=request.user)

        # Processar dados - aceitar FormData ou JSON
        if request.content_type and 'multipart/form-data' in request.content_type:
            # FormData - extrair JSON do campo 'data'
            data_json = request.POST.get('data')
            if data_json:
                data = json.loads(data_json)
            else:
                data = {}
            # Também pode ter campos diretos do FormData
            for key in request.POST.keys():
                if key not in ['csrfmiddlewaretoken', 'data']:
                    data[key] = request.POST.get(key)
        else:
            # JSON puro
            data = json.loads(request.body) if request.body else {}

        # Atualizar campos se fornecidos
        if 'data' in data:
            from datetime import datetime
            data_str = data['data']
            if '-' in data_str:
                visita.data = datetime.strptime(data_str, '%Y-%m-%d').date()
            elif '/' in data_str:
                visita.data = datetime.strptime(data_str, '%d/%m/%Y').date()

        if 'tipo' in data:
            visita.tipo = data['tipo']
        if 'local' in data:
            visita.local = data['local']
        if 'responsavel' in data:
            visita.responsavel = data['responsavel']
        if 'checklist' in data:
            visita.checklist = data['checklist']
        if 'observacoes' in data:
            visita.observacoes = data['observacoes']
        if 'status' in data:
            visita.status = data['status']
        if 'resultado_conformidade' in data:
            visita.resultado_conformidade = data['resultado_conformidade']

        visita.save()

        logger.info(f'Visita técnica atualizada: {visita.id}')

        # Sincronizar campos da Não Conformidade relacionada (se existir)
        nc_atualizada = False
        nc_id = None
        try:
            nao_conformidade = NaoConformidade.objects.get(visita_tecnica=visita)

            # Atualizar prazo se fornecido
            if 'prazo_nc' in data and data['prazo_nc']:
                prazo_nc = data['prazo_nc']
                # Converter para date se necessário
                if isinstance(prazo_nc, str):
                    from datetime import datetime
                    if '-' in prazo_nc:
                        nao_conformidade.prazo = datetime.strptime(prazo_nc, '%Y-%m-%d').date()
                    elif '/' in prazo_nc:
                        nao_conformidade.prazo = datetime.strptime(prazo_nc, '%d/%m/%Y').date()
                else:
                    nao_conformidade.prazo = prazo_nc
                nc_atualizada = True

            # Atualizar criticidade se fornecido
            if 'criticidade_nc' in data and data['criticidade_nc']:
                classificacoes_validas = ['baixa', 'media', 'alta', 'critica']
                if data['criticidade_nc'] in classificacoes_validas:
                    nao_conformidade.classificacao = data['criticidade_nc']
                    nc_atualizada = True

            if nc_atualizada:
                nao_conformidade.save()
                nc_id = str(nao_conformidade.id)
                logger.info(f'Não Conformidade {nc_id} atualizada com prazo e criticidade da visita {visita.id}')

        except NaoConformidade.DoesNotExist:
            # Não há NC relacionada, normal
            logger.debug(f'Nenhuma Não Conformidade relacionada à visita {visita.id}')
            pass

        return JsonResponse({
            'success': True,
            'message': 'Visita técnica atualizada com sucesso!',
            'id': str(visita.id),
            'nc_atualizada': nc_atualizada,
            'nc_id': nc_id
        })

    except Exception as e:
        logger.exception(f'Erro ao atualizar visita técnica: {str(e)}')
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(['DELETE'])
def deletar_visita_tecnica(request, visita_id):
    '''Deleta uma visita técnica existente'''
    import logging
    logger = logging.getLogger(__name__)

    try:
        # Verificar se a visita existe e foi criada pelo usuário
        visita = get_object_or_404(VisitaTecnica, id=visita_id, created_by=request.user)

        local = visita.local
        visita.delete()

        logger.info(f'Visita técnica deletada: {visita_id} - {local}')

        return JsonResponse({
            'success': True,
            'message': 'Visita técnica deletada com sucesso!'
        })

    except Exception as e:
        logger.exception(f'Erro ao deletar visita técnica: {str(e)}')
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


# ========== NÃO CONFORMIDADE ==========

@login_required
@require_http_methods(['POST'])
def criar_nao_conformidade_manual(request):
    '''Cria uma nova não conformidade manualmente (não vinculada a visita técnica)'''
    import logging
    logger = logging.getLogger(__name__)

    try:
        data = json.loads(request.body)

        # Validar campos obrigatórios
        campos_obrigatorios = ['descricao', 'responsavel', 'prazo']
        campos_faltantes = [campo for campo in campos_obrigatorios if not data.get(campo)]

        if campos_faltantes:
            return JsonResponse({
                'success': False,
                'error': f'Campos obrigatórios faltando: {", ".join(campos_faltantes)}'
            }, status=400)

        # Converter prazo para date se necessário
        from datetime import datetime, date
        prazo = data.get('prazo')
        if isinstance(prazo, str):
            if '-' in prazo:
                prazo = datetime.strptime(prazo, '%Y-%m-%d').date()
            elif '/' in prazo:
                prazo = datetime.strptime(prazo, '%d/%m/%Y').date()

        # Converter data_identificacao se fornecida
        data_identificacao = data.get('data_identificacao', date.today())
        if isinstance(data_identificacao, str):
            if '-' in data_identificacao:
                data_identificacao = datetime.strptime(data_identificacao, '%Y-%m-%d').date()
            elif '/' in data_identificacao:
                data_identificacao = datetime.strptime(data_identificacao, '%d/%m/%Y').date()

        nao_conformidade = NaoConformidade.objects.create(
            data_identificacao=data_identificacao,
            descricao=data.get('descricao'),
            referencia_normativa=data.get('referencia_normativa', ''),
            classificacao=data.get('classificacao', 'media'),
            responsavel=data.get('responsavel'),
            prazo=prazo,
            status=data.get('status', 'pendente'),
            visita_tecnica_id=data.get('visita_tecnica_id') if data.get('visita_tecnica_id') else None,
            created_by=request.user
        )

        logger.info(f'Não conformidade criada manualmente: {nao_conformidade.id}')

        return JsonResponse({
            'success': True,
            'message': 'Não conformidade criada com sucesso!',
            'id': str(nao_conformidade.id)
        })

    except Exception as e:
        logger.exception(f'Erro ao criar não conformidade: {str(e)}')
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(['GET'])
def obter_nao_conformidade(request, nc_id):
    '''Obtém detalhes de uma não conformidade específica'''
    import logging
    logger = logging.getLogger(__name__)

    try:
        # Verificar se a NC existe e foi criada pelo usuário
        nc = get_object_or_404(NaoConformidade, id=nc_id, created_by=request.user)

        # Buscar planos de ação relacionados
        planos_acao = PlanoAcao.objects.filter(
            nao_conformidade=nc
        ).values('id', 'tipo_acao', 'descricao', 'status', 'prazo', 'responsavel')

        # Serializar dados
        data = {
            'id': str(nc.id),
            'data_identificacao': nc.data_identificacao.strftime('%Y-%m-%d'),
            'descricao': nc.descricao,
            'referencia_normativa': nc.referencia_normativa or '',
            'classificacao': nc.classificacao,
            'responsavel': nc.responsavel,
            'prazo': nc.prazo.strftime('%Y-%m-%d'),
            'status': nc.status,
            'data_resolucao': nc.data_resolucao.strftime('%Y-%m-%d') if nc.data_resolucao else None,
            'em_atraso': nc.em_atraso,
            'visita_tecnica_id': str(nc.visita_tecnica.id) if nc.visita_tecnica else None,
            'created_at': nc.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'updated_at': nc.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
            'planos_acao': [
                {
                    'id': str(pa['id']),
                    'tipo_acao': pa['tipo_acao'],
                    'descricao': pa['descricao'],
                    'status': pa['status'],
                    'prazo': pa['prazo'].strftime('%Y-%m-%d'),
                    'responsavel': pa['responsavel']
                } for pa in planos_acao
            ]
        }

        return JsonResponse({
            'success': True,
            'data': data
        })

    except Exception as e:
        logger.exception(f'Erro ao obter não conformidade: {str(e)}')
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(['PUT'])
def atualizar_nao_conformidade(request, nc_id):
    '''Atualiza uma não conformidade existente'''
    import logging
    logger = logging.getLogger(__name__)

    try:
        # Verificar se a NC existe e foi criada pelo usuário
        nc = get_object_or_404(NaoConformidade, id=nc_id, created_by=request.user)

        # Processar dados - aceitar FormData ou JSON
        if request.content_type and 'multipart/form-data' in request.content_type:
            # FormData com possível upload de arquivos
            data_json = request.POST.get('data')
            if data_json:
                data = json.loads(data_json)
            else:
                # Extrair dados do POST
                data = {}
                for key in request.POST.keys():
                    if key not in ['csrfmiddlewaretoken', 'data']:
                        data[key] = request.POST.get(key)
        else:
            # JSON puro
            data = json.loads(request.body) if request.body else {}

        # Atualizar campos se fornecidos
        from datetime import datetime

        if 'data_identificacao' in data:
            data_str = data['data_identificacao']
            if '-' in data_str:
                nc.data_identificacao = datetime.strptime(data_str, '%Y-%m-%d').date()
            elif '/' in data_str:
                nc.data_identificacao = datetime.strptime(data_str, '%d/%m/%Y').date()

        if 'descricao' in data:
            nc.descricao = data['descricao']
        if 'referencia_normativa' in data:
            nc.referencia_normativa = data['referencia_normativa']
        if 'classificacao' in data:
            nc.classificacao = data['classificacao']
        if 'responsavel' in data:
            nc.responsavel = data['responsavel']

        if 'prazo' in data:
            data_str = data['prazo']
            if '-' in data_str:
                nc.prazo = datetime.strptime(data_str, '%Y-%m-%d').date()
            elif '/' in data_str:
                nc.prazo = datetime.strptime(data_str, '%d/%m/%Y').date()

        if 'status' in data:
            nc.status = data['status']

        if 'data_resolucao' in data:
            if data['data_resolucao']:
                data_str = data['data_resolucao']
                if '-' in data_str:
                    nc.data_resolucao = datetime.strptime(data_str, '%Y-%m-%d').date()
                elif '/' in data_str:
                    nc.data_resolucao = datetime.strptime(data_str, '%d/%m/%Y').date()
            else:
                nc.data_resolucao = None

        nc.save()

        logger.info(f'Não conformidade atualizada: {nc.id}')

        return JsonResponse({
            'success': True,
            'message': 'Não conformidade atualizada com sucesso!',
            'id': str(nc.id)
        })

    except Exception as e:
        logger.exception(f'Erro ao atualizar não conformidade: {str(e)}')
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(['DELETE'])
def deletar_nao_conformidade(request, nc_id):
    '''Deleta uma não conformidade existente'''
    import logging
    logger = logging.getLogger(__name__)

    try:
        # Verificar se a NC existe e foi criada pelo usuário
        nc = get_object_or_404(NaoConformidade, id=nc_id, created_by=request.user)

        descricao = nc.descricao[:50]
        nc.delete()

        logger.info(f'Não conformidade deletada: {nc_id} - {descricao}')

        return JsonResponse({
            'success': True,
            'message': 'Não conformidade deletada com sucesso!'
        })

    except Exception as e:
        logger.exception(f'Erro ao deletar não conformidade: {str(e)}')
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


# ========== PLANO DE AÇÃO ==========

@login_required
@require_http_methods(['GET'])
def obter_plano_acao(request, plano_id):
    '''Obtém detalhes de um plano de ação específico'''
    import logging
    logger = logging.getLogger(__name__)

    try:
        # Verificar se o plano existe e foi criado pelo usuário
        plano = get_object_or_404(PlanoAcao, id=plano_id, created_by=request.user)

        # Buscar evidências relacionadas
        evidencias = EvidenciaQualidade.objects.filter(
            plano_acao=plano
        ).values('id', 'descricao', 'arquivo', 'created_at')

        # Serializar dados
        data = {
            'id': str(plano.id),
            'nao_conformidade_id': str(plano.nao_conformidade.id),
            'tipo_acao': plano.tipo_acao,
            'descricao': plano.descricao,
            'responsavel': plano.responsavel,
            'prazo': plano.prazo.strftime('%Y-%m-%d'),
            'status': plano.status,
            'data_conclusao': plano.data_conclusao.strftime('%Y-%m-%d') if plano.data_conclusao else None,
            'observacoes': plano.observacoes or '',
            'created_at': plano.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'updated_at': plano.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
            'evidencias': [
                {
                    'id': str(e['id']),
                    'descricao': e['descricao'],
                    'arquivo': e['arquivo'],
                    'created_at': e['created_at'].strftime('%Y-%m-%d %H:%M:%S')
                } for e in evidencias
            ]
        }

        return JsonResponse({
            'success': True,
            'data': data
        })

    except Exception as e:
        logger.exception(f'Erro ao obter plano de ação: {str(e)}')
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(['PUT'])
def atualizar_plano_acao(request, plano_id):
    '''Atualiza um plano de ação existente'''
    import logging
    logger = logging.getLogger(__name__)

    try:
        # Verificar se o plano existe e foi criado pelo usuário
        plano = get_object_or_404(PlanoAcao, id=plano_id, created_by=request.user)

        # Processar dados - aceitar FormData ou JSON
        if request.content_type and 'multipart/form-data' in request.content_type:
            # FormData - extrair JSON do campo 'data'
            data_json = request.POST.get('data')
            if data_json:
                data = json.loads(data_json)
            else:
                data = {}
            # Também pode ter campos diretos do FormData
            for key in request.POST.keys():
                if key not in ['csrfmiddlewaretoken', 'data']:
                    data[key] = request.POST.get(key)
        else:
            # JSON puro
            data = json.loads(request.body) if request.body else {}

        # Atualizar campos se fornecidos
        from datetime import datetime

        if 'tipo_acao' in data:
            plano.tipo_acao = data['tipo_acao']
        if 'descricao' in data:
            plano.descricao = data['descricao']
        if 'responsavel' in data:
            plano.responsavel = data['responsavel']

        if 'prazo' in data:
            data_str = data['prazo']
            if '-' in data_str:
                plano.prazo = datetime.strptime(data_str, '%Y-%m-%d').date()
            elif '/' in data_str:
                plano.prazo = datetime.strptime(data_str, '%d/%m/%Y').date()

        if 'status' in data:
            plano.status = data['status']

        if 'data_conclusao' in data:
            if data['data_conclusao']:
                data_str = data['data_conclusao']
                if '-' in data_str:
                    plano.data_conclusao = datetime.strptime(data_str, '%Y-%m-%d').date()
                elif '/' in data_str:
                    plano.data_conclusao = datetime.strptime(data_str, '%d/%m/%Y').date()
            else:
                plano.data_conclusao = None

        if 'observacoes' in data:
            plano.observacoes = data['observacoes']

        plano.save()

        logger.info(f'Plano de ação atualizado: {plano.id}')

        return JsonResponse({
            'success': True,
            'message': 'Plano de ação atualizado com sucesso!',
            'id': str(plano.id)
        })

    except Exception as e:
        logger.exception(f'Erro ao atualizar plano de ação: {str(e)}')
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(['DELETE'])
def deletar_plano_acao(request, plano_id):
    '''Deleta um plano de ação existente'''
    import logging
    logger = logging.getLogger(__name__)

    try:
        # Verificar se o plano existe e foi criado pelo usuário
        plano = get_object_or_404(PlanoAcao, id=plano_id, created_by=request.user)

        descricao = plano.descricao[:50]
        plano.delete()

        logger.info(f'Plano de ação deletado: {plano_id} - {descricao}')

        return JsonResponse({
            'success': True,
            'message': 'Plano de ação deletado com sucesso!'
        })

    except Exception as e:
        logger.exception(f'Erro ao deletar plano de ação: {str(e)}')
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(['POST'])
def adicionar_evidencias_plano_acao(request, plano_id):
    '''Adiciona evidências (arquivos) a um plano de ação existente'''
    import logging
    logger = logging.getLogger(__name__)

    try:
        # Verificar se o plano existe e foi criado pelo usuário
        plano = get_object_or_404(PlanoAcao, id=plano_id, created_by=request.user)

        logger.info(f'=== ADICIONAR EVIDÊNCIAS PLANO DE AÇÃO ===')
        logger.info(f'Plano ID: {plano_id}')
        logger.info(f'POST data keys: {list(request.POST.keys())}')
        logger.info(f'FILES data keys: {list(request.FILES.keys())}')

        evidencias_criadas = []

        # Processar upload de evidências (múltiplos arquivos numerados)
        import os
        i = 0
        while True:
            arquivo_key = f'evidencia_{i}'
            descricao_key = f'evidencia_{i}_descricao'

            if arquivo_key not in request.FILES:
                break

            arquivo = request.FILES[arquivo_key]
            descricao = request.POST.get(descricao_key, f'Evidência {i+1} do plano de ação')

            logger.info(f'Processando evidência {i}: {arquivo.name} ({arquivo.size} bytes)')

            # Validar tipo de arquivo
            extensao = os.path.splitext(arquivo.name)[1].lower()
            extensoes_permitidas = ['.jpg', '.jpeg', '.png', '.gif', '.pdf']

            if extensao not in extensoes_permitidas:
                logger.error(f'Tipo de arquivo não permitido: {arquivo.name}')
                return JsonResponse({
                    'success': False,
                    'error': f'Tipo de arquivo não permitido: {arquivo.name}. Apenas imagens (jpg, jpeg, png, gif) e PDF são aceitos.'
                }, status=400)

            # Validar tamanho do arquivo (máximo 5MB)
            if arquivo.size > 5 * 1024 * 1024:
                logger.error(f'Arquivo muito grande: {arquivo.name} ({arquivo.size} bytes)')
                return JsonResponse({
                    'success': False,
                    'error': f'Arquivo muito grande: {arquivo.name}. Tamanho máximo: 5MB.'
                }, status=400)

            # Criar evidência
            evidencia = EvidenciaQualidade.objects.create(
                tipo='plano_acao',
                descricao=descricao,
                arquivo=arquivo,
                plano_acao=plano,
                uploaded_by=request.user
            )
            evidencias_criadas.append({
                'id': str(evidencia.id),
                'nome': arquivo.name,
                'descricao': descricao,
                'tamanho': arquivo.size
            })
            logger.info(f'Evidência criada: {evidencia.id}')

            i += 1

        # Se não encontrou no formato numerado, tentar formato antigo (evidencias)
        if not evidencias_criadas:
            arquivos = request.FILES.getlist('evidencias')
            logger.info(f'Tentando formato antigo: {len(arquivos)} arquivos')

            for arquivo in arquivos:
                # Validar tipo de arquivo
                extensao = os.path.splitext(arquivo.name)[1].lower()
                extensoes_permitidas = ['.jpg', '.jpeg', '.png', '.gif', '.pdf']

                if extensao not in extensoes_permitidas:
                    return JsonResponse({
                        'success': False,
                        'error': f'Tipo de arquivo não permitido: {arquivo.name}. Apenas imagens (jpg, jpeg, png, gif) e PDF são aceitos.'
                    }, status=400)

                if arquivo.size > 5 * 1024 * 1024:
                    return JsonResponse({
                        'success': False,
                        'error': f'Arquivo muito grande: {arquivo.name}. Tamanho máximo: 5MB.'
                    }, status=400)

                evidencia = EvidenciaQualidade.objects.create(
                    tipo='plano_acao',
                    descricao=f'Evidência do plano de ação - {arquivo.name}',
                    arquivo=arquivo,
                    plano_acao=plano,
                    uploaded_by=request.user
                )
                evidencias_criadas.append({
                    'id': str(evidencia.id),
                    'nome': arquivo.name,
                    'tamanho': arquivo.size
                })

        if not evidencias_criadas:
            return JsonResponse({
                'success': False,
                'error': 'Nenhuma evidência foi enviada.'
            }, status=400)

        logger.info(f'Total de evidências criadas: {len(evidencias_criadas)}')

        return JsonResponse({
            'success': True,
            'message': f'{len(evidencias_criadas)} evidência(s) adicionada(s) com sucesso!',
            'evidencias_count': len(evidencias_criadas),
            'evidencias': evidencias_criadas
        })

    except Exception as e:
        logger.exception(f'Erro ao adicionar evidências: {str(e)}')
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


# ================== PWA VIEWS ==================

def manifest(request):
    """
    Serve o manifest.json para PWA
    """
    import json
    import os

    # Caminho correto para o manifest
    if settings.STATIC_ROOT and os.path.exists(os.path.join(settings.STATIC_ROOT, 'manifest.json')):
        manifest_path = os.path.join(settings.STATIC_ROOT, 'manifest.json')
    else:
        manifest_path = os.path.join(settings.BASE_DIR, 'Gestao_a_Vista', 'static', 'manifest.json')

    try:
        with open(manifest_path, 'r', encoding='utf-8') as f:
            manifest_data = json.load(f)
        return JsonResponse(manifest_data, content_type='application/manifest+json')
    except FileNotFoundError:
        # Fallback: retorna manifest inline para todo o projeto
        manifest_data = {
            "name": "Gestão à Vista - RCO",
            "short_name": "Gestão à Vista",
            "description": "Sistema Integrado de Gestão à Vista - Regional Centro Oeste",
            "start_url": "/home/",
            "display": "standalone",
            "background_color": "#1e293b",
            "theme_color": "#1e293b",
            "orientation": "portrait-primary",
            "scope": "/"
        }
        return JsonResponse(manifest_data, content_type='application/manifest+json')


def service_worker(request):
    """
    Serve o service worker com o content-type correto
    """
    import os

    # Caminho correto para o service worker
    if settings.STATIC_ROOT and os.path.exists(os.path.join(settings.STATIC_ROOT, 'sw.js')):
        sw_path = os.path.join(settings.STATIC_ROOT, 'sw.js')
    else:
        sw_path = os.path.join(settings.BASE_DIR, 'Gestao_a_Vista', 'static', 'sw.js')

    try:
        with open(sw_path, 'r', encoding='utf-8') as f:
            sw_content = f.read()
        return HttpResponse(sw_content, content_type='application/javascript')
    except FileNotFoundError as e:
        # Log para debug
        print(f"[ERROR] Service Worker não encontrado em: {sw_path}")
        print(f"[ERROR] STATIC_ROOT: {settings.STATIC_ROOT}")
        print(f"[ERROR] BASE_DIR: {settings.BASE_DIR}")
        return HttpResponse(
            f'console.error("Service Worker not found at {sw_path}");',
            content_type='application/javascript',
            status=404
        )


@login_required
def calendario_2026(request):
    '''
    View para a página Calendário 2026 - Regional Centro Oeste
    '''
    import json
    from datetime import date
    from collections import defaultdict

    # Buscar todos os eventos de 2026
    eventos = EventoCalendario2026.objects.filter(
        data_inicio__year=2026
    ).order_by('data_inicio')

    # Organizar eventos por mês
    # Eventos multi-dia aparecem em todos os meses que atravessam
    eventos_por_mes = defaultdict(list)
    for evento in eventos:
        # Para eventos de múltiplos dias, adicionar em todos os meses relevantes
        dias_evento = evento.get_dias_evento()
        meses_ja_adicionados = set()

        for dia in dias_evento:
            mes = dia.month
            if mes not in meses_ja_adicionados:
                eventos_por_mes[mes].append({
                    'id': str(evento.id),
                    'data_inicio': evento.data_inicio.isoformat(),
                    'data_fim': evento.data_fim.isoformat() if evento.data_fim else None,
                    'titulo': evento.titulo,
                    'tipo': evento.tipo,
                    'tipo_display': evento.get_tipo_display(),
                    'descricao': evento.descricao or '',
                    'cor': evento.cor,
                    'legenda': evento.legenda or '',
                })
                meses_ja_adicionados.add(mes)

    # Converter defaultdict para dict normal e garantir que todos os meses existam
    eventos_organizados = {}
    meses_nomes = [
        'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
        'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro'
    ]
    for mes in range(1, 13):
        eventos_organizados[mes] = {
            'nome': meses_nomes[mes - 1],
            'eventos': eventos_por_mes.get(mes, [])
        }

    # Lista de feriados de 2026 para referência
    feriados_2026 = [
        {'data': '01/01', 'nome': 'Ano Novo'},
        {'data': '16/02', 'nome': 'Carnaval'},
        {'data': '03/04', 'nome': 'Sexta-Feira Santa'},
        {'data': '21/04', 'nome': 'Dia de Tiradentes'},
        {'data': '01/05', 'nome': 'Dia do Trabalho'},
        {'data': '04/06', 'nome': 'Corpus Christi'},
        {'data': '07/09', 'nome': 'Independência do Brasil'},
        {'data': '12/10', 'nome': 'Nossa Senhora Aparecida'},
        {'data': '24/10', 'nome': 'Aniversário Goiânia'},
        {'data': '02/11', 'nome': 'Dia de Finados'},
        {'data': '15/11', 'nome': 'Proclamação da República'},
        {'data': '20/11', 'nome': 'Dia da Consciência Negra'},
        {'data': '24/12', 'nome': 'Expediente até 12hs'},
        {'data': '25/12', 'nome': 'Natal'},
        {'data': '31/12', 'nome': 'Expediente até 12hs'},
    ]

    # Tipos de eventos disponíveis
    tipos_evento = EventoCalendario2026.TIPO_CHOICES

    context = {
        'eventos_por_mes': json.dumps(eventos_organizados),  # Serializar para JSON
        'feriados_2026': feriados_2026,
        'tipos_evento': tipos_evento,
        'total_eventos': eventos.count(),
        'ano': 2026,
    }

    return render(request, 'calendario_2026.html', context)


# ============================================================================
# API CRUD - Eventos do Calendário 2026
# ============================================================================

@login_required
@csrf_exempt
def criar_evento_calendario(request):
    '''
    API para criar um novo evento no Calendário 2026
    Método: POST
    '''
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'error': 'Método não permitido. Use POST.'
        }, status=405)

    try:
        import re
        from datetime import datetime

        # Parse dos dados JSON
        data = json.loads(request.body)

        # Validar campos obrigatórios
        # Compatibilidade: aceitar 'data' ou 'data_inicio'
        campos_obrigatorios = ['titulo', 'tipo']
        campos_faltantes = []

        for campo in campos_obrigatorios:
            valor = data.get(campo)
            if not valor or (isinstance(valor, str) and not valor.strip()):
                campos_faltantes.append(campo)

        # Verificar se ao menos 'data_inicio' ou 'data' (compatibilidade) foi fornecido
        if not data.get('data_inicio') and not data.get('data'):
            campos_faltantes.append('data_inicio')

        if campos_faltantes:
            return JsonResponse({
                'success': False,
                'error': f'Campos obrigatórios faltando ou vazios: {", ".join(campos_faltantes)}'
            }, status=400)

        # Validar título (max 255 caracteres)
        titulo = data.get('titulo', '').strip()
        if len(titulo) > 255:
            return JsonResponse({
                'success': False,
                'error': f'Título muito longo. Máximo 255 caracteres (atual: {len(titulo)})'
            }, status=400)

        # Validar tipo
        tipos_validos = [choice[0] for choice in EventoCalendario2026.TIPO_CHOICES]
        tipo = data.get('tipo')
        if tipo not in tipos_validos:
            return JsonResponse({
                'success': False,
                'error': f'Tipo inválido: "{tipo}". Valores permitidos: {", ".join(tipos_validos)}'
            }, status=400)

        # Validar e converter data_inicio (com compatibilidade para 'data')
        data_inicio_str = data.get('data_inicio') or data.get('data')
        try:
            # Aceitar formato YYYY-MM-DD (padrão HTML5 input date)
            if '-' in data_inicio_str:
                data_inicio_obj = datetime.strptime(data_inicio_str, '%Y-%m-%d').date()
            # Aceitar formato DD/MM/YYYY
            elif '/' in data_inicio_str:
                data_inicio_obj = datetime.strptime(data_inicio_str, '%d/%m/%Y').date()
            else:
                raise ValueError('Formato de data inválido')
        except (ValueError, TypeError) as e:
            return JsonResponse({
                'success': False,
                'error': f'Formato de data_inicio inválido: "{data_inicio_str}". Use YYYY-MM-DD ou DD/MM/YYYY'
            }, status=400)

        # Validar que a data_inicio está em 2026
        if data_inicio_obj.year != 2026:
            return JsonResponse({
                'success': False,
                'error': f'Data início deve ser em 2026. Data fornecida: {data_inicio_obj.strftime("%d/%m/%Y")}'
            }, status=400)

        # Validar e converter data_fim (opcional)
        data_fim_obj = None
        data_fim_str = data.get('data_fim')
        if data_fim_str:
            try:
                # Aceitar formato YYYY-MM-DD (padrão HTML5 input date)
                if '-' in data_fim_str:
                    data_fim_obj = datetime.strptime(data_fim_str, '%Y-%m-%d').date()
                # Aceitar formato DD/MM/YYYY
                elif '/' in data_fim_str:
                    data_fim_obj = datetime.strptime(data_fim_str, '%d/%m/%Y').date()
                else:
                    raise ValueError('Formato de data inválido')
            except (ValueError, TypeError) as e:
                return JsonResponse({
                    'success': False,
                    'error': f'Formato de data_fim inválido: "{data_fim_str}". Use YYYY-MM-DD ou DD/MM/YYYY'
                }, status=400)

            # Validar que a data_fim está em 2026
            if data_fim_obj.year != 2026:
                return JsonResponse({
                    'success': False,
                    'error': f'Data fim deve ser em 2026. Data fornecida: {data_fim_obj.strftime("%d/%m/%Y")}'
                }, status=400)

            # Validar que data_fim >= data_inicio
            if data_fim_obj < data_inicio_obj:
                return JsonResponse({
                    'success': False,
                    'error': f'Data fim ({data_fim_obj.strftime("%d/%m/%Y")}) não pode ser anterior à data início ({data_inicio_obj.strftime("%d/%m/%Y")})'
                }, status=400)

        # Validar cor (hexadecimal)
        cor = data.get('cor', '#3788d8').strip()
        if not re.match(r'^#[0-9A-Fa-f]{6}$', cor):
            return JsonResponse({
                'success': False,
                'error': f'Cor inválida: "{cor}". Use formato hexadecimal #RRGGBB (ex: #FF5733)'
            }, status=400)

        # Obter legenda personalizada (opcional)
        legenda = data.get('legenda', '').strip() if data.get('legenda') else None
        if legenda and len(legenda) > 10:
            return JsonResponse({
                'success': False,
                'error': f'Legenda muito longa: "{legenda}". Máximo de 10 caracteres.'
            }, status=400)

        # Criar o evento
        evento = EventoCalendario2026.objects.create(
            data_inicio=data_inicio_obj,
            data_fim=data_fim_obj,
            titulo=titulo,
            tipo=tipo,
            descricao=data.get('descricao', '').strip(),
            cor=cor,
            legenda=legenda
        )

        # Retornar sucesso com dados do evento criado
        return JsonResponse({
            'success': True,
            'message': 'Evento criado com sucesso!',
            'evento': {
                'id': str(evento.id),
                'data_inicio': evento.data_inicio.isoformat(),
                'data_fim': evento.data_fim.isoformat() if evento.data_fim else None,
                'titulo': evento.titulo,
                'tipo': evento.tipo,
                'tipo_display': evento.get_tipo_display(),
                'descricao': evento.descricao or '',
                'cor': evento.cor,
                'legenda': evento.legenda or '',
                'created_at': evento.created_at.isoformat(),
                'updated_at': evento.updated_at.isoformat()
            }
        }, status=201)

    except json.JSONDecodeError as e:
        return JsonResponse({
            'success': False,
            'error': f'Erro ao processar JSON: {str(e)}'
        }, status=400)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'Erro ao criar evento: {str(e)}')
        return JsonResponse({
            'success': False,
            'error': f'Erro ao criar evento: {str(e)}'
        }, status=500)


@login_required
def listar_eventos_calendario(request):
    '''
    API para listar eventos do Calendário 2026
    Método: GET
    Query params opcionais:
    - mes: filtrar por mês (1-12)
    - tipo: filtrar por tipo de evento
    '''
    try:
        from django.db.models import Q
        from datetime import date

        # Buscar todos os eventos de 2026
        eventos = EventoCalendario2026.objects.filter(data_inicio__year=2026)

        # Filtrar por mês se especificado
        mes = request.GET.get('mes')
        if mes:
            try:
                mes_int = int(mes)
                if 1 <= mes_int <= 12:
                    # Evento deve aparecer se:
                    # - data_inicio está no mês OU
                    # - data_fim está no mês OU
                    # - mês está entre data_inicio e data_fim (evento multi-dia que atravessa o mês)

                    # Calcular primeiro e último dia do mês
                    primeiro_dia = date(2026, mes_int, 1)
                    if mes_int == 12:
                        ultimo_dia = date(2026, 12, 31)
                    else:
                        from calendar import monthrange
                        ultimo_dia = date(2026, mes_int, monthrange(2026, mes_int)[1])

                    eventos = eventos.filter(
                        Q(data_inicio__month=mes_int) |  # Inicia no mês
                        Q(data_fim__month=mes_int) |     # Termina no mês
                        Q(data_inicio__lt=primeiro_dia, data_fim__gte=primeiro_dia)  # Atravessa o mês
                    )
                else:
                    return JsonResponse({
                        'success': False,
                        'error': 'Mês deve estar entre 1 e 12'
                    }, status=400)
            except ValueError:
                return JsonResponse({
                    'success': False,
                    'error': 'Mês deve ser um número'
                }, status=400)

        # Filtrar por tipo se especificado
        tipo = request.GET.get('tipo')
        if tipo:
            tipos_validos = [choice[0] for choice in EventoCalendario2026.TIPO_CHOICES]
            if tipo in tipos_validos:
                eventos = eventos.filter(tipo=tipo)
            else:
                return JsonResponse({
                    'success': False,
                    'error': f'Tipo inválido. Valores permitidos: {", ".join(tipos_validos)}'
                }, status=400)

        # Ordenar por data_inicio
        eventos = eventos.order_by('data_inicio', 'titulo')

        # Serializar eventos
        eventos_list = []
        for evento in eventos:
            eventos_list.append({
                'id': str(evento.id),
                'data_inicio': evento.data_inicio.isoformat(),
                'data_fim': evento.data_fim.isoformat() if evento.data_fim else None,
                'titulo': evento.titulo,
                'tipo': evento.tipo,
                'tipo_display': evento.get_tipo_display(),
                'descricao': evento.descricao or '',
                'cor': evento.cor,
                'legenda': evento.legenda or '',
                'created_at': evento.created_at.isoformat(),
                'updated_at': evento.updated_at.isoformat()
            })

        return JsonResponse({
            'success': True,
            'total': len(eventos_list),
            'eventos': eventos_list
        }, status=200)

    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'Erro ao listar eventos: {str(e)}')
        return JsonResponse({
            'success': False,
            'error': f'Erro ao listar eventos: {str(e)}'
        }, status=500)


@login_required
def obter_evento_calendario(request, evento_id):
    '''
    API para obter um evento específico do Calendário 2026
    Método: GET
    '''
    try:
        evento = get_object_or_404(EventoCalendario2026, id=evento_id)

        return JsonResponse({
            'success': True,
            'evento': {
                'id': str(evento.id),
                'data_inicio': evento.data_inicio.isoformat(),
                'data_fim': evento.data_fim.isoformat() if evento.data_fim else None,
                'titulo': evento.titulo,
                'tipo': evento.tipo,
                'tipo_display': evento.get_tipo_display(),
                'descricao': evento.descricao or '',
                'cor': evento.cor,
                'legenda': evento.legenda or '',
                'created_at': evento.created_at.isoformat(),
                'updated_at': evento.updated_at.isoformat()
            }
        }, status=200)

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Evento não encontrado: {str(e)}'
        }, status=404)


@login_required
@csrf_exempt
def atualizar_evento_calendario(request, evento_id):
    '''
    API para atualizar um evento do Calendário 2026
    Método: PUT
    '''
    if request.method != 'PUT':
        return JsonResponse({
            'success': False,
            'error': 'Método não permitido. Use PUT.'
        }, status=405)

    try:
        import re
        from datetime import datetime

        # Buscar o evento
        evento = get_object_or_404(EventoCalendario2026, id=evento_id)

        # Parse dos dados JSON
        data = json.loads(request.body)

        # Atualizar campos se fornecidos

        # Atualizar data_inicio (com compatibilidade para 'data')
        if 'data_inicio' in data or 'data' in data:
            data_inicio_str = data.get('data_inicio') or data.get('data')
            try:
                if '-' in data_inicio_str:
                    data_inicio_obj = datetime.strptime(data_inicio_str, '%Y-%m-%d').date()
                elif '/' in data_inicio_str:
                    data_inicio_obj = datetime.strptime(data_inicio_str, '%d/%m/%Y').date()
                else:
                    raise ValueError('Formato de data inválido')

                # Validar ano 2026
                if data_inicio_obj.year != 2026:
                    return JsonResponse({
                        'success': False,
                        'error': f'Data início deve ser em 2026. Data fornecida: {data_inicio_obj.strftime("%d/%m/%Y")}'
                    }, status=400)

                evento.data_inicio = data_inicio_obj
            except (ValueError, TypeError) as e:
                return JsonResponse({
                    'success': False,
                    'error': f'Formato de data_inicio inválido: "{data_inicio_str}". Use YYYY-MM-DD ou DD/MM/YYYY'
                }, status=400)

        # Atualizar data_fim (opcional)
        if 'data_fim' in data:
            data_fim_str = data.get('data_fim')
            if data_fim_str:
                try:
                    if '-' in data_fim_str:
                        data_fim_obj = datetime.strptime(data_fim_str, '%Y-%m-%d').date()
                    elif '/' in data_fim_str:
                        data_fim_obj = datetime.strptime(data_fim_str, '%d/%m/%Y').date()
                    else:
                        raise ValueError('Formato de data inválido')

                    # Validar ano 2026
                    if data_fim_obj.year != 2026:
                        return JsonResponse({
                            'success': False,
                            'error': f'Data fim deve ser em 2026. Data fornecida: {data_fim_obj.strftime("%d/%m/%Y")}'
                        }, status=400)

                    # Validar que data_fim >= data_inicio
                    if data_fim_obj < evento.data_inicio:
                        return JsonResponse({
                            'success': False,
                            'error': f'Data fim ({data_fim_obj.strftime("%d/%m/%Y")}) não pode ser anterior à data início ({evento.data_inicio.strftime("%d/%m/%Y")})'
                        }, status=400)

                    evento.data_fim = data_fim_obj
                except (ValueError, TypeError) as e:
                    return JsonResponse({
                        'success': False,
                        'error': f'Formato de data_fim inválido: "{data_fim_str}". Use YYYY-MM-DD ou DD/MM/YYYY'
                    }, status=400)
            else:
                # Se data_fim enviado como null/empty, limpar o campo
                evento.data_fim = None

        # Atualizar título
        if 'titulo' in data:
            titulo = data.get('titulo', '').strip()
            if not titulo:
                return JsonResponse({
                    'success': False,
                    'error': 'Título não pode estar vazio'
                }, status=400)
            if len(titulo) > 255:
                return JsonResponse({
                    'success': False,
                    'error': f'Título muito longo. Máximo 255 caracteres (atual: {len(titulo)})'
                }, status=400)
            evento.titulo = titulo

        # Atualizar tipo
        if 'tipo' in data:
            tipo = data.get('tipo')
            tipos_validos = [choice[0] for choice in EventoCalendario2026.TIPO_CHOICES]
            if tipo not in tipos_validos:
                return JsonResponse({
                    'success': False,
                    'error': f'Tipo inválido: "{tipo}". Valores permitidos: {", ".join(tipos_validos)}'
                }, status=400)
            evento.tipo = tipo

        # Atualizar descrição
        if 'descricao' in data:
            evento.descricao = data.get('descricao', '').strip()

        # Atualizar cor
        if 'cor' in data:
            cor = data.get('cor', '').strip()
            if not re.match(r'^#[0-9A-Fa-f]{6}$', cor):
                return JsonResponse({
                    'success': False,
                    'error': f'Cor inválida: "{cor}". Use formato hexadecimal #RRGGBB (ex: #FF5733)'
                }, status=400)
            evento.cor = cor

        # Atualizar legenda personalizada (opcional)
        if 'legenda' in data:
            legenda = data.get('legenda', '').strip() if data.get('legenda') else None
            if legenda and len(legenda) > 10:
                return JsonResponse({
                    'success': False,
                    'error': f'Legenda muito longa: "{legenda}". Máximo de 10 caracteres.'
                }, status=400)
            evento.legenda = legenda

        # Salvar as alterações
        evento.save()

        return JsonResponse({
            'success': True,
            'message': 'Evento atualizado com sucesso!',
            'evento': {
                'id': str(evento.id),
                'data_inicio': evento.data_inicio.isoformat(),
                'data_fim': evento.data_fim.isoformat() if evento.data_fim else None,
                'titulo': evento.titulo,
                'tipo': evento.tipo,
                'tipo_display': evento.get_tipo_display(),
                'descricao': evento.descricao or '',
                'cor': evento.cor,
                'legenda': evento.legenda or '',
                'created_at': evento.created_at.isoformat(),
                'updated_at': evento.updated_at.isoformat()
            }
        }, status=200)

    except json.JSONDecodeError as e:
        return JsonResponse({
            'success': False,
            'error': f'Erro ao processar JSON: {str(e)}'
        }, status=400)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'Erro ao atualizar evento: {str(e)}')
        return JsonResponse({
            'success': False,
            'error': f'Erro ao atualizar evento: {str(e)}'
        }, status=500)


@login_required
@csrf_exempt
def excluir_evento_calendario(request, evento_id):
    '''
    API para excluir um evento do Calendário 2026
    Método: DELETE
    '''
    if request.method != 'DELETE':
        return JsonResponse({
            'success': False,
            'error': 'Método não permitido. Use DELETE.'
        }, status=405)

    try:
        # Buscar o evento
        evento = get_object_or_404(EventoCalendario2026, id=evento_id)

        # Armazenar dados antes de excluir
        evento_data = {
            'id': str(evento.id),
            'titulo': evento.titulo,
            'data_inicio': evento.data_inicio.isoformat(),
            'data_fim': evento.data_fim.isoformat() if evento.data_fim else None
        }

        # Excluir o evento
        evento.delete()

        return JsonResponse({
            'success': True,
            'message': f'Evento "{evento_data["titulo"]}" excluído com sucesso!',
            'evento': evento_data
        }, status=200)

    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'Erro ao excluir evento: {str(e)}')
        return JsonResponse({
            'success': False,
            'error': f'Erro ao excluir evento: {str(e)}'
        }, status=500)
    

@login_required
def historico_item_ocorrencia_api(request):
    """
    Retorna TODO o histórico de planos de ação passados (Sem limite de data)
    para um item específico dentro de um CR específico.
    """
    cr_nome = request.GET.get('cr', '')
    item = request.GET.get('item', '')
    
    if not cr_nome or not item:
        return JsonResponse({'success': False, 'message': 'CR e Item são obrigatórios.'})
        
    gerente_nome = "Não Atribuído"
    
    estrutura = Estrutura.objects.using('default').filter(
        Q(descricao__icontains=cr_nome) | Q(cr=cr_nome)
    ).first()
    
    if estrutura and estrutura.gc:
        gerente_nome = estrutura.gc
        
    # Busca TODO o histórico (limite de data removido)
    planos = OcorrenciaPlanoAcao.objects.filter(
        cr_colaborador=cr_nome, 
        item_em_falta=item
    ).prefetch_related('historicos__usuario').select_related('criador_plano', 'aprovador', 'comprador', 'retirante').order_by('-data_criacao')
    
    historicos = []
    from django.utils.dateformat import DateFormat
    
    for p in planos:
        logs = []
        for h in p.historicos.all():
            nome_usuario = 'Sistema'
            if h.usuario:
                nome_usuario = h.usuario.name if hasattr(h.usuario, 'name') and h.usuario.name else h.usuario.username
                
            logs.append({
                'data': DateFormat(h.data_alteracao).format('d/m/Y H:i') if h.data_alteracao else None,
                'status_novo': h.status_novo,
                'status_display': dict(OcorrenciaPlanoAcao.STATUS_CHOICES).get(h.status_novo, h.status_novo),
                'usuario': nome_usuario,
                'observacao': h.observacao
            })
            
        foto_url = request.build_absolute_uri(p.foto_retirada.url) if hasattr(p, 'foto_retirada') and p.foto_retirada else None
        guia_url = request.build_absolute_uri(p.guia_de_trafego.url) if hasattr(p, 'guia_de_trafego') and p.guia_de_trafego else None
        
        historicos.append({
            'plano_id': str(p.id),
            'status': p.status,
            'status_display': p.get_status_display(),
            'criador': p.criador_plano.name if p.criador_plano else 'Desconhecido',
            'gerente_cr': gerente_nome,
            'data_criacao': DateFormat(p.data_criacao).format('d/m/Y H:i') if p.data_criacao else None,
            'justificativa': p.justificativa_aprovacao if hasattr(p, 'justificativa_aprovacao') else None,
            'is_regulatory': p.is_regulatory if hasattr(p, 'is_regulatory') else False,
            'logs': logs,
            'foto_retirada': foto_url,
            'guia_de_trafego': guia_url,
        })
        
    return JsonResponse({'success': True, 'historicos': historicos, 'gerente_atual': gerente_nome})


# ================== GESTÃO DE UNIDADES ==================

class UnidadeListView(LoginRequiredMixin, ListView):
    model = Unidade
    template_name = "unidade_list.html"
    context_object_name = "unidades"

    def get_queryset(self):
        return Unidade.objects.all().order_by("nome")

class UnidadeCriarView(LoginRequiredMixin, CreateView):
    model = Unidade
    fields = ["nome", "ativa"]
    template_name = "unidade_form.html"
    success_url = reverse_lazy("gestao_a_vista:unidade_listar")

    def form_valid(self, form):
        messages.success(self.request, f'Unidade "{form.instance.nome}" criada com sucesso.')
        return super().form_valid(form)

class UnidadeEditarView(LoginRequiredMixin, UpdateView):
    model = Unidade
    fields = ["nome", "ativa"]
    template_name = "unidade_form.html"
    success_url = reverse_lazy("gestao_a_vista:unidade_listar")

    def form_valid(self, form):
        messages.success(self.request, f'Unidade "{form.instance.nome}" atualizada com sucesso.')
        return super().form_valid(form)

class UnidadeDeletarView(LoginRequiredMixin, DeleteView):
    model = Unidade
    template_name = "unidade_confirm_delete.html"
    success_url = reverse_lazy("gestao_a_vista:unidade_listar")

    def delete(self, request, *args, **kwargs):
        unidade = self.get_object()
        messages.success(request, f'Unidade "{unidade.nome}" excluída com sucesso.')
        return super().delete(request, *args, **kwargs)


# Adicione estes imports no topo se não os tiver:
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from .models import PrestadorServico

@login_required
@require_POST
def criar_prestador_ajax(request):
    try:
        data = json.loads(request.body)
        nome = data.get('nome')
        email = data.get('email')
        area_servico = data.get('area_servico')
        unidade_id = data.get('unidade_id') or None
        
        if not nome or not email or not area_servico:
            return JsonResponse({'success': False, 'message': 'Preencha todos os campos obrigatórios.'}, status=400)
            
        # Cria e salva o novo prestador
        PrestadorServico.objects.create(
            nome=nome,
            email=email,
            area_servico=area_servico,
            unidade_id=unidade_id,
            ativo=True
        )
        
        return JsonResponse({
            'success': True, 
            'message': 'Prestador cadastrado com sucesso!'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erro: {str(e)}'}, status=500)

import json
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required
# Lembre-se de importar o seu modelo: from .models import PrestadorServico

@login_required
def listar_prestadores_ajax(request):
    try:
        from .models import Unidade
        prestadores_qs = PrestadorServico.objects.select_related('unidade').all().order_by('nome')
        resultado = []
        for p in prestadores_qs:
            resultado.append({
                'id': p.id,
                'nome': p.nome,
                'email': p.email,
                'area_servico': p.area_servico,
                'unidade_id': p.unidade_id,
                'unidade_nome': p.unidade.nome if p.unidade else '',
            })
        return JsonResponse(resultado, safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@require_POST
def editar_prestador_ajax(request, pk):
    try:
        prestador = get_object_or_404(PrestadorServico, pk=pk)
        data = json.loads(request.body)
        
        prestador.nome = data.get('nome', prestador.nome)
        prestador.email = data.get('email', prestador.email)
        prestador.area_servico = data.get('area_servico', prestador.area_servico)
        unidade_id = data.get('unidade_id')
        prestador.unidade_id = unidade_id if unidade_id else None
        prestador.save()
        
        return JsonResponse({'success': True, 'message': 'Prestador atualizado com sucesso!'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erro: {str(e)}'}, status=500)

@login_required
@require_POST
def deletar_prestador_ajax(request, pk):
    try:
        prestador = get_object_or_404(PrestadorServico, pk=pk)
        prestador.delete()
        return JsonResponse({'success': True, 'message': 'Prestador excluído com sucesso!'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erro: {str(e)}'}, status=500)



# VIEW 1: Aba Exclusiva no Sistema
# VIEW 1: Aba Exclusiva no Sistema (Agora permitindo o Iframe na mesma origem)
@method_decorator(xframe_options_sameorigin, name='dispatch')
class PainelReincidenciasView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = "painel_reincidencias.html"

    def test_func(self):
        # Aqui você define quem acessa. Ex: 'is_superuser' ou por Grupo.
        return self.request.user.username.lower() == 'daniel' or self.request.user.is_superuser or self.request.user.role == 'administrador'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .models import ReincidenciaOcorrencia
        context['reincidencias_pendentes'] = ReincidenciaOcorrencia.objects.filter(status_aprovacao='pendente').order_by('-data_reincidencia')
        context['historico_reincidencias'] = ReincidenciaOcorrencia.objects.exclude(status_aprovacao='pendente').order_by('-data_avaliacao')
        return context

# VIEW 2: Link Direto do E-mail (Aprova pelo cookie/login)
class AprovarReincidenciaView(LoginRequiredMixin, View):
    def get(self, request, pk):
        from .models import ReincidenciaOcorrencia
        
        # O LoginRequiredMixin assegura que o sistema leia o Cookie do usuário
        # Se ele não tiver cookie ativo, o Django manda pra /login?next=/reincidencias/aprovar/1/
        reincidencia = get_object_or_404(ReincidenciaOcorrencia, pk=pk)
        
        if reincidencia.status_aprovacao != 'pendente':
            messages.info(request, "Esta reincidência já foi tratada anteriormente.")
            return redirect('gestao_a_vista:livro_ocorrencias') # Ou para o Dashboard principal
            
        # Aprova e registra QUEM aprovou usando os dados da sessão (request.user)
        reincidencia.status_aprovacao = 'aprovado'
        reincidencia.avaliado_por = request.user
        reincidencia.data_avaliacao = timezone.now()
        reincidencia.save()
        
        nome_aprovador = request.user.get_full_name() or request.user.username
        messages.success(request, f"Reincidência para o CR {reincidencia.cr_colaborador} (Item: {reincidencia.item_reincidente}) foi APROVADA por {nome_aprovador}.")
        
        return redirect('gestao_a_vista:painel_reincidencias') # Redireciona para o painel restrito ou outro que preferir
    

from django.http import JsonResponse
from django.utils import timezone
from datetime import timedelta
from django.core.mail import EmailMessage
from django.template.loader import render_to_string
from django.urls import reverse
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin

class GerarDadosTesteView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        # Proteção: Apenas Daniel ou Superusuários
        if not request.user.is_superuser and request.user.username.lower() != 'daniel':
            return JsonResponse({"success": False, "message": "Acesso negado."})
        
        from .models import OcorrenciaPlanoAcao, ReincidenciaOcorrencia, EmailGerenteOcorrencia
        
        # --- 1. ROTINA DE LIMPEZA DE SEGURANÇA (Apaga testes velhos) ---
        limite_tempo = timezone.now() - timedelta(hours=24)
        
        # Apaga reincidencias de teste criadas há mais de 24h
        ReincidenciaOcorrencia.objects.filter(
            item_reincidente__startswith="[SIMULAÇÃO]", 
            data_reincidencia__lte=limite_tempo
        ).delete()
        
        # Apaga planos de teste criados há mais de 24h
        OcorrenciaPlanoAcao.objects.filter(
            item_em_falta__startswith="[SIMULAÇÃO]", 
            data_criacao__lte=limite_tempo
        ).delete()

        # --- 2. BUSCAR OS DESTINATÁRIOS NO ADMIN ---
        # Pegar um Gerente ou Coordenador que tenha e-mail preenchido
        from .models import CustomUser
        
        gestor_teste = CustomUser.objects.filter(role__in=['gerente', 'coordenador']).exclude(email__exact='').first()
        if not gestor_teste:
            return JsonResponse({
                "success": False, 
                "message": "Nenhum Gerente ou Coordenador com E-mail configurado! Crie um em 'Controle de Acessos'."
            })

        # CR fictício para o teste, usando o que ele gerencia
        lista_crs = [c.strip() for c in gestor_teste.crs.split(',')] if gestor_teste.crs else []
        cr_teste = lista_crs[0] if lista_crs else "99999"
        
        # Coleta os e-mails de quem vai RECEBER
        destinatarios = [gestor_teste.email]

        if not destinatarios:
            return JsonResponse({"success": False, "message": f"Erro de validação de e-mails para o teste."})

        # --- 3. GERAR O DADO DE TESTE (Simulando o prazo estourado) ---
        # Criando o plano como se ele tivesse sido resolvido HÁ 26 HORAS
        plano_antigo = OcorrenciaPlanoAcao.objects.create(
            cr_colaborador=cr_teste, # Usa o CR real do Admin
            item_em_falta=f"[SIMULAÇÃO] Falta de EPI", 
            status="concluido", 
            criador_plano=request.user, 
            aprovador=request.user
        )
        
        # Hackeando a data no banco para simular o passado (passou de 24h)
        data_passado = timezone.now() - timedelta(hours=26)
        OcorrenciaPlanoAcao.objects.filter(id=plano_antigo.id).update(data_criacao=data_passado)
        
        # Nomes de quem vai avaliar
        nome_gerente = getattr(configuracao, 'nome_gerente', 'Gerente')
        nome_coordenador = getattr(configuracao, 'nome_coordenador', 'Coordenador')

        # Gera a Reincidência
        reincidencia = ReincidenciaOcorrencia.objects.create(
            plano_original=plano_antigo,
            cr_colaborador=plano_antigo.cr_colaborador,
            item_reincidente=plano_antigo.item_em_falta,
            coordenador=nome_coordenador,
            gerente=nome_gerente
        )
        
        # --- 4. DISPARAR O E-MAIL (Usando os destinatários do Admin) ---
        link_painel = request.build_absolute_uri(reverse('gestao_a_vista:painel_reincidencias'))
        
        # Aqui está a correção: url_foto agora é explicitamente None
        context = {
            'plano': plano_antigo,
            'cr': reincidencia.cr_colaborador,
            'item': reincidencia.item_reincidente,
            'coordenador': nome_coordenador,
            'gerente': nome_gerente,
            'url_foto': None, 
        }
        
        # Usa o template focado em auditoria
        html_content = render_to_string('emails/alerta_reincidencia.html', context)
        
        # Django vai usar o DEFAULT_FROM_EMAIL (notificacoes@example.com) para ENVIAR
        # E vai mandar para os destinatários listados no Admin
        email_msg = EmailMessage(
            subject=f"ALERTA DE REINCIDÊNCIA (TESTE) - {reincidencia.cr_colaborador}",
            body=html_content,
            to=destinatarios, 
        )
        email_msg.content_subtype = "html"
        
        try:

            reincidencia.save()
            return JsonResponse({
                "success": True, 
                "message": f"Simulação feita!"
            })
        except Exception as e:
            print(f"Erro {e}")
            return JsonResponse({"success": False, "message": f"Erro no SMTP do Google: {e}"})




class AcaoPlanoEmailView(LoginRequiredMixin, View):
    def get(self, request, pk, acao):
        from .models import OcorrenciaPlanoAcao
        plano = get_object_or_404(OcorrenciaPlanoAcao, pk=pk)

        # Segurança: Apenas quem tem cargo de liderança pode aprovar
        if request.user.role not in ['administrador', 'gerente', 'coordenador'] and request.user.username.lower() != 'daniel':
            messages.error(request, "Sem permissão para avaliar planos de ação.")
            return redirect('gestao_a_vista:torre_controle')

        # Evita dupla aprovação
        if plano.status != 'em_aprovacao':
            messages.info(request, "Este plano de ação já foi avaliado anteriormente.")
            return redirect('gestao_a_vista:torre_controle')

        if acao == 'aprovar':
            plano.status = 'compra_cadastrar'
            plano.aprovador = request.user
            plano.save()
            messages.success(request, f"Plano de ação do item '{plano.item_em_falta}' foi APROVADO com sucesso!")
            return redirect('gestao_a_vista:torre_controle')

        elif acao == 'rejeitar':
            # Renderiza a página para o usuário preencher a justificativa
            return render(request, 'plano_acao_rejeitar.html', {'plano': plano})

        return redirect('gestao_a_vista:torre_controle')

    def post(self, request, pk, acao):
        from .models import OcorrenciaPlanoAcao
        plano = get_object_or_404(OcorrenciaPlanoAcao, pk=pk)

        # Validação de segurança novamente no POST
        if request.user.role not in ['administrador', 'gerente', 'coordenador'] and request.user.username.lower() != 'daniel':
            messages.error(request, "Sem permissão para avaliar planos de ação.")
            return redirect('gestao_a_vista:torre_controle')

        if plano.status != 'em_aprovacao':
            messages.info(request, "Este plano de ação já foi avaliado.")
            return redirect('gestao_a_vista:torre_controle')

        if acao == 'rejeitar':
            justificativa = request.POST.get('justificativa', '').strip()
            
            # Trava: não deixa recusar se não houver texto
            if not justificativa:
                messages.error(request, "A justificativa é obrigatória para recusar o plano.")
                return render(request, 'plano_acao_rejeitar.html', {'plano': plano})

            # Rejeita oficialmente
            plano.status = 'rejeitado'
            plano.aprovador = request.user
            plano.justificativa_aprovacao = justificativa
            plano.save()
            
            messages.success(request, f"Plano de ação do item '{plano.item_em_falta}' foi RECUSADO.")
            return redirect('gestao_a_vista:torre_controle')

        return redirect('gestao_a_vista:torre_controle')
    

# --- IMPORTAÇÕES NECESSÁRIAS PARA AS VIEWS DE AUDITORIA ---
import json
from django.core.mail import EmailMessage
from django.template.loader import render_to_string
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.urls import reverse
from django.conf import settings
from django.utils import timezone
from .models import ReincidenciaOcorrencia

@login_required
@require_POST
def notificar_gerente_auditoria_ajax(request, pk):
    try:
        data = json.loads(request.body)
        email_gerente = data.get('email')
        
        if not email_gerente:
            return JsonResponse({'success': False, 'message': 'O e-mail do gerente é obrigatório.'}, status=400)
            
        reincidencia = ReincidenciaOcorrencia.objects.get(id=pk)
        plano = reincidencia.plano_original
        
        # O link agora vai direto para o Painel para que a resolução seja feita via auditoria
        link_painel = request.build_absolute_uri(reverse('gestao_a_vista:painel_reincidencias'))
        
        url_foto = None
        if hasattr(plano, 'foto') and plano.foto:
            url_foto = request.build_absolute_uri(plano.foto.url)
            
        context = {
            'plano': plano,
            'cr': reincidencia.cr_colaborador,
            'item': reincidencia.item_reincidente,
            'gerente': reincidencia.gerente or 'Gestão',
            'coordenador': reincidencia.coordenador or '',
            'url_foto': url_foto,
            'link_painel': link_painel
        }
        
        # APONTA PARA O TEMPLATE CORRETO DA REINCIDÊNCIA
        html_content = render_to_string('emails/alerta_reincidencia.html', context)
        assunto = f"ALERTA DE REINCIDÊNCIA - CR {reincidencia.cr_colaborador}"
        
        email_msg = EmailMessage(
            subject=assunto,
            body=html_content,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[email_gerente]
        )
        email_msg.content_subtype = "html"
        email_msg.send(fail_silently=False)
        
        reincidencia.email_enviado = True
        reincidencia.save()
        
        return JsonResponse({'success': True, 'message': 'E-mail de notificação enviado com sucesso!'})
        
    except ReincidenciaOcorrencia.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Ocorrência de reincidência não encontrada.'}, status=404)
    except Exception as e:
        import traceback
        print("ERRO AO ENVIAR NOTIFICAÇÃO DE AUDITORIA:\n", traceback.format_exc())
        return JsonResponse({'success': False, 'message': f'Erro ao processar: {str(e)}'}, status=500)


@login_required
@require_POST
def resolver_auditoria_ajax(request, pk):
    try:
        data = json.loads(request.body)
        justificativa = data.get('justificativa')
        
        if not justificativa:
            return JsonResponse({'success': False, 'message': 'A justificativa é obrigatória.'}, status=400)
            
        reincidencia = ReincidenciaOcorrencia.objects.get(id=pk)
        
        # Atualiza a reincidência como resolvida
        reincidencia.status_aprovacao = 'aprovado'
        reincidencia.justificativa_auditoria = justificativa
        reincidencia.data_aprovacao = timezone.now()
        reincidencia.save()
        
        return JsonResponse({'success': True, 'message': 'Auditoria resolvida e finalizada com sucesso!'})
        
    except ReincidenciaOcorrencia.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Ocorrência não encontrada.'}, status=404)
    except Exception as e:
        import traceback
        print("ERRO AO RESOLVER AUDITORIA:\n", traceback.format_exc())
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
@require_POST
def voltar_nc_auditoria_ajax(request, pk):
    try:
        data = json.loads(request.body)
        justificativa = data.get('justificativa')
        
        if not justificativa:
            return JsonResponse({'success': False, 'message': 'A justificativa é obrigatória.'}, status=400)
            
        reincidencia = ReincidenciaOcorrencia.objects.get(id=pk)
        
        reincidencia.status_aprovacao = 'rejeitado'
        reincidencia.justificativa_auditoria = f"Retornado para NCs. Justificativa: {justificativa}"
        reincidencia.data_aprovacao = timezone.now()
        reincidencia.save()

        plano = reincidencia.plano_original
        if plano and plano.status == 'auditoria':
            # Foi enviado manualmente para auditoria
            plano.status = 'excluido'
            plano.save()
            from .models import HistoricoPlanoAcao
            HistoricoPlanoAcao.objects.create(
                plano=plano, status_anterior='auditoria', status_novo='excluido',
                usuario=request.user, observacao=f"Auditoria cancelada, retornado para NC. Justificativa: {justificativa}"
            )
            
        return JsonResponse({'success': True, 'message': 'Ocorrência retornada para NCs com sucesso!'})
        
    except ReincidenciaOcorrencia.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Ocorrência não encontrada.'}, status=404)
    except Exception as e:
        import traceback
        print("ERRO AO VOLTAR NC:\n", traceback.format_exc())
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


# =============================================================================
# AUDITORIA — Subpágina da Torre de Controle
# =============================================================================

# Em Gestao_a_Vista/views.py
import hashlib
import json
import unicodedata as _ud
from collections import defaultdict
from django.utils import timezone
from django.db import connections
from django.http import JsonResponse
from django.views.generic import TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required

from .models import AuditoriaOcorrenciaStatus, Estrutura

@method_decorator(xframe_options_sameorigin, name='dispatch')
class AuditoriaOcorrenciasView(LoginRequiredMixin, TemplateView):
    template_name = "auditoria_ocorrencias.html"
    # O template base já faz o front-end via AJAX chamando carregar_mais_auditoria


class AuditorRondaView(LoginRequiredMixin, TemplateView):
    template_name = "auditor_ronda.html"

import hashlib
import requests # Certifique-se de que o requests está importado

def gerar_hash_real_url(url):
    """Lê a imagem real a partir da URL em pedaços e gera a impressão digital bit a bit (MD5)"""
    hasher = hashlib.md5()
    try:
        response = requests.get(url, stream=True, timeout=5)
        response.raise_for_status()
        for chunk in response.iter_content(chunk_size=4096):
            if chunk:
                hasher.update(chunk)
        return hasher.hexdigest()
    except Exception as e:
        print(f"Erro ao baixar/processar a imagem {url}: {e}")
        # Se falhar a leitura real, faz o hash da URL como fallback
        return hashlib.md5(url.encode('utf-8')).hexdigest()

def disparar_processamento_hashes_background(itens):
    import threading
    import requests
    import hashlib
    from datetime import timedelta
    from django.db import close_old_connections
    from .models import AuditoriaOcorrenciaStatus

    def _job():
        for ocorrencia_hash, url, cr_id, cr_real, colab, item_n, dt_oc, tarefa_id_v, numero_v, inicio_v in itens:
            try:
                # Baixa a imagem com timeout de 8 segundos
                response = requests.get(url, timeout=8)
                if response.status_code == 200:
                    real_hash = hashlib.md5(response.content).hexdigest()
                else:
                    real_hash = hashlib.md5(url.encode('utf-8')).hexdigest()
            except Exception as e:
                print(f"[Auditoria Background] Erro ao baixar imagem {url}: {e}")
                real_hash = hashlib.md5(url.encode('utf-8')).hexdigest()
            
            try:
                close_old_connections()
                
                # Para evitar erros de concorrência ou registro duplicado
                obj, created = AuditoriaOcorrenciaStatus.objects.get_or_create(ocorrencia_hash=ocorrencia_hash)
                obj.image_hash = real_hash
                obj.estrutura_id = cr_id
                obj.cr_nome = cr_real
                obj.colaborador = colab
                obj.item = item_n
                if dt_oc:
                    obj.data_ocorrencia = dt_oc
                obj.evidencia_url = url
                obj.tarefa_id = tarefa_id_v
                obj.numero = numero_v
                if inicio_v:
                    obj.inicio_real = inicio_v
                
                # Se houver outra ocorrência com o mesmo image_hash no mesmo CR/estrutura, marca ambas como coincidência
                if dt_oc:
                    limite_inf = dt_oc - timedelta(days=5)
                    limite_sup = dt_oc + timedelta(days=5)
                    outros = AuditoriaOcorrenciaStatus.objects.filter(
                        image_hash=real_hash,
                        estrutura_id=cr_id,
                        data_ocorrencia__gte=limite_inf,
                        data_ocorrencia__lte=limite_sup
                    ).exclude(ocorrencia_hash=ocorrencia_hash)
                else:
                    outros = AuditoriaOcorrenciaStatus.objects.filter(
                        image_hash=real_hash,
                        estrutura_id=cr_id
                    ).exclude(ocorrencia_hash=ocorrencia_hash)

                if outros.exists():
                    obj.is_coincidencia = True
                    outros.update(is_coincidencia=True)
                
                obj.save()
            except Exception as e:
                print(f"[Auditoria Background] Erro ao salvar status no banco: {e}")

    threading.Thread(target=_job, daemon=True).start()

@login_required
def carregar_mais_auditoria(request):
    from django.db import close_old_connections, connections
    from django.utils.timezone import is_naive, make_aware
    import unicodedata as _ud
    from collections import defaultdict
    import hashlib
    
    offset = int(request.GET.get('offset', 0))
    limit = int(request.GET.get('limit', 45))
    cr_filter = request.GET.get('cr', '').strip()
    item_filter = request.GET.get('item', '').strip().upper()
    data_filter = request.GET.get('data', '').strip()

    limit_tasks = int((offset + limit) * 1.5) + 150

    try:
        close_old_connections()
        db_conn_name = 'dw_vpn' if 'dw_vpn' in connections else ('readonly' if 'readonly' in connections else 'default')
        cursor = connections[db_conn_name].cursor()

        # 1. Buscar as estruturas filtradas para usar na query
        est_query = "SELECT id FROM dbo.estrutura WHERE descricao LIKE %s"
        est_params = ['% - GO - %']
        
        if cr_filter:
            est_query += " AND descricao LIKE %s"
            est_params.append(f"%{cr_filter}%")
            
        cursor.execute(est_query, est_params)
        valid_estrutura_ids = [str(row[0]) for row in cursor.fetchall()]

        if not valid_estrutura_ids:
            cursor.close()
            return JsonResponse({'success': True, 'todas': [], 'coincidencias': [], 'has_more': False})

        # 2. Executar a Query Otimizada com CTE
        cursor.execute("SET statement_timeout = '60000';")
        
        placeholders = ','.join(['%s'] * len(valid_estrutura_ids))
        tarefa_params = list(valid_estrutura_ids)
        
        filtro_data_sql = ""
        if data_filter:
            filtro_data_sql = " AND t.terminoreal >= %s AND t.terminoreal <= %s"
            tarefa_params.extend([f"{data_filter} 00:00:00", f"{data_filter} 23:59:59"])

        tarefa_params.append(limit_tasks)

        query_detalhes = f"""
            WITH tarefas_paginadas AS (
                SELECT t.id, t.estruturaid, t.terminoreal, t.finalizadoporhash, t.numero, t.inicio
                FROM dbo.tarefa t
                INNER JOIN dbo.checklist c ON c.id = t.checklistid
                WHERE c.id = '6687b862-10d0-4144-ae30-8bdc55f22ee3'
                  AND t.status = 85
                  AND t.terminoreal >= '2026-01-01 00:00:00'
                  AND t.terminoreal <= '2026-12-31 23:59:59'
                  AND t.estruturaid IN ({placeholders})
                  {filtro_data_sql}
                ORDER BY t.terminoreal DESC
                LIMIT %s
            )
            SELECT
                tp.id AS tarefa_id,
                tp.estruturaid,
                tp.terminoreal AS data_conclusao,
                tp.numero AS tarefa_numero,
                tp.inicio AS tarefa_inicio,
                ex.perguntadescricao AS item_nome,
                ex.conteudo AS status_resposta,
                r.nome AS colaborador,
                e.descricao AS estrutura_descricao
            FROM tarefas_paginadas tp
            LEFT JOIN dbo.execucao ex ON ex.tarefaid = tp.id
            LEFT JOIN dbo.estrutura e ON e.id = tp.estruturaid
            LEFT JOIN dbo.recurso r ON r.codigohash = tp.finalizadoporhash
            ORDER BY tp.terminoreal DESC
        """

        cursor.execute(query_detalhes, tarefa_params)
        columns = [col[0] for col in cursor.description]
        all_rows = [dict(zip(columns, row)) for row in cursor.fetchall()]
        cursor.close()

        # Configurações para normalização dos dados
        TOPICOS_IGNORADOS = ["HA NAO CONFORMIDADES", "HA VEICULOS NO LOCAL", "HA POSTO ARMADO", "HA OCORRENCIAS", "HOUVE ALGUMA NAO CONFORMIDADE"]

        def _norm(texto):
            if not texto: return ""
            return ''.join(c for c in _ud.normalize('NFD', str(texto)) if _ud.category(c) != 'Mn').upper()

        grupos = defaultdict(lambda: {
            'estrutura_id': None,
            'cr_nome': '',
            'colaborador': '',
            'data': '-',
            'data_obj': None,
            'item': '',
            'gerente': 'Não Atribuído',
            'relato': '',
            'fotos': [],
            'tarefa_id': None,
            'numero': None,
            'inicio_real_obj': None,
        })

        for row in all_rows:
            item_nome = (row.get('item_nome') or '').strip()
            if not item_nome:
                continue

            if ' - ' in item_nome:
                item_nome = item_nome.split(' - ')[0].strip()
                
            item_norm = _norm(item_nome)
            if any(tp in item_norm for tp in TOPICOS_IGNORADOS):
                continue

            resp = (row.get('status_resposta') or '').strip()
            if not resp:
                continue

            tarefa_id = str(row.get('tarefa_id') or '')
            key = f"{tarefa_id}|{item_nome}"
            
            g = grupos[key]
            if not g['estrutura_id']:
                g['estrutura_id'] = str(row.get('estruturaid') or '')
                g['cr_nome'] = row.get('estrutura_descricao') or ''
                g['colaborador'] = row.get('colaborador') or ''
                g['item'] = item_nome
                g['tarefa_id'] = tarefa_id
                g['numero'] = row.get('tarefa_numero')
                g['inicio_real_obj'] = row.get('tarefa_inicio')

                g['data_obj'] = row.get('data_conclusao')
                ds = row.get('data_conclusao')
                if ds:
                    try:
                        g['data'] = ds.strftime("%d/%m/%Y - %H:%M")
                    except:
                        ds_str = str(ds)
                        g['data'] = f"{ds_str[8:10]}/{ds_str[5:7]}/{ds_str[0:4]} {ds_str[11:16]}" if len(ds_str) >= 16 else ds_str
            
            if resp.lower().startswith('http'):
                if resp not in g['fotos']:
                    g['fotos'].append(resp)
            else:
                if not g['relato']:
                    g['relato'] = resp

        valid_groups = []
        for g in grupos.values():
            relato_texto = g['relato'].upper()
            if 'CONFORME' not in relato_texto:
                continue
            
            if item_filter and item_filter not in g['item'].upper():
                continue
                
            valid_groups.append(g)

        now = timezone.now()
        def safe_date(obj):
            d = obj.get('data_obj')
            if not d: return now
            if is_naive(d):
                return make_aware(d)
            return d

        valid_groups.sort(key=safe_date, reverse=True)
        lista_paginada = valid_groups[offset : offset + limit]

        status_banco = {s.ocorrencia_hash: s for s in AuditoriaOcorrenciaStatus.objects.all()}
        hash_cache_memoria = {}
        est_ids = list(set([g['estrutura_id'] for g in lista_paginada if g['estrutura_id']]))
        mapa_gerentes = {}
        
        if est_ids:
            estruturas_local = Estrutura.objects.using('default').filter(id__in=est_ids).values('id', 'gc')
            mapa_gerentes = {str(e['id']): e['gc'] for e in estruturas_local}

        ocorrencias_finais = []
        coincidencias = []
        cr_foto_agrupamento = defaultdict(lambda: defaultdict(list))
        urls_para_processar = []

        for g in lista_paginada:
            if not g['estrutura_id']: continue
            nome_gerente = mapa_gerentes.get(g['estrutura_id'])
            g['gerente'] = nome_gerente if nome_gerente else "Sem Gerente"
            
            key = f"{g['tarefa_id']}|{g['item']}"
            g['ocorrencia_hash'] = hashlib.md5(key.encode('utf-8')).hexdigest()
            status_obj = status_banco.get(g['ocorrencia_hash'])
            
            is_realmente_auditado = False
            if status_obj and status_obj.auditado_por_id:
                is_realmente_auditado = True
                if status_obj.auditado_em and (now - status_obj.auditado_em).days > 30:
                    continue
            
            if status_obj and status_obj.tratada and status_obj.tratada_em:
                if (now - status_obj.tratada_em).days > 2:
                    continue
            
            g['is_auditado'] = is_realmente_auditado
            g['is_coincidencia'] = status_obj.is_coincidencia if status_obj else False
            g['is_tratada'] = status_obj.tratada if status_obj else False
            g['evidencia_url'] = g['fotos'][0] if g['fotos'] else None
            g['datas_coincidentes'] = []
            
            data_obj_temp = g.get('data_obj')
            if data_obj_temp and is_naive(data_obj_temp):
                data_obj_temp = make_aware(data_obj_temp)

            if 'data_obj' in g:
                del g['data_obj']
            g['data_obj'] = data_obj_temp

            inicio_obj_temp = g.get('inicio_real_obj')
            if inicio_obj_temp and is_naive(inicio_obj_temp):
                inicio_obj_temp = make_aware(inicio_obj_temp)
            g['inicio_real_obj'] = inicio_obj_temp
            g['inicio_real'] = inicio_obj_temp.strftime("%d/%m/%Y - %H:%M") if inicio_obj_temp else '-'

            if g['evidencia_url']:
                url = g['evidencia_url']
                if status_obj and status_obj.image_hash and len(status_obj.image_hash) == 32:
                    foto_hash = status_obj.image_hash
                else:
                    if url not in hash_cache_memoria:
                        # Usar hash temporário baseado no URL para evitar lentidão e timeout na requisição principal
                        foto_hash = hashlib.md5(url.encode('utf-8')).hexdigest()
                        hash_cache_memoria[url] = foto_hash
                    else:
                        foto_hash = hash_cache_memoria[url]
                    
                    urls_para_processar.append((
                        g['ocorrencia_hash'], url, g['estrutura_id'],
                        g['cr_nome'], g['colaborador'], g['item'], g['data_obj'],
                        g.get('tarefa_id'), g.get('numero'), g.get('inicio_real_obj')
                    ))
                
                cr_foto_agrupamento[g['estrutura_id']][foto_hash].append(g)

        from datetime import timedelta
        for cr_id, foto_agrupamento in cr_foto_agrupamento.items():
            for foto_hash, lista_oc in foto_agrupamento.items():
                hashes_no_lote = [x['ocorrencia_hash'] for x in lista_oc]
                
                data_ref = lista_oc[0].get('data_obj')
                if data_ref:
                    limite_inf = data_ref - timedelta(days=5)
                    limite_sup = data_ref + timedelta(days=5)
                    historico_matches = AuditoriaOcorrenciaStatus.objects.filter(
                        image_hash=foto_hash,
                        estrutura_id=cr_id,
                        data_ocorrencia__gte=limite_inf,
                        data_ocorrencia__lte=limite_sup
                    ).exclude(ocorrencia_hash__in=hashes_no_lote)
                else:
                    historico_matches = AuditoriaOcorrenciaStatus.objects.filter(
                        image_hash=foto_hash,
                        estrutura_id=cr_id
                    ).exclude(ocorrencia_hash__in=hashes_no_lote)

                tem_historico = historico_matches.exists()

                for oc in lista_oc:
                    hash_val = oc['ocorrencia_hash']
                    existentes = AuditoriaOcorrenciaStatus.objects.filter(ocorrencia_hash=hash_val)
                    if existentes.exists():
                        obj = existentes.first()
                        created = False
                        if existentes.count() > 1:
                            dados = {
                                'estrutura_id': obj.estrutura_id,
                                'auditado_em': obj.auditado_em,
                                'auditado_por': obj.auditado_por,
                                'image_hash': obj.image_hash,
                                'is_coincidencia': obj.is_coincidencia,
                                'tratada': obj.tratada,
                                'tratada_em': obj.tratada_em,
                                'tratada_por': obj.tratada_por,
                                'data_ocorrencia': obj.data_ocorrencia,
                                'colaborador': obj.colaborador,
                                'item': obj.item,
                                'cr_nome': obj.cr_nome,
                                'evidencia_url': obj.evidencia_url,
                                'tarefa_id': obj.tarefa_id,
                                'numero': obj.numero,
                                'inicio_real': obj.inicio_real,
                            }
                            AuditoriaOcorrenciaStatus.objects.filter(ocorrencia_hash=hash_val).delete()
                            obj = AuditoriaOcorrenciaStatus.objects.create(ocorrencia_hash=hash_val, **dados)
                    else:
                        try:
                            obj = AuditoriaOcorrenciaStatus.objects.create(ocorrencia_hash=hash_val)
                            created = True
                        except IntegrityError:
                            # Outra requisição concorrente já criou esse hash primeiro
                            obj = AuditoriaOcorrenciaStatus.objects.get(ocorrencia_hash=hash_val)
                            created = False
                    update_needed = False
                    
                    if not obj.image_hash or not obj.estrutura_id:
                        obj.image_hash = foto_hash
                        obj.estrutura_id = cr_id
                        update_needed = True

                    cr_real = oc.get('cr_nome') or oc.get('cr') or "Nome Indisponível"
                    
                    if not obj.data_ocorrencia and oc.get('data_obj'):
                        obj.data_ocorrencia = oc.get('data_obj')
                        update_needed = True
                        
                    if not obj.colaborador and oc.get('colaborador'):
                        obj.colaborador = oc.get('colaborador')
                        update_needed = True
                        
                    if not obj.item and oc.get('item'):
                        obj.item = oc.get('item')
                        update_needed = True

                    if not obj.cr_nome or len(str(obj.cr_nome)) == 36:
                        obj.cr_nome = cr_real
                        update_needed = True
                        
                    if oc.get('evidencia_url') and not obj.evidencia_url:
                        obj.evidencia_url = oc.get('evidencia_url')
                        update_needed = True

                    if not obj.tarefa_id and oc.get('tarefa_id'):
                        obj.tarefa_id = oc.get('tarefa_id')
                        update_needed = True

                    if not obj.numero and oc.get('numero'):
                        obj.numero = oc.get('numero')
                        update_needed = True

                    if not obj.inicio_real and oc.get('inicio_real_obj'):
                        obj.inicio_real = oc.get('inicio_real_obj')
                        update_needed = True

                    if update_needed:
                        obj.save()

                    datas_duplicadas = []
                    for outra_oc in lista_oc:
                        if outra_oc['ocorrencia_hash'] != oc['ocorrencia_hash']:
                            d_str = outra_oc.get('data_obj').strftime("%d/%m/%Y - %H:%M") if outra_oc.get('data_obj') else outra_oc['data']
                            cr_str = outra_oc.get('cr_nome') or cr_real
                            colab_str = outra_oc.get('colaborador') or 'Não identificado'
                            item_str = outra_oc.get('item') or 'Sem descrição'
                            info = f"CR: {cr_str}<br>{d_str}<br>Lançado por: {colab_str}<br>Item: {item_str}"
                            if info not in datas_duplicadas:
                                datas_duplicadas.append(info)

                    if tem_historico:
                        count = len(datas_duplicadas)
                        for match in historico_matches:
                            if count >= 9: break
                            
                            # Força a buscar o nome real caso o banco histórico tenha guardado só o ID
                            match_cr_str = match.cr_nome
                            if not match_cr_str or len(str(match_cr_str)) == 36:
                                try:
                                    est = Estrutura.objects.using('default').filter(id=match.estrutura_id).first()
                                    match_cr_str = est.descricao if est else "Nome Indisponível"
                                except:
                                    match_cr_str = "Nome Indisponível"
                                    
                            d_str = match.data_ocorrencia.strftime("%d/%m/%Y - %H:%M") if match.data_ocorrencia else "Data Desconhecida"
                            colab_str = match.colaborador or 'Não identificado'
                            item_str = match.item or 'Sem descrição'
                            
                            info = f"CR: {match_cr_str}<br>{d_str}<br>Lançado por: {colab_str}<br>Item: {item_str}"
                            
                            if info not in datas_duplicadas:
                                datas_duplicadas.append(info)
                            count += 1

                    is_duplicada_agora = len(lista_oc) > 1 or tem_historico
                    if is_duplicada_agora or obj.is_coincidencia:
                        oc['is_coincidencia'] = True
                        oc['datas_coincidentes'] = datas_duplicadas
                        if oc not in coincidencias:
                            coincidencias.append(oc)
                        if not obj.is_coincidencia:
                            obj.is_coincidencia = True
                            obj.save()
                    
                    if 'data_obj' in oc:
                        del oc['data_obj']
                    if 'inicio_real_obj' in oc:
                        del oc['inicio_real_obj']

        # CARREGAMENTO INSTANTÂNEO DE HISTÓRICO
        if offset == 0 and not data_filter:
            busca_cr = request.GET.get('cr', '').strip()
            busca_item = request.GET.get('item', '').strip()
            busca_data = request.GET.get('data', '').strip()
            
            from django.db.models import Q
            salvas_db = AuditoriaOcorrenciaStatus.objects.filter(is_coincidencia=True, tratada=False).order_by('-data_ocorrencia')
            if busca_cr:
                salvas_db = salvas_db.filter(Q(cr_nome__icontains=busca_cr) | Q(estrutura_id__icontains=busca_cr))
            if busca_item:
                salvas_db = salvas_db.filter(item__icontains=busca_item)
            if busca_data:
                salvas_db = salvas_db.filter(data_ocorrencia__date=busca_data)

            for obj_salvo in salvas_db:
                ja_existe = any(x['ocorrencia_hash'] == obj_salvo.ocorrencia_hash for x in coincidencias)
                if not ja_existe:
                    gemas = AuditoriaOcorrenciaStatus.objects.filter(image_hash=obj_salvo.image_hash).exclude(ocorrencia_hash=obj_salvo.ocorrencia_hash)[:9]
                    descricoes_gemeas = []
                    for gm in gemas:
                        d_str = gm.data_ocorrencia.strftime("%d/%m/%Y - %H:%M") if gm.data_ocorrencia else "Data Desconhecida"
                        colab_str = gm.colaborador or "Não identificado"
                        item_str = gm.item or "Sem descrição"
                        
                        cr_str = gm.cr_nome
                        if not cr_str or len(str(cr_str)) == 36:
                            try:
                                est = Estrutura.objects.using('default').filter(id=gm.estrutura_id).first()
                                cr_str = est.descricao if est else "Nome Indisponível"
                            except:
                                cr_str = "Nome Indisponível"
                                
                        descricoes_gemeas.append(f"CR: {cr_str}<br>{d_str}<br>Lançado por: {colab_str}<br>Item: {item_str}")
                    
                    data_str_principal = obj_salvo.data_ocorrencia.strftime("%d/%m/%Y - %H:%M") if obj_salvo.data_ocorrencia else "Data Desconhecida"
                    
                    cr_principal = obj_salvo.cr_nome
                    if not cr_principal or len(str(cr_principal)) == 36:
                        try:
                            est = Estrutura.objects.using('default').filter(id=obj_salvo.estrutura_id).first()
                            cr_principal = est.descricao if est else "Nome Indisponível"
                        except:
                            cr_principal = "Nome Indisponível"
                    
                    coincidencias.append({
                        'ocorrencia_hash': obj_salvo.ocorrencia_hash,
                        'is_auditado': False,
                        'is_tratada': obj_salvo.tratada,
                        'is_coincidencia': True,
                        'cr_nome': cr_principal,
                        'data': data_str_principal,
                        'colaborador': obj_salvo.colaborador or "Não identificado",
                        'gerente': "Gestão",
                        'item': obj_salvo.item or "Sem descrição",
                        'tarefa_id': obj_salvo.tarefa_id,
                        'numero': obj_salvo.numero,
                        'inicio_real': obj_salvo.inicio_real.strftime("%d/%m/%Y - %H:%M") if obj_salvo.inicio_real else '-',
                        'relato': "Ocorrência vinda do banco de dados tratada como coincidência.",
                        'fotos': [obj_salvo.evidencia_url] if obj_salvo.evidencia_url else [],
                        'evidencia_url': obj_salvo.evidencia_url,
                        'datas_coincidentes': descricoes_gemeas
                    })

        for g in lista_paginada:
            if not g.get('is_coincidencia'):
                ocorrencias_finais.append(g)

        # Disparar o download e processamento de hashes das fotos em background (assíncrono)
        if urls_para_processar:
            disparar_processamento_hashes_background(urls_para_processar)

        return JsonResponse({
            'success': True,
            'todas': ocorrencias_finais,
            'coincidencias': coincidencias,
            'has_more': (offset + limit) < len(valid_groups)
        })

    except Exception as e:
        import traceback
        print("Erro em carregar_mais_auditoria:", e)
        print(traceback.format_exc())
        return JsonResponse({'success': False, 'error': str(e), 'message': str(e)})

        

def is_verificacao_geral_rodando(limit):
    try:
        return int(limit) < 40
    except:
        return False


@login_required
@require_POST
def marcar_auditoria_ajax(request):
    try:
        data = json.loads(request.body)
        hash_id = data.get('ocorrencia_hash')
        
        existentes = AuditoriaOcorrenciaStatus.objects.filter(ocorrencia_hash=hash_id)
        if existentes.exists():
            status_obj = existentes.first()
            if existentes.count() > 1:
                dados = {
                    'estrutura_id': status_obj.estrutura_id,
                    'auditado_em': status_obj.auditado_em,
                    'auditado_por': status_obj.auditado_por,
                    'image_hash': status_obj.image_hash,
                    'is_coincidencia': status_obj.is_coincidencia,
                    'tratada': status_obj.tratada,
                    'tratada_em': status_obj.tratada_em,
                    'tratada_por': status_obj.tratada_por,
                    'data_ocorrencia': status_obj.data_ocorrencia,
                    'colaborador': status_obj.colaborador,
                    'item': status_obj.item,
                    'cr_nome': status_obj.cr_nome,
                    'evidencia_url': status_obj.evidencia_url,
                    'tarefa_id': status_obj.tarefa_id,
                    'numero': status_obj.numero,
                    'inicio_real': status_obj.inicio_real,
                }
                AuditoriaOcorrenciaStatus.objects.filter(ocorrencia_hash=hash_id).delete()
                status_obj = AuditoriaOcorrenciaStatus.objects.create(ocorrencia_hash=hash_id, **dados)
        else:
            try:
                status_obj = AuditoriaOcorrenciaStatus.objects.create(ocorrencia_hash=hash_id)
            except IntegrityError:
                # Outra requisição concorrente já criou esse hash primeiro
                status_obj = AuditoriaOcorrenciaStatus.objects.get(ocorrencia_hash=hash_id)
        
        # AQUI RESOLVE O ERRO 400: Agora ele baseia o toggle pelo Usuário e trata Nulls de forma segura
        if status_obj.auditado_por_id:
            status_obj.auditado_por = None
            status_obj.auditado_em = None
        else:
            status_obj.auditado_por = request.user
            status_obj.auditado_em = timezone.now()
            
        status_obj.save()

        return JsonResponse({'success': True, 'is_auditado': bool(status_obj.auditado_por_id)})
    except Exception as e:
        import traceback
        print("Erro Marcar Auditoria:", traceback.format_exc())
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@login_required
@require_POST
def tratar_coincidencia_ajax(request):
    try:
        data = json.loads(request.body)
        hash_id = data.get('ocorrencia_hash')
        
        existentes = AuditoriaOcorrenciaStatus.objects.filter(ocorrencia_hash=hash_id)
        if existentes.exists():
            status_obj = existentes.first()
            if existentes.count() > 1:
                dados = {
                    'estrutura_id': status_obj.estrutura_id,
                    'auditado_em': status_obj.auditado_em,
                    'auditado_por': status_obj.auditado_por,
                    'image_hash': status_obj.image_hash,
                    'is_coincidencia': status_obj.is_coincidencia,
                    'tratada': status_obj.tratada,
                    'tratada_em': status_obj.tratada_em,
                    'tratada_por': status_obj.tratada_por,
                    'data_ocorrencia': status_obj.data_ocorrencia,
                    'colaborador': status_obj.colaborador,
                    'item': status_obj.item,
                    'cr_nome': status_obj.cr_nome,
                    'evidencia_url': status_obj.evidencia_url,
                    'tarefa_id': status_obj.tarefa_id,
                    'numero': status_obj.numero,
                    'inicio_real': status_obj.inicio_real,
                }
                AuditoriaOcorrenciaStatus.objects.filter(ocorrencia_hash=hash_id).delete()
                status_obj = AuditoriaOcorrenciaStatus.objects.create(ocorrencia_hash=hash_id, **dados)
        else:
            try:
                status_obj = AuditoriaOcorrenciaStatus.objects.create(ocorrencia_hash=hash_id)
            except IntegrityError:
                # Outra requisição concorrente já criou esse hash primeiro
                status_obj = AuditoriaOcorrenciaStatus.objects.get(ocorrencia_hash=hash_id)
        status_obj.tratada = True
        status_obj.tratada_em = timezone.now()
        status_obj.tratada_por = request.user
        status_obj.save()
        
        return JsonResponse({'success': True})
    except Exception as e:
        import traceback
        print("Erro Tratar Coincidencia:", traceback.format_exc())
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


import hashlib
import requests
from django.db import close_old_connections

def gerar_hash_imagem(url):
    """
    Faz o download da imagem e gera o hash.
    Timeout aumentado para 30 segundos para garantir a leitura de fotos pesadas.
    """
    try:
        close_old_connections() # Evita o erro de "SSL connection has been closed"
        
        # Timeout de 30 segundos (ou até mais) para esperar imagens grandes
        response = requests.get(url, timeout=30)
        if response.status_code == 200:
            return hashlib.md5(response.content).hexdigest()
        else:
            print(f"Erro ao baixar imagem. Status: {response.status_code} - URL: {url}")
            return None
    except Exception as e:
        print(f"Erro ao baixar a imagem (mesmo com 30s de espera): {e}")
        return None

def baixar_qr_code_sala(request, sala_id):
    from django.shortcuts import get_object_or_404
    from django.http import HttpResponse
    from django.urls import reverse
    import qrcode
    from io import BytesIO
    from .models import GestaoSala
    
    sala = get_object_or_404(GestaoSala, id_sala=sala_id)
    
    # URL to the reserva_salas page with ?sala=<sala_id> auto-select param
    path_url = reverse("gestao_a_vista:reserva_salas", kwargs={"unidade_slug": sala.unidade.slug})
    url_reserva = request.build_absolute_uri(path_url) + f"?sala={sala.id_sala}"
    
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(url_reserva)
    qr.make(fit=True)

    img = qr.make_image(fill_color="black", back_color="white")
    
    buffer = BytesIO()
    img.save(buffer, format="PNG")
    
    response = HttpResponse(buffer.getvalue(), content_type="image/png")
    response['Content-Disposition'] = f'attachment; filename="qrcode_sala_{sala.nome}.png"'
    response['Access-Control-Allow-Origin'] = '*'
    return response

@login_required
@require_GET
def retrospectiva_ocorrencias_api(request):
    """
    API para alimentar o Dashboard de Retrospectiva do Livro de Ocorrências.
    Acesso restrito a administradores.
    """
    if request.user.role != 'administrador' and not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'Acesso negado'}, status=403)

    from django.utils import timezone
    from datetime import timedelta
    from django.db.models import Count, F
    from collections import defaultdict
    from .models import OcorrenciaPlanoAcao, ReincidenciaOcorrencia, HistoricoPlanoAcao, CustomUser

    agora = timezone.now()
    uma_semana_atras = agora - timedelta(days=7)

    # 1. Total Concluídas e Rejeitadas
    total_concluidas = OcorrenciaPlanoAcao.objects.filter(status='concluido').count()
    total_auditoria = ReincidenciaOcorrencia.objects.count()

    # 2. Tarefas Paradas
    status_parados = ['em_aprovacao', 'compra_cadastrar', 'compra_pedido', 'compra_entregue']
    planos_parados = OcorrenciaPlanoAcao.objects.filter(status__in=status_parados).select_related('criador_plano')

    tarefas_paradas_por_setor = defaultdict(int)
    tarefas_paradas_cr = defaultdict(int)
    
    tarefas_alerta = []
    total_paradas = 0
    soma_dias_parados = 0

    # Dicionário mapeando os IDs de plano para a sua última alteração de status
    # Usamos distinct('plano') + order_by('plano', '-data_alteracao') no Postgres,
    # mas para compatibilidade usamos uma iteração
    ultimos_historicos = {}
    for h in HistoricoPlanoAcao.objects.filter(plano__status__in=status_parados).order_by('plano_id', 'data_alteracao'):
        ultimos_historicos[h.plano_id] = h

    for plano in planos_parados:
        total_paradas += 1
        tarefas_paradas_por_setor[plano.get_status_display()] += 1
        tarefas_paradas_cr[plano.cr_colaborador] += 1

        ultimo_hist = ultimos_historicos.get(plano.id)
        if ultimo_hist:
            data_base = ultimo_hist.data_alteracao
        else:
            data_base = plano.data_criacao

        dias_parado = (agora - data_base).days
        soma_dias_parados += dias_parado

        # Regra de Alerta: > 5 dias normal, > 10 dias regulatório
        # Exceções: Compras (pedido realizado) demora naturalmente. Vamos alertar as demais.
        if plano.status in ['em_aprovacao', 'compra_cadastrar', 'compra_entregue']:
            limite = 10 if getattr(plano, 'is_regulatory', False) else 5
            if dias_parado > limite:
                tarefas_alerta.append({
                    'cr': plano.cr_colaborador,
                    'item': plano.item_em_falta,
                    'setor': plano.get_status_display(),
                    'dias': dias_parado,
                    'is_reg': getattr(plano, 'is_regulatory', False)
                })

    tempo_medio_parado = round(soma_dias_parados / total_paradas, 1) if total_paradas > 0 else 0

    # 3. Gráfico Semanal de Permanência (Últimos 7 dias)
    # Qual foi a frequência de dias parados ANTES de movimentar?
    movimentacoes_semana = HistoricoPlanoAcao.objects.filter(data_alteracao__gte=uma_semana_atras).order_by('plano_id', 'data_alteracao')
    
    historico_por_plano = defaultdict(list)
    # Pega mais histórico para calcular a diferença correta
    historicos_geral = HistoricoPlanoAcao.objects.filter(
        plano__id__in=movimentacoes_semana.values_list('plano_id', flat=True)
    ).order_by('plano_id', 'data_alteracao')

    for h in historicos_geral:
        historico_por_plano[h.plano_id].append(h)

    frequencia_dias = defaultdict(int)
    for plano_id, hists in historico_por_plano.items():
        for i in range(1, len(hists)):
            # A movimentação ocorreu na última semana?
            if hists[i].data_alteracao >= uma_semana_atras:
                dias_no_status_anterior = (hists[i].data_alteracao - hists[i-1].data_alteracao).days
                frequencia_dias[f"{dias_no_status_anterior} dias"] += 1

    # 4. Rankings de Usuários
    # Quem mais cria planos
    top_criadores = list(OcorrenciaPlanoAcao.objects.values(nome=F('criador_plano__name')).annotate(total=Count('id')).order_by('-total')[:5])
    
    # Quem mais rejeita planos
    top_rejeitadores = list(HistoricoPlanoAcao.objects.filter(status_novo='rejeitado').values(nome=F('usuario__name')).annotate(total=Count('id')).order_by('-total')[:5])
    
    # Quem mais teve planos rejeitados
    top_rejeitados = list(OcorrenciaPlanoAcao.objects.filter(status='rejeitado').values(nome=F('criador_plano__name')).annotate(total=Count('id')).order_by('-total')[:5])

    # 5. Feed de Movimentações
    ultimas_mov = HistoricoPlanoAcao.objects.select_related('usuario', 'plano').order_by('-data_alteracao')[:20]
    feed = []
    for m in ultimas_mov:
        nome_user = m.usuario.name if m.usuario and m.usuario.name else (m.usuario.username if m.usuario else "Sistema")
        feed.append({
            'usuario': nome_user,
            'acao': m.get_status_novo_display() if hasattr(m, 'get_status_novo_display') else m.status_novo,
            'item': m.plano.item_em_falta,
            'cr': m.plano.cr_colaborador,
            'data': m.data_alteracao.strftime("%d/%m/%Y %H:%M"),
            'observacao': m.observacao
        })

    # Ordenar dados para gráficos
    tarefas_alerta = sorted(tarefas_alerta, key=lambda x: x['dias'], reverse=True)
    
    top_crs = sorted(tarefas_paradas_cr.items(), key=lambda x: x[1], reverse=True)[:5]
    top_crs_labels = [c[0] for c in top_crs]
    top_crs_values = [c[1] for c in top_crs]

    return JsonResponse({
        'kpis': {
            'total_concluidas': total_concluidas,
            'total_auditoria': total_auditoria,
            'total_paradas': total_paradas,
            'tempo_medio_parado': tempo_medio_parado,
        },
        'alertas': tarefas_alerta,
        'graficos': {
            'paradas_por_setor': {
                'labels': list(tarefas_paradas_por_setor.keys()),
                'values': list(tarefas_paradas_por_setor.values()),
            },
            'top_crs': {
                'labels': top_crs_labels,
                'values': top_crs_values,
            },
            'frequencia_permanencia': {
                'labels': list(frequencia_dias.keys()),
                'values': list(frequencia_dias.values()),
            },
            'top_criadores': [c for c in top_criadores if c['nome']],
            'top_rejeitadores': [c for c in top_rejeitadores if c['nome']],
            'top_rejeitados': [c for c in top_rejeitados if c['nome']],
        },
        'feed': feed
    })


import csv
from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib import colors

@login_required
@require_GET
def exportar_retrospectiva_excel(request):
    """
    Exporta a lista de tarefas paradas e alertas para Excel (formato CSV compatível).
    Acesso restrito a administradores.
    """
    if request.user.role != 'administrador' and not request.user.is_superuser:
        return HttpResponse('Acesso negado', status=403)

    from django.utils import timezone
    from collections import defaultdict
    from .models import OcorrenciaPlanoAcao, HistoricoPlanoAcao

    agora = timezone.now()
    status_parados = ['em_aprovacao', 'compra_cadastrar', 'compra_pedido', 'compra_entregue']
    planos_parados = OcorrenciaPlanoAcao.objects.filter(status__in=status_parados).select_related('criador_plano')

    ultimos_historicos = {}
    for h in HistoricoPlanoAcao.objects.filter(plano__status__in=status_parados).order_by('plano_id', 'data_alteracao'):
        ultimos_historicos[h.plano_id] = h

    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="Retrospectiva_Ocorrencias.csv"'
    
    # Adiciona BOM para o Excel abrir com a codificação correta
    response.write('\ufeff')
    writer = csv.writer(response, delimiter=';')
    
    writer.writerow(['RELATORIO DE TAREFAS PARADAS - LIVRO DE OCORRENCIAS'])
    writer.writerow(['CR', 'Item em Falta', 'Setor Atual', 'Dias Parado', 'É Regulatório', 'Criador do Plano'])
    
    for plano in planos_parados:
        ultimo_hist = ultimos_historicos.get(plano.id)
        data_base = ultimo_hist.data_alteracao if ultimo_hist else plano.data_criacao
        dias_parado = (agora - data_base).days
        is_reg = 'Sim' if getattr(plano, 'is_regulatory', False) else 'Não'
        criador = plano.criador_plano.name if plano.criador_plano else 'Desconhecido'

        writer.writerow([
            plano.cr_colaborador,
            plano.item_em_falta,
            plano.get_status_display(),
            dias_parado,
            is_reg,
            criador
        ])

    return response

@login_required
@require_GET
def exportar_retrospectiva_pdf(request):
    """
    Gera um relatório bonito em PDF usando ReportLab.
    Acesso restrito a administradores.
    """
    if request.user.role != 'administrador' and not request.user.is_superuser:
        return HttpResponse('Acesso negado', status=403)

    from django.utils import timezone
    from collections import defaultdict
    from .models import OcorrenciaPlanoAcao, ReincidenciaOcorrencia, HistoricoPlanoAcao
    from reportlab.lib.units import inch

    agora = timezone.now()
    
    # Calculando os mesmos KPIs da API
    total_concluidas = OcorrenciaPlanoAcao.objects.filter(status='concluido').count()
    total_auditoria = ReincidenciaOcorrencia.objects.count()
    
    status_parados = ['em_aprovacao', 'compra_cadastrar', 'compra_pedido', 'compra_entregue']
    planos_parados = OcorrenciaPlanoAcao.objects.filter(status__in=status_parados)
    
    ultimos_historicos = {}
    for h in HistoricoPlanoAcao.objects.filter(plano__status__in=status_parados).order_by('plano_id', 'data_alteracao'):
        ultimos_historicos[h.plano_id] = h

    tarefas_alerta = []
    total_paradas = planos_parados.count()
    soma_dias_parados = 0

    for plano in planos_parados:
        ultimo_hist = ultimos_historicos.get(plano.id)
        data_base = ultimo_hist.data_alteracao if ultimo_hist else plano.data_criacao
        dias_parado = (agora - data_base).days
        soma_dias_parados += dias_parado

        if plano.status in ['em_aprovacao', 'compra_cadastrar', 'compra_entregue']:
            limite = 10 if getattr(plano, 'is_regulatory', False) else 5
            if dias_parado > limite:
                tarefas_alerta.append({
                    'cr': plano.cr_colaborador,
                    'item': plano.item_em_falta,
                    'setor': plano.get_status_display(),
                    'dias': dias_parado,
                    'reg': 'Sim' if getattr(plano, 'is_regulatory', False) else 'Não'
                })

    tarefas_alerta = sorted(tarefas_alerta, key=lambda x: x['dias'], reverse=True)
    tempo_medio = round(soma_dias_parados / total_paradas, 1) if total_paradas > 0 else 0

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="Relatorio_Retrospectiva.pdf"'

    c = canvas.Canvas(response, pagesize=A4)
    largura, altura = A4

    # Header corporativo (Azul Escuro)
    c.setFillColorRGB(15/255, 23/255, 42/255) # #0f172a
    c.rect(0, altura - 80, largura, 80, stroke=0, fill=1)
    
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 18)
    c.drawString(40, altura - 45, "GRUPO EXEMPLO - RETROSPECTIVA DO LIVRO DE OCORRENCIAS")
    c.setFont("Helvetica", 10)
    c.drawString(40, altura - 65, f"Relatório gerado em: {agora.strftime('%d/%m/%Y %H:%M')}")

    # KPIs Section
    y = altura - 130
    c.setFillColorRGB(30/255, 64/255, 175/255) # #1e40af
    c.setFont("Helvetica-Bold", 14)
    c.drawString(40, y, "RESUMO GERAL (KPIS)")

    y -= 30
    c.setFillColor(colors.black)
    c.setFont("Helvetica", 12)
    c.drawString(50, y, f"Ocorrências Concluídas: {total_concluidas}")
    c.drawString(300, y, f"Enviadas para Auditoria: {total_auditoria}")
    
    y -= 25
    c.drawString(50, y, f"Tarefas Paradas Atualmente: {total_paradas}")
    c.drawString(300, y, f"Tempo Médio de Fila: {tempo_medio} dias")

    # Linha separadora
    y -= 20
    c.setStrokeColorRGB(200/255, 200/255, 200/255)
    c.line(40, y, largura - 40, y)

    # Alertas Section
    y -= 40
    c.setFillColorRGB(220/255, 38/255, 38/255) # #dc2626 (Red)
    c.setFont("Helvetica-Bold", 14)
    c.drawString(40, y, "TAREFAS GARGALADAS (ACIMA DO LIMITE)")

    y -= 30
    c.setFillColorRGB(15/255, 23/255, 42/255)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(40, y, "DIAS")
    c.drawString(90, y, "CR")
    c.drawString(160, y, "ITEM")
    c.drawString(360, y, "SETOR")
    c.drawString(500, y, "REGULATORIO")

    y -= 10
    c.setStrokeColorRGB(15/255, 23/255, 42/255)
    c.line(40, y, largura - 40, y)

    c.setFont("Helvetica", 10)
    c.setFillColor(colors.black)
    
    y -= 20
    if not tarefas_alerta:
        c.drawString(40, y, "Nenhuma tarefa gargalada no momento. Ótimo trabalho!")
    else:
        for t in tarefas_alerta[:25]: # Max 25 na primeira página para não quebrar layout básico
            c.drawString(40, y, str(t['dias']))
            c.drawString(90, y, str(t['cr'])[:10])
            c.drawString(160, y, str(t['item'])[:35])
            c.drawString(360, y, str(t['setor'])[:20])
            c.drawString(500, y, t['reg'])
            y -= 20
            
            if y < 50:
                c.showPage()
                y = altura - 50
                c.setFont("Helvetica", 10)

    # Footer
    c.setFont("Helvetica-Oblique", 8)
    c.setFillColor(colors.gray)
    c.drawString(40, 30, "Documento interno gerado pelo sistema de Gestão à Vista - Grupo Exemplo.")

    c.showPage()
    c.save()

    return response


# Procure a linha que tem o SolicitacaoCadastro e adicione o UserProfile e CustomUser:
from .models import SolicitacaoCadastro, UserProfile, CustomUser
from .models import UserProfile
from django.contrib.auth.hashers import make_password

# Adicione esta linha de importação aqui em cima:
from .models import SolicitacaoCadastro

def solicitar_cadastro(request):
    from .models import Regional
    regionais = Regional.objects.all().order_by('nome')

    if request.method == 'POST':
        nome = request.POST.get('nome')
        email = request.POST.get('email')
        telefone = request.POST.get('telefone')
        senha = request.POST.get('senha')
        regional_id = request.POST.get('regional')

        # Verifica se o email já existe
        if SolicitacaoCadastro.objects.filter(email=email).exists() or CustomUser.objects.filter(email=email).exists():
            messages.error(request, "Este e-mail já está em uso ou possui solicitação pendente.")
            return render(request, "registration/cadastro.html", {"regionais": regionais})

        if not regional_id or not regionais.filter(id=regional_id).exists():
            messages.error(request, "Selecione a Regional/Unidade a que você pertence.")
            return render(request, "registration/cadastro.html", {"regionais": regionais})

        SolicitacaoCadastro.objects.create(
            nome_completo=nome,
            email=email,
            telefone=telefone,
            senha=make_password(senha), # Segurança: Hash da senha imediatamente
            regional_id=regional_id,
        )
        messages.success(request, "Solicitação enviada com sucesso! Aguarde a aprovação de um administrador.")
        return redirect('gestao_a_vista:login')

    return render(request, "registration/cadastro.html", {"regionais": regionais})


@login_required
def listar_solicitacoes(request):
    # Apenas administradores veem esta tela
    if request.user.role != 'administrador':
        return redirect('gestao_a_vista:home')
        
    solicitacoes = SolicitacaoCadastro.objects.filter(status='pendente')

    # Mesmo filtro por regional usado em controle_acessos: admin normal so
    # ve solicitacoes da propria regional; admin global respeita o Filtro
    # Regional ativo no topo (sem filtro selecionado = ve todas).
    if not getattr(request.user, "is_global_admin", False):
        if request.user.regional_id:
            solicitacoes = solicitacoes.filter(regional_id=request.user.regional_id)
        else:
            solicitacoes = SolicitacaoCadastro.objects.none()
    else:
        active_regional_slug = request.session.get("active_regional")
        if active_regional_slug:
            solicitacoes = solicitacoes.filter(regional__db_slug=active_regional_slug)

    if request.method == 'POST':
        solicitacao_id = request.POST.get('solicitacao_id')
        acao = request.POST.get('acao')
        solicitacao = get_object_or_404(SolicitacaoCadastro, id=solicitacao_id)
        
        if acao == 'aprovar':
            role = request.POST.get('role', 'publico')
            crs = request.POST.get('crs', '')
            is_general = request.POST.get('is_general') == 'on'
            
            # Gera um username a partir do email (antes do @)
            username_base = solicitacao.email.split('@')[0]
            
            # A regional pretendida (escolhida pelo solicitante no formulario
            # publico) e a fonte de verdade. Adm supremo (global) ainda pode
            # trocar via dropdown, se o solicitante tiver errado a escolha.
            if getattr(request.user, "is_global_admin", False):
                regional_id = request.POST.get('regional') or solicitacao.regional_id
            else:
                regional_id = solicitacao.regional_id or request.user.regional_id

            # Cria o usuário real no banco
            novo_usuario = CustomUser.objects.create(
                username=username_base,
                email=solicitacao.email,
                name=solicitacao.nome_completo,
                password=solicitacao.senha, # A senha já estava criptografada na solicitação
                role=role,
                crs=crs,
                is_general=is_general,
                regional_id=regional_id,
            )
            
            # Atualiza perfil com o telefone
            UserProfile.objects.update_or_create(
                user=novo_usuario,
                defaults={'phone': solicitacao.telefone}
            )
            
            solicitacao.status = 'aprovado'
            solicitacao.save()
            messages.success(request, f"O usuário {solicitacao.nome_completo} foi aprovado com sucesso!")
            
        elif acao == 'rejeitar':
            solicitacao.status = 'rejeitado'
            solicitacao.save()
            messages.error(request, "Solicitação de cadastro rejeitada.")
            
        return redirect('gestao_a_vista:solicitacoes_cadastro')
        
    from .models import Regional
    regionais = Regional.objects.all().order_by('nome')

    context = {
        "solicitacoes": solicitacoes, 
        "role_choices": CustomUser.ROLE_CHOICES,
        "regionais": regionais,
        "is_global_admin": getattr(request.user, "is_global_admin", False),
    }
    return render(request, "solicitacoes_cadastro.html", context)


class ExplorerView(LoginRequiredMixin, TemplateView):
    """
    View para a aba do Explorador de Arquivos (Projetos)
    """
    template_name = "explorer.html"

    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        from django.http import JsonResponse
        from django.db import transaction
        import uuid
        from .models import ExplorerNode, ExplorerAuditLog, CustomUser

        action = request.GET.get("action")

        # 1. LISTAR NÓS DO EXPLORER
        if action == "explorer_list":
            try:
                # Se não houver nenhum nó, criar dados iniciais padrão
                if not ExplorerNode.objects.exists():
                    self.create_initial_explorer_data(request.user)

                nodes = ExplorerNode.objects.select_related('parent').all().order_by('is_folder', 'name')
                nodes_data = [{
                    'id': str(n.id),
                    'name': n.name,
                    'is_folder': n.is_folder,
                    'parent_id': str(n.parent_id) if n.parent else None,
                    'file_type': n.file_type,
                    'language': n.language,
                    'created_by': n.created_by_name,
                    'updated_at': n.updated_at.strftime('%d/%m/%Y %H:%M')
                } for n in nodes]
                return JsonResponse({'status': 'success', 'nodes': nodes_data})
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': str(e)})

        # 2. CARREGAR CONTEÚDO E DETALHES DE ARQUIVO
        elif action == "explorer_file_content":
            try:
                node_id = request.GET.get("node_id")
                node = ExplorerNode.objects.get(id=node_id, is_folder=False)
                
                # Registrar Auditoria
                ExplorerAuditLog.objects.create(
                    user_name=request.user.name or request.user.username,
                    action='OPEN',
                    node_name=node.name,
                    node_type='file',
                    details=f"Abriu o arquivo '{node.name}'"
                )

                return JsonResponse({
                    'status': 'success',
                    'node': {
                        'id': str(node.id),
                        'name': node.name,
                        'content': node.content,
                        'file_type': node.file_type,
                        'language': node.language,
                        'created_by': node.created_by_name,
                        'updated_at': node.updated_at.strftime('%d/%m/%Y %H:%M')
                    }
                })
            except ExplorerNode.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Arquivo não encontrado'})
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': str(e)})

        # 3. CARREGAR LOGS DE AUDITORIA (Restrito a Administradores)
        elif action == "explorer_audit_logs":
            if request.user.role != 'administrador' and not request.user.is_superuser:
                return JsonResponse({'status': 'error', 'message': 'Acesso negado. Apenas administradores podem ver a auditoria.'}, status=403)
            
            try:
                logs = ExplorerAuditLog.objects.all().order_by('-timestamp')[:500]
                logs_data = [{
                    'id': str(l.id),
                    'user': l.user_name,
                    'action': l.action,
                    'node_name': l.node_name,
                    'node_type': l.node_type,
                    'details': l.details,
                    'timestamp': l.timestamp.strftime('%d/%m/%Y %H:%M:%S')
                } for l in logs]
                return JsonResponse({'status': 'success', 'logs': logs_data})
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': str(e)})

        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        from django.http import JsonResponse
        from django.db import transaction
        import json
        import uuid
        from .models import ExplorerNode, ExplorerAuditLog

        action = request.POST.get("action")
        if not action:
            try:
                # Tenta ler do body JSON
                body_data = json.loads(request.body)
                action = body_data.get("action")
            except Exception:
                body_data = {}
        else:
            body_data = request.POST

        # 1. CRIAR PASTA OU ARQUIVO
        if action == "explorer_create":
            try:
                name = body_data.get("name", "").strip()
                is_folder = body_data.get("is_folder") in [True, "true", "on", 1]
                parent_id = body_data.get("parent_id")
                file_type = body_data.get("file_type", "text")

                if not name:
                    return JsonResponse({'status': 'error', 'message': 'O nome é obrigatório'})

                parent = None
                if parent_id and parent_id != 'null':
                    parent = ExplorerNode.objects.get(id=parent_id)

                with transaction.atomic():
                    node = ExplorerNode.objects.create(
                        name=name,
                        is_folder=is_folder,
                        parent=parent,
                        file_type=file_type,
                        created_by_name=request.user.name or request.user.username,
                        content=""
                    )
                    
                    # Log de Auditoria
                    tipo_str = "Pasta" if is_folder else "Arquivo"
                    ExplorerAuditLog.objects.create(
                        user_name=request.user.name or request.user.username,
                        action='CREATE',
                        node_name=name,
                        node_type='folder' if is_folder else 'file',
                        details=f"Criou a pasta '{name}'" if is_folder else f"Criou o arquivo '{name}' (Tipo: {file_type})"
                    )

                return JsonResponse({
                    'status': 'success',
                    'node': {
                        'id': str(node.id),
                        'name': node.name,
                        'is_folder': node.is_folder,
                        'parent_id': str(node.parent_id) if node.parent else None
                    }
                })
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': str(e)})

        # 2. SALVAR CONTEÚDO DO ARQUIVO
        elif action == "explorer_update":
            try:
                node_id = body_data.get("node_id")
                content = body_data.get("content", "")
                file_type = body_data.get("file_type", "text")
                language = body_data.get("language", "plain")

                node = ExplorerNode.objects.get(id=node_id, is_folder=False)
                
                with transaction.atomic():
                    node.content = content
                    node.file_type = file_type
                    node.language = language
                    node.save()

                    # Log de Auditoria
                    ExplorerAuditLog.objects.create(
                        user_name=request.user.name or request.user.username,
                        action='EDIT',
                        node_name=node.name,
                        node_type='file',
                        details=f"Editou o arquivo '{node.name}'"
                    )

                return JsonResponse({'status': 'success'})
            except ExplorerNode.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Arquivo não encontrado'})
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': str(e)})

        # 3. RENOMEAR NÓ
        elif action == "explorer_rename":
            try:
                node_id = body_data.get("node_id")
                new_name = body_data.get("new_name", "").strip()

                if not new_name:
                    return JsonResponse({'status': 'error', 'message': 'O novo nome é obrigatório'})

                node = ExplorerNode.objects.get(id=node_id)
                old_name = node.name

                if old_name != new_name:
                    with transaction.atomic():
                        node.name = new_name
                        node.save()

                        # Log de Auditoria
                        tipo_str = "folder" if node.is_folder else "file"
                        tipo_pt = "pasta" if node.is_folder else "arquivo"
                        ExplorerAuditLog.objects.create(
                            user_name=request.user.name or request.user.username,
                            action='RENAME',
                            node_name=new_name,
                            node_type=tipo_str,
                            details=f"Renomeou a {tipo_pt} de '{old_name}' para '{new_name}'"
                        )

                return JsonResponse({'status': 'success', 'name': node.name})
            except ExplorerNode.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Item não encontrado'})
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': str(e)})

        # 4. EXCLUIR NÓ (RECURSIVO)
        elif action == "explorer_delete":
            try:
                node_id = body_data.get("node_id")
                node = ExplorerNode.objects.get(id=node_id)
                name = node.name
                is_folder = node.is_folder

                with transaction.atomic():
                    node.delete()

                    # Log de Auditoria
                    tipo_str = "folder" if is_folder else "file"
                    tipo_pt = "pasta" if is_folder else "arquivo"
                    ExplorerAuditLog.objects.create(
                        user_name=request.user.name or request.user.username,
                        action='DELETE',
                        node_name=name,
                        node_type=tipo_str,
                        details=f"Excluiu a {tipo_pt} '{name}' e todo o seu conteúdo recursivamente"
                    )

                return JsonResponse({'status': 'success'})
            except ExplorerNode.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Item não encontrado'})
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': str(e)})

        return JsonResponse({'status': 'error', 'message': 'Ação desconhecida'})

    def create_initial_explorer_data(self, user):
        """Gera dados iniciais de pastas e arquivos de exemplo"""
        from .models import ExplorerNode
        username = user.name or user.username
        
        # 1. Pasta Documentação
        doc_folder = ExplorerNode.objects.create(
            name="Documentação",
            is_folder=True,
            created_by_name=username
        )
        ExplorerNode.objects.create(
            name="manual_usuario.txt",
            is_folder=False,
            parent=doc_folder,
            file_type="text",
            content="BEM-VINDO AO GESTÃO À VISTA!\n\nEste é o manual básico do usuário do sistema.\n\nInstruções:\n1. Use o menu lateral para navegar entre os módulos.\n2. A Torre de Controle serve para auditar e gerenciar ocorrências.\n3. A nova pasta Projetos no menu lateral agrupa o Planner (Kanban) e o Explorer de Arquivos.\n4. Para criar novos arquivos e pastas, utilize os botões da barra lateral do Explorer.\n5. O modo 'Código' conta com tema escuro e visual profissional tipo IDE.",
            created_by_name=username
        )

        # 2. Pasta Scripts
        scripts_folder = ExplorerNode.objects.create(
            name="Scripts",
            is_folder=True,
            created_by_name=username
        )
        ExplorerNode.objects.create(
            name="backup_db.py",
            is_folder=False,
            parent=scripts_folder,
            file_type="code",
            language="python",
            content="import os\nimport sys\nfrom datetime import datetime\n\ndef run_backup():\n    print(f'[{datetime.now()}] Iniciando backup do banco de dados...')\n    # TODO: Integrar com AWS S3\n    db_host = os.getenv(\"DB_HOST\", \"localhost\")\n    print(f'Conectando ao host {db_host}...')\n    print('Backup concluído com sucesso!')\n\nif __name__ == \"__main__\":\n    run_backup()",
            created_by_name=username
        )
        ExplorerNode.objects.create(
            name="check_services.js",
            is_folder=False,
            parent=scripts_folder,
            file_type="code",
            language="javascript",
            content="// Script Node.js para monitoramento de portas e endpoints\nconst http = require('http');\n\nconst services = [\n    { name: 'Portal Gestão', url: 'http://localhost:8000/health/' },\n    { name: 'Banco de Dados', url: 'http://localhost:5432/' }\n];\n\nservices.forEach(service => {\n    console.log(`Checando status do serviço: ${service.name}...`);\n    // Teste de conexão http simulado\n});",
            created_by_name=username
        )

        # 3. Pasta Configurações
        config_folder = ExplorerNode.objects.create(
            name="Configurações",
            is_folder=True,
            created_by_name=username
        )
        ExplorerNode.objects.create(
            name="config.env",
            is_folder=False,
            parent=config_folder,
            file_type="text",
            content="PORT=8000\nDEBUG=True\nALLOWED_HOSTS=localhost\nCACHE_TIMEOUT=600\nDATABASE_URL=postgres://usuario:senha@localhost:5432/gestao_a_vista",
            created_by_name=username
        )


# =============================================================================
# HELPER FUNCTIONS FOR EMAIL NOTIFICATIONS
# =============================================================================

def enviar_email_implantacao(card, etapa):
    """
    Envia e-mails de notificação baseados na etapa atual do Fluxo de Implantação.
    """
    from django.core.mail import EmailMessage
    from django.template.loader import render_to_string
    from django.conf import settings
    from .models import CustomUser
    
    destinatarios = []
    etapas_info = {
        1: {"label": "Solicitação IMPLANTAÇÃO", "cargo": "Coordenador"},
        2: {"label": "Tipo de implantação", "cargo": "Coordenador"},
        3: {"label": "Mapeamento dos locais", "cargo": "Supervisor"},
        4: {"label": "Criação do Checklist", "cargo": "Projetos"},
        5: {"label": "Criação das rotinas (RONDA, LIVRO)", "cargo": "Projetos"},
        6: {"label": "Criação dos QR codes", "cargo": "Projetos"},
        7: {"label": "Treinamento", "cargo": "Projetos"},
        8: {"label": "Entrega do projeto", "cargo": "Projetos"},
        9: {"label": "Link e painel do BI", "cargo": "BI"},
        10: {"label": "Finalizado", "cargo": "Concluído"}
    }
    
    info = etapas_info.get(etapa, {"label": f"Etapa {etapa}", "cargo": ""})
    cargo = info["cargo"]
    
    if cargo == "Coordenador":
        destinatarios = list(CustomUser.objects.filter(role="coordenador", is_active=True).values_list("email", flat=True))
    elif cargo == "Supervisor":
        destinatarios = list(CustomUser.objects.filter(role="supervisor", is_active=True).values_list("email", flat=True))
    elif cargo == "Projetos":
        destinatarios = list(CustomUser.objects.filter(setor="PROJETOS", is_active=True).values_list("email", flat=True))
    elif cargo == "BI":
        destinatarios = list(CustomUser.objects.filter(setor="BI", is_active=True).values_list("email", flat=True))
        
    destinatarios = [email for email in destinatarios if email]
    if not destinatarios:
        return
        
    host = settings.ALLOWED_HOSTS[0] if (hasattr(settings, 'ALLOWED_HOSTS') and settings.ALLOWED_HOSTS) else 'localhost:8000'
    link = f"http://{host}/implantacoes-fluxo/"
    
    context = {
        'card': card,
        'etapa_label': info["label"],
        'link': link
    }
    
    try:
        html_content = render_to_string('emails/alerta_implantacao.html', context)
        assunto = f"FLUXO DE IMPLANTAÇÃO - AÇÃO REQUERIDA - {card.nome}"
        
        email_msg = EmailMessage(
            subject=assunto,
            body=html_content,
            to=destinatarios
        )
        email_msg.content_subtype = "html"
        email_msg.send(fail_silently=True)
    except Exception as e:
        print(f"Erro ao enviar email de implantacao: {e}")


def enviar_email_desmobilizacao(card):
    """
    Envia e-mails de notificação para todos os usuários de todas as áreas de desmobilização.
    """
    from django.core.mail import EmailMessage
    from django.template.loader import render_to_string
    from django.conf import settings
    from .models import CustomUser
    
    setores_desmobilizacao = ['TI', 'PEC', 'QUALIDADE', 'PROJETOS', 'SESMT', 'SUPRIMENTOS']
    destinatarios = list(CustomUser.objects.filter(
        setor__in=setores_desmobilizacao, 
        is_active=True
    ).values_list("email", flat=True))
    
    destinatarios = [email for email in destinatarios if email]
    if not destinatarios:
        return
        
    host = settings.ALLOWED_HOSTS[0] if (hasattr(settings, 'ALLOWED_HOSTS') and settings.ALLOWED_HOSTS) else 'localhost:8000'
    link = f"http://{host}/desmobilizacoes-fluxo/"
    
    context = {
        'card': card,
        'link': link
    }
    
    try:
        html_content = render_to_string('emails/alerta_desmobilizacao.html', context)
        assunto = f"FLUXO DE DESMOBILIZAÇÃO DE CR ABERTO - {card.cr} - {card.cr_descricao or ''}"
        
        email_msg = EmailMessage(
            subject=assunto,
            body=html_content,
            to=destinatarios
        )
        email_msg.content_subtype = "html"
        email_msg.send(fail_silently=True)
    except Exception as e:
        print(f"Erro ao enviar email de desmobilizacao: {e}")


# =============================================================================
# VIEWS FOR FLUXO DE DESMOBILIZAÇÃO DE CR
# =============================================================================

@login_required
@check_page_permission("desmobilizacoes_fluxo")
def desmobilizacoes_fluxo(request):
    """View principal para o Kanban de Fluxo de Desmobilização de CR"""
    from .models import CardDesmobilizacao
    
    cards = CardDesmobilizacao.objects.all().select_related('created_by').prefetch_related('respostas')
    
    # Enriquecer os cards e corrigir status se houver divergência
    cards_list = []
    for card in cards:
        respostas = list(card.respostas.all())
        total_perguntas = len(respostas)
        concluidas = sum(1 for r in respostas if r.concluido)
        
        # Correção automática se o card estiver 'concluida' mas não estiver 100% verde
        if card.status == 'concluida' and (total_perguntas == 0 or concluidas != total_perguntas):
            card.status = 'em_andamento'
            card.save()
        
        # Mapeamento de status de checklist das 6 áreas em paralelo
        areas = ['TI', 'PEC', 'QUALIDADE', 'PROJETOS', 'SESMT', 'SUPRIMENTOS']
        areas_status = {}
        for area in areas:
            perguntas_area = [r for r in respostas if r.area == area]
            if perguntas_area:
                concluidas_area = sum(1 for p in perguntas_area if p.concluido)
                is_concluido = concluidas_area == len(perguntas_area)
                concluido_em_dt = max(p.updated_at for p in perguntas_area) if is_concluido and perguntas_area else None
                concluido_em = concluido_em_dt.strftime("%d/%m/%Y %H:%M") if concluido_em_dt else None
                concluido_em_curto = concluido_em_dt.strftime("%d/%m") if concluido_em_dt else None
                areas_status[area] = {
                    'concluido': is_concluido,
                    'concluidas_count': concluidas_area,
                    'total_count': len(perguntas_area),
                    'concluido_em': concluido_em,
                    'concluido_em_curto': concluido_em_curto
                }
            else:
                areas_status[area] = {
                    'concluido': False, 
                    'concluidas_count': 0, 
                    'total_count': 0,
                    'concluido_em': None,
                    'concluido_em_curto': None
                }
                
        cards_list.append({
            "id": card.id,
            "cr": card.cr,
            "cr_descricao": card.cr_descricao,
            "status": card.status,
            "total_perguntas": total_perguntas,
            "concluidas": concluidas,
            "porcentagem": int((concluidas / total_perguntas * 100)) if total_perguntas > 0 else 0,
            "areas_status": areas_status,
            "created_at": card.created_at,
            "created_by_name": card.created_by.name or card.created_by.username if card.created_by else "Sistema",
            "data_desmobilizacao": card.data_desmobilizacao,
        })
        
    cards_andamento = [c for c in cards_list if c["status"] == 'em_andamento']
    cards_pausada = [c for c in cards_list if c["status"] == 'pausada']
    cards_concluida = [c for c in cards_list if c["status"] == 'concluida']
    
    total_cards = len(cards_list)
    em_andamento_count = len(cards_andamento)
    pausada_count = len(cards_pausada)
    concluida_count = len(cards_concluida)
    
    context = {
        "cards_andamento": cards_andamento,
        "cards_pausada": cards_pausada,
        "cards_concluida": cards_concluida,
        "total_cards": total_cards,
        "em_andamento_count": em_andamento_count,
        "pausada_count": pausada_count,
        "concluida_count": concluida_count,
    }
    
    return render(request, "desmobilizacoes_fluxo.html", context)


@login_required
@check_page_permission("desmobilizacoes_fluxo")
def desmobilizacoes_fluxo_criar(request):
    """API para criar um novo card de desmobilização e pré-inicializar as 25 perguntas das 6 áreas em paralelo"""
    from django.http import JsonResponse
    from .models import CardDesmobilizacao, DesmobilizacaoPerguntaResposta
    
    if request.method != "POST":
        return JsonResponse({"success": False, "message": "Método inválido"}, status=405)
        
    try:
        cr = request.POST.get("cr", "").strip()
        cr_descricao = request.POST.get("cr_descricao", "").strip()
        data_desmobilizacao_str = request.POST.get("data_desmobilizacao", "").strip()
        data_desmobilizacao = data_desmobilizacao_str if data_desmobilizacao_str else None
        
        if not cr:
            return JsonResponse({"success": False, "message": "O código do CR é obrigatório."}, status=400)
            
        card = CardDesmobilizacao.objects.create(
            cr=cr,
            cr_descricao=cr_descricao,
            status="em_andamento",
            created_by=request.user,
            data_desmobilizacao=data_desmobilizacao
        )
        
        # Lista estruturada com as 25 perguntas fornecidas
        perguntas_by_area = {
            'SUPRIMENTOS': [
                "Verificar se há contratos de locação e checar sobre vigência e cláusula de encerramento",
                "Notificar fornecedores de locação sobre o encerramento do contrato",
                "Acompanhar a devolução/transferência de veículos/equipamentos locados",
                "Acompanhar recolhimento/transferência de veículos locados",
                "Atualizar planilha de controle de frota, conforme informação do GC",
                "Atualizar controle de ativo, conforme PEC e informação do GC"
            ],
            'SESMT': [
                "Receber/recolher documentação SSMA física ou digital para subir para o card do colaborador",
                "Atualizar planilha de controle de PPRA/PCMSO",
                "Atualizar planilha de controle de CIPA dos contratos"
            ],
            'QUALIDADE': [
                "Aplicação de Pesquisa de Encerramento do contrato",
                "Registro e consolidação do feedback do Cliente para a diretoria"
            ],
            'PEC': [
                "Contrato",
                "Enviar WF informando o encerramento",
                "Confirmar data de encerramento junto a operação",
                "Alinhamento com o GC da cobrança das medições em aberto junto com o último faturamento",
                "Informar as pendências de medição e contas a receber ao jurídico"
            ],
            'PROJETOS': [
                "Desativação de rotinas"
            ],
            'TI': [
                "Recolhimento/movimentação de notebooks",
                "Recolhimento/movimentação de celulares corporativos",
                "Recolhimento de periféricos (mouse, teclado, carregadores)",
                "Atualização do Sistema60 / SistemaT / Inventario",
                "Cancelamento ou transferência de linhas telefônicas",
                "Bloqueio de acessos ao e-mail e sistemas",
                "Desvinculação do CR no sistema de ponto"
            ]
        }
        
        to_create = []
        for area, perguntas in perguntas_by_area.items():
            for i, txt in enumerate(perguntas, 1):
                to_create.append(DesmobilizacaoPerguntaResposta(
                    card=card,
                    area=area,
                    pergunta_key=f"pergunta_{i}",
                    texto_pergunta=txt
                ))
                
        DesmobilizacaoPerguntaResposta.objects.bulk_create(to_create)
        
        # Enviar notificações paralelas por email
        try:
            enviar_email_desmobilizacao(card)
        except Exception as mail_err:
            print(f"Erro ao disparar alertas por email de desmobilização: {mail_err}")
            
        return JsonResponse({
            "success": True,
            "message": "Fluxo de desmobilização aberto com sucesso! E-mails de alerta enviados para todas as áreas.",
            "id": str(card.id),
            "cr": card.cr,
            "cr_descricao": card.cr_descricao,
            "status": card.status
        })
    except Exception as e:
        return JsonResponse({"success": False, "message": f"Erro ao criar desmobilização: {str(e)}"}, status=400)


@login_required
@check_page_permission("desmobilizacoes_fluxo")
def desmobilizacoes_fluxo_card_details(request):
    """API para obter detalhes completos do card, suas perguntas e múltiplos anexos"""
    from django.http import JsonResponse
    from .models import CardDesmobilizacao
    
    if request.method != "GET":
        return JsonResponse({"success": False, "message": "Método inválido"}, status=405)
        
    card_id = request.GET.get("id")
    if not card_id:
        return JsonResponse({"success": False, "message": "ID do card não fornecido"}, status=400)
        
    try:
        card = CardDesmobilizacao.objects.select_related('created_by').prefetch_related('respostas', 'respostas__anexos').get(id=card_id)
        
        respostas_list = []
        respostas = list(card.respostas.all())
        for r in card.respostas.all().order_by('area', 'pergunta_key'):
            anexos = [{
                "id": str(aneco.id),
                "nome": aneco.nome_original,
                "url": aneco.arquivo.url
            } for aneco in r.anexos.all()]
            
            respostas_list.append({
                "id": str(r.id),
                "area": r.area,
                "pergunta_key": r.pergunta_key,
                "texto_pergunta": r.texto_pergunta,
                "concluido": r.concluido,
                "resposta_texto": r.resposta_texto or "",
                "respondido_por_nome": r.respondido_por_nome or "",
                "anexos": anexos
            })
            
        # Calcular areas_status com data de conclusao
        areas = ['TI', 'PEC', 'QUALIDADE', 'PROJETOS', 'SESMT', 'SUPRIMENTOS']
        areas_status = {}
        for area in areas:
            perguntas_area = [r for r in respostas if r.area == area]
            if perguntas_area:
                concluidas_area = sum(1 for p in perguntas_area if p.concluido)
                is_concluido = concluidas_area == len(perguntas_area)
                concluido_em_dt = max(p.updated_at for p in perguntas_area) if is_concluido and perguntas_area else None
                concluido_em = concluido_em_dt.strftime("%d/%m/%Y %H:%M") if concluido_em_dt else None
                concluido_em_curto = concluido_em_dt.strftime("%d/%m") if concluido_em_dt else None
                areas_status[area] = {
                    'concluido': is_concluido,
                    'concluidas_count': concluidas_area,
                    'total_count': len(perguntas_area),
                    'concluido_em': concluido_em,
                    'concluido_em_curto': concluido_em_curto
                }
            else:
                areas_status[area] = {
                    'concluido': False, 
                    'concluidas_count': 0, 
                    'total_count': 0,
                    'concluido_em': None,
                    'concluido_em_curto': None
                }

        return JsonResponse({
            "success": True,
            "id": str(card.id),
            "cr": card.cr,
            "cr_descricao": card.cr_descricao,
            "status": card.status,
            "created_at": card.created_at.strftime("%d/%m/%Y %H:%M"),
            "created_by_name": card.created_by.name or card.created_by.username if card.created_by else "Sistema",
            "data_desmobilizacao": card.data_desmobilizacao.strftime("%d/%m/%Y") if card.data_desmobilizacao else None,
            "areas_status": areas_status,
            "respostas": respostas_list
        })
    except Exception as e:
        return JsonResponse({"success": False, "message": f"Erro ao obter detalhes do card: {str(e)}"}, status=400)


@login_required
@check_page_permission("desmobilizacoes_fluxo")
def desmobilizacoes_fluxo_status(request):
    """API para atualizar o status do Kanban do card de desmobilização"""
    from django.http import JsonResponse
    from .models import CardDesmobilizacao
    
    if request.method != "POST":
        return JsonResponse({"success": False, "message": "Método inválido"}, status=405)
        
    card_id = request.POST.get("id")
    novo_status = request.POST.get("status")
    
    if not card_id or not novo_status:
        return JsonResponse({"success": False, "message": "Parâmetros incompletos"}, status=400)
        
    if novo_status not in ['em_andamento', 'pausada', 'concluida']:
        return JsonResponse({"success": False, "message": "Status inválido"}, status=400)
        
    try:
        card = CardDesmobilizacao.objects.get(id=card_id)
        
        # Se tentar mudar para concluida, validar se todas as perguntas foram respondidas
        if novo_status == 'concluida':
            total_perguntas = card.respostas.count()
            concluidas_perguntas = card.respostas.filter(concluido=True).count()
            if total_perguntas == 0 or total_perguntas != concluidas_perguntas:
                return JsonResponse({
                    "success": False, 
                    "message": "Não é possível concluir o fluxo: existem itens pendentes no checklist."
                }, status=400)
                
        card.status = novo_status
        card.save()
        return JsonResponse({
            "success": True, 
            "message": "Status do fluxo atualizado com sucesso!",
            "card_id": str(card.id),
            "status": card.status
        })
    except Exception as e:
        return JsonResponse({"success": False, "message": f"Erro ao atualizar status: {str(e)}"}, status=400)


@login_required
@check_page_permission("desmobilizacoes_fluxo")
def desmobilizacoes_fluxo_excluir(request):
    """API para excluir um fluxo de desmobilização (Somente administradores)"""
    from django.http import JsonResponse
    from .models import CardDesmobilizacao
    
    if request.method != "POST":
        return JsonResponse({"success": False, "message": "Método inválido"}, status=405)
        
    if request.user.role != 'administrador':
        return JsonResponse({"success": False, "message": "Apenas administradores podem excluir fluxos."}, status=403)
        
    card_id = request.POST.get("id")
    if not card_id:
        return JsonResponse({"success": False, "message": "ID do card não fornecido"}, status=400)
        
    try:
        card = CardDesmobilizacao.objects.get(id=card_id)
        card.delete()
        return JsonResponse({"success": True, "message": "Fluxo de desmobilização excluído com sucesso!"})
    except CardDesmobilizacao.DoesNotExist:
        return JsonResponse({"success": False, "message": "Fluxo não encontrado"}, status=404)
    except Exception as e:
        return JsonResponse({"success": False, "message": f"Erro ao excluir fluxo: {str(e)}"}, status=400)


@login_required
@check_page_permission("desmobilizacoes_fluxo")
def desmobilizacoes_fluxo_salvar(request):
    """API para salvar as respostas de uma área/pergunta, anexar múltiplos arquivos em lista ou deletá-los"""
    from django.http import JsonResponse
    from .models import DesmobilizacaoPerguntaResposta, DesmobilizacaoAnexo, CardDesmobilizacao
    import json
    
    if request.method != "POST":
        return JsonResponse({"success": False, "message": "Método inválido"}, status=405)
        
    try:
        # Lista de IDs de perguntas que estão sendo salvas nesta submissão
        pergunta_ids = request.POST.getlist("pergunta_ids")
        if not pergunta_ids:
            return JsonResponse({"success": False, "message": "Nenhuma pergunta enviada para salvar."}, status=400)
            
        username = request.user.name or request.user.username
        card = None
        
        for pid in pergunta_ids:
            try:
                pergunta = DesmobilizacaoPerguntaResposta.objects.get(id=pid)
                if not card:
                    card = pergunta.card
                    
                concluido_str = request.POST.get(f"concluido_{pid}", "false")
                concluido = concluido_str.lower() in ["true", "1", "on"]
                resposta_texto = request.POST.get(f"observacoes_{pid}", "").strip()
                
                # Salvar dados de checklist e notas
                pergunta.concluido = concluido
                pergunta.resposta_texto = resposta_texto
                pergunta.respondido_por_nome = username
                pergunta.save()
                
                # Tratar deleções de anexos existentes
                deletar_anexos_str = request.POST.get(f"delete_anexos_{pid}", "[]")
                try:
                    deletar_anexos_list = json.loads(deletar_anexos_str)
                    if deletar_anexos_list:
                        DesmobilizacaoAnexo.objects.filter(id__in=deletar_anexos_list, pergunta_resposta=pergunta).delete()
                except Exception as del_err:
                    print(f"Erro ao processar deleção de anexos: {del_err}")
                    
                # Tratar múltiplos novos uploads (anexos_{pid})
                novos_arquivos = request.FILES.getlist(f"anexos_{pid}")
                for arq in novos_arquivos:
                    DesmobilizacaoAnexo.objects.create(
                        pergunta_resposta=pergunta,
                        arquivo=arq,
                        nome_original=arq.name
                    )
            except DesmobilizacaoPerguntaResposta.DoesNotExist:
                continue
                
        # Verificar se todas as 25 perguntas do card foram respondidas/concluídas
        if card:
            total_perguntas = card.respostas.count()
            concluidas_perguntas = card.respostas.filter(concluido=True).count()
            if total_perguntas > 0 and total_perguntas == concluidas_perguntas:
                card.status = "concluida"
            else:
                if card.status == "concluida":
                    card.status = "em_andamento"
            card.save()
                
            # Mapeamento de status de checklist das 6 áreas em paralelo
            respostas = list(card.respostas.all())
            areas = ['TI', 'PEC', 'QUALIDADE', 'PROJETOS', 'SESMT', 'SUPRIMENTOS']
            areas_status = {}
            for area in areas:
                perguntas_area = [r for r in respostas if r.area == area]
                if perguntas_area:
                    concluidas_area = sum(1 for p in perguntas_area if p.concluido)
                    is_concluido = concluidas_area == len(perguntas_area)
                    concluido_em_dt = max(p.updated_at for p in perguntas_area) if is_concluido and perguntas_area else None
                    concluido_em = concluido_em_dt.strftime("%d/%m/%Y %H:%M") if concluido_em_dt else None
                    concluido_em_curto = concluido_em_dt.strftime("%d/%m") if concluido_em_dt else None
                    areas_status[area] = {
                        'concluido': is_concluido,
                        'concluidas_count': concluidas_area,
                        'total_count': len(perguntas_area),
                        'concluido_em': concluido_em,
                        'concluido_em_curto': concluido_em_curto
                    }
                else:
                    areas_status[area] = {
                        'concluido': False, 
                        'concluidas_count': 0, 
                        'total_count': 0,
                        'concluido_em': None,
                        'concluido_em_curto': None
                    }
                    
            porcentagem = int((concluidas_perguntas / total_perguntas * 100)) if total_perguntas > 0 else 0
            
            return JsonResponse({
                "success": True,
                "message": "Checklist e anexos salvos com sucesso!",
                "card_id": str(card.id),
                "status": card.status,
                "total_perguntas": total_perguntas,
                "concluidas": concluidas_perguntas,
                "porcentagem": porcentagem,
                "areas_status": areas_status
            })
    except Exception as e:
        return JsonResponse({"success": False, "message": f"Erro ao salvar: {str(e)}"}, status=400)


from .models import PsicossocialProjeto, ColaboradorSRA

@login_required
def psicossocial_list(request):
    from django.db.models import Q
    if not (request.user.role == "administrador" or request.user.has_page_permission("psicossocial")):
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied("Você não tem permissão para acessar esta página.")
        
    projects = PsicossocialProjeto.objects.all().order_by('-created_at')

    active_sra = ColaboradorSRA.objects.filter(
        Q(dt_demissao__isnull=True) & (
            Q(situacao__iexact="normal") |
            Q(situacao__iexact="ativo") |
            Q(situacao__iexact="ativa") |
            Q(situacao__icontains="férias") |
            Q(situacao__icontains="ferias") |
            Q(situacao__isnull=True) |
            Q(situacao="")
        )
    )
    sra_count_qs = ColaboradorSRA.objects.all()
    sra_count = sra_count_qs.count()
    clientes = list(active_sra.exclude(cliente__isnull=True).exclude(cliente='').values_list('cliente', flat=True).distinct().order_by('cliente'))
    unidades = list(active_sra.exclude(nome_unidade__isnull=True).exclude(nome_unidade='').values_list('nome_unidade', flat=True).distinct().order_by('nome_unidade'))
    empresas_codigo = list(active_sra.exclude(empresa_codigo__isnull=True).exclude(empresa_codigo='').values_list('empresa_codigo', flat=True).distinct().order_by('empresa_codigo'))
    filiais_codigo = list(active_sra.exclude(filial_codigo__isnull=True).exclude(filial_codigo='').values_list('filial_codigo', flat=True).distinct().order_by('filial_codigo'))
    colaboradores_json = "[]"
    
    def normalize_code(val):
        if val is None:
            return ""
        if isinstance(val, (int, float)):
            val = int(val)
        s = str(val).strip()
        s_clean = s.lstrip('0')
        if not s_clean and s:
            return "0"
        return s_clean

    # Pre-fill mappings: company_code | branch_code -> company_name & cnpj
    mappings_dict = {}
    qs = ColaboradorSRA.objects.exclude(empresa_codigo__isnull=True).exclude(empresa_codigo='').values('empresa_codigo', 'filial_codigo', 'empresa_nome', 'cnpj')
    for item in qs:
        emp_cod = normalize_code(item['empresa_codigo'])
        fil_cod = normalize_code(item['filial_codigo'])
        if not emp_cod:
            continue
        key = f"{emp_cod}|{fil_cod}".lower()
        if key not in mappings_dict:
            mappings_dict[key] = {
                'empresa_nome': item['empresa_nome'] or '',
                'cnpj': format_cnpj(item['cnpj'])
            }
        else:
            if not mappings_dict[key]['empresa_nome'] and item['empresa_nome']:
                mappings_dict[key]['empresa_nome'] = item['empresa_nome']
            if not mappings_dict[key]['cnpj'] and item['cnpj']:
                mappings_dict[key]['cnpj'] = format_cnpj(item['cnpj'])
    company_mappings_json = json.dumps(mappings_dict)

    for project in projects:
        project.cnpj = format_cnpj(project.cnpj)

    context = {
        "projects": projects,
        "sra_count": sra_count,
        "clientes": clientes,
        "unidades": unidades,
        "empresas_codigo": empresas_codigo,
        "filiais_codigo": filiais_codigo,
        "colaboradores_json": colaboradores_json,
        "company_mappings_json": company_mappings_json,
    }
    return render(request, "psicossocial.html", context)


@login_required
def psicossocial_delete(request, project_id):
    if not (request.user.role == "administrador" or request.user.has_page_permission("psicossocial")):
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied("Você não tem permissão para esta ação.")
        
    projeto = get_object_or_404(PsicossocialProjeto, id=project_id)
    # Delete physical files
    if projeto.planilha_respostas and os.path.exists(projeto.planilha_respostas.path):
        os.remove(projeto.planilha_respostas.path)
    if projeto.planilha_workforce and os.path.exists(projeto.planilha_workforce.path):
        os.remove(projeto.planilha_workforce.path)
    if projeto.planilha_resultado and os.path.exists(projeto.planilha_resultado.path):
        os.remove(projeto.planilha_resultado.path)
    if projeto.relatorio_word and os.path.exists(projeto.relatorio_word.path):
        os.remove(projeto.relatorio_word.path)
        
    projeto.delete()
    messages.success(request, "Projeto excluído com sucesso.")
    return redirect("gestao_a_vista:psicossocial")


@login_required
def psicossocial_create(request):
    if not (request.user.role == "administrador" or request.user.has_page_permission("psicossocial")):
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied("Você não tem permissão para acessar esta página.")
        
    if request.method == "POST":
        # Extract inputs
        nome = request.POST.get("nome", "").strip()
        empresa = request.POST.get("empresa", "").strip()
        cnpj = request.POST.get("cnpj", "").strip()
        localidade = request.POST.get("localidade", "").strip()
        responsavel_tecnico = request.POST.get("responsavel_tecnico", "").strip()
        detalhamento_fatores_protetivos = request.POST.get("detalhamento_fatores_protetivos", "").strip()
        
        empresa_codigo = request.POST.get("empresa_codigo", "").strip()
        filial_codigo = request.POST.get("filial_codigo", "").strip()
        
        # Parse Dates safely
        def parse_date(dt_str):
            if not dt_str:
                return None
            from datetime import datetime as dt_class
            try:
                return dt_class.strptime(dt_str, "%Y-%m-%d").date()
            except ValueError:
                return None
                
        periodo_inicio = parse_date(request.POST.get("periodo_inicio"))
        periodo_fim = parse_date(request.POST.get("periodo_fim"))
        data_aplicacao = parse_date(request.POST.get("data_aplicacao"))
        data_emissao = parse_date(request.POST.get("data_emissao"))
        
        planilha_respostas = request.FILES.get("planilha_respostas")

        try:
            total_colaboradores = int(request.POST.get("total_colaboradores", "0") or 0)
        except ValueError:
            total_colaboradores = 0

        # Resolve Company Name and CNPJ based on Empresa Code and Filial Code
        resolved_empresa = None
        resolved_cnpj = None
        if empresa_codigo and filial_codigo:
            # Look in existing database first
            match_qs = ColaboradorSRA.objects.filter(
                empresa_codigo=empresa_codigo,
                filial_codigo=filial_codigo
            )
            match = match_qs.exclude(empresa_nome="").exclude(cnpj="").first()
            if match:
                resolved_empresa = match.empresa_nome
                resolved_cnpj = match.cnpj

        if resolved_empresa:
            empresa = resolved_empresa
        if resolved_cnpj:
            cnpj = resolved_cnpj
            
        if not planilha_respostas:
            messages.error(request, "A planilha de respostas (Forms) é obrigatória.")
            return redirect("gestao_a_vista:psicossocial")
        try:
            # Create project record
            projeto = PsicossocialProjeto(
                nome=nome,
                empresa=empresa,
                cnpj=cnpj,
                localidade=localidade,
                periodo_inicio=periodo_inicio,
                periodo_fim=periodo_fim,
                responsavel_tecnico=responsavel_tecnico,
                data_aplicacao=data_aplicacao,
                data_emissao=data_emissao,
                planilha_respostas=planilha_respostas,
                created_by=request.user,
                empresa_codigo=empresa_codigo,
                filial_codigo=filial_codigo,
                detalhamento_fatores_protetivos=detalhamento_fatores_protetivos,
                total_colaboradores=total_colaboradores
            )
            projeto.save()

            # Avaliação psicossocial é anônima: não há mais seleção de colaboradores
            # nomeados (Layout 1/SRA); o total de elegíveis vem do total_colaboradores
            # e da tabela de headcount por GHE informados no formulário.
            workforce_file_path = None

            # --- RUN SCORING ENGINE ---
            import os
            import uuid
            from psicossocial.metodologia import load_metodologia
            from psicossocial.excel_export import export_excel
            from psicossocial.processing import process_excel
            from psicossocial.report_docx import generate_word_report
            
            metodologia_path = os.path.join(settings.BASE_DIR, "psicossocial", "config", "metodologia_v1.yaml")
            metodologia = load_metodologia(metodologia_path)
            
            # 1. Output spreadsheet
            out_xlsx_name = f"resultado_tecnico_{projeto.id}_{uuid.uuid4().hex[:8]}.xlsx"
            out_xlsx_path = os.path.join(settings.MEDIA_ROOT, "psicossocial", "resultados", out_xlsx_name)
            os.makedirs(os.path.dirname(out_xlsx_path), exist_ok=True)
            
            export_excel(
                input_path=projeto.planilha_respostas.path,
                output_path=out_xlsx_path,
                metodologia=metodologia,
                workforce_path=workforce_file_path,
                filter_unit=projeto.nome
            )
            projeto.planilha_resultado = f"psicossocial/resultados/{out_xlsx_name}"
            projeto.save()
            
            # 2. Build summary for Word Report
            processed_survey = process_excel(projeto.planilha_respostas.path, metodologia, filter_unit=projeto.nome)
            representatividade_summary = []

            # Extract headcounts from POST
            from psicossocial.report_docx import normalize_activity_q6
            headcounts = {}
            # Try to get list of manual inputs first
            ghe_setores = request.POST.getlist("ghe_setor[]")
            tipos_atividade = request.POST.getlist("tipo_atividade[]")
            headcounts_total = request.POST.getlist("headcount_total[]")
            if ghe_setores and tipos_atividade:
                for g, t, h in zip(ghe_setores, tipos_atividade, headcounts_total):
                    if g and t:
                        t_norm = normalize_activity_q6(t.strip())
                        key = f"{g.strip()} | {t_norm}"
                        try:
                            headcounts[key] = int(h)
                        except ValueError:
                            headcounts[key] = 0
            
            # Fallback for key-based parameters (e.g. from tests or old UI forms)
            for key, val in request.POST.items():
                if key.startswith("headcount_") and key != "headcount_total[]":
                    area_name = key[len("headcount_"):]
                    if " | " in area_name:
                        g, t = area_name.split(" | ", 1)
                        t_norm = normalize_activity_q6(t.strip())
                        area_name = f"{g.strip()} | {t_norm}"
                    if area_name not in headcounts:
                        try:
                            headcounts[area_name] = int(val)
                        except ValueError:
                            headcounts[area_name] = 0


            # 3. Output Word Report
            word_rel_path = generate_word_report(projeto, processed_survey, metodologia, representatividade_summary, headcounts=headcounts)
            projeto.relatorio_word = word_rel_path
            projeto.save()

            messages.success(request, f"Projeto '{nome}' criado e processado com sucesso! Arquivos gerados para download.")
        except Exception as e:
            import traceback
            print("ERROR IN PSICOSOCIAL ENGINE:")
            traceback.print_exc()
            messages.error(request, f"Erro ao processar projeto: {str(e)}")
            
    return redirect("gestao_a_vista:psicossocial")


@login_required
def psicossocial_sra(request):
    if not (request.user.role == "administrador" or request.user.has_page_permission("psicossocial")):
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied("Você não tem permissão para acessar esta página.")
        
    from django.db.models import Q
    from django.db import transaction
    import openpyxl
    import tempfile
    import os
    from datetime import date
    
    # Handle POST
    if request.method == "POST":
        action = request.POST.get("action")
        
        if action == "import_sra":
            sra_file = request.FILES.get("sra_file")
            if not sra_file:
                messages.error(request, "Nenhum arquivo enviado.")
                return redirect("gestao_a_vista:psicossocial_sra")
                
            # Save uploaded file temporarily
            fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
            try:
                with os.fdopen(fd, 'wb') as tmp:
                    for chunk in sra_file.chunks():
                        tmp.write(chunk)
                        
                wb = openpyxl.load_workbook(tmp_path, data_only=True)
                sheet = wb["SRA 11-06-26"] if "SRA 11-06-26" in wb.sheetnames else wb.worksheets[0]
                
                # Identify headers
                headers = [str(cell or "").strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
                col_map = {}
                for col_name in ["Dt Demissao", "Municipio", "Estado", "Negocio", "CPF", "Nome", "Matricula", "CR", "Cliente", "Empresa", "Filial"]:
                    col_index = None
                    for idx, h in enumerate(headers):
                        if h.strip().lower() == col_name.strip().lower():
                            col_index = idx
                            break
                    col_map[col_name] = col_index

                # Check for "Situacao" column (usually column AH / index 33)
                situacao_index = None
                for idx, h in enumerate(headers):
                    h_lower = h.strip().lower()
                    if h_lower in ["situacao", "situação", "status", "sit. cta.", "sit. cta", "sit cta", "situação do contrato", "situação do colaborador"]:
                        situacao_index = idx
                        break
                if situacao_index is None and len(headers) > 33:
                    situacao_index = 33
                col_map["Situacao"] = situacao_index

                cnpj_index = None
                for idx, h in enumerate(headers):
                    h_clean = h.strip().lower()
                    if h_clean in ["cnpj", "c.n.p.j", "cnpj empresa", "cnpj/mf"]:
                        cnpj_index = idx
                        break
                if cnpj_index is None:
                    for idx, h in enumerate(headers):
                        h_clean = h.strip().lower()
                        if "cnpj" in h_clean and "cliente" not in h_clean:
                            cnpj_index = idx
                            break
                if cnpj_index is None:
                    for idx, h in enumerate(headers):
                        h_clean = h.strip().lower()
                        if "cnpj" in h_clean or "inscricao" in h_clean or "c.n.p.j" in h_clean:
                            cnpj_index = idx
                            break

                empresa_nome_index = None
                for idx, h in enumerate(headers):
                    h_lower = h.strip().lower()
                    if "razao social" in h_lower or "razão social" in h_lower or "empresa contratada" in h_lower or "empresa nome" in h_lower or "nome empresa" in h_lower:
                        empresa_nome_index = idx

                nome_unidade_index = None
                for idx, h in enumerate(headers):
                    h_lower = h.strip().lower()
                    if "unidade de trabalho" in h_lower or h_lower in ["nome da unidade", "unidade"]:
                        nome_unidade_index = idx
                        break

                grupo_cliente_index = None
                for idx, h in enumerate(headers):
                    h_lower = h.strip().lower()
                    if "grupo cliente" in h_lower or ("grupo" in h_lower and "cliente" in h_lower):
                        grupo_cliente_index = idx
                        break

                # Verify required columns for calculation (apenas o essencial para identificar o colaborador)
                required = ["CPF", "Nome"]
                missing = [r for r in required if col_map[r] is None]
                if missing:
                    os.remove(tmp_path)
                    messages.error(request, f"Colunas obrigatórias ausentes na planilha Layout 1: {', '.join(missing)}")
                    return redirect("gestao_a_vista:psicossocial_sra")
                    
                # Parse rows
                colaboradores = []
                row_count = 0
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if not any(val is not None for val in row):
                        continue
                        
                    cpf_val = str(row[col_map["CPF"]] or "").strip() if col_map["CPF"] is not None else ""
                    if not cpf_val:
                        continue
                        
                    # Format/cleanup CPF (remove dots, dashes, spaces)
                    cpf_cleaned = "".join(filter(str.isdigit, cpf_val))
                    if not cpf_cleaned:
                        continue
                        
                    nome_val = str(row[col_map["Nome"]] or "").strip() if col_map["Nome"] is not None else ""
                    
                    # Check for demobilized employees
                    dt_demissao_val = row[col_map["Dt Demissao"]] if col_map["Dt Demissao"] is not None else None
                    dt_demissao_date = None
                    if dt_demissao_val:
                        if isinstance(dt_demissao_val, datetime):
                            dt_demissao_date = dt_demissao_val.date()
                        elif isinstance(dt_demissao_val, date):
                            dt_demissao_date = dt_demissao_val
                        else:
                            # try parsing
                            try:
                                dt_demissao_date = datetime.strptime(str(dt_demissao_val).strip(), "%Y-%m-%d").date()
                            except ValueError:
                                try:
                                    dt_demissao_date = datetime.strptime(str(dt_demissao_val).strip(), "%d/%m/%Y").date()
                                except ValueError:
                                    dt_demissao_date = None
                                    
                    # Extract values by mapped column names (novo padrão de layout)
                    empresa_codigo_val = str(row[col_map["Empresa"]] or "").strip() if (col_map["Empresa"] is not None and len(row) > col_map["Empresa"]) else ""
                    filial_codigo_val = str(row[col_map["Filial"]] or "").strip() if (col_map["Filial"] is not None and len(row) > col_map["Filial"]) else ""
                    nome_unidade_val = str(row[nome_unidade_index] or "").strip() if (nome_unidade_index is not None and len(row) > nome_unidade_index) else ""
                    grupo_cliente_val = str(row[grupo_cliente_index] or "").strip() if (grupo_cliente_index is not None and len(row) > grupo_cliente_index) else ""

                    cnpj_val = str(row[cnpj_index] or "").strip() if (cnpj_index is not None and len(row) > cnpj_index) else ""
                    empresa_nome_val = str(row[empresa_nome_index] or "").strip() if (empresa_nome_index is not None and len(row) > empresa_nome_index) else ""

                    matricula_val = str(row[col_map["Matricula"]] or "").strip() if col_map["Matricula"] is not None else ""
                    negocio_val = str(row[col_map["Negocio"]] or "").strip() if col_map["Negocio"] is not None else ""
                    municipio_val = normalize_city_name(row[col_map["Municipio"]]) if col_map["Municipio"] is not None else ""
                    estado_val = str(row[col_map["Estado"]] or "").strip() if col_map["Estado"] is not None else ""
                    cr_val = str(row[col_map["CR"]] or "").strip() if col_map["CR"] is not None else ""
                    cliente_val = str(row[col_map["Cliente"]] or "").strip() if col_map["Cliente"] is not None else ""

                    situacao_val = str(row[col_map["Situacao"]] or "").strip() if col_map["Situacao"] is not None else ""
                    if not situacao_val:
                        situacao_val = "Ativo"

                    colaboradores.append(ColaboradorSRA(
                        cpf=cpf_cleaned,
                        nome=nome_val,
                        matricula=matricula_val,
                        negocio=negocio_val,
                        municipio=municipio_val,
                        estado=estado_val,
                        dt_demissao=dt_demissao_date,
                        cr=cr_val,
                        cliente=cliente_val,
                        situacao=situacao_val,
                        grupo_cliente=grupo_cliente_val,
                        empresa_codigo=empresa_codigo_val,
                        filial_codigo=filial_codigo_val,
                        nome_unidade=nome_unidade_val,
                        cnpj=cnpj_val,
                        empresa_nome=empresa_nome_val
                    ))
                    row_count += 1
                    
                # Save to partitioned databases
                save_colaboradores_to_databases(colaboradores)
                    
                # Save the physical SRA file to media/psicossocial/sra_base_latest.xlsx
                sra_dir = os.path.join(settings.MEDIA_ROOT, "psicossocial")
                os.makedirs(sra_dir, exist_ok=True)
                sra_dest = os.path.join(sra_dir, "sra_base_latest.xlsx")
                if os.path.exists(sra_dest):
                    os.remove(sra_dest)
                import shutil
                shutil.copy(tmp_path, sra_dest)
                
                messages.success(request, f"Base Layout 1 importada com sucesso! {row_count} registros cadastrados.")
            except Exception as e:
                messages.error(request, f"Erro ao processar arquivo Layout 1: {str(e)}")
            finally:
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)
            return redirect("gestao_a_vista:psicossocial_sra")
            
        elif action == "import_cnpj":
            cnpj_file = request.FILES.get("cnpj_file")
            if not cnpj_file:
                messages.error(request, "Nenhum arquivo de empresas enviado.")
                return redirect("gestao_a_vista:psicossocial_sra")
                
            def normalize_code(val):
                if val is None:
                    return ""
                if isinstance(val, (int, float)):
                    val = int(val)
                s = str(val).strip()
                s_clean = s.lstrip('0')
                if not s_clean and s:
                    return "0"
                return s_clean
                
            fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
            try:
                with os.fdopen(fd, 'wb') as tmp:
                    for chunk in cnpj_file.chunks():
                        tmp.write(chunk)
                        
                wb = openpyxl.load_workbook(tmp_path, data_only=True)
                sheet = None
                sheet_name = "2.Layout_BD_CNPJ"
                if sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                else:
                    for name in wb.sheetnames:
                        name_lower = name.lower()
                        if "cnpj" in name_lower or "empresa" in name_lower or "layout" in name_lower:
                            sheet = wb[name]
                            break
                    if sheet is None:
                        sheet = wb.worksheets[1] if len(wb.worksheets) > 1 else wb.worksheets[0]
                    
                header_row_idx = 1
                best_match_count = -1
                best_col_map = {
                    "empresa_codigo": None,
                    "filial_codigo": None,
                    "cnpj": None,
                    "empresa_nome": None,
                }
                
                for r_idx in range(1, min(11, sheet.max_row + 1)):
                    row_vals = next(sheet.iter_rows(min_row=r_idx, max_row=r_idx, values_only=True))
                    if not row_vals or not any(val is not None for val in row_vals):
                        continue
                    row_cells = [str(cell or "").strip() for cell in row_vals]
                    
                    temp_map = {
                        "empresa_codigo": None,
                        "filial_codigo": None,
                        "cnpj": None,
                        "empresa_nome": None,
                    }
                    for idx, h in enumerate(row_cells):
                        h_clean = h.strip().lower()
                        if h_clean in ["empresa", "codigo empresa", "cod. empresa", "cod empresa", "código da empresa", "código empresa", "cd_empresa", "cód. empresa", "cód empresa", "empresa_cod"]:
                            temp_map["empresa_codigo"] = idx
                        elif h_clean in ["filial", "codigo filial", "cod. filial", "cod filial", "código da filial", "código filial", "cd_filial", "cód. filial", "cód filial", "filial_cod"]:
                            temp_map["filial_codigo"] = idx
                        elif "cnpj" in h_clean or h_clean in ["cnpj", "c.n.p.j", "cnpj empresa", "cnpj/mf", "cnpj principal", "c.n.p.j.", "nr_cnpj", "cnpj_empresa"]:
                            temp_map["cnpj"] = idx
                        elif h_clean in ["razao social", "razão social", "empresa contratada", "empresa nome", "nome empresa", "nome da empresa", "nm_razao_social", "razao_social", "descrição empresa"]:
                            temp_map["empresa_nome"] = idx
                            
                    for idx, h in enumerate(row_cells):
                        h_clean = h.strip().lower()
                        if temp_map["empresa_codigo"] is None and ("codigo" in h_clean or "cod" in h_clean) and "empresa" in h_clean:
                            temp_map["empresa_codigo"] = idx
                        if temp_map["filial_codigo"] is None and ("codigo" in h_clean or "cod" in h_clean) and "filial" in h_clean:
                            temp_map["filial_codigo"] = idx
                        if temp_map["cnpj"] is None and "cnpj" in h_clean:
                            temp_map["cnpj"] = idx
                        if temp_map["empresa_nome"] is None and ("razao" in h_clean or "razão" in h_clean or "social" in h_clean or "desc" in h_clean):
                            temp_map["empresa_nome"] = idx
                            
                    matches = sum(1 for k, v in temp_map.items() if v is not None)
                    if matches > best_match_count:
                        best_match_count = matches
                        best_col_map = temp_map
                        header_row_idx = r_idx
                
                col_map = best_col_map
                
                missing = []
                if col_map["empresa_codigo"] is None: missing.append("Código Empresa")
                if col_map["filial_codigo"] is None: missing.append("Código Filial")
                if col_map["cnpj"] is None and col_map["empresa_nome"] is None:
                    missing.append("CNPJ ou Razão Social")
                    
                if missing:
                    os.remove(tmp_path)
                    messages.error(request, f"Colunas necessárias ausentes na planilha base: {', '.join(missing)}")
                    return redirect("gestao_a_vista:psicossocial_sra")

                coluna_avisos = []
                if col_map["cnpj"] is None:
                    coluna_avisos.append("CNPJ")
                if col_map["empresa_nome"] is None:
                    coluna_avisos.append("Razão Social")

                update_map = {}
                for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
                    if not any(val is not None for val in row):
                        continue
                    emp_cod_raw = row[col_map["empresa_codigo"]]
                    fil_cod_raw = row[col_map["filial_codigo"]]
                    if emp_cod_raw is None:
                        continue
                    emp_cod = normalize_code(emp_cod_raw)
                    fil_cod = normalize_code(fil_cod_raw)
                    if not emp_cod:
                        continue
                    cnpj_val = str(row[col_map["cnpj"]] or "").strip() if col_map["cnpj"] is not None else ""
                    emp_name_val = str(row[col_map["empresa_nome"]] or "").strip() if col_map["empresa_nome"] is not None else ""
                    
                    key = (emp_cod, fil_cod)
                    update_map[key] = (cnpj_val, emp_name_val)
                    
                # Update records in active DB
                from Gestao_a_Vista.thread_local import get_current_db
                active_db = get_current_db() or 'default'
                
                updated_count = 0
                with transaction.atomic(using=active_db):
                    target_qs = ColaboradorSRA.objects.using(active_db).all()
                    if not getattr(request.user, 'is_global_admin', False):
                        user_crs = [c.strip() for c in request.user.crs.split(',') if c.strip()] if getattr(request.user, 'crs', None) else []
                        if user_crs:
                            target_qs = target_qs.filter(cr__in=user_crs)

                    colabs_to_update = []
                    for colab in target_qs:
                        emp_cod = normalize_code(colab.empresa_codigo)
                        fil_cod = normalize_code(colab.filial_codigo)
                        if not emp_cod:
                            continue
                        key = (emp_cod, fil_cod)
                        if key in update_map:
                            cnpj_val, emp_name_val = update_map[key]
                            if cnpj_val:
                                colab.cnpj = cnpj_val
                            if emp_name_val:
                                colab.empresa_nome = emp_name_val
                            colabs_to_update.append(colab)
                                
                    if colabs_to_update:
                        ColaboradorSRA.objects.using(active_db).bulk_update(colabs_to_update, ['cnpj', 'empresa_nome'])
                        updated_count = len(colabs_to_update)

                if coluna_avisos:
                    messages.warning(request, f"Atenção: não encontrei a coluna de {' e '.join(coluna_avisos)} na planilha enviada — esse(s) campo(s) não foi(ram) atualizado(s). Verifique o cabeçalho da planilha.")

                if updated_count == 0:
                    messages.warning(request, "Nenhum colaborador foi atualizado: nenhum Código Empresa/Filial da planilha bateu com os cadastrados na base atual. Confira se os códigos da planilha de empresas correspondem aos da Base Layout 1 desta regional.")
                else:
                    messages.success(request, f"Empresas importadas com sucesso! {updated_count} colaboradores atualizados com CNPJ e Razão Social.")
            except Exception as e:
                messages.error(request, f"Erro ao processar aba 2.Layout_BD_CNPJ: {str(e)}")
            finally:
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)
            return redirect("gestao_a_vista:psicossocial_sra")
            
        elif action == "create_colaborador":
            cpf_val = request.POST.get("cpf", "").strip()
            nome_val = request.POST.get("nome", "").strip()
            matricula_val = request.POST.get("matricula", "").strip()
            negocio_val = request.POST.get("negocio", "").strip()
            municipio_val = normalize_city_name(request.POST.get("municipio", ""))
            estado_val = request.POST.get("estado", "").strip()
            dt_demissao_val = request.POST.get("dt_demissao", "").strip()
            cr_val = request.POST.get("cr", "").strip()
            cliente_val = request.POST.get("cliente", "").strip()
            situacao_val = request.POST.get("situacao", "Ativo").strip()
            
            grupo_cliente_val = request.POST.get("grupo_cliente", "").strip()
            empresa_codigo_val = request.POST.get("empresa_codigo", "").strip()
            filial_codigo_val = request.POST.get("filial_codigo", "").strip()
            nome_unidade_val = request.POST.get("nome_unidade", "").strip()
            cnpj_val = request.POST.get("cnpj", "").strip()
            empresa_nome_val = request.POST.get("empresa_nome", "").strip()
            
            if not cpf_val or not nome_val:
                messages.error(request, "CPF e Nome são campos obrigatórios.")
                return redirect("gestao_a_vista:psicossocial_sra")
                
            cpf_cleaned = "".join(filter(str.isdigit, cpf_val))
            if not cpf_cleaned:
                messages.error(request, "CPF inválido.")
                return redirect("gestao_a_vista:psicossocial_sra")
                
            # Parse Date
            dt_demissao_date = None
            if dt_demissao_val:
                try:
                    dt_demissao_date = datetime.strptime(dt_demissao_val, "%Y-%m-%d").date()
                except ValueError:
                    messages.error(request, "Formato de data de demissão inválido.")
                    return redirect("gestao_a_vista:psicossocial_sra")
                    
            try:
                ColaboradorSRA.objects.create(
                    cpf=cpf_cleaned,
                    nome=nome_val,
                    matricula=matricula_val,
                    negocio=negocio_val,
                    municipio=municipio_val,
                    estado=estado_val,
                    dt_demissao=dt_demissao_date,
                    cr=cr_val,
                    cliente=cliente_val,
                    situacao=situacao_val,
                    grupo_cliente=grupo_cliente_val,
                    empresa_codigo=empresa_codigo_val,
                    filial_codigo=filial_codigo_val,
                    nome_unidade=nome_unidade_val,
                    cnpj=cnpj_val,
                    empresa_nome=empresa_nome_val
                )
                messages.success(request, f"Colaborador '{nome_val}' cadastrado com sucesso na base Layout 1.")
            except Exception as e:
                messages.error(request, f"Erro ao salvar colaborador: {str(e)}")
            return redirect("gestao_a_vista:psicossocial_sra")
            
    # GET
    q = request.GET.get("q", "").strip()
    cliente_filter = request.GET.get("cliente_filter", "").strip()
    status_filter = request.GET.get("status_filter", "").strip()
    show_companies = request.GET.get("show_companies") == "1"
    
    colaboradores_list = ColaboradorSRA.objects.all().order_by("nome")
    clientes_qs = ColaboradorSRA.objects.exclude(cliente__isnull=True).exclude(cliente='')
    
    # Filter list and client dropdown by user's regional/CRs if not a global admin
    if not getattr(request.user, 'is_global_admin', False):
        user_crs = [c.strip() for c in request.user.crs.split(',') if c.strip()] if getattr(request.user, 'crs', None) else []
        if user_crs:
            colaboradores_list = colaboradores_list.filter(cr__in=user_crs)
            clientes_qs = clientes_qs.filter(cr__in=user_crs)

    if q:
        colaboradores_list = colaboradores_list.filter(
            Q(nome__icontains=q) | Q(cpf__icontains=q) | Q(matricula__icontains=q) | Q(negocio__icontains=q)
        )
    if cliente_filter:
        colaboradores_list = colaboradores_list.filter(cliente=cliente_filter)
        
    if status_filter == "ativos":
        colaboradores_list = colaboradores_list.filter(dt_demissao__isnull=True).filter(Q(situacao__iexact="ativo") | Q(situacao__isnull=True) | Q(situacao=""))
    elif status_filter == "inativos":
        colaboradores_list = colaboradores_list.filter(Q(dt_demissao__isnull=False) | ~Q(situacao__iexact="ativo"))
    elif status_filter == "ferias":
        colaboradores_list = colaboradores_list.filter(Q(situacao__icontains="férias") | Q(situacao__icontains="ferias"))
    elif status_filter == "demitidos":
        colaboradores_list = colaboradores_list.filter(dt_demissao__isnull=False)
        
    clientes = clientes_qs.values_list('cliente', flat=True).distinct().order_by('cliente')
    sra_count = colaboradores_list.count()
    
    # Extract unique companies query set
    empresas_qs = ColaboradorSRA.objects.exclude(Q(empresa_codigo__isnull=True) | Q(empresa_codigo=''))
    if not getattr(request.user, 'is_global_admin', False):
        user_crs = [c.strip() for c in request.user.crs.split(',') if c.strip()] if getattr(request.user, 'crs', None) else []
        if user_crs:
            empresas_qs = empresas_qs.filter(cr__in=user_crs)

    if q:
        empresas_qs = empresas_qs.filter(
            Q(empresa_nome__icontains=q) | Q(cnpj__icontains=q) | Q(empresa_codigo__icontains=q) | Q(filial_codigo__icontains=q)
        )
        
    empresas_list = empresas_qs.values(
        'empresa_codigo', 'filial_codigo', 'empresa_nome', 'cnpj'
    ).distinct().order_by('empresa_codigo', 'filial_codigo')
    
    empresas_count = empresas_list.count()
    
    from django.core.paginator import Paginator
    if show_companies:
        paginator = Paginator(empresas_list, 50)
        page_number = request.GET.get("page")
        page_obj = paginator.get_page(page_number)
        for emp in page_obj:
            emp['cnpj'] = format_cnpj(emp['cnpj'])
        colaboradores = []
        empresas = page_obj
    else:
        paginator = Paginator(colaboradores_list, 50)
        page_number = request.GET.get("page")
        page_obj = paginator.get_page(page_number)
        colaboradores = page_obj
        empresas = []
        
    return render(request, "psicossocial_sra.html", {
        "colaboradores": colaboradores,
        "empresas": empresas,
        "show_companies": show_companies,
        "page_obj": page_obj,
        "q": q,
        "cliente_filter": cliente_filter,
        "status_filter": status_filter,
        "clientes": clientes,
        "sra_count": sra_count,
        "empresas_count": empresas_count
    })


@login_required
def psicossocial_sra_delete(request, pk):
    if not (request.user.role == "administrador" or request.user.has_page_permission("psicossocial")):
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied("Você não tem permissão para acessar esta página.")
        
    colaborador = get_object_or_404(ColaboradorSRA, pk=pk)
    colaborador.delete()
    messages.success(request, f"Colaborador {colaborador.nome} excluído com sucesso da base Layout 1.")
    return redirect("gestao_a_vista:psicossocial_sra")


@login_required
def psicossocial_autocomplete(request):
    if not (request.user.role == "administrador" or request.user.has_page_permission("psicossocial")):
        from django.http import JsonResponse
        return JsonResponse({"error": "Unauthorized"}, status=403)
        
    field = request.GET.get("field", "").strip()
    q = request.GET.get("q", "").strip()
    
    if not field or len(q) < 2:
        from django.http import JsonResponse
        return JsonResponse([], safe=False)
        
    from django.http import JsonResponse
    
    queries_starts = []
    queries_contains = []
    
    if field == 'projeto_nome':
        # Search project names, client, nome_unidade, negocio and CR
        queries_starts = [
            (PsicossocialProjeto, 'nome', q),
            (ColaboradorSRA, 'cliente', q),
            (ColaboradorSRA, 'nome_unidade', q),
            (ColaboradorSRA, 'negocio', q),
            (ColaboradorSRA, 'cr', q)
        ]
        queries_contains = [
            (PsicossocialProjeto, 'nome', q),
            (ColaboradorSRA, 'cliente', q),
            (ColaboradorSRA, 'nome_unidade', q),
            (ColaboradorSRA, 'negocio', q),
            (ColaboradorSRA, 'cr', q)
        ]
    elif field == 'empresa':
        # Search company names, client, negocio and CR
        queries_starts = [
            (PsicossocialProjeto, 'empresa', q),
            (ColaboradorSRA, 'cliente', q),
            (ColaboradorSRA, 'negocio', q),
            (ColaboradorSRA, 'cr', q)
        ]
        queries_contains = [
            (PsicossocialProjeto, 'empresa', q),
            (ColaboradorSRA, 'cliente', q),
            (ColaboradorSRA, 'negocio', q),
            (ColaboradorSRA, 'cr', q)
        ]
    elif field == 'localidade':
        # Search localidade, and municipio/estado from SRA
        queries_starts = [
            (PsicossocialProjeto, 'localidade', q),
            (ColaboradorSRA, 'localidade_sra', q)
        ]
        queries_contains = [
            (PsicossocialProjeto, 'localidade', q),
            (ColaboradorSRA, 'localidade_sra', q)
        ]
    elif field == 'cnpj':
        queries_starts = [(PsicossocialProjeto, 'cnpj', q)]
        queries_contains = [(PsicossocialProjeto, 'cnpj', q)]
    elif field == 'responsavel_tecnico':
        queries_starts = [(PsicossocialProjeto, 'responsavel_tecnico', q)]
        queries_contains = [(PsicossocialProjeto, 'responsavel_tecnico', q)]
    elif field == 'nome':
        queries_starts = [(ColaboradorSRA, 'nome', q)]
        queries_contains = [(ColaboradorSRA, 'nome', q)]
    elif field == 'cpf':
        queries_starts = [(ColaboradorSRA, 'cpf', q)]
        queries_contains = [(ColaboradorSRA, 'cpf', q)]
    elif field == 'matricula':
        queries_starts = [(ColaboradorSRA, 'matricula', q)]
        queries_contains = [(ColaboradorSRA, 'matricula', q)]
    elif field == 'negocio':
        queries_starts = [(ColaboradorSRA, 'negocio', q), (ColaboradorSRA, 'cr', q)]
        queries_contains = [(ColaboradorSRA, 'negocio', q), (ColaboradorSRA, 'cr', q)]
    elif field == 'municipio':
        queries_starts = [(ColaboradorSRA, 'municipio', q)]
        queries_contains = [(ColaboradorSRA, 'municipio', q)]
    elif field == 'estado':
        queries_starts = [(ColaboradorSRA, 'estado', q)]
        queries_contains = [(ColaboradorSRA, 'estado', q)]
    elif field == 'cr':
        queries_starts = [(ColaboradorSRA, 'cr', q)]
        queries_contains = [(ColaboradorSRA, 'cr', q)]
    elif field == 'cliente':
        queries_starts = [(ColaboradorSRA, 'cliente', q)]
        queries_contains = [(ColaboradorSRA, 'cliente', q)]
    elif field == 'empresa_codigo':
        queries_starts = [(ColaboradorSRA, 'empresa_codigo', q), (PsicossocialProjeto, 'empresa_codigo', q)]
        queries_contains = [(ColaboradorSRA, 'empresa_codigo', q), (PsicossocialProjeto, 'empresa_codigo', q)]
    elif field == 'filial_codigo':
        queries_starts = [(ColaboradorSRA, 'filial_codigo', q), (PsicossocialProjeto, 'filial_codigo', q)]
        queries_contains = [(ColaboradorSRA, 'filial_codigo', q), (PsicossocialProjeto, 'filial_codigo', q)]
        
    limit = 10
    results = []
    seen = set()
    
    from django.db.models import Q
    
    active_filter = Q(dt_demissao__isnull=True) & (
        Q(situacao__iexact="normal") |
        Q(situacao__iexact="ativo") |
        Q(situacao__iexact="ativa") |
        Q(situacao__icontains="férias") |
        Q(situacao__icontains="ferias") |
        Q(situacao__isnull=True) |
        Q(situacao="")
    )
    
    # 1. Process starts-with queries
    for model, db_field, q_val in queries_starts:
        if db_field == 'localidade_sra':
            items = model.objects.filter(active_filter if model == ColaboradorSRA else Q()).filter(municipio__istartswith=q_val)
            items = items.values('municipio', 'estado').distinct()[:limit]
            vals = [f"{item['municipio']} - {item['estado']}".strip() for item in items if item['municipio'] and item['estado']]
        else:
            qs = model.objects
            if model == ColaboradorSRA:
                qs = qs.filter(active_filter)
            vals = list(qs.filter(**{f"{db_field}__istartswith": q_val}).values_list(db_field, flat=True).distinct()[:limit])
            vals = [str(v).strip() for v in vals if v]
            
        for v in vals:
            if v.lower() not in seen:
                seen.add(v.lower())
                results.append(v)
                if len(results) >= limit:
                    break
        if len(results) >= limit:
            break
            
    # 2. Process contains queries
    if len(results) < limit:
        for model, db_field, q_val in queries_contains:
            if db_field == 'localidade_sra':
                items = model.objects.filter(active_filter if model == ColaboradorSRA else Q()).filter(municipio__icontains=q_val).exclude(municipio__istartswith=q_val)
                items = items.values('municipio', 'estado').distinct()[:limit - len(results)]
                vals = [f"{item['municipio']} - {item['estado']}".strip() for item in items if item['municipio'] and item['estado']]
            else:
                qs = model.objects
                if model == ColaboradorSRA:
                    qs = qs.filter(active_filter)
                vals = list(qs.filter(**{f"{db_field}__icontains": q_val}).exclude(**{f"{db_field}__istartswith": q_val}).values_list(db_field, flat=True).distinct()[:limit - len(results)])
                vals = [str(v).strip() for v in vals if v]
                
            for v in vals:
                if v.lower() not in seen:
                    seen.add(v.lower())
                    results.append(v)
                    if len(results) >= limit:
                        break
            if len(results) >= limit:
                break
                
    return JsonResponse(results, safe=False)


@login_required
def psicossocial_sra_list_api(request):
    if not (request.user.role == "administrador" or request.user.has_page_permission("psicossocial")):
        from django.http import JsonResponse
        return JsonResponse({"error": "Unauthorized"}, status=403)
        
    q = request.GET.get("q", "").strip()
    cliente = request.GET.get("cliente", "").strip()
    unidade = request.GET.get("unidade", "").strip()
    
    from django.db.models import Q
    active_sra = ColaboradorSRA.objects.filter(
        Q(dt_demissao__isnull=True) & (
            Q(situacao__iexact="normal") |
            Q(situacao__iexact="ativo") |
            Q(situacao__iexact="ativa") |
            Q(situacao__icontains="férias") |
            Q(situacao__icontains="ferias") |
            Q(situacao__isnull=True) |
            Q(situacao="")
        )
    )
    if cliente:
        active_sra = active_sra.filter(cliente=cliente)
    if unidade:
        active_sra = active_sra.filter(Q(nome_unidade=unidade) | Q(cliente=unidade))
        
    if q:
        active_sra = active_sra.filter(Q(nome__icontains=q) | Q(cpf__icontains=q))
        
    colabs_list = list(active_sra.values('pk', 'nome', 'cpf', 'cliente', 'nome_unidade', 'situacao').order_by('nome')[:500])
    
    from django.http import JsonResponse
    return JsonResponse(colabs_list, safe=False)


def install_and_copy_financeiro():
    import os
    import sys
    import shutil
    import subprocess
    from pathlib import Path

    # 1. Install pip requirements inside the current venv python
    try:
        import pandas as pd
        import openpyxl
        import ollama
    except ImportError:
        print("Installing financial requirements (pandas, openpyxl, ollama, truststore)...")
        try:
            subprocess.check_call([
                sys.executable, "-m", "pip", "install", 
                "--trusted-host", "pypi.org", 
                "--trusted-host", "files.pythonhosted.org", 
                "--trusted-host", "pypi.python.org", 
                "pandas", "openpyxl", "ollama", "truststore"
            ])
            print("Installation completed!")
        except Exception as err:
            print(f"Error executing pip install: {err}")

    # 2. Define source and destination paths
    root_dir = Path(__file__).resolve().parent.parent
    src_dir = root_dir / "Projeto-Financeiro"
    dest_dir = root_dir / "Gestao_a_Vista" / "financeiro"

    # Create directories
    for sub in ["", "config", "data", "data/input", "data/processed", "Base de dados", "Base de dados/Compras Produto", "Base de dados/Resultado RE X OR"]:
        (dest_dir / sub).mkdir(parents=True, exist_ok=True)
    
    (dest_dir / "__init__.py").touch(exist_ok=True)

    # 3. Copy python files from src/ to financeiro/
    src_code = src_dir / "src"
    if src_code.exists():
        for item in src_code.iterdir():
            if item.is_file() and item.suffix == ".py":
                dest_file = dest_dir / item.name
                if not dest_file.exists() or dest_file.stat().st_mtime < item.stat().st_mtime:
                    shutil.copy2(item, dest_file)
                    print(f"Copied {item.name} to {dest_file}")

    # 4. Copy config files
    src_config = src_dir / "config" / "piloto.json"
    if src_config.exists():
        dest_config = dest_dir / "config" / "piloto.json"
        if not dest_config.exists() or dest_config.stat().st_mtime < src_config.stat().st_mtime:
            shutil.copy2(src_config, dest_config)
            print("Copied piloto.json configuration.")

# Run the installer/copier once on module load
try:
    install_and_copy_financeiro()
except Exception as e:
    print(f"Error in install_and_copy_financeiro: {e}")


def filter_dataframe(df, filters):
    if df.empty:
        return df
    
    filtered = df
    selected_months = filters.get("meses", [])
    if selected_months:
        filtered = filtered[filtered["mes"].astype(str).isin(selected_months)]
        
    selected_alerts = filters.get("alertas", [])
    if selected_alerts:
        filtered = filtered[filtered["nivel_alerta"].isin(selected_alerts)]
        
    selected_pecs = filters.get("pecs", [])
    if selected_pecs:
        filtered = filtered[filtered["pec"].astype(str).isin(selected_pecs)]
        
    selected_cc_sups = filters.get("cc_sups", [])
    if selected_cc_sups and "cc_sup" in filtered:
        filtered = filtered[filtered["cc_sup"].astype(str).isin(selected_cc_sups)]
        
    selected_accounts = filters.get("contas", [])
    if selected_accounts:
        filtered = filtered[filtered["conta_contabil"].astype(str).isin(selected_accounts)]
        
    selected_suppliers = filters.get("fornecedores", [])
    if selected_suppliers and "fornecedores" in filtered:
        filtered = filtered[
            filtered["fornecedores"].fillna("").apply(
                lambda value: any(str(supplier) in str(value) for supplier in selected_suppliers)
            )
        ]
    return filtered


def filter_closing_dataframe(df, filters):
    if df.empty:
        return df
    
    filtered = df
    selected_months = filters.get("meses", [])
    if selected_months and "mes" in filtered:
        filtered = filtered[filtered["mes"].astype(str).isin(selected_months)]
        
    selected_alerts = filters.get("alertas", [])
    if selected_alerts and "nivel_alerta_fechamento" in filtered:
        filtered = filtered[filtered["nivel_alerta_fechamento"].isin(selected_alerts)]
        
    selected_pecs = filters.get("pecs", [])
    if selected_pecs and "pec" in filtered:
        filtered = filtered[filtered["pec"].astype(str).isin(selected_pecs)]
        
    selected_cc_sups = filters.get("cc_sups", [])
    if selected_cc_sups and "cc_sup" in filtered:
        filtered = filtered[filtered["cc_sup"].astype(str).isin(selected_cc_sups)]
        
    selected_accounts = filters.get("contas", [])
    if selected_accounts and "conta_contabil" in filtered:
        selected_account_text = {str(account) for account in selected_accounts}
        filtered = filtered[
            filtered["conta_contabil"].fillna("").apply(
                lambda value: any(str(account) in str(value) for account in selected_account_text)
            )
        ]
    return filtered


@login_required
@check_page_permission("financeiro")
def financeiro(request):
    """
    Renderiza o dashboard do Financeiro integrado nativamente
    """
    from Gestao_a_Vista.financeiro.config import load_config
    config = load_config()
    
    return render(request, "financeiro.html", {
        "nome_carteira": config.nome_carteira,
        "destinatario_padrao": config.destinatario_padrao,
    })


@login_required
@check_page_permission("financeiro")
def financeiro_api_data(request):
    import json
    import pandas as pd
    import traceback
    from django.http import JsonResponse
    
    try:
        from Gestao_a_Vista.financeiro.pipeline import load_processed_accounts, load_processed_closing
        from Gestao_a_Vista.financeiro.finance import summarize_kpis, summarize_alert_counts, top_costs_by_dimension, month_variation
        
        # Load request parameters
        filters = {}
        if request.method == "POST":
            try:
                filters = json.loads(request.body)
            except Exception:
                pass
                
        raw_accounts = load_processed_accounts()
        raw_closing = load_processed_closing()
        
        # Dynamically extract filter options from raw loaded data
        filter_options = {
            "meses": [],
            "pecs": [],
            "cc_sups": [],
            "contas": [],
            "fornecedores": [],
            "alertas": ["sem_orcamento", "estourado", "critico", "atencao", "sem_alerta"]
        }
        
        if not raw_accounts.empty:
            filter_options["meses"] = sorted([str(x) for x in raw_accounts["mes"].dropna().unique() if str(x).strip()])
            filter_options["pecs"] = sorted([str(x) for x in raw_accounts["pec"].dropna().unique() if str(x).strip()])
            if "cc_sup" in raw_accounts:
                filter_options["cc_sups"] = sorted([str(x) for x in raw_accounts["cc_sup"].dropna().unique() if str(x).strip()])
            filter_options["contas"] = sorted([str(x) for x in raw_accounts["conta_contabil"].dropna().unique() if str(x).strip()])
            
            suppliers_set = set()
            if "fornecedores" in raw_accounts:
                for val in raw_accounts["fornecedores"].dropna():
                    for sup in str(val).split(","):
                        sup_clean = sup.strip()
                        if sup_clean:
                            suppliers_set.add(sup_clean)
            filter_options["fornecedores"] = sorted(list(suppliers_set))
            
        filtered_accounts = filter_dataframe(raw_accounts, filters)
        filtered_closing = filter_closing_dataframe(raw_closing, filters)
        
        # KPI Calculations
        kpis = {}
        if not filtered_accounts.empty:
            costs = filtered_accounts[filtered_accounts["tipo"] == "custo"]
            revenue = filtered_accounts[filtered_accounts["tipo"] == "receita"]
            
            budgeted_revenue = float(revenue["orcado"].clip(lower=0).sum()) if not revenue.empty else 0.0
            cost_budget = abs(float(costs["orcado"].sum())) if not costs.empty else 0.0
            realized_costs = float(costs["realizado_abs"].sum()) if not costs.empty else 0.0
            
            kpis["receita_orcada"] = budgeted_revenue
            kpis["orcamento_custos"] = cost_budget
            kpis["realizado_custos"] = realized_costs
            kpis["saldo_custos"] = cost_budget - realized_costs
            kpis["excesso_total"] = float(costs["excesso_abs"].sum()) if not costs.empty else 0.0
            kpis["alertas_total"] = int(costs[costs["nivel_alerta"] != "sem_alerta"].shape[0]) if not costs.empty else 0
            kpis["margem_orcada"] = budgeted_revenue - cost_budget
            kpis["margem_atual_orcada"] = budgeted_revenue - realized_costs
        else:
            kpis = {
                "receita_orcada": 0, "orcamento_custos": 0, "realizado_custos": 0,
                "saldo_custos": 0, "excesso_total": 0, "alertas_total": 0,
                "margem_orcada": 0, "margem_atual_orcada": 0
            }
            
        closing_kpis = {}
        if not filtered_closing.empty:
            closing_costs = filtered_closing[filtered_closing["tipo"] == "custo"]
            total_budget = float(closing_costs["orcado"].sum()) if not closing_costs.empty else 0.0
            total_realized = float(closing_costs["realizado_fechado_abs"].sum()) if not closing_costs.empty else 0.0
            total_purchased = float(closing_costs["comprado_ate_agora_abs"].sum()) if not closing_costs.empty else 0.0
            
            closing_kpis["orcado_custos"] = abs(total_budget)
            closing_kpis["realizado_fechado"] = total_realized
            closing_kpis["comprado_exportado"] = total_purchased
            closing_kpis["diferenca"] = total_realized - total_purchased
            closing_kpis["excesso"] = float(closing_costs["excesso_fechamento"].sum()) if not closing_costs.empty else 0.0
            closing_kpis["alertas"] = int((closing_costs["nivel_alerta_fechamento"] != "sem_alerta").sum()) if not closing_costs.empty else 0
        else:
            closing_kpis = {
                "orcado_custos": 0, "realizado_fechado": 0, "comprado_exportado": 0,
                "diferenca": 0, "excesso": 0, "alertas": 0
            }
            
        # Alert Counts
        alert_counts = {"sem_orcamento": 0, "estourado": 0, "critico": 0, "atencao": 0}
        if not filtered_accounts.empty:
            counts = filtered_accounts[filtered_accounts["tipo"] == "custo"]["nivel_alerta"].value_counts().to_dict()
            for key in alert_counts:
                alert_counts[key] = int(counts.get(key, 0))
                
        closing_alert_counts = {"sem_orcamento": 0, "estourado": 0, "critico": 0, "atencao": 0}
        if not filtered_closing.empty:
            counts = filtered_closing[filtered_closing["tipo"] == "custo"]["nivel_alerta_fechamento"].value_counts().to_dict()
            for key in closing_alert_counts:
                closing_alert_counts[key] = int(counts.get(key, 0))
                
        # Chart coordinates
        charts = {
            "alert_distribution": {
                "labels": ["Sem orçamento", "Estourado", "Crítico", "Atenção", "Sem alerta"],
                "values": [0, 0, 0, 0, 0]
            },
            "budget_vs_spent": {
                "months": [],
                "budgeted": [],
                "spent": []
            },
            "top_excesses": {
                "labels": [],
                "values": [],
                "colors": []
            }
        }
        
        if not filtered_accounts.empty:
            # Alert distribution values
            counts = filtered_accounts["nivel_alerta"].value_counts().to_dict()
            charts["alert_distribution"]["values"] = [
                int(counts.get("sem_orcamento", 0)),
                int(counts.get("estourado", 0)),
                int(counts.get("critico", 0)),
                int(counts.get("atencao", 0)),
                int(counts.get("sem_alerta", 0))
            ]
            
            # Budget vs spent by month
            monthly = (
                filtered_accounts[filtered_accounts["tipo"] == "custo"]
                .groupby("mes", dropna=False)
                .agg(orcado_abs=("orcado_abs", "sum"), realizado_abs=("realizado_abs", "sum"))
                .reset_index()
                .sort_values("mes")
            )
            charts["budget_vs_spent"]["months"] = [str(x) for x in monthly["mes"]]
            charts["budget_vs_spent"]["budgeted"] = [float(x) for x in monthly["orcado_abs"]]
            charts["budget_vs_spent"]["spent"] = [float(x) for x in monthly["realizado_abs"]]
            
            # Top 12 Excesses
            excess = (
                filtered_accounts[(filtered_accounts["tipo"] == "custo") & (filtered_accounts["excesso_abs"] > 0)]
                .sort_values("excesso_abs", ascending=False)
                .head(12)
            )
            COLOR_MAP = {"sem_orcamento": "#7c2d12", "estourado": "#b91c1c", "critico": "#c2410c", "atencao": "#a16207", "sem_alerta": "#047857"}
            if not excess.empty:
                charts["top_excesses"]["labels"] = [str(x)[:35] for x in excess["conta_contabil"]]
                charts["top_excesses"]["values"] = [float(x) for x in excess["excesso_abs"]]
                charts["top_excesses"]["colors"] = [COLOR_MAP.get(x, "#047857") for x in excess["nivel_alerta"]]
                
        # Serialize Tables safely
        def serialize_df(df):
            if df.empty:
                return []
            cleaned = df.copy()
            for col in cleaned.columns:
                cleaned[col] = cleaned[col].astype(object).where(cleaned[col].notnull(), None)
            return cleaned.to_dict(orient="records")
            
        tables = {
            "sem_orcamento": [],
            "estouradas": [],
            "perto_de_estourar": [],
            "base_detalhada": [],
            "closing_estouradas": [],
            "closing_perto_de_estourar": [],
            "closing_so_re_x_or": [],
            "closing_diferencas": [],
            "closing_base_detalhada": []
        }
        
        if not filtered_accounts.empty:
            tables["sem_orcamento"] = serialize_df(filtered_accounts[(filtered_accounts["tipo"] == "custo") & (filtered_accounts["nivel_alerta"] == "sem_orcamento")].sort_values("realizado_abs", ascending=False))
            tables["estouradas"] = serialize_df(filtered_accounts[(filtered_accounts["tipo"] == "custo") & (filtered_accounts["nivel_alerta"] == "estourado")].sort_values("excesso_abs", ascending=False))
            tables["perto_de_estourar"] = serialize_df(filtered_accounts[(filtered_accounts["tipo"] == "custo") & (filtered_accounts["nivel_alerta"].isin(["critico", "atencao"]))].sort_values("consumo_pct", ascending=False))
            tables["base_detalhada"] = serialize_df(filtered_accounts.sort_values("realizado_abs", ascending=False))
            
        if not filtered_closing.empty:
            tables["closing_estouradas"] = serialize_df(filtered_closing[(filtered_closing["tipo"] == "custo") & (filtered_closing["nivel_alerta_fechamento"] == "estourado")].sort_values("excesso_fechamento", ascending=False))
            tables["closing_perto_de_estourar"] = serialize_df(filtered_closing[(filtered_closing["tipo"] == "custo") & (filtered_closing["nivel_alerta_fechamento"].isin(["critico", "atencao"])) & (filtered_closing["consumo_pct_fechamento"] >= 0.8) & (filtered_closing["consumo_pct_fechamento"] < 1.0)].sort_values("consumo_pct_fechamento", ascending=False))
            tables["closing_so_re_x_or"] = serialize_df(filtered_closing[filtered_closing["origem_observacao"] == "somente_realizado_re_x_or"].sort_values("realizado_fechado_abs", ascending=False))
            tables["closing_diferencas"] = serialize_df(filtered_closing[filtered_closing["diferenca_realizado_vs_compras_abs"] > 0.01].sort_values("diferenca_realizado_vs_compras_abs", ascending=False))
            tables["closing_base_detalhada"] = serialize_df(filtered_closing.sort_values("realizado_fechado_abs", ascending=False))
        
        def sanitize_json_data(data):
            import math
            if isinstance(data, dict):
                return {k: sanitize_json_data(v) for k, v in data.items()}
            elif isinstance(data, list):
                return [sanitize_json_data(x) for x in data]
            elif isinstance(data, float):
                if math.isnan(data) or math.isinf(data):
                    return None
                return data
            elif pd.isna(data):
                return None
            return data

        response_data = {
            "kpis": kpis,
            "closing_kpis": closing_kpis,
            "filter_options": filter_options,
            "alert_counts": alert_counts,
            "closing_alert_counts": closing_alert_counts,
            "charts": charts,
            "tables": tables
        }
        return JsonResponse(sanitize_json_data(response_data))
    except Exception as e:
        tb = traceback.format_exc()
        try:
            with open("dashboard_error_debug.txt", "a", encoding="utf-8") as f:
                f.write(f"\n--- API DATA ERROR ---\n{tb}")
        except Exception:
            pass
        return JsonResponse({"error": f"{str(e)}\n{tb}"}, status=500)
    
    
@login_required
@check_page_permission("financeiro")
def financeiro_api_chat(request):
    import json
    from django.http import JsonResponse
    
    if request.method != "POST":
        return JsonResponse({"error": "POST method required"}, status=405)
        
    try:
        data = json.loads(request.body)
    except Exception:
        data = {}
        
    prompt = data.get("prompt", "").strip()
    history = data.get("history", [])
    answer_mode = data.get("answer_mode", "auto")
    filters = data.get("filters", {})
    
    from Gestao_a_Vista.financeiro.pipeline import load_processed_accounts, load_processed_closing
    from Gestao_a_Vista.financeiro.assistant import answer_question
    
    accounts = filter_dataframe(load_processed_accounts(), filters)
    closing = filter_closing_dataframe(load_processed_closing(), filters)
    
    try:
        response = answer_question(prompt, accounts, closing, history=history, answer_mode=answer_mode)
        return JsonResponse({"response": response})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@login_required
@check_page_permission("financeiro")
def financeiro_api_email(request):
    import json
    from django.http import JsonResponse
    
    if request.method != "POST":
        return JsonResponse({"error": "POST method required"}, status=405)
        
    try:
        data = json.loads(request.body)
    except Exception:
        data = {}
        
    filters = data.get("filters", {})
    
    from Gestao_a_Vista.financeiro.pipeline import load_processed_accounts
    from Gestao_a_Vista.financeiro.assistant import draft_email
    
    accounts = filter_dataframe(load_processed_accounts(), filters)
    
    try:
        html = draft_email(accounts)
        return JsonResponse({"html": html})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@login_required
@check_page_permission("financeiro")
def financeiro_api_upload(request):
    from django.http import JsonResponse
    import traceback
    
    if request.method != "POST":
        return JsonResponse({"error": "POST method required"}, status=405)
        
    dataset = request.POST.get("dataset", "").strip()
    month = request.POST.get("month", "").strip()
    uploaded_file = request.FILES.get("file")
    
    if not uploaded_file:
        return JsonResponse({"error": "No file uploaded"}, status=400)
    if dataset not in ("compras", "resultado"):
        return JsonResponse({"error": "Invalid dataset type"}, status=400)
    if not month:
        return JsonResponse({"error": "Month label required"}, status=400)
        
    from Gestao_a_Vista.financeiro.uploads import save_uploaded_planilha
    
    try:
        file_bytes = uploaded_file.read()
        saved_path = save_uploaded_planilha(
            content=file_bytes,
            original_name=uploaded_file.name,
            dataset=dataset,
            month_label=month,
        )
        return JsonResponse({"success": True, "path": str(saved_path)})
    except PermissionError as pe:
        import os
        import stat
        tb = traceback.format_exc()
        diag_lines = [f"PermissionError: {str(pe)}", f"Traceback:\n{tb}"]
        try:
            diag_lines.append(f"Current Process UID: {os.getuid()} GID: {os.getgid()}")
        except Exception:
            diag_lines.append("Could not retrieve UID/GID (non-posix platform)")
            
        from Gestao_a_Vista.financeiro.pipeline import get_input_dir
        active_input_dir = get_input_dir()
        compras_dir = active_input_dir / "Compras Produto"
        resultado_dir = active_input_dir / "Resultado RE X OR"
        
        for name, p in [("INPUT_DIR", active_input_dir), ("Compras Produto", compras_dir), ("Resultado RE X OR", resultado_dir)]:
            if p.exists():
                try:
                    st = p.stat()
                    diag_lines.append(f"Path: {p} | Mode: {oct(st.st_mode)} | Owner UID: {st.st_uid} GID: {st.st_gid}")
                except Exception as ex:
                    diag_lines.append(f"Path: {p} | Stat failed: {str(ex)}")
            else:
                diag_lines.append(f"Path: {p} | Does not exist")
                
        diag_str = "\n".join(diag_lines)
        try:
            with open("upload_debug.txt", "w", encoding="utf-8") as f:
                f.write(diag_str)
        except Exception:
            pass
            
        return JsonResponse({
            "error": f"O arquivo de destino está bloqueado ou a pasta de uploads no servidor não tem permissão de escrita. Detalhes: {diag_str[:200]}...",
            "details": diag_str
        }, status=500)
    except ValueError as ve:
        return JsonResponse({"error": str(ve)}, status=400)
    except Exception as e:
        tb = traceback.format_exc()
        try:
            with open("upload_debug.txt", "w", encoding="utf-8") as f:
                f.write(f"--- UPLOAD ERROR ---\n{tb}")
        except Exception:
            pass
        return JsonResponse({"error": f"{str(e)}\n{tb}"}, status=500)


@login_required
@check_page_permission("financeiro")
def financeiro_api_run_pipeline(request):
    from django.http import JsonResponse
    import traceback
    
    if request.method != "POST":
        return JsonResponse({"error": "POST method required"}, status=405)
        
    from Gestao_a_Vista.financeiro.config import load_config
    from Gestao_a_Vista.financeiro.pipeline import run_pipeline
    
    try:
        config = load_config()
        result = run_pipeline(config)
        
        serialized = {
            "raw_rows": result.get("raw_rows", 0),
            "account_rows": result.get("account_rows", 0),
            "closing_rows": result.get("closing_rows", 0),
            "budget_rows": result.get("budget_rows", 0),
        }
        return JsonResponse({"success": True, "result": serialized})
    except ValueError as ve:
        return JsonResponse({"error": str(ve)}, status=400)
    except Exception as e:
        tb = traceback.format_exc()
        try:
            with open("upload_debug.txt", "a", encoding="utf-8") as f:
                f.write(f"\n--- RUN PIPELINE ERROR ---\n{tb}")
        except Exception:
            pass
        return JsonResponse({"error": f"{str(e)}\n{tb}"}, status=500)


@login_required
@check_page_permission("financeiro")
def financeiro_api_clear(request):
    from django.http import JsonResponse
    import traceback
    import shutil
    import os
    import stat
    from Gestao_a_Vista.financeiro.pipeline import get_input_dir, get_processed_dir
    
    if request.method != "POST":
        return JsonResponse({"error": "POST method required"}, status=405)
        
    log_messages = []
    
    def clear_dir_contents(dir_path):
        if not dir_path.exists():
            log_messages.append(f"Directory does not exist: {dir_path}")
            return
            
        log_messages.append(f"Clearing directory: {dir_path}")
        
        def remove_readonly(func, p, excinfo):
            try:
                os.chmod(p, stat.S_IWRITE)
                func(p)
                log_messages.append(f"Successfully cleared readonly path: {p}")
            except Exception as ex:
                log_messages.append(f"Failed to clear readonly path: {p} (Error: {str(ex)})")

        for item in dir_path.iterdir():
            try:
                os.chmod(str(item), stat.S_IWRITE)
            except Exception:
                pass
                
            if item.is_file():
                try:
                    item.unlink()
                    log_messages.append(f"Unlinked file: {item}")
                except Exception as ex:
                    log_messages.append(f"Failed to unlink file: {item} (Error: {str(ex)})")
                    try:
                        with open(str(item), 'wb') as f:
                            f.write(b"")
                        log_messages.append(f"Emptied file content as fallback: {item}")
                    except Exception as ex2:
                        log_messages.append(f"Failed to empty file as fallback: {item} (Error: {str(ex2)})")
            elif item.is_dir():
                try:
                    shutil.rmtree(str(item), onerror=remove_readonly)
                    log_messages.append(f"Removed directory: {item}")
                except TypeError:
                    try:
                        shutil.rmtree(str(item), onexc=remove_readonly)
                        log_messages.append(f"Removed directory (onexc): {item}")
                    except Exception as ex:
                        log_messages.append(f"Failed to remove directory (onexc): {item} (Error: {str(ex)})")
                except Exception as ex:
                    log_messages.append(f"Failed to remove directory: {item} (Error: {str(ex)})")

    try:
        active_input_dir = get_input_dir()
        active_processed_dir = get_processed_dir()
        
        compras_dir = active_input_dir / "Compras Produto"
        resultado_dir = active_input_dir / "Resultado RE X OR"
        
        # Clear contents instead of deleting the base directory itself
        compras_dir.mkdir(parents=True, exist_ok=True)
        resultado_dir.mkdir(parents=True, exist_ok=True)
        active_processed_dir.mkdir(parents=True, exist_ok=True)
        
        clear_dir_contents(compras_dir)
        clear_dir_contents(resultado_dir)
        clear_dir_contents(active_processed_dir)
        
        # Write debug log
        try:
            with open("clear_debug.txt", "w", encoding="utf-8") as f:
                f.write("\n".join(log_messages))
        except Exception:
            pass
            
        return JsonResponse({
            "success": True, 
            "message": "Todos os dados financeiros e planilhas foram limpos com sucesso.",
            "details": log_messages
        })
    except Exception as e:
        tb = traceback.format_exc()
        return JsonResponse({"error": f"{str(e)}\n{tb}", "details": log_messages}, status=500)


@login_required
@check_page_permission("financeiro")
def test_ollama_connection(request):
    from django.http import JsonResponse
    import os
    import socket
    import httpx
    import requests
    from ollama import Client
    from Gestao_a_Vista.financeiro.assistant import build_ollama_client_config, OLLAMA_MODEL
    import traceback
    
    results = {}
    
    # 1. Environment variables
    results["env"] = {
        "OLLAMA_MODEL": os.getenv("OLLAMA_MODEL"),
        "OLLAMA_HOST": os.getenv("OLLAMA_HOST"),
        "OLLAMA_API_KEY_LEN": len(os.getenv("OLLAMA_API_KEY", "")),
        "OLLAMA_ENABLED": os.getenv("OLLAMA_ENABLED"),
    }
    
    # 2. DNS resolution of ollama.com
    try:
        results["dns_ollama_com"] = socket.gethostbyname("ollama.com")
    except Exception as e:
        results["dns_ollama_com"] = f"DNS Error: {str(e)}"
        
    # 3. HTTP connection with requests (no verify)
    try:
        r = requests.get("https://ollama.com", timeout=5, verify=False)
        results["http_requests_no_verify"] = f"Success (Status: {r.status_code})"
    except Exception as e:
        results["http_requests_no_verify"] = f"Failed: {str(e)}"
        
    # 4. HTTP connection with requests (verify=True)
    try:
        r = requests.get("https://ollama.com", timeout=5)
        results["http_requests_verify"] = f"Success (Status: {r.status_code})"
    except Exception as e:
        results["http_requests_verify"] = f"Failed: {str(e)}"
        
    # 5. HTTP connection with httpx (no verify)
    try:
        with httpx.Client(verify=False) as client:
            r = client.get("https://ollama.com", timeout=5)
            results["http_httpx_no_verify"] = f"Success (Status: {r.status_code})"
    except Exception as e:
        results["http_httpx_no_verify"] = f"Failed: {str(e)}"

    # 6. HTTP connection with httpx (verify=True)
    try:
        with httpx.Client() as client:
            r = client.get("https://ollama.com", timeout=5)
            results["http_httpx_verify"] = f"Success (Status: {r.status_code})"
    except Exception as e:
        results["http_httpx_verify"] = f"Failed: {str(e)}"
        
    # 7. Test direct Ollama Client
    try:
        config = build_ollama_client_config()
        results["ollama_client_config"] = {k: (v if k != "headers" else "present") for k, v in config.items()}
        
        client = Client(**config)
        response = client.chat(
            model=OLLAMA_MODEL,
            messages=[{"role": "user", "content": "ping"}],
            options={"num_predict": 1}
        )
        results["ollama_chat_test"] = f"Success response keys: {list(response.keys()) if hasattr(response, 'keys') else str(response)}"
    except Exception as e:
        results["ollama_chat_test"] = f"Failed: {str(e)}\n{traceback.format_exc()}"
        
    return JsonResponse(results)


@login_required
def gerenciar_regionais(request):
    """
    Tela de gerenciamento de regionais - Exclusiva do Administrador Supremo (Global)
    """
    if not getattr(request.user, "is_global_admin", False):
        return redirect("gestao_a_vista:home")

    from .models import Regional, CustomUser
    from .db_manager import check_and_create_regional_db
    from django.utils.text import slugify

    if request.method == "POST":
        nome = request.POST.get("nome")
        estado = request.POST.get("estado")
        cidade = request.POST.get("cidade")
        diretor_regional = request.POST.get("diretor_regional")
        diretor_executivo = request.POST.get("diretor_executivo")

        if not nome or not estado:
            messages.error(request, "Nome e Estado são campos obrigatórios.")
        else:
            try:
                estado = estado.strip().upper()
                estado_lower = estado.lower()

                # A primeira Regional de um estado mantém o alias/banco já
                # usado hoje (db_<estado>). Qualquer Regional adicional no
                # mesmo estado ganha um banco próprio, isolado, derivado do
                # nome (ex: 2ª regional em SP não compartilha mais o mesmo
                # banco físico da 1ª).
                if Regional.objects.filter(estado=estado).exists():
                    base_slug = slugify(f"{estado_lower}-{nome}")[:45] or estado_lower
                    db_slug = base_slug
                    suffix = 1
                    while Regional.objects.filter(db_slug=db_slug).exists():
                        suffix += 1
                        db_slug = f"{base_slug}-{suffix}"[:50]
                else:
                    db_slug = estado_lower

                # Tenta criar o banco de dados dinamicamente no PostgreSQL
                check_and_create_regional_db(db_slug)

                # Salva a regional no banco de dados central (default)
                Regional.objects.create(
                    nome=nome,
                    estado=estado,
                    db_slug=db_slug,
                    cidade=cidade if cidade else None,
                    diretor_regional=diretor_regional if diretor_regional else None,
                    diretor_executivo=diretor_executivo if diretor_executivo else None
                )
                messages.success(request, f"Regional '{nome}' criada com sucesso e banco '{db_slug}' provisionado!")
            except Exception as e:
                messages.error(request, f"Erro ao criar a regional: {str(e)}")
        return redirect("gestao_a_vista:gerenciar_regionais")

    regionais = Regional.objects.all().order_by("estado", "nome")
    
    # Contagem de usuários por regional
    from django.db.models import Count
    user_counts = CustomUser.objects.values('regional_id').annotate(count=Count('id'))
    counts_dict = {item['regional_id']: item['count'] for item in user_counts if item['regional_id']}

    # Vincular contagens ao objeto de regional
    for reg in regionais:
        reg.user_count = counts_dict.get(reg.id, 0)

    context = {
        "regionais": regionais,
    }
    return render(request, "gerenciar_regionais.html", context)


@login_required
def alterar_regional_ativa(request):
    """
    Altera a regional ativa do Administrador Supremo na sessão.
    Isso altera o banco de dados dinâmico das requisições subsequentes.
    """
    if not getattr(request.user, "is_global_admin", False):
        return redirect("gestao_a_vista:home")

    from .models import Regional

    regional_id = request.GET.get("regional_id")
    if regional_id and regional_id != "GLOBAL":
        try:
            regional = Regional.objects.filter(id=regional_id).first()
        except (ValueError, ValidationError):
            regional = None
        if regional:
            request.session['active_regional'] = regional.db_slug
            messages.success(request, f"Visualização da regional {regional.nome} ativada.")
        else:
            messages.error(request, "Regional não encontrada.")
    else:
        if 'active_regional' in request.session:
            del request.session['active_regional']
        messages.success(request, "Visualização global ativada.")

    referer = request.META.get('HTTP_REFERER')
    if referer:
        return redirect(referer)
    return redirect("gestao_a_vista:home")


# === MÓDULO CMO EFETIVO (Conformidade de Efetivo) ===
#
# Implementa o PRD "Central de Controle" (substituição da PDM), replicando
# o padrão visual/arquitetural da Torre de Controle (LivroOcorrenciasView):
# uma única TemplateView monta o contexto de todas as abas, e cada ação do
# usuário é um endpoint JSON separado (mesmo padrão de /api/ocorrencias/*).

def _cmo_efetivo_pode_ver_cr(user, cr_value):
    """Mesma regra de acesso por CR já usada na Torre de Controle
    (LivroOcorrenciasView.get_context_data): admin/superuser/geral vê tudo;
    usuário sem CR configurado também vê tudo (comportamento permissivo já
    estabelecido no restante do sistema); caso contrário, só vê registros
    cujo CR bate com algum CR do usuário.
    """
    if user.role == 'administrador' or getattr(user, 'is_superuser', False) or getattr(user, 'is_general', False):
        return True
    user_crs = (getattr(user, 'crs', '') or '').strip()
    if not user_crs:
        return True
    import re
    crs_nums = [m.group() for c in user_crs.split(',') if (m := re.search(r'\d+', c))]
    cr_str = str(cr_value or '')
    cr_num_match = re.search(r'\d+', cr_str)
    if cr_num_match and cr_num_match.group() in crs_nums:
        return True
    return any(c.strip() and c.strip() in cr_str for c in user_crs.split(','))


def _cmo_efetivo_filtrar_lista_por_cr(user, itens):
    return [item for item in itens if _cmo_efetivo_pode_ver_cr(user, getattr(item, 'cr', None))]


def _cmo_efetivo_pode_agir_torre(user):
    """RN08: perfis autorizados -- Torre registra conformidade, contato,
    cobertura e solicita troca."""
    return user.role == 'administrador' or user.has_page_permission('cmo_efetivo_torre')


def _cmo_efetivo_pode_agir_cmo(user):
    """RN08: perfis autorizados -- CMO aprova/reprova troca e lança em folha."""
    return user.role == 'administrador' or user.has_page_permission('cmo_efetivo_cmo')


def registrar_log_cmo_efetivo(entidade, registro_id, acao, usuario, valor_anterior=None, valor_novo=None, observacao=""):
    """RN09: log de auditoria (usuário, data, campo alterado, valor anterior/novo)."""
    from .models import CMOEfetivoLog
    try:
        CMOEfetivoLog.objects.create(
            entidade=entidade,
            registro_id=str(registro_id),
            acao=acao,
            usuario=usuario if getattr(usuario, 'is_authenticated', False) else None,
            valor_anterior=valor_anterior or {},
            valor_novo=valor_novo or {},
            observacao=observacao,
        )
    except Exception:
        logger.exception("Erro ao registrar log de auditoria do CMO Efetivo")


def _cmo_efetivo_criar_lancamento_se_necessario(ocorrencia_tipo, ocorrencia, cliente_nome, colaborador_nome):
    """RN04: registros com impacto em ponto/folha geram pendência para a
    CMO. Evita duplicar a pendência se já existir uma para essa ocorrência."""
    from .models import CMOEfetivoLancamento
    ja_existe = CMOEfetivoLancamento.objects.filter(
        ocorrencia_tipo=ocorrencia_tipo, ocorrencia_id=str(ocorrencia.id)
    ).exists()
    if ja_existe:
        return None
    return CMOEfetivoLancamento.objects.create(
        ocorrencia_tipo=ocorrencia_tipo,
        ocorrencia_id=str(ocorrencia.id),
        cliente_nome=cliente_nome,
        colaborador_nome=colaborador_nome,
    )


def _cmo_efetivo_resolver_cliente_por_cr(cr):
    """Resolve o Nome do Cliente a partir do CR -- o OpsVista manda só o
    número do CR, não o nome da empresa. Usa ColaboradorSRA como fonte,
    mesmo padrão de lookup já usado pelo autocomplete deste módulo (ver
    comentário no bloco de models CMO Efetivo)."""
    from .models import ColaboradorSRA
    if not cr:
        return ''
    registro = ColaboradorSRA.objects.filter(cr=cr).exclude(cliente='').values('cliente').first()
    return registro['cliente'] if registro else ''


# TODO(OpsVista): a pergunta de ronda "O efetivo está completo? (Sim/Não)"
# ainda não existe no Vista -- ficou definido na reunião de 14/07/2026 que a
# equipe do Vista vai criar essa pergunta no checklist de abertura de ronda,
# e a resposta deve alimentar automaticamente CMOEfetivoConformidade
# (status_efetivo = 'conforme' se Sim, 'nao_conforme' se Não), no mesmo
# mecanismo já usado pelo Livro de Ocorrências (LivroOcorrenciasView) --
# provavelmente um novo "item" de checklist com esse nome. Enquanto isso não
# existe, a aba Conformidade de Efetivo fica vazia (não há mais lançamento
# manual pela Torre) -- todo registro nela mostra "Aguardando Confirmação"
# por padrão até essa sincronização ser implementada.
CMO_EFETIVO_ITEM_CHECKLIST_VISTA = "Efetivo Completo"


def _cmo_efetivo_sincronizar_conformidade_vista():
    """Placeholder: quando o OpsVista passar a enviar a resposta da
    pergunta acima (provavelmente via Livro de Ocorrências, item
    CMO_EFETIVO_ITEM_CHECKLIST_VISTA), este é o lugar para ler essas
    respostas e fazer upsert de CMOEfetivoConformidade por (cr,
    data_referencia), usando _cmo_efetivo_resolver_cliente_por_cr(cr) para
    preencher cliente_nome. Não é chamada em lugar nenhum ainda -- ver
    comentário do TODO acima."""
    raise NotImplementedError(
        "Integração com o OpsVista ainda não implementada (ver reunião de 14/07/2026)."
    )


@method_decorator(never_cache, name='dispatch')
class CMOEfetivoView(LoginRequiredMixin, TemplateView):
    """
    Módulo de Conformidade Efetivo (CMO) -- PRD "Central de Controle".
    Uma única view monta o contexto de todas as abas (Dashboard,
    Conformidade de Efetivo, Coberturas, Lançamento CMO, Previsão de
    Cobertura), no mesmo padrão de LivroOcorrenciasView. Lançamento CMO e
    Previsão de Cobertura mostram a mesma lista (registros_movimentacao),
    ver reunião de 14/07/2026.
    """
    template_name = "cmo_efetivo.html"

    def get(self, request, *args, **kwargs):
        if not (request.user.role == 'administrador' or request.user.has_page_permission('cmo_efetivo')):
            messages.error(request, "Você não tem permissão para acessar esta página.")
            return redirect('gestao_a_vista:home')
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .models import (
            CMOEfetivoConformidade, CMOEfetivoCobertura,
            CMOEfetivoRegistroMovimentacao, CMO_EFETIVO_TIPO_SERVICO_CHOICES,
            CMO_EFETIVO_SITUACAO_CHOICES, CMO_EFETIVO_SITUACAO_COBERTURA_CHOICES,
        )
        from collections import Counter

        user = self.request.user

        # --- Filtros (Data | Cliente | Tipo Serviço | Status) ---
        dias_filtro = self.request.GET.get('dias', '30')
        try:
            dias_int = int(dias_filtro) if dias_filtro in ('7', '15', '30', '60', '90') else 30
        except ValueError:
            dias_int = 30
        data_limite = timezone.now() - timedelta(days=dias_int)

        filtro_cliente = self.request.GET.get('filter_cliente', '').strip()
        filtro_tipo_servico = self.request.GET.get('filter_tipo_servico', '').strip()
        filtro_status = self.request.GET.get('filter_status', 'all')

        LIMITE_REGISTROS = 300

        # --- Conformidade de Efetivo ---
        # Alimentada pelo Livro de Ocorrências do OpsVista (ver
        # CMO_EFETIVO_ITEM_CHECKLIST_VISTA / _cmo_efetivo_sincronizar_conformidade_vista
        # mais abaixo) -- não há mais lançamento manual pela Torre nesta aba.
        conformidades_qs = CMOEfetivoConformidade.objects.select_related(
            'responsavel_contato'
        ).filter(criado_em__gte=data_limite, cancelado=False)
        if filtro_cliente:
            conformidades_qs = conformidades_qs.filter(cliente_nome__icontains=filtro_cliente)
        if filtro_tipo_servico:
            conformidades_qs = conformidades_qs.filter(tipo_servico_nome=filtro_tipo_servico)
        if filtro_status and filtro_status != 'all':
            conformidades_qs = conformidades_qs.filter(status_efetivo=filtro_status)
        conformidades = _cmo_efetivo_filtrar_lista_por_cr(
            user, list(conformidades_qs.order_by('-data_referencia')[:LIMITE_REGISTROS])
        )

        # --- Coberturas ---
        coberturas_qs = CMOEfetivoCobertura.objects.select_related(
            'conformidade_origem'
        ).filter(criado_em__gte=data_limite)
        if filtro_cliente:
            coberturas_qs = coberturas_qs.filter(cliente_nome__icontains=filtro_cliente)
        if filtro_tipo_servico:
            coberturas_qs = coberturas_qs.filter(tipo_servico_nome=filtro_tipo_servico)
        coberturas = _cmo_efetivo_filtrar_lista_por_cr(
            user, list(coberturas_qs.order_by('-data_cobertura')[:LIMITE_REGISTROS])
        )

        # --- Lançamentos CMO / Previsão de Cobertura ---
        # Mesma lista/colunas nas duas abas (ver "Incluir Registro" na
        # Conformidade de Efetivo, que é quem cria estes registros).
        registros_movimentacao_qs = CMOEfetivoRegistroMovimentacao.objects.select_related(
            'conformidade_origem'
        ).filter(criado_em__gte=data_limite)
        if filtro_cliente:
            registros_movimentacao_qs = registros_movimentacao_qs.filter(cliente_nome__icontains=filtro_cliente)
        if filtro_tipo_servico:
            registros_movimentacao_qs = registros_movimentacao_qs.filter(tipo_servico_nome=filtro_tipo_servico)
        registros_movimentacao = _cmo_efetivo_filtrar_lista_por_cr(
            user, list(registros_movimentacao_qs.order_by('-data_referencia')[:LIMITE_REGISTROS])
        )

        # --- KPIs do Dashboard da Torre ---
        kpis = {
            'total_clientes': len({c.cliente_nome for c in conformidades}),
            'efetivo_conforme': sum(1 for c in conformidades if c.status_efetivo == 'conforme'),
            'efetivo_nao_conforme': sum(1 for c in conformidades if c.status_efetivo == 'nao_conforme'),
            'aguardando_confirmacao': sum(1 for c in conformidades if c.status_efetivo == 'aguardando_confirmacao'),
            'faltas_registradas': sum(
                1 for c in conformidades
                if c.status_efetivo == 'nao_conforme' and c.motivo_nao_conformidade in ('falta_nao_justificada', 'falta_justificada')
            ),
            'coberturas_pendentes': sum(1 for c in coberturas if c.status == 'pendente'),
            'coberturas_confirmadas': sum(1 for c in coberturas if c.status == 'confirmada'),
            'registros_movimentacao_total': len(registros_movimentacao),
        }

        # --- Gráficos sugeridos pelo PRD ---
        nc_por_cliente = Counter(c.cliente_nome for c in conformidades if c.status_efetivo == 'nao_conforme')
        faltas_por_dia = Counter(
            c.data_referencia.strftime('%d/%m') for c in conformidades if c.status_efetivo == 'nao_conforme'
        )
        coberturas_por_tipo = Counter((c.tipo_servico_nome or 'Não informado') for c in coberturas)
        top_clientes = nc_por_cliente.most_common(10)
        faltas_labels = sorted(faltas_por_dia.keys(), key=lambda d: (d.split('/')[1], d.split('/')[0]))

        charts = {
            'nc_por_cliente': {
                'labels': [k for k, _v in nc_por_cliente.most_common(15)],
                'values': [v for _k, v in nc_por_cliente.most_common(15)],
            },
            'faltas_por_periodo': {
                'labels': faltas_labels,
                'values': [faltas_por_dia[k] for k in faltas_labels],
            },
            'coberturas_por_tipo': {
                'labels': list(coberturas_por_tipo.keys()),
                'values': list(coberturas_por_tipo.values()),
            },
            'top_clientes': {
                'labels': [k for k, _v in top_clientes],
                'values': [v for _k, v in top_clientes],
            },
        }

        context.update({
            'dias_atuais': str(dias_int),
            'conformidades': conformidades,
            'coberturas': coberturas,
            'registros_movimentacao': registros_movimentacao,
            'kpis': kpis,
            'charts_json': json.dumps(charts, ensure_ascii=False),
            'tipo_servico_choices': CMO_EFETIVO_TIPO_SERVICO_CHOICES,
            'situacao_choices': CMO_EFETIVO_SITUACAO_CHOICES,
            'situacao_cobertura_choices': CMO_EFETIVO_SITUACAO_COBERTURA_CHOICES,
            'filtros': {
                'cliente': filtro_cliente,
                'tipo_servico': filtro_tipo_servico,
                'status': filtro_status,
            },
            'pode_agir_torre': _cmo_efetivo_pode_agir_torre(user),
            'pode_agir_cmo': _cmo_efetivo_pode_agir_cmo(user),
        })
        return context


@login_required
@check_page_permission("cmo_efetivo")
@require_GET
def cmo_efetivo_autocomplete(request):
    """Autocomplete de Cliente/Colaborador com base em ColaboradorSRA (fonte
    de dados já existente e sincronizada -- não é FK, ver comentário no
    bloco de models CMO Efetivo em models.py)."""
    from .models import ColaboradorSRA

    tipo = request.GET.get('tipo', 'colaborador')
    termo = request.GET.get('q', '').strip()
    if len(termo) < 2:
        return JsonResponse({'success': True, 'resultados': []})

    qs = ColaboradorSRA.objects.filter(situacao='Ativo')
    user = request.user
    if not (user.role == 'administrador' or getattr(user, 'is_superuser', False) or getattr(user, 'is_general', False)):
        user_crs = (getattr(user, 'crs', '') or '').strip()
        if user_crs:
            crs_list = [c.strip() for c in user_crs.split(',') if c.strip()]
            cr_q = Q()
            for c in crs_list:
                cr_q |= Q(cr__icontains=c)
            if crs_list:
                qs = qs.filter(cr_q)

    if tipo == 'cliente':
        linhas = qs.exclude(cliente__isnull=True).exclude(cliente='').filter(
            cliente__icontains=termo
        ).values_list('cliente', 'grupo_cliente', 'cr').distinct()[:20]
        resultados = [{'cliente': nome, 'grupo_cliente': grupo, 'cr': cr} for nome, grupo, cr in linhas]
    else:
        resultados = list(qs.filter(nome__icontains=termo).values('nome', 'matricula', 'cliente', 'grupo_cliente', 'cr')[:20])

    return JsonResponse({'success': True, 'resultados': resultados})


@login_required
@check_page_permission("cmo_efetivo")
@require_POST
def cmo_efetivo_salvar_conformidade(request):
    """Cria ou atualiza (upsert via 'id' opcional) um registro de
    Conformidade de Efetivo. Implementa RN01, RN02 e RN04."""
    from .models import CMOEfetivoConformidade, CMO_EFETIVO_TIPO_SERVICO_CHOICES

    if not _cmo_efetivo_pode_agir_torre(request.user):
        return JsonResponse({'success': False, 'message': 'Você não tem permissão para registrar conformidade.'}, status=403)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Dados inválidos.'}, status=400)

    registro_id = data.get('id')
    cliente_nome = (data.get('cliente_nome') or '').strip()
    tipo_servico_nome = data.get('tipo_servico') or None
    status_efetivo = data.get('status_efetivo')
    colaborador_nome = (data.get('colaborador_nome') or '').strip()

    # RN01: todo registro vinculado a cliente + tipo de serviço
    if not cliente_nome or not tipo_servico_nome or not colaborador_nome:
        return JsonResponse({'success': False, 'message': 'Cliente, tipo de serviço e colaborador são obrigatórios.'}, status=400)

    if tipo_servico_nome not in dict(CMO_EFETIVO_TIPO_SERVICO_CHOICES):
        return JsonResponse({'success': False, 'message': 'Tipo de serviço inválido.'}, status=400)

    if status_efetivo not in ('conforme', 'nao_conforme'):
        return JsonResponse({'success': False, 'message': 'Status do efetivo inválido.'}, status=400)

    motivo = data.get('motivo_nao_conformidade') or None
    # RN02: motivo obrigatório se não conforme
    if status_efetivo == 'nao_conforme' and not motivo:
        return JsonResponse({'success': False, 'message': 'Informe o motivo da não conformidade.'}, status=400)

    previsao_cobertura = data.get('previsao_cobertura') or None
    if previsao_cobertura:
        try:
            previsao_cobertura = datetime.fromisoformat(previsao_cobertura)
        except ValueError:
            previsao_cobertura = None

    data_referencia_str = data.get('data_referencia')
    try:
        data_referencia = datetime.strptime(data_referencia_str, '%Y-%m-%d').date() if data_referencia_str else timezone.localdate()
    except ValueError:
        data_referencia = timezone.localdate()

    impacto_folha = bool(data.get('impacto_folha', False)) if status_efetivo == 'nao_conforme' else False

    campos = dict(
        cliente_nome=cliente_nome,
        grupo_cliente=(data.get('grupo_cliente') or '').strip() or None,
        cr=(data.get('cr') or '').strip() or None,
        tipo_servico_nome=tipo_servico_nome,
        colaborador_nome=colaborador_nome,
        colaborador_matricula=(data.get('colaborador_matricula') or '').strip() or None,
        colaborador_telefone=(data.get('colaborador_telefone') or '').strip() or None,
        data_referencia=data_referencia,
        status_efetivo=status_efetivo,
        motivo_nao_conformidade=motivo if status_efetivo == 'nao_conforme' else None,
        motivo_outro_detalhe=(data.get('motivo_outro_detalhe') or '').strip() or None,
        previsao_cobertura=previsao_cobertura,
        impacto_folha=impacto_folha,
        observacoes=(data.get('observacoes') or '').strip() or None,
    )
    if impacto_folha:
        campos['status_lancamento'] = 'pendente'
    elif status_efetivo == 'conforme':
        campos['status_lancamento'] = 'nao_aplica'

    acao = 'criacao'
    valor_anterior = {}
    if registro_id:
        try:
            registro = CMOEfetivoConformidade.objects.get(id=registro_id, cancelado=False)
        except (CMOEfetivoConformidade.DoesNotExist, ValidationError):
            return JsonResponse({'success': False, 'message': 'Registro não encontrado.'}, status=404)
        valor_anterior = {'status_efetivo': registro.status_efetivo, 'status_lancamento': registro.status_lancamento}
        for campo, valor in campos.items():
            setattr(registro, campo, valor)
        registro.save()
        acao = 'atualizacao'
    else:
        registro = CMOEfetivoConformidade.objects.create(
            criado_por=request.user, responsavel_contato=request.user, **campos
        )

    # RN04: impacto em folha gera pendência para a CMO
    if impacto_folha:
        _cmo_efetivo_criar_lancamento_se_necessario('conformidade', registro, registro.cliente_nome, registro.colaborador_nome)

    registrar_log_cmo_efetivo(
        'conformidade', registro.id, acao, request.user,
        valor_anterior=valor_anterior, valor_novo={k: str(v) for k, v in campos.items()},
    )

    return JsonResponse({'success': True, 'message': 'Conformidade registrada com sucesso.', 'id': str(registro.id)})


@login_required
@check_page_permission("cmo_efetivo")
@require_POST
def cmo_efetivo_salvar_registro_movimentacao(request):
    """"Incluir Registro": cria um Registro de Movimentação a partir de uma
    linha Não Conforme de Conformidade de Efetivo, reproduzindo os campos
    da Planilha Diária de Movimentações. Alimenta, com as mesmas colunas,
    as abas Lançamentos CMO e Previsão de Cobertura."""
    from .models import (
        CMOEfetivoConformidade, CMOEfetivoRegistroMovimentacao,
        CMO_EFETIVO_SITUACAO_CHOICES, CMO_EFETIVO_SITUACAO_COBERTURA_CHOICES,
    )

    if not _cmo_efetivo_pode_agir_torre(request.user):
        return JsonResponse({'success': False, 'message': 'Você não tem permissão para incluir este registro.'}, status=403)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Dados inválidos.'}, status=400)

    conformidade = get_object_or_404(CMOEfetivoConformidade, id=data.get('conformidade_id'))
    if conformidade.status_efetivo != 'nao_conforme':
        return JsonResponse({'success': False, 'message': 'Só é possível incluir registro em linhas Não Conforme.'}, status=400)

    nome = (data.get('nome') or '').strip()
    situacao = data.get('situacao')
    if not nome or not situacao:
        return JsonResponse({'success': False, 'message': 'Nome e Situação são obrigatórios.'}, status=400)
    if situacao not in dict(CMO_EFETIVO_SITUACAO_CHOICES):
        return JsonResponse({'success': False, 'message': 'Situação inválida.'}, status=400)

    situacao_cobertura = data.get('situacao_cobertura') or None
    if situacao_cobertura and situacao_cobertura not in dict(CMO_EFETIVO_SITUACAO_COBERTURA_CHOICES):
        situacao_cobertura = None

    registro = CMOEfetivoRegistroMovimentacao.objects.create(
        conformidade_origem=conformidade,
        cliente_nome=conformidade.cliente_nome,
        cr=conformidade.cr,
        tipo_servico_nome=conformidade.tipo_servico_nome,
        data_referencia=conformidade.data_referencia,
        gerente=(data.get('gerente') or '').strip(),
        nome=nome,
        cargo=(data.get('cargo') or '').strip(),
        posto=(data.get('posto') or '').strip(),
        posto_cr=(data.get('posto_cr') or '').strip(),
        horario=(data.get('horario') or '').strip(),
        intervalo=(data.get('intervalo') or '').strip(),
        cobertura=(data.get('cobertura') or '').strip(),
        cargo_cobertura=(data.get('cargo_cobertura') or '').strip(),
        observacao=(data.get('observacao') or '').strip() or None,
        situacao=situacao,
        situacao_cobertura=situacao_cobertura,
        criado_por=request.user,
    )

    registrar_log_cmo_efetivo(
        'movimentacao', registro.id, 'criacao', request.user,
        valor_novo={'situacao': situacao, 'situacao_cobertura': situacao_cobertura or ''},
    )

    return JsonResponse({'success': True, 'message': 'Registro incluído com sucesso.', 'id': str(registro.id)})


@login_required
@check_page_permission("cmo_efetivo")
@require_POST
def cmo_efetivo_salvar_cobertura(request):
    """Cria ou atualiza (upsert via 'id' opcional) uma Cobertura."""
    from .models import CMOEfetivoCobertura, CMOEfetivoConformidade, CMO_EFETIVO_TIPO_SERVICO_CHOICES

    if not _cmo_efetivo_pode_agir_torre(request.user):
        return JsonResponse({'success': False, 'message': 'Você não tem permissão para registrar cobertura.'}, status=403)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Dados inválidos.'}, status=400)

    registro_id = data.get('id')
    cliente_nome = (data.get('cliente_nome') or '').strip()
    nome_cobertura = (data.get('nome_cobertura') or '').strip()
    colaborador_substituido = (data.get('colaborador_substituido') or '').strip()
    justificativa = (data.get('justificativa') or '').strip()
    data_cobertura_str = data.get('data_cobertura')

    if not cliente_nome or not nome_cobertura or not justificativa or not data_cobertura_str:
        return JsonResponse({'success': False, 'message': 'Cliente, colaborador da cobertura, justificativa e data são obrigatórios.'}, status=400)

    try:
        data_cobertura = datetime.strptime(data_cobertura_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'success': False, 'message': 'Data da cobertura inválida.'}, status=400)

    tipo_servico_nome = data.get('tipo_servico') or None
    if tipo_servico_nome and tipo_servico_nome not in dict(CMO_EFETIVO_TIPO_SERVICO_CHOICES):
        tipo_servico_nome = None

    conformidade_origem = None
    if data.get('conformidade_origem_id'):
        conformidade_origem = CMOEfetivoConformidade.objects.filter(id=data['conformidade_origem_id']).first()

    horario_previsto = data.get('horario_previsto') or None
    if horario_previsto:
        try:
            horario_previsto = datetime.strptime(horario_previsto, '%H:%M').time()
        except ValueError:
            horario_previsto = None

    status = data.get('status', 'pendente')
    if status not in dict(CMOEfetivoCobertura.STATUS_CHOICES):
        status = 'pendente'

    campos = dict(
        conformidade_origem=conformidade_origem,
        nome_cobertura=nome_cobertura,
        matricula_cobertura=(data.get('matricula_cobertura') or '').strip() or None,
        cliente_nome=cliente_nome,
        grupo_cliente=(data.get('grupo_cliente') or '').strip() or None,
        cr=(data.get('cr') or '').strip() or None,
        tipo_servico_nome=tipo_servico_nome,
        colaborador_substituido=colaborador_substituido,
        justificativa=justificativa,
        data_cobertura=data_cobertura,
        horario_previsto=horario_previsto,
        status=status,
        observacoes=(data.get('observacoes') or '').strip() or None,
    )

    acao = 'criacao'
    if registro_id:
        registro = get_object_or_404(CMOEfetivoCobertura, id=registro_id)
        for campo, valor in campos.items():
            setattr(registro, campo, valor)
        registro.save()
        acao = 'atualizacao'
    else:
        registro = CMOEfetivoCobertura.objects.create(criado_por=request.user, **campos)
        # RN03/RN04: cobertura registrada representa impacto operacional -> pendência CMO
        _cmo_efetivo_criar_lancamento_se_necessario('cobertura', registro, registro.cliente_nome, registro.nome_cobertura)

    registrar_log_cmo_efetivo('cobertura', registro.id, acao, request.user, valor_novo={k: str(v) for k, v in campos.items()})

    return JsonResponse({'success': True, 'message': 'Cobertura salva com sucesso.', 'id': str(registro.id)})


@login_required
@check_page_permission("cmo_efetivo")
@require_POST
def cmo_efetivo_solicitar_troca(request):
    """Solicita uma Autorização de Troca de Serviço. Implementa a regra de
    mês vigente e RN07 (alerta de conflito não-bloqueante)."""
    from .models import CMOEfetivoTroca, CMO_EFETIVO_TIPO_SERVICO_CHOICES

    if not _cmo_efetivo_pode_agir_torre(request.user):
        return JsonResponse({'success': False, 'message': 'Você não tem permissão para solicitar troca.'}, status=403)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Dados inválidos.'}, status=400)

    cliente_nome = (data.get('cliente_nome') or '').strip()
    colaborador_atual = (data.get('colaborador_atual') or '').strip()
    colaborador_substituto = (data.get('colaborador_substituto') or '').strip()
    justificativa = (data.get('justificativa') or '').strip()
    data_troca_str = data.get('data_troca')

    if not all([cliente_nome, colaborador_atual, colaborador_substituto, justificativa, data_troca_str]):
        return JsonResponse({'success': False, 'message': 'Preencha todos os campos obrigatórios.'}, status=400)

    try:
        data_troca = datetime.strptime(data_troca_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'success': False, 'message': 'Data da troca inválida.'}, status=400)

    # Obs do formulário no PRD: troca deve ser realizada dentro do mês vigente.
    hoje = timezone.localdate()
    if data_troca.year != hoje.year or data_troca.month != hoje.month:
        return JsonResponse({'success': False, 'message': 'A troca deve ser realizada dentro do mês vigente.'}, status=400)

    tipo_servico_nome = data.get('tipo_servico') or None
    if tipo_servico_nome and tipo_servico_nome not in dict(CMO_EFETIVO_TIPO_SERVICO_CHOICES):
        tipo_servico_nome = None

    forcar = bool(data.get('forcar', False))

    # RN07: alerta de troca duplicada (mesmo cliente + colaborador + data) --
    # não bloqueia o cadastro no MVP, só exige confirmação/justificativa.
    ja_existe_conflito = CMOEfetivoTroca.objects.filter(
        cliente_nome=cliente_nome, colaborador_atual=colaborador_atual, data_troca=data_troca, cancelado=False,
    ).exists()
    if ja_existe_conflito and not forcar:
        return JsonResponse({
            'success': False,
            'conflito': True,
            'message': 'Já existe uma troca registrada para este cliente, colaborador e data. Confirme para continuar mesmo assim.',
        })

    registro = CMOEfetivoTroca.objects.create(
        solicitante=request.user,
        cliente_nome=cliente_nome,
        grupo_cliente=(data.get('grupo_cliente') or '').strip() or None,
        cr=(data.get('cr') or '').strip() or None,
        tipo_servico_nome=tipo_servico_nome,
        colaborador_atual=colaborador_atual,
        colaborador_substituto=colaborador_substituto,
        justificativa=justificativa,
        data_troca=data_troca,
        status_cmo='pendente',
        conflito_detectado=bool(forcar and ja_existe_conflito),
        conflito_justificativa=(data.get('conflito_justificativa') or '').strip() or None,
        criado_por=request.user,
    )
    registrar_log_cmo_efetivo('troca', registro.id, 'criacao', request.user, valor_novo={'status_cmo': 'pendente'})

    return JsonResponse({'success': True, 'message': 'Solicitação de troca registrada. Aguardando decisão da CMO.', 'id': str(registro.id)})


@login_required
@check_page_permission("cmo_efetivo")
@require_POST
def cmo_efetivo_decidir_troca(request):
    """CMO aprova ou reprova uma solicitação de troca. RN06 e RN08."""
    from .models import CMOEfetivoTroca

    if not _cmo_efetivo_pode_agir_cmo(request.user):
        return JsonResponse({'success': False, 'message': 'Somente a CMO pode aprovar ou reprovar trocas.'}, status=403)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Dados inválidos.'}, status=400)

    troca = get_object_or_404(CMOEfetivoTroca, id=data.get('id'))
    acao = data.get('acao')
    if acao not in ('aprovar', 'reprovar'):
        return JsonResponse({'success': False, 'message': 'Ação inválida.'}, status=400)

    status_anterior = troca.status_cmo
    troca.status_cmo = 'aprovado' if acao == 'aprovar' else 'reprovado'
    troca.observacao_cmo = (data.get('observacao_cmo') or '').strip() or None
    troca.aprovado_por = request.user
    troca.aprovado_em = timezone.now()
    troca.save()

    if troca.status_cmo == 'aprovado':
        _cmo_efetivo_criar_lancamento_se_necessario('troca', troca, troca.cliente_nome, troca.colaborador_atual)

    registrar_log_cmo_efetivo(
        'troca', troca.id, f'decisao_{troca.status_cmo}', request.user,
        valor_anterior={'status_cmo': status_anterior}, valor_novo={'status_cmo': troca.status_cmo},
    )

    return JsonResponse({'success': True, 'message': f'Troca {troca.get_status_cmo_display().lower()}.'})


@login_required
@check_page_permission("cmo_efetivo")
@require_POST
def cmo_efetivo_marcar_lancado(request):
    """CMO marca uma pendência como lançada em folha. RN04 e RN05."""
    from .models import CMOEfetivoLancamento, CMOEfetivoConformidade

    if not _cmo_efetivo_pode_agir_cmo(request.user):
        return JsonResponse({'success': False, 'message': 'Somente a CMO pode marcar um lançamento como concluído.'}, status=403)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Dados inválidos.'}, status=400)

    lancamento = get_object_or_404(CMOEfetivoLancamento, id=data.get('id'))

    # RN05: CMO só encerra a pendência informando data e responsável (ambos setados aqui)
    lancamento.lancado = True
    lancamento.data_lancamento = timezone.now()
    lancamento.responsavel_cmo = request.user
    if data.get('observacoes'):
        lancamento.observacoes = data['observacoes'].strip()
    lancamento.save()

    # "Após lançamentos na página de PDM, essas informações irão para a Previsão de Cobertura"
    if lancamento.ocorrencia_tipo == 'conformidade' and lancamento.ocorrencia_id:
        CMOEfetivoConformidade.objects.filter(id=lancamento.ocorrencia_id).update(status_lancamento='lancado')

    registrar_log_cmo_efetivo(
        'lancamento', lancamento.id, 'lancado_em_folha', request.user,
        valor_novo={'lancado': True, 'responsavel_cmo': str(request.user.id)},
    )

    return JsonResponse({'success': True, 'message': 'Lançamento marcado como concluído.'})


@login_required
@check_page_permission("cmo_efetivo")
@require_POST
def cmo_efetivo_cancelar(request):
    """RN10: registros finalizados não são excluídos por usuários comuns,
    apenas cancelados de forma controlada (com motivo obrigatório)."""
    from .models import CMOEfetivoConformidade, CMOEfetivoCobertura, CMOEfetivoTroca

    if not (request.user.role == 'administrador' or _cmo_efetivo_pode_agir_cmo(request.user)):
        return JsonResponse({'success': False, 'message': 'Você não tem permissão para cancelar registros.'}, status=403)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Dados inválidos.'}, status=400)

    entidade = data.get('entidade')
    registro_id = data.get('id')
    motivo = (data.get('motivo') or '').strip()

    if not motivo:
        return JsonResponse({'success': False, 'message': 'Informe o motivo do cancelamento.'}, status=400)

    if entidade == 'conformidade':
        registro = get_object_or_404(CMOEfetivoConformidade, id=registro_id)
        registro.cancelado = True
        registro.cancelado_por = request.user
        registro.cancelado_em = timezone.now()
        registro.motivo_cancelamento = motivo
        registro.save()
    elif entidade == 'troca':
        registro = get_object_or_404(CMOEfetivoTroca, id=registro_id)
        registro.cancelado = True
        registro.cancelado_por = request.user
        registro.cancelado_em = timezone.now()
        registro.motivo_cancelamento = motivo
        registro.save()
    elif entidade == 'cobertura':
        registro = get_object_or_404(CMOEfetivoCobertura, id=registro_id)
        registro.status = 'cancelada'
        registro.observacoes = f"{registro.observacoes or ''}\nCancelado: {motivo}".strip()
        registro.save()
    else:
        return JsonResponse({'success': False, 'message': 'Entidade inválida.'}, status=400)

    registrar_log_cmo_efetivo(entidade, registro_id, 'cancelamento', request.user, observacao=motivo)

    return JsonResponse({'success': True, 'message': 'Registro cancelado.'})


# === MÓDULO LINKS IMPORTANTES ===

class LinksImportantesView(LoginRequiredMixin, TemplateView):
    """Aba de atalhos de fácil acesso a sistemas usados no dia a dia (Sistema360,
    OpsVista, Prisma, etc.). Cadastro é feito pelos próprios usuários com
    acesso à aba, ver links_importantes_salvar/excluir logo abaixo."""
    template_name = "links_importantes.html"

    def get(self, request, *args, **kwargs):
        if not (request.user.role == 'administrador' or request.user.has_page_permission('links_importantes')):
            messages.error(request, "Você não tem permissão para acessar esta página.")
            return redirect('gestao_a_vista:home')
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .models import LinkImportante
        context['links'] = LinkImportante.objects.all().order_by('ordem', 'titulo')
        return context


@login_required
@check_page_permission("links_importantes")
@require_POST
def links_importantes_salvar(request):
    """Cria ou atualiza (upsert via 'id' opcional) um link importante."""
    from .models import LinkImportante

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Dados inválidos.'}, status=400)

    link_id = data.get('id')
    titulo = (data.get('titulo') or '').strip()
    url = (data.get('url') or '').strip()
    descricao = (data.get('descricao') or '').strip()

    if not titulo or not url:
        return JsonResponse({'success': False, 'message': 'Título e URL são obrigatórios.'}, status=400)

    if link_id:
        link = get_object_or_404(LinkImportante, id=link_id)
    else:
        link = LinkImportante(criado_por=request.user)
        maior_ordem = LinkImportante.objects.aggregate(m=models.Max('ordem'))['m'] or 0
        link.ordem = maior_ordem + 1

    link.titulo = titulo
    link.url = url
    link.descricao = descricao
    link.ativo = bool(data.get('ativo', True))

    ordem_recebida = data.get('ordem')
    if ordem_recebida not in (None, ''):
        try:
            link.ordem = int(ordem_recebida)
        except (TypeError, ValueError):
            pass

    try:
        link.full_clean()
        link.save()
    except ValidationError as e:
        mensagens = sum(e.message_dict.values(), []) if hasattr(e, 'message_dict') else e.messages
        return JsonResponse({'success': False, 'message': '; '.join(mensagens)}, status=400)

    return JsonResponse({
        'success': True,
        'link': {
            'id': str(link.id),
            'titulo': link.titulo,
            'url': link.url,
            'descricao': link.descricao,
            'ordem': link.ordem,
            'ativo': link.ativo,
        },
    })


@login_required
@check_page_permission("links_importantes")
@require_POST
def links_importantes_excluir(request):
    from .models import LinkImportante

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Dados inválidos.'}, status=400)

    link_id = data.get('id')
    link = get_object_or_404(LinkImportante, id=link_id)
    link.delete()
    return JsonResponse({'success': True})






# === MÓDULO AJUDA / TUTORIAL DE PRIMEIRO ACESSO ===

class AjudaView(LoginRequiredMixin, TemplateView):
    """Página de Ajuda com o tour visual por todas as abas do sistema.
    Sem permissão de página de propósito: todo usuário autenticado pode
    rever o tutorial aqui. O mesmo deck de slides é exibido automaticamente
    (uma única vez, obrigatório) no primeiro acesso, via modal incluído no
    base.html -- ver partials/tutorial_slides.html e CustomUser.tutorial_visto."""
    template_name = "ajuda.html"


@login_required
@require_POST
def tutorial_marcar_visto(request):
    """Marca que o usuário já viu o tutorial de primeiro acesso. Chamado no
    momento em que o modal é EXIBIDO (não ao concluir): o requisito é mostrar
    apenas uma vez, mesmo que a pessoa feche no meio -- para rever, existe a
    página Ajuda, sempre no fim do menu."""
    if not request.user.tutorial_visto:
        request.user.tutorial_visto = True
        request.user.save(update_fields=["tutorial_visto"])
    return JsonResponse({"success": True})


# === MÓDULO GESTÃO DE ORDENS DE MANUTENÇÃO / PCM (DEMO — Innova Industrial) ===

def controle_ordens(request):
    """DEMO front-end do Sistema de Gestão de Ordens de Manutenção (PCM),
    baseada na especificação de requisitos do PCM e na planilha real
    '1. Planejamento_ordens_Gestão_v3.xlsx' (base de 14/07/2026), cujos
    agregados/CPS/amostra estão embutidos no JS do template. Nesta fase é só
    demonstração para validação com a gerência: não há modelos nem gravação
    no banco. Backend real (carga SAP 3x/dia com upsert, apontamento da
    operação e perfis) fica para depois do aval."""
    return render(request, "controle_ordens.html")
