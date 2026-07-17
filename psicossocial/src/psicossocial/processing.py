from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from psicossocial.excel_import import (
    ExcelImportError,
    _build_expected_columns_by_key,
    build_expected_columns,
    diagnose_excel,
)
from psicossocial.metodologia import Metodologia
from psicossocial.score import (
    DimensionScore,
    classify_risk_score,
    normalize_label,
    score_dimension,
)


PROBABILITY_LABELS = {
    1: "Eventual",
    2: "Moderada",
    3: "Frequente",
    4: "Intensa / Continua",
}
SEVERITY_LABELS = {
    1: "Baixa",
    2: "Moderada",
    3: "Alta",
    4: "Critica",
}
LEVEL_RANGES = {
    1: "1,00 - 1,99",
    2: "2,00 - 2,99",
    3: "3,00 - 3,99",
    4: "4,00 - 5,00",
}
MATRIX_RISK = {
    severity: {probability: severity * probability for probability in range(1, 5)}
    for severity in range(1, 5)
}
MATRIX_ACTIONS = {
    "Baixo": "Risco controlado ou com baixa relevancia ocupacional. Manter monitoramento.",
    "Moderado": "Necessita monitoramento e implementacao de acoes preventivas.",
    "Alto": "Requer priorizacao de medidas de controle e plano de acao. Monitoramento continuo.",
    "Critico": "Intervencao prioritaria/imediata. Acoes corretivas urgentes e acompanhamento gerencial.",
}
MATRIX_CLASSIFICATION_RANGES = (
    (1, 3, "Baixo"),
    (4, 6, "Moderado"),
    (8, 9, "Alto"),
    (12, 16, "Critico"),
)


@dataclass(frozen=True)
class RespondentResult:
    row_number: int
    respondent_id: str | None
    characterization: dict[str, str]
    dimensions: dict[str, DimensionScore]
    vco: dict[str, str]
    causes: tuple[str, ...]
    improvements: tuple[str, ...]

    def to_dict(self) -> dict[str, Any]:
        return {
            "linha": self.row_number,
            "id": self.respondent_id,
            "caracterizacao": self.characterization,
            "dimensoes": {
                code: {
                    "raw_score": score.raw_score,
                    "risk_score": score.risk_score,
                    "classificacao": score.classification,
                }
                for code, score in self.dimensions.items()
            },
            "vco": self.vco,
            "causas": list(self.causes),
            "melhorias": list(self.improvements),
        }


@dataclass(frozen=True)
class ProcessedSurvey:
    path: Path
    sheet_name: str
    respondents: tuple[RespondentResult, ...]

    def summary(self, metodologia: Metodologia) -> dict[str, Any]:
        return {
            "arquivo": str(self.path),
            "aba": self.sheet_name,
            "total_respondentes": len(self.respondents),
            "dimensoes": _dimension_summary(self.respondents, metodologia),
            "vco": _vco_summary(self.respondents),
            "causas_top": _top_occurrences(
                respondent.causes for respondent in self.respondents
            ),
            "melhorias_top": _top_occurrences(
                respondent.improvements for respondent in self.respondents
            ),
        }


@dataclass(frozen=True)
class DimensionExposure:
    dimension_code: str
    respondent_count: int
    mean_risk_score: float
    probability_level: int
    probability_label: str
    distribution_by_level: dict[int, int]
    impacted_count: int
    impacted_percent: float
    severity_level: int
    severity_label: str
    matrix_risk_level: int
    matrix_risk_classification: str
    matrix_recommended_action: str


def unit_matches(survey_unit: str, project_unit: str) -> bool:
    if not survey_unit or not project_unit:
        return False
        
    from unicodedata import combining, normalize

    def normalize_unit(name: str) -> str:
        without_accents = "".join(
            char for char in normalize("NFKD", name) if not combining(char)
        )
        cleaned = "".join(c if c.isalnum() else " " for c in without_accents)
        val = cleaned.lower().replace("minarecao", "mineracao")
        return " ".join(val.split())

    norm_survey = normalize_unit(survey_unit)
    norm_project = normalize_unit(project_unit)
    
    if not norm_survey or not norm_project:
        return False
        
    if norm_survey == norm_project:
        return True
        
    words_survey = set(norm_survey.split())
    words_project = set(norm_project.split())
    
    noise = {"de", "do", "da", "em", "e", "para", "o", "a", "os", "as", "rumo", "contrato"}
    words_survey -= noise
    words_project -= noise
    
    if not words_survey or not words_project:
        return False
        
    conflicting_subunits = {"fabrica", "loja", "eixo", "vias", "mineracao"}
    survey_subunit = words_survey.intersection(conflicting_subunits)
    project_subunit = words_project.intersection(conflicting_subunits)
    
    if survey_subunit or project_subunit:
        if survey_subunit != project_subunit:
            return False
            
    if words_survey.issubset(words_project) or words_project.issubset(words_survey):
        return True
        
    return False


def process_excel(
    path: str | Path,
    metodologia: Metodologia,
    sheet_name: str | None = None,
    header_row: int = 1,
    filter_unit: str | None = None,
) -> ProcessedSurvey:
    workbook_path = Path(path)
    if not workbook_path.exists():
        raise ExcelImportError(f"Arquivo Excel nao encontrado: {workbook_path}")

    diagnostic = diagnose_excel(
        workbook_path,
        metodologia,
        sheet_name=sheet_name,
        header_row=header_row,
    )
    _raise_if_incompatible(diagnostic)

    workbook = load_workbook(workbook_path, read_only=False, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]
    headers = _read_headers(worksheet, header_row)
    column_to_code, metadata_columns = _build_column_maps(headers, metodologia)

    respondents: list[RespondentResult] = []
    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        if not any(value is not None for value in row):
            continue

        values_by_code = {
            code: _cell_to_text(row[column_index - 1])
            for column_index, code in column_to_code.items()
            if column_index - 1 < len(row)
        }
        
        # Check if we should filter by unit (Q2)
        if filter_unit:
            survey_unit = values_by_code.get("Q2", "")
            if not unit_matches(survey_unit, filter_unit):
                continue

        metadata = {
            metadata_key: _cell_to_text(row[column_index - 1])
            for column_index, metadata_key in metadata_columns.items()
            if column_index - 1 < len(row)
        }

        respondents.append(
            _process_row(
                row_number=row_number,
                values_by_code=values_by_code,
                metadata=metadata,
                metodologia=metodologia,
            )
        )

    return ProcessedSurvey(
        path=workbook_path,
        sheet_name=worksheet.title,
        respondents=tuple(respondents),
    )


def _raise_if_incompatible(diagnostic: Any) -> None:
    problems = []

    if diagnostic.missing_columns:
        missing = "; ".join(
            f"{column.code} ({column.text})" for column in diagnostic.missing_columns
        )
        problems.append(f"colunas esperadas ausentes: {missing}")

    if diagnostic.unknown_likert_values:
        unknown = "; ".join(
            f"{code}: {', '.join(values)}"
            for code, values in diagnostic.unknown_likert_values.items()
        )
        problems.append(f"respostas Likert desconhecidas: {unknown}")

    if problems:
        raise ExcelImportError(
            "Planilha incompativel com a metodologia atual. "
            + " | ".join(problems)
            + ". Rode diagnosticar-excel para ver o detalhe."
        )


def _process_row(
    row_number: int,
    values_by_code: dict[str, str],
    metadata: dict[str, str],
    metodologia: Metodologia,
) -> RespondentResult:
    characterization = {
        question_code: values_by_code.get(question_code, "")
        for question_code in ("Q1", "Q2", "Q3", "Q4", "Q5", "Q6", "Q7", "Q8")
    }

    dimensions: dict[str, DimensionScore] = {}
    for question_code in ("Q9", "Q10", "Q11", "Q12", "Q13", "Q14", "Q15"):
        question_config = metodologia.perguntas[question_code]
        responses_by_item = {
            item["codigo"]: values_by_code.get(item["codigo"], "")
            for item in question_config["itens"]
        }
        score = score_dimension(
            question_config,
            responses_by_item,
            metodologia.data["escala_likert"],
            metodologia.data["score"],
        )
        dimensions[score.dimension_code] = score

    q16 = metodologia.perguntas["Q16"]
    vco = {
        item["codigo"]: values_by_code.get(item["codigo"], "")
        for item in q16["itens"]
    }

    return RespondentResult(
        row_number=row_number,
        respondent_id=metadata.get("id") or None,
        characterization=characterization,
        dimensions=dimensions,
        vco=vco,
        causes=_split_multi_select(values_by_code.get("Q17", "")),
        improvements=_split_multi_select(values_by_code.get("Q18", "")),
    )


def _read_headers(worksheet: Any, header_row: int) -> list[str]:
    header_values = next(
        worksheet.iter_rows(
            min_row=header_row,
            max_row=header_row,
            values_only=True,
        )
    )
    return ["" if value is None else str(value) for value in header_values]


def _build_column_maps(
    headers: list[str],
    metodologia: Metodologia,
) -> tuple[dict[int, str], dict[int, str]]:
    expected_by_key = _build_expected_columns_by_key(build_expected_columns(metodologia))
    column_to_code: dict[int, str] = {}
    metadata_columns: dict[int, str] = {}

    for index, header in enumerate(headers, start=1):
        canonical = _canonical_text(header)
        if canonical in expected_by_key:
            column_to_code[index] = expected_by_key[canonical].code
            continue

        normalized = normalize_label(header)
        if normalized in {"id", "email", "nome"}:
            metadata_columns[index] = normalized

    return column_to_code, metadata_columns


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def _split_multi_select(value: str) -> tuple[str, ...]:
    if not value:
        return ()
    parts = []
    for part in value.split(";"):
        cleaned = part.strip()
        if cleaned.endswith("."):
            cleaned = cleaned[:-1].strip()
        if cleaned:
            parts.append(cleaned)
    return tuple(parts)


def _dimension_summary(
    respondents: tuple[RespondentResult, ...],
    metodologia: Metodologia,
) -> dict[str, dict[str, Any]]:
    summary: dict[str, dict[str, Any]] = {}
    for dimension_code in metodologia.dimensoes:
        scores = [
            respondent.dimensions[dimension_code]
            for respondent in respondents
            if dimension_code in respondent.dimensions
        ]
        if not scores:
            continue

        raw_score = _average(score.raw_score for score in scores)
        risk_score = _average(score.risk_score for score in scores)
        summary[dimension_code] = {
            "raw_score_medio": raw_score,
            "risk_score_medio": risk_score,
            "classificacao": classify_risk_score(risk_score, metodologia.data["score"]),
        }

    return summary


def dimension_exposure_summary(
    respondents: tuple[RespondentResult, ...],
    dimension_code: str,
) -> DimensionExposure:
    scores = [
        respondent.dimensions[dimension_code].risk_score
        for respondent in respondents
        if dimension_code in respondent.dimensions
    ]
    distribution = {level: 0 for level in LEVEL_RANGES}
    for score in scores:
        distribution[score_to_level(score)] += 1

    mean_risk_score = _average(scores)
    probability_level = score_to_level(mean_risk_score)
    impacted_count = sum(1 for score in scores if score >= 2.0)
    impacted_percent = round(impacted_count / len(scores), 4) if scores else 0.0
    severity_level = representativeness_to_severity_level(impacted_percent)
    matrix_risk_level, matrix_risk_classification = classify_matrix_risk(
        probability_level=probability_level,
        severity_level=severity_level,
    )

    return DimensionExposure(
        dimension_code=dimension_code,
        respondent_count=len(scores),
        mean_risk_score=mean_risk_score,
        probability_level=probability_level,
        probability_label=PROBABILITY_LABELS[probability_level],
        distribution_by_level=distribution,
        impacted_count=impacted_count,
        impacted_percent=impacted_percent,
        severity_level=severity_level,
        severity_label=SEVERITY_LABELS[severity_level],
        matrix_risk_level=matrix_risk_level,
        matrix_risk_classification=matrix_risk_classification,
        matrix_recommended_action=MATRIX_ACTIONS[matrix_risk_classification],
    )


def score_to_level(score: float) -> int:
    rounded_score = round(float(score), 2)
    if 1.0 <= rounded_score <= 1.99:
        return 1
    if 2.0 <= rounded_score <= 2.99:
        return 2
    if 3.0 <= rounded_score <= 3.99:
        return 3
    if 4.0 <= rounded_score <= 5.0:
        return 4
    raise ExcelImportError(f"Score fora das faixas de nivel: {score}")


def classify_matrix_risk(probability_level: int, severity_level: int) -> tuple[int, str]:
    try:
        risk_result = MATRIX_RISK[severity_level][probability_level]
    except KeyError as exc:
        raise ExcelImportError(
            "Niveis invalidos para matriz 4x4: "
            f"probabilidade={probability_level}, severidade={severity_level}"
        ) from exc

    for minimum, maximum, classification in MATRIX_CLASSIFICATION_RANGES:
        if minimum <= risk_result <= maximum:
            return risk_result, classification

    raise ExcelImportError(f"Resultado fora das faixas da matriz 4x4: {risk_result}")


def representativeness_to_severity_level(percent: float) -> int:
    rounded_percent = round(float(percent), 4)
    if 0 <= rounded_percent <= 0.25:
        return 1
    if rounded_percent <= 0.50:
        return 2
    if rounded_percent <= 0.75:
        return 3
    if rounded_percent <= 1.0:
        return 4
    raise ExcelImportError(f"Representatividade fora da faixa esperada: {percent}")


def _vco_summary(respondents: tuple[RespondentResult, ...]) -> dict[str, Any]:
    total_yes = 0
    by_item: dict[str, int] = {}

    for respondent in respondents:
        for item_code, answer in respondent.vco.items():
            if normalize_label(answer) == "sim":
                total_yes += 1
                by_item[item_code] = by_item.get(item_code, 0) + 1

    return {
        "total_respostas_sim": total_yes,
        "por_item": by_item,
    }


def _top_occurrences(values_per_respondent: Any) -> list[dict[str, Any]]:
    counts: dict[str, int] = {}
    for values in values_per_respondent:
        for value in values:
            counts[value] = counts.get(value, 0) + 1

    return [
        {"valor": value, "frequencia": count}
        for value, count in sorted(counts.items(), key=lambda item: (-item[1], item[0]))
    ]


def _average(values: Any) -> float:
    numbers = tuple(float(value) for value in values)
    if not numbers:
        return 0.0
    return round(sum(numbers) / len(numbers), 4)


def _canonical_text(value: str) -> str:
    normalized = normalize_label(value.replace("\xa0", " "))
    chars = [char if char.isalnum() else " " for char in normalized]
    return " ".join("".join(chars).split())
