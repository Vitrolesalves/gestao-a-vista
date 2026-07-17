from __future__ import annotations

from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from psicossocial.excel_import import diagnose_excel
from psicossocial.metodologia import Metodologia
from psicossocial.processing import (
    LEVEL_RANGES,
    MATRIX_ACTIONS,
    MATRIX_CLASSIFICATION_RANGES,
    MATRIX_RISK,
    PROBABILITY_LABELS,
    ProcessedSurvey,
    RespondentResult,
    SEVERITY_LABELS,
    dimension_exposure_summary,
    process_excel,
    unit_matches,
)
from psicossocial.score import classify_risk_score, normalize_label
from psicossocial.workforce import build_representativity_summary


TECHNICAL_SHEETS = (
    "01_BASE_FORMS",
    "02_DE_PARA",
    "03_PARAMETROS",
    "04_BASE_TRATADA",
    "05_SCORE_DIMENSOES",
    "06_CRUZAMENTO_AMBIENTE",
    "07_CRUZAMENTO_COMPLEMENTAR",
    "08_VCO_TRATATIVA",
    "09_CAUSAS_MELHORIAS",
    "10_PLANO_ACAO",
    "12_INTEGRACAO_PGR_PCMSO",
    "13_LOG_AUDITORIA",
)

DIMENSION_CODES = ("DEM", "EST", "CTV", "REL", "ORG", "ITI", "SAG")

TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
SUBTLE_FILL = PatternFill("solid", fgColor="F3F6FA")
SECTION_FILL = PatternFill("solid", fgColor="EAF3F8")
WHITE_FONT = Font(color="FFFFFF", bold=True, size=12)
HEADER_FONT = Font(color="1F1F1F", bold=True)
TITLE_FONT = Font(color="FFFFFF", bold=True, size=13)
DESCRIPTION_FONT = Font(color="44546A", italic=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)
CLASSIFICATION_FILLS = {
    "Favoravel": PatternFill("solid", fgColor="E2F0D9"),
    "Atencao": PatternFill("solid", fgColor="FFF2CC"),
    "Critico": PatternFill("solid", fgColor="FCE4D6"),
    "Grave": PatternFill("solid", fgColor="F4CCCC"),
}
PRIORITY_FILLS = {
    "Baixa": PatternFill("solid", fgColor="E2F0D9"),
    "Media": PatternFill("solid", fgColor="FFF2CC"),
    "Alta": PatternFill("solid", fgColor="FCE4D6"),
    "Muito alta": PatternFill("solid", fgColor="F4CCCC"),
}
SHEET_TAB_COLORS = {
    "01_BASE_FORMS": "9EADCC",
    "02_DE_PARA": "5B9BD5",
    "03_PARAMETROS": "70AD47",
    "04_BASE_TRATADA": "9EADCC",
    "05_ELEGIBILIDADE": "70AD47",
    "05_SCORE_DIMENSOES": "4472C4",
    "06_CRUZAMENTO_AMBIENTE": "4472C4",
    "07_CRUZAMENTO_COMPLEMENTAR": "4472C4",
    "08_VCO_TRATATIVA": "C00000",
    "09_CAUSAS_MELHORIAS": "ED7D31",
    "10_PLANO_ACAO": "A5A5A5",
    "12_INTEGRACAO_PGR_PCMSO": "70AD47",
    "13_LOG_AUDITORIA": "A5A5A5",
}


def export_excel(
    input_path: str | Path,
    output_path: str | Path,
    metodologia: Metodologia,
    sheet_name: str | None = None,
    workforce_path: str | Path | None = None,
    filter_unit: str | None = None,
) -> Path:
    """Gera a primeira planilha tecnica de resultados do backend."""

    input_file = Path(input_path)
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    processed = process_excel(input_file, metodologia, sheet_name=sheet_name, filter_unit=filter_unit)
    workbook = Workbook()
    workbook.remove(workbook.active)

    _write_base_forms(workbook, input_file, sheet_name, metodologia, filter_unit)
    _write_de_para(workbook, metodologia)
    _write_parametros(workbook, metodologia)
    _write_base_tratada(workbook, processed, metodologia)
    if workforce_path:
        _write_elegibilidade(workbook, processed, workforce_path)
    _write_score_dimensoes(workbook, processed, metodologia)
    _write_cruzamento_ambiente(workbook, processed, metodologia)
    _write_cruzamento_complementar(workbook, processed, metodologia)
    _write_vco(workbook, processed)
    _write_causas_melhorias(workbook, processed)
    _write_plano_acao(workbook, processed, metodologia)
    _write_integracao(workbook)
    _write_log(workbook, input_file, processed)

    workbook.save(output_file)
    return output_file


def _write_base_forms(
    workbook: Workbook,
    input_path: Path,
    sheet_name: str | None,
    metodologia: Metodologia,
    filter_unit: str | None = None,
) -> None:
    source = load_workbook(input_path, read_only=False, data_only=True)
    source_ws = source[sheet_name] if sheet_name else source.worksheets[0]
    ws = _new_sheet(workbook, "01_BASE_FORMS", "Base exportada do Forms copiada para referência.")

    q2_index = None
    if filter_unit:
        diagnostic = diagnose_excel(input_path, metodologia, sheet_name=sheet_name)
        for match in diagnostic.matched_columns:
            if match.code == "Q2":
                q2_index = match.column_index - 1
                break

    header_processed = False
    for row in source_ws.iter_rows(values_only=True):
        if not any(value is not None for value in row):
            continue
        if not header_processed:
            ws.append(list(row))
            header_processed = True
            continue

        if filter_unit and q2_index is not None and q2_index < len(row):
            survey_unit = str(row[q2_index] or "").strip()
            if not unit_matches(survey_unit, filter_unit):
                continue

        ws.append(list(row))

    _style_table(ws)


def _write_de_para(workbook: Workbook, metodologia: Metodologia) -> None:
    ws = _new_sheet(workbook, "02_DE_PARA", "Mapa oficial de uso tecnico de cada questao.")
    ws.append(["Questao", "Campo no formulario", "Uso tecnico na analise", "Tratamento"])

    for question_code, question in metodologia.perguntas.items():
        question_type = question["tipo"]
        if question_type == "dimensao_score":
            treatment = f"Dimensao {question['codigo_dimensao']} - {question['regra_padrao']}"
            usage = "Score psicossocial quantitativo"
            field = question["titulo"]
        elif question_type == "vco":
            treatment = "Fluxo separado; nao compoe media"
            usage = "Violencia e comportamentos ofensivos"
            field = question["titulo"]
        elif question_type == "qualitativa_estruturada":
            treatment = "Frequencia; apoio ao plano de acao"
            usage = ", ".join(question.get("uso_backend", []))
            field = question["titulo"]
        else:
            treatment = "Caracterizacao e cruzamento"
            usage = ", ".join(question.get("uso_backend", []))
            field = question["titulo"]
        ws.append([question_code, field, usage, treatment])

    _style_table(ws)


def _write_parametros(workbook: Workbook, metodologia: Metodologia) -> None:
    ws = _new_sheet(workbook, "03_PARAMETROS", "Regras de pontuacao, criticidade e listas de apoio.")
    ws.append(["Resposta Likert", "Pontuacao", None, "Score minimo", "Score maximo", "Classificacao"])

    respostas = metodologia.data["escala_likert"]["respostas"]
    faixas = metodologia.data["score"]["regra_classificacao"]["faixas"]
    max_rows = max(len(respostas), len(faixas))
    for index in range(max_rows):
        likert = list(respostas.items())[index] if index < len(respostas) else ("", "")
        faixa = faixas[index] if index < len(faixas) else {}
        ws.append(
            [
                likert[0],
                likert[1],
                None,
                faixa.get("minimo", ""),
                faixa.get("maximo", ""),
                faixa.get("classe", ""),
            ]
        )

    ws.append([])
    ws.append(["Dimensao", "Pergunta", "Regra"])
    for code, dimension in metodologia.dimensoes.items():
        ws.append([code, dimension["pergunta"], dimension["regra"]])

    ws.append([])
    ws.append([
        "Nivel",
        "Faixa score medio",
        "Probabilidade / Exposicao",
        "Representatividade do GHE",
        "Severidade / Impacto",
    ])
    for level in LEVEL_RANGES:
        representativeness = {
            1: "Ate 25%",
            2: ">25% a 50%",
            3: ">50% a 75%",
            4: "Acima de 75%",
        }[level]
        ws.append(
            [
                level,
                LEVEL_RANGES[level],
                PROBABILITY_LABELS[level],
                representativeness,
                SEVERITY_LABELS[level],
            ]
        )

    ws.append([])
    ws.append(["Matriz 4x4 - Severidade", "Probabilidade 1", "Probabilidade 2", "Probabilidade 3", "Probabilidade 4"])
    for severity in sorted(MATRIX_RISK, reverse=True):
        row = [f"{severity} - {SEVERITY_LABELS[severity]}"]
        for probability in sorted(MATRIX_RISK[severity]):
            risk_result = MATRIX_RISK[severity][probability]
            _, risk_classification = _classify_matrix_result_for_export(risk_result)
            row.append(f"{risk_result} - {risk_classification}")
        ws.append(row)

    ws.append([])
    ws.append(["Resultado da matriz", "Classificacao do risco", "Acao recomendada"])
    for minimum, maximum, classification in MATRIX_CLASSIFICATION_RANGES:
        result_range = f"{minimum} a {maximum}"
        ws.append([result_range, classification, MATRIX_ACTIONS[classification]])

    _style_table(ws, extra_header_rows=(11, 20, 27, 33))


def _write_base_tratada(
    workbook: Workbook,
    processed: ProcessedSurvey,
    metodologia: Metodologia,
) -> None:
    ws = _new_sheet(workbook, "04_BASE_TRATADA", "Base padronizada com scores de risco por item.")
    headers = [
        "ID_Resposta",
        "Q1_Cargo_Funcao",
        "Q2_Unidade_Contrato",
        "Q3_Turno",
        "Q4_Escala",
        "Q5_Area_Atividade",
        "Q6_Tipo_Ambiente",
        "Q7_Tempo_Funcao",
        "Q8_Grupo_Idade",
    ]
    item_codes = _score_item_codes(metodologia)
    vco_codes = _vco_item_codes(metodologia)
    headers.extend([f"{code}_ScoreRisco" for code in item_codes])
    headers.extend([f"{code}_Flag" for code in vco_codes])
    headers.extend(["Q17_Desgaste", "Q18_Melhoria"])
    ws.append(headers)

    for respondent in processed.respondents:
        row = [
            respondent.respondent_id,
            respondent.characterization.get("Q1", ""),
            respondent.characterization.get("Q2", ""),
            respondent.characterization.get("Q3", ""),
            respondent.characterization.get("Q4", ""),
            respondent.characterization.get("Q5", ""),
            respondent.characterization.get("Q6", ""),
            respondent.characterization.get("Q7", ""),
            respondent.characterization.get("Q8", ""),
        ]
        item_scores = {
            item.item_code: item.risk_score
            for dimension in respondent.dimensions.values()
            for item in dimension.items
        }
        row.extend(item_scores.get(code, "") for code in item_codes)
        row.extend(_vco_flag(respondent.vco.get(code, "")) for code in vco_codes)
        row.extend(["; ".join(respondent.causes), "; ".join(respondent.improvements)])
        ws.append(row)

    _style_table(ws)
    _format_numeric_columns(ws, score_columns=range(10, 40), flag_columns=range(40, 45))


def _write_elegibilidade(
    workbook: Workbook,
    processed: ProcessedSurvey,
    workforce_path: str | Path,
) -> None:
    ws = _new_sheet(
        workbook,
        "05_ELEGIBILIDADE",
        "Elegibilidade, adesao e representatividade por local e GHE com base na SRA.",
    )
    ws.append([
        "Recorte",
        "Local",
        "GHE/Grupo",
        "Colaboradores_Total_SRA",
        "Respondentes_Validos_Forms",
        "Adesao_%",
        "Representativo_60%?",
        "Analise_Segmentada_Permitida?",
    ])
    respondent_characterizations = [
        respondent.characterization for respondent in processed.respondents
    ]
    for row in build_representativity_summary(workforce_path, respondent_characterizations):
        ws.append([
            row["recorte"],
            row["local"],
            row["ghe"],
            row["colaboradores_total"],
            row["respondentes_validos"],
            row["adesao_percentual"],
            row["representativo_60"],
            row["analise_segmentada_permitida"],
        ])
    _style_table(ws)
    _format_percent_columns(ws, (6,))


def _write_score_dimensoes(
    workbook: Workbook,
    processed: ProcessedSurvey,
    metodologia: Metodologia,
) -> None:
    ws = _new_sheet(workbook, "05_SCORE_DIMENSOES", "Score de risco por dimensao e classificacao por resposta.")
    headers = [
        "ID_Resposta",
        "Unidade",
        "Ambiente",
        "Area",
        "Funcao",
        "Turno",
        "Escala",
        "Tempo_Funcao",
        "Grupo_Idade",
    ]
    for code in DIMENSION_CODES:
        headers.extend([f"{code}_Score", f"{code}_Classificacao"])
    headers.extend(["Score_Geral", "Criticidade_Geral"])
    ws.append(headers)

    for respondent in processed.respondents:
        row = [
            respondent.respondent_id,
            respondent.characterization.get("Q2", ""),
            respondent.characterization.get("Q6", ""),
            respondent.characterization.get("Q5", ""),
            respondent.characterization.get("Q1", ""),
            respondent.characterization.get("Q3", ""),
            respondent.characterization.get("Q4", ""),
            respondent.characterization.get("Q7", ""),
            respondent.characterization.get("Q8", ""),
        ]
        risk_scores = []
        for code in DIMENSION_CODES:
            dimension = respondent.dimensions[code]
            risk_scores.append(dimension.risk_score)
            row.extend([dimension.risk_score, dimension.classification])
        general = _average(risk_scores)
        row.extend([general, classify_risk_score(general, metodologia.data["score"])])
        ws.append(row)

    _style_table(ws)
    _format_score_sheet(ws, classification_columns=(11, 13, 15, 17, 19, 21, 23, 25))
    _format_numeric_columns(ws, score_columns=(10, 12, 14, 16, 18, 20, 22, 24))


def _write_cruzamento_ambiente(
    workbook: Workbook,
    processed: ProcessedSurvey,
    metodologia: Metodologia,
) -> None:
    ws = _new_sheet(
        workbook,
        "06_CRUZAMENTO_AMBIENTE",
        "Cruzamento Q6 x dimensoes com probabilidade pela media e severidade pela representatividade do GHE.",
    )
    headers = [
        "Ambiente",
        "Dimensao",
        "Respondentes",
        "Score_Medio",
        "Classificacao_COPSOQ",
        "Probabilidade_Nivel",
        "Probabilidade",
        "Nivel_1_Qtd",
        "Nivel_2_Qtd",
        "Nivel_3_Qtd",
        "Nivel_4_Qtd",
        "Percentual_Atencao_Risco_GHE",
        "Severidade_Nivel",
        "Severidade",
        "Matriz_Resultado",
        "Matriz_Classificacao",
        "Matriz_Acao_Recomendada",
        "Plano_Acao?",
    ]
    ws.append(headers)

    for ambiente, respondents in _group_by(processed.respondents, "Q6").items():
        for code in DIMENSION_CODES:
            exposure = dimension_exposure_summary(tuple(respondents), code)
            classification = classify_risk_score(
                exposure.mean_risk_score,
                metodologia.data["score"],
            )
            ws.append(
                [
                    ambiente,
                    code,
                    exposure.respondent_count,
                    exposure.mean_risk_score,
                    classification,
                    exposure.probability_level,
                    exposure.probability_label,
                    exposure.distribution_by_level[1],
                    exposure.distribution_by_level[2],
                    exposure.distribution_by_level[3],
                    exposure.distribution_by_level[4],
                    exposure.impacted_percent,
                    exposure.severity_level,
                    exposure.severity_label,
                    exposure.matrix_risk_level,
                    exposure.matrix_risk_classification,
                    exposure.matrix_recommended_action,
                    "Sim" if exposure.matrix_risk_classification != "Baixo" else "Nao",
                ]
            )

    _style_table(ws)
    _format_score_sheet(ws, classification_columns=(5, 16), priority_columns=())
    _format_numeric_columns(ws, score_columns=(4,), flag_columns=(6, 8, 9, 10, 11, 13, 15))
    _format_percent_columns(ws, (12,))


def _write_cruzamento_complementar(
    workbook: Workbook,
    processed: ProcessedSurvey,
    metodologia: Metodologia,
) -> None:
    ws = _new_sheet(
        workbook,
        "07_CRUZAMENTO_COMPLEMENTAR",
        "Cruzamentos complementares por unidade, area, funcao, tempo, turno e escala.",
    )
    headers = [
        "Recorte",
        "Grupo",
        "Dimensao",
        "Respondentes",
        "Score_Medio",
        "Classificacao_COPSOQ",
        "Probabilidade_Nivel",
        "Probabilidade",
        "Nivel_1_Qtd",
        "Nivel_2_Qtd",
        "Nivel_3_Qtd",
        "Nivel_4_Qtd",
        "Percentual_Atencao_Risco_GHE",
        "Severidade_Nivel",
        "Severidade",
        "Matriz_Resultado",
        "Matriz_Classificacao",
        "Matriz_Acao_Recomendada",
        "Plano_Acao?",
    ]
    ws.append(headers)

    complementary_cuts = (
        ("Q2_Unidade_Contrato", "Q2"),
        ("Q1_GHE_Negocio_Atividade", "Q1"),
        ("Q5_Area_Atividade", "Q5"),
        ("Q7_Tempo_Funcao", "Q7"),
        ("Q3_Turno", "Q3"),
        ("Q4_Escala", "Q4"),
    )
    for cut_label, question_code in complementary_cuts:
        for group, respondents in _group_by(processed.respondents, question_code).items():
            for code in DIMENSION_CODES:
                exposure = dimension_exposure_summary(tuple(respondents), code)
                classification = classify_risk_score(
                    exposure.mean_risk_score,
                    metodologia.data["score"],
                )
                ws.append(
                    [
                        cut_label,
                        group,
                        code,
                        exposure.respondent_count,
                        exposure.mean_risk_score,
                        classification,
                        exposure.probability_level,
                        exposure.probability_label,
                        exposure.distribution_by_level[1],
                        exposure.distribution_by_level[2],
                        exposure.distribution_by_level[3],
                        exposure.distribution_by_level[4],
                        exposure.impacted_percent,
                        exposure.severity_level,
                        exposure.severity_label,
                        exposure.matrix_risk_level,
                        exposure.matrix_risk_classification,
                        exposure.matrix_recommended_action,
                        "Sim" if exposure.matrix_risk_classification != "Baixo" else "Nao",
                    ]
                )

    _style_table(ws)
    _format_score_sheet(ws, classification_columns=(6, 17), priority_columns=())
    _format_numeric_columns(ws, score_columns=(5,), flag_columns=(7, 9, 10, 11, 12, 14, 16))
    _format_percent_columns(ws, (13,))


def _write_vco(workbook: Workbook, processed: ProcessedSurvey) -> None:
    ws = _new_sheet(workbook, "08_VCO_TRATATIVA", "Tratamento institucional da Q16. VCO nao compoe media.")
    headers = [
        "Ambiente",
        "Respondentes",
        "Desrespeito grave",
        "Ameaca/intimidacao",
        "Assedio moral",
        "Assedio sexual",
        "Violencia fisica",
        "Total Sim",
        "% Sim",
        "Criticidade VCO",
        "Encaminhamento recomendado",
        "Status",
    ]
    ws.append(headers)
    vco_codes = ["Q16_VCO_1", "Q16_VCO_2", "Q16_VCO_3", "Q16_VCO_4", "Q16_VCO_5"]
    for ambiente, respondents in _group_by(processed.respondents, "Q6").items():
        counts = [_count_yes(r.vco.get(code, "") for r in respondents) for code in vco_codes]
        total_yes = sum(counts)
        total_possible = max(len(respondents) * len(vco_codes), 1)
        percent = round(total_yes / total_possible, 4)
        criticality = "Critico" if total_yes else "Favoravel"
        status = "Encaminhamento restrito recomendado" if total_yes else "Monitorar"
        ws.append(
            [
                ambiente,
                len(respondents),
                *counts,
                total_yes,
                percent,
                criticality,
                "Avaliar indicadores agregados e reforcar canais internos.",
                status,
            ]
        )

    _style_table(ws)
    _format_percent_columns(ws, (9,))
    _format_score_sheet(ws, classification_columns=(10,))


def _write_causas_melhorias(workbook: Workbook, processed: ProcessedSurvey) -> None:
    ws = _new_sheet(workbook, "09_CAUSAS_MELHORIAS", "Q17 e Q18 para causas percebidas e melhorias.")
    ws.append(["Tipo", "Valor", "Quantidade", "% respondentes"])
    total = max(len(processed.respondents), 1)
    for item in _top_occurrences(r.causes for r in processed.respondents):
        ws.append(["Q17_Causa", item[0], item[1], round(item[1] / total, 4)])
    for item in _top_occurrences(r.improvements for r in processed.respondents):
        ws.append(["Q18_Melhoria", item[0], item[1], round(item[1] / total, 4)])
    _style_table(ws)
    _format_percent_columns(ws, (4,))


def _write_plano_acao(
    workbook: Workbook,
    processed: ProcessedSurvey,
    metodologia: Metodologia,
) -> None:
    ws = _new_sheet(workbook, "10_PLANO_ACAO", "Plano de acao preliminar por criticidade. Validar tecnicamente.")
    ws.append(
        [
            "Ambiente",
            "Dimensao",
            "Score",
            "Criticidade",
            "Causa percebida principal",
            "Acao sugerida",
            "Responsavel",
            "Data inicio",
            "Prazo final",
            "Status",
            "Evidencia",
            "Integracao PGR/AEP/PCMSO",
            "Prioridade",
        ]
    )
    today = datetime.today().date()
    for ambiente, respondents in _group_by(processed.respondents, "Q6").items():
        main_cause = _top_occurrences(r.causes for r in respondents)
        cause = main_cause[0][0] if main_cause else ""
        for code in DIMENSION_CODES:
            risk = _average(r.dimensions[code].risk_score for r in respondents)
            classification = classify_risk_score(risk, metodologia.data["score"])
            if classification == "Favoravel":
                continue
            ws.append(
                [
                    ambiente,
                    code,
                    risk,
                    classification,
                    cause,
                    _suggest_action(code),
                    "Gestor do contrato + SSMA/SESMT + RH",
                    today,
                    today + timedelta(days=60),
                    "Pendente de validacao tecnica",
                    "Ata, plano executado, comunicacao, lista de presenca ou evidencia equivalente.",
                    "PGR/AEP: avaliar organizacao do trabalho; PCMSO: monitorar coletivamente quando aplicavel.",
                    _priority(classification),
                ]
            )
    _style_table(ws)
    _format_score_sheet(ws, classification_columns=(4,), priority_columns=(13,))
    _format_numeric_columns(ws, score_columns=(3,))
    _format_date_columns(ws, (8, 9))


def _write_integracao(workbook: Workbook) -> None:
    ws = _new_sheet(workbook, "12_INTEGRACAO_PGR_PCMSO", "Evidencia de integracao ocupacional para PGR, AEP e PCMSO.")
    ws.append(["Achado", "Criterio de disparo", "Encaminhamento tecnico", "Evidencia esperada"])
    ws.append([
        "Dimensao em Atencao/Critico/Grave",
        "risk_score >= 2.00 em Q9 a Q15",
        "Avaliar plano de acao e integracao na AEP/PGR",
        "Relatorio, matriz, plano e evidencias",
    ])
    ws.append([
        "VCO com resposta Sim",
        "Qualquer indicador consolidado positivo em Q16",
        "Tratativa restrita institucional conforme canais internos",
        "Registro de encaminhamento restrito",
    ])
    _style_table(ws)


def _write_log(workbook: Workbook, input_path: Path, processed: ProcessedSurvey) -> None:
    ws = _new_sheet(workbook, "13_LOG_AUDITORIA", "Controle de versao, execucao e evidencias.")
    ws.append(["Data", "Responsavel", "Acao realizada", "Documento/Evidencia", "Observacao"])
    ws.append([
        datetime.now(),
        "Backend Psicossocial NR-01",
        "Geracao automatica da planilha tecnica",
        str(input_path),
        f"{len(processed.respondents)} respondentes processados.",
    ])
    _style_table(ws)


def _new_sheet(workbook: Workbook, name: str, description: str):
    ws = workbook.create_sheet(name)
    ws.append([name])
    ws.append([description])
    ws.append([])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    ws["A1"].fill = TITLE_FILL
    ws["A1"].font = TITLE_FONT
    ws["A2"].fill = SUBTLE_FILL
    ws["A2"].font = DESCRIPTION_FONT
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.sheet_properties.tabColor = SHEET_TAB_COLORS.get(name, "5B9BD5")
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 34
    return ws


def _style_table(ws, header_row: int = 4, extra_header_rows: tuple[int, ...] = ()) -> None:
    if ws.max_row < header_row:
        return
    for row_number in (header_row, *extra_header_rows):
        if ws.max_row < row_number:
            continue
        for cell in ws[row_number]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[row_number].height = 32

    for row in ws.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if cell.row % 2 == 1 and cell.value is not None:
                cell.fill = SECTION_FILL

    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = ws.dimensions
    max_col = min(ws.max_column, 30)
    for column in range(1, max_col + 1):
        letter = get_column_letter(column)
        width = max(
            12,
            min(
                42,
                max(
                    len(str(ws.cell(row=row, column=column).value or ""))
                    for row in range(1, min(ws.max_row, 50) + 1)
                )
                + 2,
            ),
        )
        ws.column_dimensions[letter].width = width

    for row_number in range(header_row + 1, ws.max_row + 1):
        ws.row_dimensions[row_number].height = 28


def _format_score_sheet(
    ws,
    classification_columns: tuple[int, ...],
    priority_columns: tuple[int, ...] = (),
    header_row: int = 4,
) -> None:
    for row in range(header_row + 1, ws.max_row + 1):
        for column in classification_columns:
            cell = ws.cell(row=row, column=column)
            fill = CLASSIFICATION_FILLS.get(str(cell.value))
            if fill:
                cell.fill = fill
                cell.font = HEADER_FONT
        for column in priority_columns:
            cell = ws.cell(row=row, column=column)
            fill = PRIORITY_FILLS.get(str(cell.value))
            if fill:
                cell.fill = fill
                cell.font = HEADER_FONT


def _format_numeric_columns(
    ws,
    score_columns: Any = (),
    flag_columns: Any = (),
    header_row: int = 4,
) -> None:
    for row in range(header_row + 1, ws.max_row + 1):
        for column in score_columns:
            cell = ws.cell(row=row, column=column)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"
        for column in flag_columns:
            cell = ws.cell(row=row, column=column)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0"


def _format_percent_columns(ws, columns: tuple[int, ...], header_row: int = 4) -> None:
    for row in range(header_row + 1, ws.max_row + 1):
        for column in columns:
            cell = ws.cell(row=row, column=column)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00%"


def _format_date_columns(ws, columns: tuple[int, ...], header_row: int = 4) -> None:
    for row in range(header_row + 1, ws.max_row + 1):
        for column in columns:
            ws.cell(row=row, column=column).number_format = "dd/mm/yyyy"


def _classify_matrix_result_for_export(risk_result: int) -> tuple[int, str]:
    for minimum, maximum, classification in MATRIX_CLASSIFICATION_RANGES:
        if minimum <= risk_result <= maximum:
            return risk_result, classification
    return risk_result, "Nao classificado"


def _score_item_codes(metodologia: Metodologia) -> list[str]:
    return [
        item["codigo"]
        for question_code in ("Q9", "Q10", "Q11", "Q12", "Q13", "Q14", "Q15")
        for item in metodologia.perguntas[question_code]["itens"]
    ]


def _vco_item_codes(metodologia: Metodologia) -> list[str]:
    return [item["codigo"] for item in metodologia.perguntas["Q16"]["itens"]]


def _vco_flag(value: str) -> int:
    return 1 if normalize_label(value) == "sim" else 0


def _count_yes(values: Any) -> int:
    return sum(1 for value in values if normalize_label(value) == "sim")


def _group_by(respondents: tuple[RespondentResult, ...], question_code: str) -> dict[str, list[RespondentResult]]:
    groups: dict[str, list[RespondentResult]] = defaultdict(list)
    for respondent in respondents:
        key = respondent.characterization.get(question_code) or "Nao informado"
        groups[key].append(respondent)
    return dict(sorted(groups.items(), key=lambda item: item[0]))


def _top_occurrences(values_per_respondent: Any) -> list[tuple[str, int]]:
    counts: dict[str, int] = {}
    for values in values_per_respondent:
        for value in values:
            counts[value] = counts.get(value, 0) + 1
    return sorted(counts.items(), key=lambda item: (-item[1], item[0]))


def _average(values: Any) -> float:
    numbers = tuple(float(value) for value in values)
    if not numbers:
        return 0.0
    return round(sum(numbers) / len(numbers), 4)


def _suggest_action(dimension_code: str) -> str:
    actions = {
        "DEM": "Revisar carga, ritmo, prioridades e distribuicao de demandas.",
        "EST": "Avaliar sinais coletivos de esgotamento e medidas de recuperacao.",
        "CTV": "Avaliar jornada, escala e interferencias trabalho-vida pessoal.",
        "REL": "Reforcar apoio da lideranca, comunicacao e tratamento de conflitos.",
        "ORG": "Melhorar clareza, autonomia, informacoes e organizacao das atividades.",
        "ITI": "Avaliar adequacao entre trabalho, expectativas, habilidades e identidade profissional.",
        "SAG": "Monitorar saude geral percebida e fatores do trabalho associados.",
    }
    return actions.get(dimension_code, "Definir acao tecnica conforme analise.")


def _priority(classification: str) -> str:
    return {
        "Atencao": "Media",
        "Critico": "Alta",
        "Grave": "Muito alta",
    }.get(classification, "Baixa")
