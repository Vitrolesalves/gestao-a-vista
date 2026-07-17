from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from psicossocial.metodologia import Metodologia
from psicossocial.score import ScoreError, build_response_scale, normalize_label


METADATA_HEADERS = {
    "id",
    "hora de inicio",
    "hora de conclusao",
    "email",
    "nome",
    "hora da ultima modificacao",
}

HEADER_ALIASES = {
    "Q1": (
        "Qual negócio e atividade você se enquadra?",
    ),
    "Q2": (
        "Em qual unidade ou contrato você trabalha?",
    ),
    "Q13_ORG_2": (
        "Tenho autonomia (liberdade) para realizar minhas atividades diárias. (Cumprimento de procedimentos operacionais)",
        "Tenho autonomia/liberdade para realizar minhas atividades diárias.",
    ),
    "Q13_ORG_3": (
        "A empresa me proporciona oportunidade de crescimento profissional (Ex. cursos/treinamentos  de capacitação)",
        "A empresa me proporciona oportunidade de crescimento profissional.",
    ),
    "Q14_SAG_4": (
        "Tenho dificuldade para dormir por causa do trabalho.",
    ),
    "Q15_ITI_1": (
        "Meu trabalho não corresponde às minhas expectativas.",
        "Meu trabalho  corresponde às minhas expectativas.",
    ),
    "Q15_ITI_2": (
        "Sinto que minhas habilidades não são bem aproveitadas.",
    ),
    "Q15_ITI_3": (
        "Tenho pouca identificação com o trabalho que realizo",
    ),
    "Q16_VCO_1": (
        "Já sofreu desrespeito grave no trabalho?",
    ),
    "Q16_VCO_2": (
        "Já sofreu ameaça verbal ou intimidação?",
    ),
    "Q16_VCO_3": (
        "Já sofreu assédio moral?",
    ),
    "Q16_VCO_4": (
        "Já sofreu assédio sexual?",
    ),
}


class ExcelImportError(ValueError):
    """Erro de leitura ou diagnostico do Excel de entrada."""


@dataclass(frozen=True)
class ExpectedColumn:
    code: str
    text: str
    group: str


@dataclass(frozen=True)
class ColumnMatch:
    code: str
    group: str
    header: str
    column_index: int


@dataclass(frozen=True)
class ExcelDiagnostic:
    path: Path
    sheet_name: str
    total_rows: int
    total_data_rows: int
    total_columns: int
    matched_columns: tuple[ColumnMatch, ...]
    missing_columns: tuple[ExpectedColumn, ...]
    metadata_columns: tuple[str, ...]
    unmapped_columns: tuple[str, ...]
    unknown_likert_values: dict[str, tuple[str, ...]]

    def to_dict(self) -> dict[str, Any]:
        return {
            "arquivo": str(self.path),
            "aba": self.sheet_name,
            "total_linhas": self.total_rows,
            "total_respostas": self.total_data_rows,
            "total_colunas": self.total_columns,
            "colunas_mapeadas": len(self.matched_columns),
            "colunas_esperadas_ausentes": [
                {"codigo": column.code, "grupo": column.group, "texto": column.text}
                for column in self.missing_columns
            ],
            "colunas_metadata": list(self.metadata_columns),
            "colunas_nao_mapeadas": list(self.unmapped_columns),
            "valores_likert_desconhecidos": {
                code: list(values)
                for code, values in self.unknown_likert_values.items()
            },
        }


def build_expected_columns(metodologia: Metodologia) -> tuple[ExpectedColumn, ...]:
    expected: list[ExpectedColumn] = []

    for question_code, question in metodologia.perguntas.items():
        question_type = question["tipo"]
        if question_type == "caracterizacao":
            expected.append(
                ExpectedColumn(
                    code=question_code,
                    text=str(question["texto"]),
                    group="caracterizacao",
                )
            )
        elif question_type == "dimensao_score":
            for item in question["itens"]:
                expected.append(
                    ExpectedColumn(
                        code=str(item["codigo"]),
                        text=str(item["texto"]),
                        group=str(question["codigo_dimensao"]),
                    )
                )
        elif question_type == "vco":
            for item in question["itens"]:
                expected.append(
                    ExpectedColumn(
                        code=str(item["codigo"]),
                        text=str(item["texto"]),
                        group="VCO",
                    )
                )
        elif question_type == "qualitativa_estruturada":
            expected.append(
                ExpectedColumn(
                    code=question_code,
                    text=str(question["texto"]),
                    group=str(question["codigo_campo"]),
                )
            )

    return tuple(expected)


def diagnose_excel(
    path: str | Path,
    metodologia: Metodologia,
    sheet_name: str | None = None,
    header_row: int = 1,
) -> ExcelDiagnostic:
    workbook_path = Path(path)
    if not workbook_path.exists():
        raise ExcelImportError(f"Arquivo Excel nao encontrado: {workbook_path}")

    workbook = load_workbook(workbook_path, read_only=False, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]

    header_values = next(
        worksheet.iter_rows(
            min_row=header_row,
            max_row=header_row,
            values_only=True,
        )
    )
    headers = ["" if value is None else str(value) for value in header_values]
    expected_columns = build_expected_columns(metodologia)
    expected_by_key = _build_expected_columns_by_key(expected_columns)

    matched: list[ColumnMatch] = []
    matched_codes: set[str] = set()
    metadata_columns: list[str] = []
    unmapped_columns: list[str] = []
    code_by_column_index: dict[int, str] = {}
    score_codes = {
        column.code
        for column in expected_columns
        if column.group in metodologia.dimensoes
    }

    for index, header in enumerate(headers, start=1):
        if not header:
            continue

        canonical = _canonical_text(header)
        if canonical in expected_by_key:
            expected = expected_by_key[canonical]
            matched.append(
                ColumnMatch(
                    code=expected.code,
                    group=expected.group,
                    header=header,
                    column_index=index,
                )
            )
            matched_codes.add(expected.code)
            code_by_column_index[index] = expected.code
        elif normalize_label(header) in METADATA_HEADERS:
            metadata_columns.append(header)
        else:
            unmapped_columns.append(header)

    missing = tuple(
        column for column in expected_columns if column.code not in matched_codes
    )
    unknown_likert_values = _find_unknown_likert_values(
        worksheet=worksheet,
        header_row=header_row,
        code_by_column_index=code_by_column_index,
        score_codes=score_codes,
        metodologia=metodologia,
    )

    return ExcelDiagnostic(
        path=workbook_path,
        sheet_name=worksheet.title,
        total_rows=worksheet.max_row,
        total_data_rows=_count_non_empty_data_rows(worksheet, header_row),
        total_columns=worksheet.max_column,
        matched_columns=tuple(matched),
        missing_columns=missing,
        metadata_columns=tuple(metadata_columns),
        unmapped_columns=tuple(unmapped_columns),
        unknown_likert_values=unknown_likert_values,
    )


def _find_unknown_likert_values(
    worksheet: Any,
    header_row: int,
    code_by_column_index: dict[int, str],
    score_codes: set[str],
    metodologia: Metodologia,
) -> dict[str, tuple[str, ...]]:
    try:
        response_scale = build_response_scale(metodologia.data["escala_likert"])
    except ScoreError as exc:
        raise ExcelImportError(str(exc)) from exc

    unknown: dict[str, set[str]] = {}
    for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
        for index, value in enumerate(row, start=1):
            code = code_by_column_index.get(index)
            if code not in score_codes or value is None:
                continue

            raw_value = str(value)
            if normalize_label(raw_value) not in response_scale:
                unknown.setdefault(code, set()).add(raw_value)

    return {code: tuple(sorted(values)) for code, values in sorted(unknown.items())}


def _count_non_empty_data_rows(worksheet: Any, header_row: int) -> int:
    total = 0
    for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
        if any(value is not None for value in row):
            total += 1
    return total


def _canonical_text(value: str) -> str:
    normalized = normalize_label(value.replace("\xa0", " "))
    chars = [char if char.isalnum() else " " for char in normalized]
    return " ".join("".join(chars).split())


def _build_expected_columns_by_key(
    expected_columns: tuple[ExpectedColumn, ...]
) -> dict[str, ExpectedColumn]:
    by_key = {_canonical_text(column.text): column for column in expected_columns}
    by_code = {column.code: column for column in expected_columns}
    for code, aliases in HEADER_ALIASES.items():
        column = by_code.get(code)
        if not column:
            continue
        for alias in aliases:
            by_key[_canonical_text(alias)] = column
    return by_key
