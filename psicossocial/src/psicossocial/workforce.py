from __future__ import annotations

from collections import Counter
from pathlib import Path
from typing import Any
from unicodedata import combining, normalize

from openpyxl import load_workbook


GHE_ALIASES = (
    ("SEGURANCA ELETRONICA", ("SEG ELETRONICA", "SEGURANCA ELETRONICA")),
    ("SEGURANCA HUMANA", ("SEGURANCA HUMANA", "VIGILANCIA")),
    ("LOGISTICA", ("LOGISTICA",)),
    ("INDIRETO (ADMINISTRATIVO / CORPORATIVO)", ("INDIRETO", "ADMINISTRATIVO", "CORPORATIVO")),
    ("INFRA-SERVICOS - BOMBEIRO", ("BOMBEIRO",)),
    ("INFRA-SERVICOS", ("INFRA SERVICOS", "INFRA-SERVICOS", "LIMPEZA", "PORTARIA", "RECEPCAO", "CONTROLE DE ACESSO")),
    ("MANUTENCAO DE REDES - ENGENHARIA", ("MANUTENCAO DE REDES",)),
    ("MANUTENCAO INDUSTRIAL - ENGENHARIA", ("MANUTENCAO INDUSTRIAL",)),
    ("MANUTENCAO PREDIAL - ENGENHARIA", ("MANUTENCAO PREDIAL",)),
    ("CATERING", ("CATERING", "COZINHA")),
)


def build_representativity_summary(
    workforce_path: str | Path,
    respondent_characterizations: list[dict[str, str]],
) -> list[dict[str, Any]]:
    workforce = _load_workforce(workforce_path)
    workforce_local_ghe = Counter((row["local"], row["ghe"]) for row in workforce)
    workforce_local = Counter(row["local"] for row in workforce)
    workforce_ghe = Counter(row["ghe"] for row in workforce)

    respondents = [
        {
            "local": canonical_local(row.get("Q2", "")),
            "ghe": canonical_ghe(row.get("Q1", "")),
        }
        for row in respondent_characterizations
        if row.get("Q1") or row.get("Q2")
    ]
    respondent_local_ghe = Counter((row["local"], row["ghe"]) for row in respondents)
    respondent_local = Counter(row["local"] for row in respondents)
    respondent_ghe = Counter(row["ghe"] for row in respondents)

    rows: list[dict[str, Any]] = []
    for local, ghe in sorted(set(workforce_local_ghe) | set(respondent_local_ghe)):
        rows.append(_summary_row("Local_GHE", local, ghe, workforce_local_ghe[(local, ghe)], respondent_local_ghe[(local, ghe)]))
    for local in sorted(set(workforce_local) | set(respondent_local)):
        rows.append(_summary_row("Local", local, "Todos", workforce_local[local], respondent_local[local]))
    for ghe in sorted(set(workforce_ghe) | set(respondent_ghe)):
        rows.append(_summary_row("GHE", "Todos", ghe, workforce_ghe[ghe], respondent_ghe[ghe]))
    return rows


def canonical_local(value: Any) -> str:
    text = str(value or "").strip().replace("\xa0", " ")
    if not text:
        return "Nao informado"
    if " - " in text:
        city, state = text.rsplit(" - ", 1)
        return f"{_title_ascii(city)} - {normalize_label(state).upper()}"
    return _title_ascii(text)


def canonical_ghe(value: Any) -> str:
    text = str(value or "").replace("\xa0", " ").strip()
    normalized = normalize_label(text)
    if not normalized:
        return "Nao informado"
    for canonical, aliases in GHE_ALIASES:
        if any(alias in normalized for alias in aliases):
            return canonical
    return normalized


def normalize_label(value: str) -> str:
    without_accents = "".join(
        char for char in normalize("NFKD", value) if not combining(char)
    )
    chars = [char if char.isalnum() else " " for char in without_accents.upper()]
    normalized = " ".join("".join(chars).split())
    return normalized


def _load_workforce(path: str | Path) -> list[dict[str, str]]:
    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    worksheet = workbook["SRA 11-06-26"] if "SRA 11-06-26" in workbook.sheetnames else workbook.worksheets[0]
    headers = [str(cell or "").strip() for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    index = {header: position for position, header in enumerate(headers)}
    required = ("Dt Demissao", "Municipio", "Estado", "Negocio")
    missing = [header for header in required if header not in index]
    if missing:
        raise ValueError(f"Colunas ausentes na base SRA: {', '.join(missing)}")

    # Dynamically find the "Situacao" column index if it exists
    situacao_col = None
    for pos, header in enumerate(headers):
        h_lower = header.strip().lower()
        if h_lower in ["situacao", "situação", "status", "sit. cta.", "sit. cta", "sit cta", "situação do contrato", "situação do colaborador"]:
            situacao_col = pos
            break

    rows: list[dict[str, str]] = []
    for values in worksheet.iter_rows(min_row=2, values_only=True):
        if not any(value is not None for value in values):
            continue
        if values[index["Dt Demissao"]]:
            continue
        
        # Filter active situation - accept normal, ativo, ativa, ferias, férias, normal/férias
        if situacao_col is not None and situacao_col < len(values):
            sit_val = str(values[situacao_col] or "").strip().lower()
            if sit_val and sit_val not in ["normal", "ativo", "ativa", "ferias", "férias", "normal/férias"]:
                continue

        city = values[index["Municipio"]]
        state = values[index["Estado"]]
        
        # Check Column V (index 21) first for unit/contract name
        unit_name = values[21] if (len(values) > 21 and values[21]) else None
        local_val = unit_name if unit_name else f"{city} - {state}"
        
        rows.append(
            {
                "local": canonical_local(local_val),
                "ghe": canonical_ghe(values[index["Negocio"]]),
            }
        )
    return rows


def _summary_row(recorte: str, local: str, ghe: str, total: int, respondents: int) -> dict[str, Any]:
    percent = round(respondents / total, 4) if total else 0.0
    return {
        "recorte": recorte,
        "local": local,
        "ghe": ghe,
        "colaboradores_total": total,
        "respondentes_validos": respondents,
        "adesao_percentual": percent,
        "representativo_60": "Sim" if percent >= 0.6 else "Nao",
        "analise_segmentada_permitida": "Sim" if respondents >= 5 else "Nao - agrupar para preservar anonimato",
    }


def _title_ascii(value: str) -> str:
    normalized = normalize_label(value)
    return " ".join(part.capitalize() for part in normalized.split())
