from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from psicossocial.excel_export import TECHNICAL_SHEETS, export_excel
from psicossocial.metodologia import load_metodologia


ROOT_DIR = Path(__file__).resolve().parents[1]
CONFIG_PATH = ROOT_DIR / "config" / "metodologia_v1.yaml"


def test_export_excel_cria_planilha_com_abas_tecnicas(tmp_path: Path) -> None:
    metodologia = load_metodologia(CONFIG_PATH)
    input_path = _create_minimal_input(tmp_path, metodologia)
    output_path = tmp_path / "resultado.xlsx"

    generated = export_excel(input_path, output_path, metodologia)

    assert generated == output_path
    workbook = load_workbook(output_path, read_only=True, data_only=True)
    assert workbook.sheetnames == list(TECHNICAL_SHEETS)
    assert workbook["05_SCORE_DIMENSOES"].max_row == 5
    assert workbook["06_CRUZAMENTO_AMBIENTE"].max_row == 11
    assert workbook["06_CRUZAMENTO_AMBIENTE"]["F4"].value == "Probabilidade_Nivel"
    assert workbook["06_CRUZAMENTO_AMBIENTE"]["M4"].value == "Severidade_Nivel"
    assert workbook["06_CRUZAMENTO_AMBIENTE"]["O4"].value == "Matriz_Resultado"
    assert workbook["06_CRUZAMENTO_AMBIENTE"]["P4"].value == "Matriz_Classificacao"

    complementar = workbook["07_CRUZAMENTO_COMPLEMENTAR"]
    assert complementar.max_row == 46
    assert complementar["A4"].value == "Recorte"
    assert complementar["B4"].value == "Grupo"
    assert complementar["C4"].value == "Dimensao"
    assert complementar["D4"].value == "Respondentes"
    assert complementar["E4"].value == "Score_Medio"
    assert complementar["F4"].value == "Classificacao_COPSOQ"
    assert complementar["G4"].value == "Probabilidade_Nivel"
    assert complementar["M4"].value == "Percentual_Atencao_Risco_GHE"
    assert complementar["P4"].value == "Matriz_Resultado"
    assert complementar["Q4"].value == "Matriz_Classificacao"
    assert complementar["S4"].value == "Plano_Acao?"
    assert complementar["A5"].value == "Q2_Unidade_Contrato"
    assert complementar["B5"].value == "valor Q2"
    assert complementar["C5"].value == "DEM"
    assert complementar["D5"].value == 1
    assert complementar["E5"].value == 2
    assert complementar["F5"].value == "Atencao"
    assert complementar["G5"].value == 2
    assert complementar["Q5"].value == "Alto"
    assert complementar["S5"].value == "Sim"
    assert complementar["A12"].value == "Q1_GHE_Negocio_Atividade"


def _create_minimal_input(tmp_path: Path, metodologia) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    headers = []
    values = []

    for question_code, question in metodologia.perguntas.items():
        if question["tipo"] == "caracterizacao":
            headers.append(question["texto"])
            values.append("Ambiente administrativo (escritorio)" if question_code == "Q6" else f"valor {question_code}")
        elif question["tipo"] == "dimensao_score":
            for item in question["itens"]:
                headers.append(item["texto"])
                values.append("Sempre" if item["direcao"] == "protetivo_inverter" else "Raramente")
        elif question["tipo"] == "vco":
            for item in question["itens"]:
                headers.append(item["texto"])
                values.append("Nao")
        elif question["tipo"] == "qualitativa_estruturada":
            headers.append(question["texto"])
            values.append("Melhor comunicacao.;")

    worksheet.append(["ID", *headers])
    worksheet.append([1, *values])
    input_path = tmp_path / "input.xlsx"
    workbook.save(input_path)
    return input_path
